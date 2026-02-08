"""Tests for skip-embedded behavior when Embedded_Data is invalid."""

from __future__ import annotations

import json
import tempfile
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from qsync import excel_io
from qsync.dimensions import items as items_dimension
from qsync.dimensions import edf as edf_dimension
from qsync.sync_core import preview_changes
from qsync.workbook_resolver import WorkbookResolver


def _payload_with_question_and_embedded() -> dict:
    return {
        "result": {
            "SurveyFlow": {
                "Flow": [
                    {"Type": "Standard", "ID": "BL_1"},
                    {
                        "Type": "EmbeddedData",
                        "FlowID": "FL_1",
                        "EmbeddedData": [
                            {
                                "Field": "DEBUG",
                                "Type": "Custom",
                                "Value": "X",
                                "Description": "DEBUG",
                                "DataVisibility": [],
                                "AnalyzeText": False,
                            }
                        ],
                    },
                ]
            },
            "Blocks": {
                "BL_1": {
                    "Type": "Standard",
                    "ID": "BL_1",
                    "Description": "Block 1",
                    "BlockElements": [{"Type": "Question", "QuestionID": "QID1"}],
                }
            },
            "Questions": {
                "QID1": {
                    "QuestionText": "Old text",
                    "QuestionType": "TE",
                    "Selector": "SL",
                    "SubSelector": "TX",
                    "DataExportTag": "Q1",
                    "QuestionDescription": "Q1",
                    "Configuration": {},
                    "Choices": {},
                    "Answers": {},
                }
            },
        }
    }


def _write_cached_survey(root: Path, survey_id: str, payload: dict) -> None:
    surveys_dir = root / "surveys"
    backups_dir = surveys_dir / "backups"
    surveys_dir.mkdir(parents=True, exist_ok=True)
    backups_dir.mkdir(parents=True, exist_ok=True)
    cached_path = surveys_dir / f"TEST__{survey_id}.json"
    backup_path = backups_dir / f"TEST__{survey_id}.json"
    cached_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    backup_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def _remove_embedded_row(xlsx_path: Path, field: str) -> None:
    wb = load_workbook(xlsx_path)
    ws = wb[excel_io.EMBEDDED_DATA_SHEET]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: i for i, name in enumerate(headers)}
    field_idx = idx["Field"]
    for row in ws.iter_rows(min_row=2, values_only=False):
        if str(row[field_idx].value or "").strip() == field:
            ws.delete_rows(row[0].row, 1)
            break
    wb.save(xlsx_path)


def _update_question_text(xlsx_path: Path, qid: str, text: str) -> None:
    wb = load_workbook(xlsx_path)
    ws = wb[excel_io.QUESTION_SHEET]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: i for i, name in enumerate(headers)}
    qid_idx = idx["QID"]
    text_idx = idx["Text_en_MD"]
    for row in ws.iter_rows(min_row=2, values_only=False):
        if str(row[qid_idx].value or "").strip() == qid:
            row[text_idx].value = text
            break
    wb.save(xlsx_path)


def test_preview_changes_skip_embedded_with_missing_rows() -> None:
    payload = _payload_with_question_and_embedded()
    with tempfile.TemporaryDirectory(prefix="qsync_test_") as tmpdir:
        root = Path(tmpdir)
        survey_id = "SV_TEST"
        _write_cached_survey(root, survey_id, payload)

        excel_dir = root / "excel"
        excel_dir.mkdir(parents=True, exist_ok=True)
        xlsx_path = WorkbookResolver(root=root).resolve(survey_id)
        excel_io.init_workbook_from_survey(survey_id, payload, xlsx_path)

        _remove_embedded_row(xlsx_path, "DEBUG")
        _update_question_text(xlsx_path, "QID1", "Updated text")

        with patch("qsync.qualtrics_client._workspace_root", return_value=root):
            changes = preview_changes(
                survey_id,
                xlsx_path,
                check_drift=False,
                skip_embedded=True,
            )

        assert changes
        assert all(change.kind != "embedded" for change in changes)


def test_stage_builds_payload_when_skip_embedded() -> None:
    payload = _payload_with_question_and_embedded()
    with tempfile.TemporaryDirectory(prefix="qsync_test_") as tmpdir:
        root = Path(tmpdir)
        survey_id = "SV_TEST"
        _write_cached_survey(root, survey_id, payload)

        excel_dir = root / "excel"
        excel_dir.mkdir(parents=True, exist_ok=True)
        xlsx_path = WorkbookResolver(root=root).resolve(survey_id)
        excel_io.init_workbook_from_survey(survey_id, payload, xlsx_path)

        _remove_embedded_row(xlsx_path, "DEBUG")
        _update_question_text(xlsx_path, "QID1", "Updated text")

        with (
            patch("qsync.qualtrics_client._workspace_root", return_value=root),
            patch("qsync.dimensions.items.enforce_no_drift", lambda *a, **k: None),
        ):
            payload = items_dimension._build_pending_payload_from_workbook(
                survey_id,
                xlsx_path,
                scope_expr=None,
                ignore_embedded=True,
                allow_drift=True,
                interactive=False,
            )

        assert payload is not None
        assert list(payload.embedded_fields or []) == []
        assert payload.changes


def test_detect_changes_warns_but_not_items_changes_on_edf_only() -> None:
    payload = _payload_with_question_and_embedded()
    with tempfile.TemporaryDirectory(prefix="qsync_test_") as tmpdir:
        root = Path(tmpdir)
        survey_id = "SV_TEST"
        _write_cached_survey(root, survey_id, payload)

        excel_dir = root / "excel"
        excel_dir.mkdir(parents=True, exist_ok=True)
        xlsx_path = WorkbookResolver(root=root).resolve(survey_id)
        excel_io.init_workbook_from_survey(survey_id, payload, xlsx_path)

        _remove_embedded_row(xlsx_path, "DEBUG")

        with (
            patch("qsync.qualtrics_client._workspace_root", return_value=root),
            patch("qsync.workbook_resolver.resolve_root", lambda required=False: root),
        ):
            result = items_dimension.detect_changes(survey_id)
            edf_result = edf_dimension.detect_changes(survey_id)

        assert result.has_changes is False
        assert not result.warning_detail
        assert edf_result.warning_detail
