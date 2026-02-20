"""Test that Trash block questions are properly filtered from all Excel sheets."""

from openpyxl import load_workbook

from qsync.excel_io import (
    OPTIONS_SHEET,
    QUESTION_SHEET,
    SBS_COLUMN_ANSWERS_SHEET,
    SBS_COLUMNS_SHEET,
    SUBITEMS_SHEET,
    SYSTEM_SHEET,
    build_option_rows,
    build_question_rows,
    build_subitem_rows,
    init_workbook_from_survey,
)


def test_trash_questions_excluded_from_all_sheets():
    """Verify that questions in Trash blocks are excluded from Questions, Options, and Subitems sheets."""

    # Create a minimal survey payload with both Standard and Trash blocks
    survey_payload = {
        "result": {
            "SurveyID": "SV_TEST",
            "Blocks": {
                "BL_STANDARD": {
                    "Type": "Standard",
                    "Description": "Main Questions",
                    "BlockElements": [
                        {"Type": "Question", "QuestionID": "QID1"},
                        {"Type": "Question", "QuestionID": "QID2"},
                    ],
                },
                "BL_TRASH": {
                    "Type": "Trash",
                    "Description": "Trash / Unused Questions",
                    "BlockElements": [
                        {"Type": "Question", "QuestionID": "QID_TRASH_1"},
                        {"Type": "Question", "QuestionID": "QID_TRASH_2"},
                    ],
                },
            },
            "Questions": {
                "QID1": {
                    "QuestionType": "MC",
                    "DataExportTag": "Q1",
                    "QuestionText": "Active Question 1",
                    "Choices": {
                        "1": {"Display": "Choice 1"},
                        "2": {"Display": "Choice 2"},
                    },
                },
                "QID2": {
                    "QuestionType": "Matrix",
                    "DataExportTag": "Q2",
                    "QuestionText": "Active Matrix Question",
                    "Choices": {
                        "1": {"Display": "Row 1"},
                        "2": {"Display": "Row 2"},
                    },
                    "Answers": {
                        "1": {"Display": "Strongly Disagree"},
                        "2": {"Display": "Disagree"},
                        "3": {"Display": "Agree"},
                    },
                },
                "QID_TRASH_1": {
                    "QuestionType": "MC",
                    "DataExportTag": "TRASH_Q1",
                    "QuestionText": "Trashed Question 1",
                    "Choices": {
                        "1": {"Display": "Trash Choice 1"},
                        "2": {"Display": "Trash Choice 2"},
                        "3": {"Display": "Trash Choice 3"},
                    },
                },
                "QID_TRASH_2": {
                    "QuestionType": "Matrix",
                    "DataExportTag": "TRASH_Q2",
                    "QuestionText": "Trashed Matrix Question",
                    "Choices": {
                        "1": {"Display": "Trash Row 1"},
                    },
                    "Answers": {
                        "1": {"Display": "Trash Answer 1"},
                        "2": {"Display": "Trash Answer 2"},
                    },
                },
            },
            "SurveyFlow": {
                "Flow": [
                    {"ID": "BL_STANDARD", "Type": "Block"},
                ],
            },
        }
    }

    # Build all sheets
    question_rows = build_question_rows("SV_TEST", survey_payload)
    option_rows = build_option_rows("SV_TEST", survey_payload)
    subitem_rows = build_subitem_rows("SV_TEST", survey_payload)

    # Verify Questions sheet includes only non-Trash questions
    assert "QID1" in question_rows, "Active question QID1 should be in Questions sheet"
    assert "QID2" in question_rows, "Active question QID2 should be in Questions sheet"
    assert (
        "QID_TRASH_1" not in question_rows
    ), "Trash question QID_TRASH_1 should NOT be in Questions sheet"
    assert (
        "QID_TRASH_2" not in question_rows
    ), "Trash question QID_TRASH_2 should NOT be in Questions sheet"

    # Verify Options sheet includes only options for non-Trash questions
    option_qids = {qid for qid, _ in option_rows.keys()}

    assert "QID1" in option_qids, "Options for QID1 should be present"
    assert "QID2" in option_qids, "Options for QID2 (Matrix answers) should be present"
    assert (
        "QID_TRASH_1" not in option_qids
    ), "Options for QID_TRASH_1 should NOT be present"
    assert (
        "QID_TRASH_2" not in option_qids
    ), "Options for QID_TRASH_2 should NOT be present"

    # Count options for active questions
    qid1_options = [key for key in option_rows.keys() if key[0] == "QID1"]
    qid2_options = [key for key in option_rows.keys() if key[0] == "QID2"]

    assert len(qid1_options) == 2, "QID1 should have 2 choices"
    assert len(qid2_options) == 3, "QID2 Matrix should have 3 answers"

    # Verify Subitems sheet includes only subitems for non-Trash questions
    subitem_qids = {qid for qid, _, _ in subitem_rows.keys()}

    assert "QID2" in subitem_qids, "Subitems for QID2 (Matrix rows) should be present"
    assert (
        "QID_TRASH_2" not in subitem_qids
    ), "Subitems for QID_TRASH_2 should NOT be present"

    # Count subitems for active Matrix question
    qid2_subitems = [key for key in subitem_rows.keys() if key[0] == "QID2"]
    assert len(qid2_subitems) == 2, "QID2 Matrix should have 2 subitems (rows)"

    # Verify total counts
    assert len(question_rows) == 2, "Should have exactly 2 questions (not 4)"

    # Options: QID1 has 2 choices + QID2 has 3 answers = 5 total
    assert (
        len(option_rows) == 5
    ), "Should have 5 option rows (not including trash questions)"

    # Subitems: QID2 has 2 matrix rows
    assert (
        len(subitem_rows) == 2
    ), "Should have 2 subitem rows (not including trash questions)"


def test_trash_filtering_with_no_trash_blocks():
    """Verify that filtering works correctly when there are no Trash blocks."""

    survey_payload = {
        "result": {
            "SurveyID": "SV_TEST2",
            "Blocks": {
                "BL_STANDARD": {
                    "Type": "Standard",
                    "Description": "Main Questions",
                    "BlockElements": [
                        {"Type": "Question", "QuestionID": "QID1"},
                    ],
                },
            },
            "Questions": {
                "QID1": {
                    "QuestionType": "MC",
                    "DataExportTag": "Q1",
                    "QuestionText": "Question 1",
                    "Choices": {
                        "1": {"Display": "Choice 1"},
                    },
                },
            },
            "SurveyFlow": {
                "Flow": [
                    {"ID": "BL_STANDARD", "Type": "Block"},
                ],
            },
        }
    }

    question_rows = build_question_rows("SV_TEST2", survey_payload)
    option_rows = build_option_rows("SV_TEST2", survey_payload)

    assert len(question_rows) == 1, "Should have 1 question"
    assert len(option_rows) == 1, "Should have 1 option"
    assert "QID1" in question_rows


def test_trash_filtering_empty_survey():
    """Verify that filtering works correctly with an empty survey."""

    survey_payload = {
        "result": {
            "SurveyID": "SV_EMPTY",
            "Blocks": {},
            "Questions": {},
            "SurveyFlow": {"Flow": []},
        }
    }

    question_rows = build_question_rows("SV_EMPTY", survey_payload)
    option_rows = build_option_rows("SV_EMPTY", survey_payload)
    subitem_rows = build_subitem_rows("SV_EMPTY", survey_payload)

    assert len(question_rows) == 0
    assert len(option_rows) == 0
    assert len(subitem_rows) == 0


def test_referential_integrity():
    """Verify that all QIDs in Options and Subitems sheets exist in Questions sheet."""

    survey_payload = {
        "result": {
            "SurveyID": "SV_INTEGRITY",
            "Blocks": {
                "BL_1": {
                    "Type": "Standard",
                    "Description": "Block 1",
                    "BlockElements": [
                        {"Type": "Question", "QuestionID": "QID1"},
                        {"Type": "Question", "QuestionID": "QID2"},
                    ],
                },
                "BL_TRASH": {
                    "Type": "Trash",
                    "Description": "Trash",
                    "BlockElements": [
                        {"Type": "Question", "QuestionID": "QID_TRASH"},
                    ],
                },
            },
            "Questions": {
                "QID1": {
                    "QuestionType": "MC",
                    "QuestionText": "Q1",
                    "Choices": {"1": {"Display": "C1"}},
                },
                "QID2": {
                    "QuestionType": "Matrix",
                    "QuestionText": "Q2",
                    "Choices": {"1": {"Display": "R1"}},
                    "Answers": {"1": {"Display": "A1"}},
                },
                "QID_TRASH": {
                    "QuestionType": "MC",
                    "QuestionText": "Trash",
                    "Choices": {"1": {"Display": "TC1"}, "2": {"Display": "TC2"}},
                },
            },
            "SurveyFlow": {"Flow": [{"ID": "BL_1", "Type": "Block"}]},
        }
    }

    question_rows = build_question_rows("SV_INTEGRITY", survey_payload)
    option_rows = build_option_rows("SV_INTEGRITY", survey_payload)
    subitem_rows = build_subitem_rows("SV_INTEGRITY", survey_payload)

    question_qids = set(question_rows.keys())
    option_qids = {qid for qid, _ in option_rows.keys()}
    subitem_qids = {qid for qid, _, _ in subitem_rows.keys()}

    # All QIDs in Options must exist in Questions
    for qid in option_qids:
        assert qid in question_qids, f"Option QID {qid} not found in Questions sheet"

    # All QIDs in Subitems must exist in Questions
    for qid in subitem_qids:
        assert qid in question_qids, f"Subitem QID {qid} not found in Questions sheet"

    # Verify trash question is not in any sheet
    assert "QID_TRASH" not in question_qids
    assert "QID_TRASH" not in option_qids
    assert "QID_TRASH" not in subitem_qids


def test_questions_and_subitems_are_flow_scoped_and_flow_ordered():
    """Questions/Subitems should include only in-flow QIDs and follow SurveyFlow order."""

    survey_payload = {
        "result": {
            "SurveyID": "SV_FLOW_SCOPE",
            "Blocks": {
                "BL_FLOW_A": {
                    "Type": "Standard",
                    "Description": "Flow A",
                    "BlockElements": [
                        {"Type": "Question", "QuestionID": "QID1"},
                        {"Type": "Question", "QuestionID": "QID3"},
                    ],
                },
                "BL_FLOW_B": {
                    "Type": "Standard",
                    "Description": "Flow B",
                    "BlockElements": [
                        {"Type": "Question", "QuestionID": "QID2"},
                    ],
                },
                "BL_OUT_OF_FLOW": {
                    "Type": "Standard",
                    "Description": "Out of flow",
                    "BlockElements": [
                        {"Type": "Question", "QuestionID": "QID_OUT"},
                    ],
                },
            },
            # Keep dict order intentionally different from SurveyFlow order.
            "Questions": {
                "QID2": {
                    "QuestionType": "Matrix",
                    "DataExportTag": "Q2",
                    "QuestionText": "Flow question 2",
                    "Choices": {"1": {"Display": "Row 2"}},
                    "Answers": {"1": {"Display": "Agree"}},
                },
                "QID_OUT": {
                    "QuestionType": "Matrix",
                    "DataExportTag": "Q_OUT",
                    "QuestionText": "Out-of-flow question",
                    "Choices": {"1": {"Display": "Out row"}},
                    "Answers": {"1": {"Display": "Out answer"}},
                },
                "QID3": {
                    "QuestionType": "Matrix",
                    "DataExportTag": "Q3",
                    "QuestionText": "Flow question 3",
                    "Choices": {"1": {"Display": "Row 3"}},
                    "Answers": {"1": {"Display": "Neutral"}},
                },
                "QID1": {
                    "QuestionType": "Matrix",
                    "DataExportTag": "Q1",
                    "QuestionText": "Flow question 1",
                    "ChoiceOrder": ["2", "1"],
                    "Choices": {
                        "1": {"Display": "Row 1"},
                        "2": {"Display": "Row 1B"},
                    },
                    "Answers": {"1": {"Display": "Disagree"}},
                },
            },
            "SurveyFlow": {
                "Flow": [
                    {"Type": "Block", "ID": "BL_FLOW_A"},
                    {"Type": "Block", "ID": "BL_FLOW_B"},
                ],
            },
        }
    }

    question_rows = build_question_rows("SV_FLOW_SCOPE", survey_payload)
    option_rows = build_option_rows("SV_FLOW_SCOPE", survey_payload)
    subitem_rows = build_subitem_rows("SV_FLOW_SCOPE", survey_payload)

    assert list(question_rows.keys()) == ["QID1", "QID3", "QID2"]
    assert "QID_OUT" not in question_rows

    option_qids = {qid for qid, _ in option_rows.keys()}
    assert option_qids == {"QID1", "QID3", "QID2"}

    subitem_keys = list(subitem_rows.keys())
    assert [qid for qid, _field, _aid in subitem_keys] == ["QID1", "QID1", "QID3", "QID2"]
    assert subitem_keys[0] == ("QID1", "Answer", "2")
    assert subitem_keys[1] == ("QID1", "Answer", "1")
    assert "QID_OUT" not in {qid for qid, _field, _aid in subitem_keys}


def test_workbook_adds_flow_metadata_columns_on_item_sheets(tmp_path):
    survey_payload = {
        "result": {
            "SurveyID": "SV_FLOW_COLS",
            "Blocks": {
                "BL_A": {
                    "Type": "Standard",
                    "Description": "Block A",
                    "BlockElements": [{"Type": "Question", "QuestionID": "QID2"}],
                },
            },
            "Questions": {
                "QID2": {
                    "QuestionType": "MC",
                    "DataExportTag": "Q2",
                    "QuestionText": "Question 2",
                    "Choices": {"1": {"Display": "Yes"}},
                }
            },
            "SurveyFlow": {"Flow": [{"Type": "Block", "ID": "BL_A"}]},
        }
    }

    xlsx_path = tmp_path / "SV_FLOW_COLS.xlsx"
    init_workbook_from_survey("SV_FLOW_COLS", survey_payload, xlsx_path)
    wb = load_workbook(xlsx_path)

    for sheet_name in (
        QUESTION_SHEET,
        OPTIONS_SHEET,
        SUBITEMS_SHEET,
        SBS_COLUMNS_SHEET,
        SBS_COLUMN_ANSWERS_SHEET,
    ):
        headers = [cell.value for cell in next(wb[sheet_name].iter_rows(max_row=1))]
        assert "BlockID" in headers
        assert "BlockOrder" in headers
        assert "QuestionOrder" in headers
        assert "QuestionOrderInBlock" in headers


def test_workbook_orders_by_flow_question_order_not_lexical_qid(tmp_path):
    survey_payload = {
        "result": {
            "SurveyID": "SV_FLOW_SORT",
            "Blocks": {
                "BL_A": {
                    "Type": "Standard",
                    "Description": "Block A",
                    "BlockElements": [
                        {"Type": "Question", "QuestionID": "QID2"},
                        {"Type": "Question", "QuestionID": "QID100"},
                    ],
                },
            },
            "Questions": {
                "QID2": {
                    "QuestionType": "MC",
                    "DataExportTag": "Q2",
                    "QuestionText": "Question 2",
                    "Choices": {"1": {"Display": "Yes"}},
                },
                "QID100": {
                    "QuestionType": "MC",
                    "DataExportTag": "Q100",
                    "QuestionText": "Question 100",
                    "Choices": {"1": {"Display": "No"}},
                },
            },
            "SurveyFlow": {"Flow": [{"Type": "Block", "ID": "BL_A"}]},
        }
    }

    xlsx_path = tmp_path / "SV_FLOW_SORT.xlsx"
    init_workbook_from_survey("SV_FLOW_SORT", survey_payload, xlsx_path)
    wb = load_workbook(xlsx_path)

    q_ws = wb[QUESTION_SHEET]
    q_headers = [cell.value for cell in next(q_ws.iter_rows(max_row=1))]
    qid_idx = q_headers.index("QID") + 1
    q_qids = [
        str(q_ws.cell(row=row, column=qid_idx).value or "").strip()
        for row in range(2, q_ws.max_row + 1)
        if str(q_ws.cell(row=row, column=qid_idx).value or "").strip()
    ]
    assert q_qids[:2] == ["QID2", "QID100"]

    o_ws = wb[OPTIONS_SHEET]
    o_headers = [cell.value for cell in next(o_ws.iter_rows(max_row=1))]
    o_qid_idx = o_headers.index("QID") + 1
    o_qids = [
        str(o_ws.cell(row=row, column=o_qid_idx).value or "").strip()
        for row in range(2, o_ws.max_row + 1)
        if str(o_ws.cell(row=row, column=o_qid_idx).value or "").strip()
    ]
    assert o_qids[:2] == ["QID2", "QID100"]


def _system_routing_payload() -> dict:
    return {
        "result": {
            "SurveyID": "SV_SYSTEM",
            "Blocks": {
                "BL_MAIN": {
                    "Type": "Standard",
                    "Description": "Main Block",
                    "BlockElements": [
                        {"Type": "Question", "QuestionID": "QID1"},
                        {"Type": "Question", "QuestionID": "QID31"},
                        {"Type": "Question", "QuestionID": "QID44"},
                        {"Type": "Question", "QuestionID": "QID55"},
                    ],
                },
            },
            "Questions": {
                "QID1": {
                    "QuestionType": "Matrix",
                    "DataExportTag": "Q1",
                    "QuestionText": "Main question",
                    "Choices": {
                        "1": {"Display": "Row 1"},
                        "2": {"Display": "Row 2"},
                    },
                    "Answers": {
                        "1": {"Display": "Yes"},
                        "2": {"Display": "No"},
                    },
                },
                "QID31": {
                    "QuestionType": "Meta",
                    "DataExportTag": "MetaInfo",
                    "QuestionText": "Meta block",
                    "Choices": {"1": {"Display": "Meta choice"}},
                },
                "QID44": {
                    "QuestionType": "CAPTCHA",
                    "DataExportTag": "CaptchaInfo",
                    "QuestionText": "Captcha block",
                    "Choices": {"1": {"Display": "Captcha prompt"}},
                },
                "QID55": {
                    "QuestionType": "Timing",
                    "DataExportTag": "Q_Time",
                    "QuestionText": "Timing block",
                    "Choices": {"1": {"Display": "Page Submit"}},
                },
            },
            "SurveyFlow": {"Flow": [{"Type": "Block", "ID": "BL_MAIN"}]},
        }
    }


def _sheet_qids(wb, sheet_name: str) -> list[str]:
    ws = wb[sheet_name]
    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    qid_idx = headers.index("QID") + 1
    return [
        str(ws.cell(row=row, column=qid_idx).value or "").strip()
        for row in range(2, ws.max_row + 1)
        if str(ws.cell(row=row, column=qid_idx).value or "").strip()
    ]


def test_system_question_types_are_routed_out_of_item_sheets(tmp_path):
    payload = _system_routing_payload()

    question_rows = build_question_rows("SV_SYSTEM", payload)
    option_rows = build_option_rows("SV_SYSTEM", payload)
    subitem_rows = build_subitem_rows("SV_SYSTEM", payload)

    assert set(question_rows.keys()) == {"QID1"}
    assert {qid for qid, _ in option_rows.keys()} == {"QID1"}
    assert {qid for qid, _field, _answer in subitem_rows.keys()} == {"QID1"}

    xlsx_path = tmp_path / "SV_SYSTEM.xlsx"
    init_workbook_from_survey("SV_SYSTEM", payload, xlsx_path)
    wb = load_workbook(xlsx_path)

    assert set(_sheet_qids(wb, QUESTION_SHEET)) == {"QID1"}
    assert set(_sheet_qids(wb, OPTIONS_SHEET)) == {"QID1"}
    assert set(_sheet_qids(wb, SUBITEMS_SHEET)) == {"QID1"}

    system_qids = set(_sheet_qids(wb, SYSTEM_SHEET))
    assert {"QID31", "QID44", "QID55"}.issubset(system_qids)
    assert "QID1" not in system_qids


def test_init_workbook_prunes_legacy_system_rows_from_item_sheets(tmp_path):
    payload = _system_routing_payload()
    xlsx_path = tmp_path / "SV_SYSTEM_LEGACY.xlsx"
    init_workbook_from_survey("SV_SYSTEM", payload, xlsx_path)

    wb = load_workbook(xlsx_path)

    q_ws = wb[QUESTION_SHEET]
    q_headers = [cell.value for cell in next(q_ws.iter_rows(max_row=1))]
    q_row = [None] * len(q_headers)
    q_row[q_headers.index("QID")] = "QID44"
    q_row[q_headers.index("QuestionType")] = "CAPTCHA"
    q_row[q_headers.index("DataExportTag")] = "CaptchaInfo"
    q_ws.append(q_row)

    o_ws = wb[OPTIONS_SHEET]
    o_headers = [cell.value for cell in next(o_ws.iter_rows(max_row=1))]
    o_row = [None] * len(o_headers)
    o_row[o_headers.index("QID")] = "QID44"
    o_row[o_headers.index("ChoiceId")] = "1"
    o_row[o_headers.index("QuestionType")] = "CAPTCHA"
    o_row[o_headers.index("ExportTag")] = "CaptchaInfo"
    o_ws.append(o_row)

    s_ws = wb[SUBITEMS_SHEET]
    s_headers = [cell.value for cell in next(s_ws.iter_rows(max_row=1))]
    s_row = [None] * len(s_headers)
    s_row[s_headers.index("QID")] = "QID44"
    s_row[s_headers.index("AnswerId")] = "1"
    s_row[s_headers.index("Field")] = "Answer"
    s_row[s_headers.index("QuestionType")] = "CAPTCHA"
    s_row[s_headers.index("ExportTag")] = "CaptchaInfo"
    s_ws.append(s_row)

    wb.save(xlsx_path)

    init_workbook_from_survey("SV_SYSTEM", payload, xlsx_path)
    wb = load_workbook(xlsx_path)

    assert "QID44" not in _sheet_qids(wb, QUESTION_SHEET)
    assert "QID44" not in _sheet_qids(wb, OPTIONS_SHEET)
    assert "QID44" not in _sheet_qids(wb, SUBITEMS_SHEET)
