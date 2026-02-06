from __future__ import annotations

import io
import json
from contextlib import redirect_stdout
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from qsync import excel_io
from qsync.sync_orchestrator import (
    DimensionSyncResult,
    _detect_unstaged_changes,
    _display_survey_overview,
    _sync_dimensions_once,
)
from qsync.workbook_resolver import WorkbookResolver


def _survey_payload() -> dict:
    return {
        "result": {
            "SurveyID": "SV_TEST",
            "SurveyOptions": {
                "SurveyLanguage": "EN",
                "AvailableLanguages": {"EN": True, "FR": True},
            },
            "SurveyFlow": {
                "Flow": [
                    {"Type": "Standard", "ID": "BL_1"},
                    {
                        "Type": "EmbeddedData",
                        "FlowID": "FL_1",
                        "EmbeddedData": [
                            {
                                "Field": "DEBUG",
                                "Type": "Custom",
                                "Value": "F",
                                "Description": "DEBUG",
                                "DataVisibility": [],
                                "AnalyzeText": False,
                            }
                        ],
                    },
                ]
            },
            "Blocks": {
                "BL_1": {
                    "Type": "Standard",
                    "ID": "BL_1",
                    "Description": "Block 1",
                    "BlockElements": [{"Type": "Question", "QuestionID": "QID1"}],
                }
            },
            "Questions": {
                "QID1": {
                    "QuestionType": "TE",
                    "Selector": "SL",
                    "SubSelector": "TX",
                    "QuestionText": "Old text",
                    "DataExportTag": "Q1",
                    "QuestionDescription": "Q1",
                    "Configuration": {},
                    "Choices": {},
                    "Answers": {},
                    "Language": {"FR": {"QuestionText": "Ancien texte"}},
                }
            },
        }
    }


def _write_cached_survey(root: Path, survey_id: str, payload: dict) -> None:
    surveys_dir = root / "surveys"
    backups_dir = surveys_dir / "backups"
    surveys_dir.mkdir(parents=True, exist_ok=True)
    backups_dir.mkdir(parents=True, exist_ok=True)
    (surveys_dir / f"TEST__{survey_id}.json").write_text(
        json.dumps(payload, indent=2), encoding="utf-8"
    )
    (backups_dir / f"TEST__{survey_id}.json").write_text(
        json.dumps(payload, indent=2), encoding="utf-8"
    )


def _remove_embedded_row(xlsx_path: Path, field: str) -> None:
    wb = load_workbook(xlsx_path)
    ws = wb[excel_io.EMBEDDED_DATA_SHEET]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: i for i, name in enumerate(headers)}
    field_idx = idx["Field"]
    for row in ws.iter_rows(min_row=2, values_only=False):
        if str(row[field_idx].value or "").strip() == field:
            ws.delete_rows(row[0].row, 1)
            break
    wb.save(xlsx_path)


def _update_question_text(
    xlsx_path: Path, qid: str, text_en: str, text_fr: str
) -> None:
    wb = load_workbook(xlsx_path)
    ws = wb[excel_io.QUESTION_SHEET]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: i for i, name in enumerate(headers)}
    qid_idx = idx["QID"]
    en_idx = idx["Text_en_MD"]
    fr_idx = idx["Text_fr_MD"]
    for row in ws.iter_rows(min_row=2, values_only=False):
        if str(row[qid_idx].value or "").strip() == qid:
            row[en_idx].value = text_en
            row[fr_idx].value = text_fr
            break
    wb.save(xlsx_path)


def test_edf_invalid_still_allows_items_and_translations_sync_flow(
    tmp_path: Path, monkeypatch
) -> None:
    survey_id = "SV_TEST"
    payload = _survey_payload()
    _write_cached_survey(tmp_path, survey_id, payload)

    monkeypatch.setenv("QSYNC_ROOT", str(tmp_path))

    xlsx_path = WorkbookResolver(root=tmp_path).resolve(survey_id)
    excel_io.init_workbook_from_survey(survey_id, payload, xlsx_path, languages=["FR"])
    _remove_embedded_row(xlsx_path, "DEBUG")
    _update_question_text(xlsx_path, "QID1", "Updated text", "Texte FR mis a jour")

    with patch("qsync.qualtrics_client._workspace_root", return_value=tmp_path):
        unstaged = _detect_unstaged_changes(survey_id)

    assert unstaged["items"].has_changes is True
    assert unstaged["translations"].has_changes is True
    assert unstaged["edf"].has_changes is False
    assert unstaged["edf"].warning_detail
    assert "repair-edf" in unstaged["edf"].warning_detail

    buf = io.StringIO()
    with redirect_stdout(buf):
        _display_survey_overview(
            survey_id,
            f"Test Survey ({survey_id})",
            staged={
                "items": "none",
                "edf": "none",
                "js": "none",
                "translations": "none",
                "eos": "none",
            },
            unstaged=unstaged,
            has_pending=False,
        )
    overview = buf.getvalue()
    assert "Repair workbook issues only" in overview
    assert f"qsync items repair-edf --survey-id {survey_id}" in overview

    sync_result = DimensionSyncResult(
        dimension="items", success=True, applied_changes=True
    )
    with (
        patch("qsync.qualtrics_client._workspace_root", return_value=tmp_path),
        patch("qsync.sync_orchestrator.detect_conflicts", return_value=[]),
        patch("qsync.sync_orchestrator._display_push_report"),
        patch("qsync.sync_orchestrator._orchestrated_publish"),
        patch(
            "qsync.sync_orchestrator.sync_dimension", return_value=sync_result
        ) as mock_sync,
    ):
        summary = _sync_dimensions_once(
            survey_id=survey_id,
            dimensions=["items", "translations"],
            interactive=False,
            force_live=False,
            force_preview=False,
            auto_yes=True,
            allow_drift=False,
            skip_publish=True,
            scope=None,
            per_dimension=False,
            allow_skip_embedded=True,
        )

    assert summary is not None
    pushed_dims = [call.args[1] for call in mock_sync.call_args_list]
    assert pushed_dims == ["items", "translations"]
    assert mock_sync.call_args_list[0].kwargs["ignore_embedded"] is True
    assert mock_sync.call_args_list[1].kwargs["ignore_embedded"] is False
