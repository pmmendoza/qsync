from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from qsync import excel_io
from qsync.dimensions import items as items_dimension
from qsync.sync_orchestrator import _autofix_command
from qsync.workbook_resolver import WorkbookResolver


def _payload_with_single_mc() -> dict:
    return {
        "result": {
            "SurveyID": "SV_TEST",
            "SurveyOptions": {"SurveyLanguage": "EN", "AvailableLanguages": ["EN"]},
            "SurveyFlow": {"Flow": [{"Type": "Standard", "ID": "BL_1"}]},
            "Blocks": {
                "BL_1": {
                    "Type": "Standard",
                    "ID": "BL_1",
                    "Description": "Block 1",
                    "BlockElements": [{"Type": "Question", "QuestionID": "QID1"}],
                }
            },
            "Questions": {
                "QID1": {
                    "QuestionText": "Pick one",
                    "QuestionType": "MC",
                    "Selector": "SAVR",
                    "SubSelector": "TX",
                    "DataExportTag": "Q1",
                    "QuestionDescription": "Q1",
                    "Configuration": {},
                    "Choices": {
                        "1": {"Display": "Yes", "Recode": "1"},
                        "2": {"Display": "No", "Recode": "2"},
                    },
                    "ChoiceOrder": ["1", "2"],
                    "Answers": {},
                }
            },
        }
    }


def _write_cached_survey(root: Path, survey_id: str, payload: dict) -> None:
    surveys_dir = root / "surveys"
    backups_dir = surveys_dir / "backups"
    surveys_dir.mkdir(parents=True, exist_ok=True)
    backups_dir.mkdir(parents=True, exist_ok=True)
    data = json.dumps(payload, indent=2)
    (surveys_dir / f"TEST__{survey_id}.json").write_text(data, encoding="utf-8")
    (backups_dir / f"TEST__{survey_id}.json").write_text(data, encoding="utf-8")


def _append_orphan_rows(xlsx_path: Path) -> None:
    wb = load_workbook(xlsx_path)

    ws_q = wb[excel_io.QUESTION_SHEET]
    q_headers = [cell.value for cell in next(ws_q.iter_rows(min_row=1, max_row=1))]
    q_idx = {name: i for i, name in enumerate(q_headers)}
    row_q = [None] * len(q_headers)
    row_q[q_idx["SurveyID"]] = "SV_TEST"
    row_q[q_idx["QID"]] = "QID999"
    row_q[q_idx["QuestionType"]] = "MC"
    row_q[q_idx["DataExportTag"]] = "orphan_q"
    ws_q.append(row_q)

    ws_o = wb[excel_io.OPTIONS_SHEET]
    o_headers = [cell.value for cell in next(ws_o.iter_rows(min_row=1, max_row=1))]
    o_idx = {name: i for i, name in enumerate(o_headers)}
    row_o = [None] * len(o_headers)
    row_o[o_idx["SurveyID"]] = "SV_TEST"
    row_o[o_idx["QID"]] = "QID999"
    row_o[o_idx["ChoiceId"]] = "1"
    row_o[o_idx["QuestionType"]] = "MC"
    row_o[o_idx["ExportTag"]] = "orphan_q"
    ws_o.append(row_o)

    wb.save(xlsx_path)


def test_init_with_prune_orphans_removes_stale_rows(tmp_path: Path) -> None:
    payload = _payload_with_single_mc()
    survey_id = "SV_TEST"
    xlsx_path = WorkbookResolver(root=tmp_path).resolve(survey_id)

    excel_io.init_workbook_from_survey(survey_id, payload, xlsx_path)
    _append_orphan_rows(xlsx_path)

    before = excel_io.inspect_workbook_orphan_rows(survey_id, payload, xlsx_path)
    assert before.total_rows == 2

    report = excel_io.init_workbook_from_survey(
        survey_id,
        payload,
        xlsx_path,
        prune_orphans=True,
    )
    assert report is not None
    assert report.total_rows == 2
    assert report.counts_by_sheet() == {"Options": 1, "Questions": 1}

    after = excel_io.inspect_workbook_orphan_rows(survey_id, payload, xlsx_path)
    assert after.total_rows == 0


def test_items_detect_changes_reports_fixable_orphan_warning(
    tmp_path: Path, monkeypatch
) -> None:
    payload = _payload_with_single_mc()
    survey_id = "SV_TEST"
    _write_cached_survey(tmp_path, survey_id, payload)
    xlsx_path = WorkbookResolver(root=tmp_path).resolve(survey_id)
    excel_io.init_workbook_from_survey(survey_id, payload, xlsx_path)
    _append_orphan_rows(xlsx_path)

    monkeypatch.setenv("QSYNC_ROOT", str(tmp_path))
    result = items_dimension.detect_changes(survey_id)

    assert result.has_changes is False
    assert result.warning_detail is not None
    assert "orphan item rows" in result.warning_detail.lower()
    assert "--prune-orphans" in result.warning_detail
    assert result.safe_to_autofix is True


def test_items_autofix_command_uses_prune_orphans() -> None:
    command = _autofix_command("items", "SV_TEST")
    assert command == "qsync items pull --survey-id SV_TEST --prune-orphans"
