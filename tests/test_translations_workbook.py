from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from qsync.excel_io import (
    OPTIONS_SHEET,
    QUESTION_SHEET,
    SUBITEMS_SHEET,
    TRANSLATION_KEY_SHEET,
    init_workbook_from_survey,
)
from qsync.translations import run_translation_doctor


def _write_cached_survey(tmp_path: Path, survey_id: str, payload: dict) -> Path:
    surveys_dir = tmp_path / "surveys"
    surveys_dir.mkdir(parents=True, exist_ok=True)
    path = surveys_dir / f"Test__{survey_id}.json"
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    return path


def _survey_payload() -> dict:
    return {
        "result": {
            "SurveyFlow": {"Flow": [{"Type": "Standard", "ID": "BL_1"}]},
            "Blocks": {
                "BL_1": {
                    "Type": "Default",
                    "Description": "Block 1",
                    "BlockElements": [
                        {"Type": "Question", "QuestionID": "QID1"},
                    ],
                }
            },
            "Questions": {
                "QID1": {
                    "QuestionType": "Matrix",
                    "Selector": "Likert",
                    "QuestionText": "<p>Hello</p>",
                    "DataExportTag": "q1",
                    "Choices": {
                        "1": {"Display": "Row 1"},
                        "2": {"Display": "Row 2"},
                    },
                    "ChoiceOrder": ["1", "2"],
                    "Answers": {
                        "1": {"Display": "A1"},
                        "2": {"Display": "A2"},
                    },
                    "AnswerOrder": ["1", "2"],
                }
            },
        }
    }


def _survey_payload_with_labels() -> dict:
    return {
        "result": {
            "SurveyID": "SV_TEST",
            "SurveyOptions": {
                "SurveyLanguage": "EN",
                "AvailableLanguages": ["EN", "FR"],
            },
            "SurveyFlow": {"Flow": [{"Type": "Standard", "ID": "BL_1"}]},
            "Blocks": {
                "BL_1": {
                    "Type": "Default",
                    "Description": "Block 1",
                    "BlockElements": [
                        {"Type": "Question", "QuestionID": "QID1"},
                    ],
                }
            },
            "Questions": {
                "QID1": {
                    "QuestionType": "Slider",
                    "Selector": "SL",
                    "QuestionText": "<p>Hello</p>",
                    "DataExportTag": "q1",
                    "Answers": {
                        "1": {"Display": "A1"},
                        "2": {"Display": "A2"},
                    },
                    "Labels": {
                        "1": {"Display": "Low"},
                        "2": {"Display": "High"},
                    },
                    "Language": {
                        "FR": {
                            "Labels": {
                                "1": {"Display": "Bas"},
                                "2": {"Display": "Haut"},
                            }
                        }
                    },
                }
            },
        }
    }


def _survey_payload_with_translation_blocks() -> dict:
    return {
        "result": {
            "SurveyID": "SV_TEST",
            "SurveyOptions": {
                "SurveyLanguage": "EN",
                "AvailableLanguages": ["EN", "FR"],
            },
            "SurveyFlow": {"Flow": [{"Type": "Standard", "ID": "BL_1"}]},
            "Blocks": {
                "BL_1": {
                    "Type": "Default",
                    "Description": "Block 1",
                    "BlockElements": [
                        {"Type": "Question", "QuestionID": "QID1"},
                        {"Type": "Question", "QuestionID": "QID2"},
                    ],
                }
            },
            "Questions": {
                "QID1": {
                    "QuestionType": "MC",
                    "Selector": "SAVR",
                    "QuestionText": "Base Q1",
                    "DataExportTag": "q1",
                    "Choices": {
                        "1": {"Display": "Yes"},
                        "2": {"Display": "No"},
                    },
                    "ChoiceOrder": ["1", "2"],
                    "Language": {
                        "FR": {
                            "QuestionText": "FR Q1",
                            "Choices": {
                                "1": {"Display": "Oui"},
                                "2": {"Display": "Non"},
                            },
                        }
                    },
                },
                "QID2": {
                    "QuestionType": "Matrix",
                    "Selector": "Likert",
                    "QuestionText": "Base Q2",
                    "DataExportTag": "q2",
                    "Choices": {"1": {"Display": "Row 1"}},
                    "ChoiceOrder": ["1"],
                    "Answers": {"1": {"Display": "Col 1"}},
                    "AnswerOrder": ["1"],
                    "Language": {
                        "FR": {
                            "QuestionText": "FR Q2",
                            "Choices": {"1": {"Display": "Ligne 1"}},
                            "Answers": {"1": {"Display": "Colonne 1"}},
                        }
                    },
                },
            },
        }
    }


def test_init_workbook_adds_translation_columns(tmp_path: Path) -> None:
    payload = _survey_payload()
    workbook_path = tmp_path / "workbook.xlsx"

    init_workbook_from_survey(
        "SV_TEST",
        payload,
        workbook_path,
        languages=["FR"],
    )

    wb = load_workbook(workbook_path)
    assert QUESTION_SHEET in wb.sheetnames
    assert OPTIONS_SHEET in wb.sheetnames
    assert SUBITEMS_SHEET in wb.sheetnames

    q_headers = [cell.value for cell in next(wb[QUESTION_SHEET].iter_rows(max_row=1))]
    o_headers = [cell.value for cell in next(wb[OPTIONS_SHEET].iter_rows(max_row=1))]
    s_headers = [cell.value for cell in next(wb[SUBITEMS_SHEET].iter_rows(max_row=1))]

    assert "text_fr" in q_headers
    assert "ishtml_fr" in q_headers
    assert "Label_fr_MD" in o_headers
    assert "Label_fr_IsHTML" in o_headers
    assert "Label_fr_MD" in s_headers
    assert "Label_fr_IsHTML" in s_headers

    assert TRANSLATION_KEY_SHEET in wb.sheetnames
    ws = wb[TRANSLATION_KEY_SHEET]
    assert ws.sheet_state == "hidden"
    keys = [row[6].value for row in ws.iter_rows(min_row=2, max_row=5)]
    assert "QID1_QuestionText" in keys

    # Ensure the Questions table definition matches the sheet width so Excel does not
    # attempt to "repair" stale table metadata after adding translation columns.
    q_ws = wb[QUESTION_SHEET]
    q_table = q_ws._tables.get("QuestionsTable")
    assert q_table is not None
    assert q_table.ref == f"A1:{get_column_letter(q_ws.max_column)}{q_ws.max_row}"


def test_init_workbook_populates_translation_cells(tmp_path: Path) -> None:
    payload = _survey_payload_with_translation_blocks()
    workbook_path = tmp_path / "workbook.xlsx"

    init_workbook_from_survey("SV_TEST", payload, workbook_path, languages=["FR"])

    wb = load_workbook(workbook_path)

    q_ws = wb[QUESTION_SHEET]
    q_headers = [cell.value for cell in next(q_ws.iter_rows(max_row=1))]
    qid_idx = q_headers.index("QID") + 1
    text_fr_idx = q_headers.index("text_fr") + 1
    q1_row = None
    q2_row = None
    for row in range(2, q_ws.max_row + 1):
        qid = str(q_ws.cell(row=row, column=qid_idx).value or "").strip()
        if qid == "QID1":
            q1_row = row
        elif qid == "QID2":
            q2_row = row
    assert q1_row is not None
    assert q2_row is not None
    assert q_ws.cell(row=q1_row, column=text_fr_idx).value == "FR Q1"
    assert q_ws.cell(row=q2_row, column=text_fr_idx).value == "FR Q2"

    o_ws = wb[OPTIONS_SHEET]
    o_headers = [cell.value for cell in next(o_ws.iter_rows(max_row=1))]
    o_qid_idx = o_headers.index("QID") + 1
    o_choice_idx = o_headers.index("ChoiceId") + 1
    label_fr_idx = o_headers.index("Label_fr_MD") + 1
    q1_choice1_row = None
    q2_answer1_row = None
    for row in range(2, o_ws.max_row + 1):
        qid = str(o_ws.cell(row=row, column=o_qid_idx).value or "").strip()
        choice_id = str(o_ws.cell(row=row, column=o_choice_idx).value or "").strip()
        if qid == "QID1" and choice_id == "1":
            q1_choice1_row = row
        elif qid == "QID2" and choice_id == "1":
            q2_answer1_row = row
    assert q1_choice1_row is not None
    assert q2_answer1_row is not None
    assert o_ws.cell(row=q1_choice1_row, column=label_fr_idx).value == "Oui"
    assert o_ws.cell(row=q2_answer1_row, column=label_fr_idx).value == "Colonne 1"

    s_ws = wb[SUBITEMS_SHEET]
    s_headers = [cell.value for cell in next(s_ws.iter_rows(max_row=1))]
    s_qid_idx = s_headers.index("QID") + 1
    s_answer_idx = s_headers.index("AnswerId") + 1
    s_field_idx = s_headers.index("Field") + 1
    s_label_fr_idx = s_headers.index("Label_fr_MD") + 1
    q2_row1_row = None
    for row in range(2, s_ws.max_row + 1):
        qid = str(s_ws.cell(row=row, column=s_qid_idx).value or "").strip()
        answer_id = str(s_ws.cell(row=row, column=s_answer_idx).value or "").strip()
        field = str(s_ws.cell(row=row, column=s_field_idx).value or "").strip()
        if qid == "QID2" and field == "Answer" and answer_id == "1":
            q2_row1_row = row
            break
    assert q2_row1_row is not None
    assert s_ws.cell(row=q2_row1_row, column=s_label_fr_idx).value == "Ligne 1"

    q_ws.cell(row=q1_row, column=text_fr_idx).value = "Custom FR Q1"
    o_ws.cell(row=q1_choice1_row, column=label_fr_idx).value = "Custom Oui"
    wb.save(workbook_path)

    init_workbook_from_survey("SV_TEST", payload, workbook_path, languages=["FR"])
    wb = load_workbook(workbook_path)
    q_ws = wb[QUESTION_SHEET]
    q_headers = [cell.value for cell in next(q_ws.iter_rows(max_row=1))]
    qid_idx = q_headers.index("QID") + 1
    text_fr_idx = q_headers.index("text_fr") + 1
    q1_row = None
    for row in range(2, q_ws.max_row + 1):
        qid = str(q_ws.cell(row=row, column=qid_idx).value or "").strip()
        if qid == "QID1":
            q1_row = row
            break
    assert q1_row is not None
    assert q_ws.cell(row=q1_row, column=text_fr_idx).value == "Custom FR Q1"

    o_ws = wb[OPTIONS_SHEET]
    o_headers = [cell.value for cell in next(o_ws.iter_rows(max_row=1))]
    o_qid_idx = o_headers.index("QID") + 1
    o_choice_idx = o_headers.index("ChoiceId") + 1
    label_fr_idx = o_headers.index("Label_fr_MD") + 1
    q1_choice1_row = None
    for row in range(2, o_ws.max_row + 1):
        qid = str(o_ws.cell(row=row, column=o_qid_idx).value or "").strip()
        choice_id = str(o_ws.cell(row=row, column=o_choice_idx).value or "").strip()
        if qid == "QID1" and choice_id == "1":
            q1_choice1_row = row
            break
    assert q1_choice1_row is not None
    assert o_ws.cell(row=q1_choice1_row, column=label_fr_idx).value == "Custom Oui"


def test_workbook_doctor_flags_placeholder(tmp_path: Path, monkeypatch) -> None:
    payload = _survey_payload()
    payload["result"]["SurveyOptions"] = {
        "SurveyLanguage": "EN",
        "AvailableLanguages": ["EN", "FR"],
    }
    payload["result"]["Questions"]["QID1"][
        "QuestionText"
    ] = "Hello ${e://Field/COUNTRY}"

    monkeypatch.setenv("QSYNC_ROOT", str(tmp_path))
    _write_cached_survey(tmp_path, "SV_TEST", payload)

    workbook_path = tmp_path / "workbook.xlsx"
    init_workbook_from_survey(
        "SV_TEST",
        payload,
        workbook_path,
        languages=["FR"],
    )

    wb = load_workbook(workbook_path)
    ws = wb[QUESTION_SHEET]
    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    col_idx = headers.index("text_fr") + 1
    ws.cell(row=2, column=col_idx).value = "Bonjour"
    wb.save(workbook_path)

    report = run_translation_doctor(
        "SV_TEST",
        ["FR"],
        base_language="EN",
        workbook_path=workbook_path,
    )

    assert any("missing placeholders" in err for err in report.errors)


def test_workbook_doctor_allows_empty_when_base_empty(
    tmp_path: Path, monkeypatch
) -> None:
    payload = _survey_payload()
    payload["result"]["SurveyOptions"] = {
        "SurveyLanguage": "EN",
        "AvailableLanguages": ["EN", "FR"],
    }
    payload["result"]["Questions"]["QID1"]["QuestionText"] = ""

    monkeypatch.setenv("QSYNC_ROOT", str(tmp_path))
    _write_cached_survey(tmp_path, "SV_TEST", payload)

    workbook_path = tmp_path / "workbook.xlsx"
    init_workbook_from_survey(
        "SV_TEST",
        payload,
        workbook_path,
        languages=["FR"],
    )

    wb = load_workbook(workbook_path)
    # Populate all FR cells with non-empty values, except the one key that is
    # intentionally empty in the base language.
    q_ws = wb[QUESTION_SHEET]
    q_headers = [cell.value for cell in next(q_ws.iter_rows(max_row=1))]
    q_fr_idx = q_headers.index("text_fr") + 1
    q_ws.cell(row=2, column=q_fr_idx).value = ""

    o_ws = wb[OPTIONS_SHEET]
    o_headers = [cell.value for cell in next(o_ws.iter_rows(max_row=1))]
    o_qid_idx = o_headers.index("QID") + 1
    o_choice_idx = o_headers.index("ChoiceId") + 1
    o_fr_idx = o_headers.index("Label_fr_MD") + 1
    for row in range(2, o_ws.max_row + 1):
        if str(o_ws.cell(row=row, column=o_qid_idx).value or "").strip() != "QID1":
            continue
        choice = str(o_ws.cell(row=row, column=o_choice_idx).value or "").strip()
        o_ws.cell(row=row, column=o_fr_idx).value = f"FR opt {choice}"

    s_ws = wb[SUBITEMS_SHEET]
    s_headers = [cell.value for cell in next(s_ws.iter_rows(max_row=1))]
    s_qid_idx = s_headers.index("QID") + 1
    s_fr_idx = s_headers.index("Label_fr_MD") + 1
    for row in range(2, s_ws.max_row + 1):
        if str(s_ws.cell(row=row, column=s_qid_idx).value or "").strip() != "QID1":
            continue
        s_ws.cell(row=row, column=s_fr_idx).value = f"FR row {row}"
    wb.save(workbook_path)

    report = run_translation_doctor(
        "SV_TEST",
        ["FR"],
        base_language="EN",
        workbook_path=workbook_path,
    )

    assert not report.errors
    assert report.coverage["FR"]["empty"] == 0
    assert not any("Coverage incomplete" in warn for warn in report.warnings)


def test_subitems_field_backfill_for_labels(tmp_path: Path) -> None:
    payload = _survey_payload_with_labels()
    workbook_path = tmp_path / "workbook.xlsx"

    init_workbook_from_survey(
        "SV_TEST",
        payload,
        workbook_path,
        languages=["FR"],
    )

    wb = load_workbook(workbook_path)
    ws = wb[SUBITEMS_SHEET]
    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    qid_idx = headers.index("QID") + 1
    answer_idx = headers.index("AnswerId") + 1
    field_idx = headers.index("Field") + 1
    label_en_idx = next(
        headers.index(h) + 1
        for h in headers
        if str(h).startswith("Label_") and str(h).endswith("_MD")
    )
    label_fr_idx = headers.index("Label_fr_MD") + 1

    label_rows = {}
    answer_fields = []
    for row in range(2, ws.max_row + 1):
        qid = str(ws.cell(row=row, column=qid_idx).value or "").strip()
        if qid != "QID1":
            continue
        field = str(ws.cell(row=row, column=field_idx).value or "").strip()
        answer_id = str(ws.cell(row=row, column=answer_idx).value or "").strip()
        if field == "Label":
            label_rows[answer_id] = row
        else:
            answer_fields.append(field)

    assert set(label_rows.keys()) == {"1", "2"}
    assert all(field == "Answer" for field in answer_fields)

    low_row = label_rows["1"]
    assert ws.cell(row=low_row, column=label_en_idx).value == "Low"
    assert ws.cell(row=low_row, column=label_fr_idx).value == "Bas"

    ws.cell(row=low_row, column=label_fr_idx).value = "Bas custom"
    wb.save(workbook_path)

    init_workbook_from_survey(
        "SV_TEST",
        payload,
        workbook_path,
        languages=["FR"],
    )

    wb = load_workbook(workbook_path)
    ws = wb[SUBITEMS_SHEET]
    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    field_idx = headers.index("Field") + 1
    label_fr_idx = headers.index("Label_fr_MD") + 1

    label_row_count = 0
    for row in range(2, ws.max_row + 1):
        field = str(ws.cell(row=row, column=field_idx).value or "").strip()
        if field == "Label":
            label_row_count += 1
    assert label_row_count == 2

    qid_idx = headers.index("QID") + 1
    answer_idx = headers.index("AnswerId") + 1
    low_row = None
    for row in range(2, ws.max_row + 1):
        qid = str(ws.cell(row=row, column=qid_idx).value or "").strip()
        answer_id = str(ws.cell(row=row, column=answer_idx).value or "").strip()
        field = str(ws.cell(row=row, column=field_idx).value or "").strip()
        if qid == "QID1" and field == "Label" and answer_id == "1":
            low_row = row
            break
    assert low_row is not None
    assert ws.cell(row=low_row, column=label_fr_idx).value == "Bas custom"


def test_workbook_doctor_warns_large_delta(tmp_path: Path, monkeypatch) -> None:
    payload = _survey_payload()
    payload["result"]["SurveyOptions"] = {
        "SurveyLanguage": "EN",
        "AvailableLanguages": ["EN", "FR"],
    }
    payload["result"]["Questions"]["QID1"][
        "QuestionText"
    ] = "This is a base sentence that is long enough to trigger deltas."

    monkeypatch.setenv("QSYNC_ROOT", str(tmp_path))
    _write_cached_survey(tmp_path, "SV_TEST", payload)

    workbook_path = tmp_path / "workbook.xlsx"
    init_workbook_from_survey(
        "SV_TEST",
        payload,
        workbook_path,
        languages=["FR"],
    )

    wb = load_workbook(workbook_path)
    ws = wb[QUESTION_SHEET]
    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    col_idx = headers.index("text_fr") + 1
    ws.cell(row=2, column=col_idx).value = "Court."
    wb.save(workbook_path)

    report = run_translation_doctor(
        "SV_TEST",
        ["FR"],
        base_language="EN",
        workbook_path=workbook_path,
    )

    assert any("large delta" in warn for warn in report.warnings)


def test_init_preserves_non_empty_cells(tmp_path: Path) -> None:
    payload = _survey_payload()
    payload["result"]["SurveyOptions"] = {
        "SurveyLanguage": "EN",
        "AvailableLanguages": ["EN", "FR"],
    }

    workbook_path = tmp_path / "workbook.xlsx"
    init_workbook_from_survey(
        "SV_TEST",
        payload,
        workbook_path,
        languages=["FR"],
    )

    wb = load_workbook(workbook_path)
    q_ws = wb[QUESTION_SHEET]
    q_headers = [cell.value for cell in next(q_ws.iter_rows(max_row=1))]
    en_idx = next(
        q_headers.index(h) + 1
        for h in q_headers
        if str(h).startswith("text_")
    )
    fr_idx = q_headers.index("text_fr") + 1
    q_ws.cell(row=2, column=en_idx).value = "Custom EN"
    q_ws.cell(row=2, column=fr_idx).value = "Custom FR"

    o_ws = wb[OPTIONS_SHEET]
    o_headers = [cell.value for cell in next(o_ws.iter_rows(max_row=1))]
    o_fr_idx = o_headers.index("Label_fr_MD") + 1
    o_ws.cell(row=2, column=o_fr_idx).value = "Custom FR Opt"

    wb.save(workbook_path)

    init_workbook_from_survey(
        "SV_TEST",
        payload,
        workbook_path,
        languages=["FR"],
    )

    wb = load_workbook(workbook_path)
    q_ws = wb[QUESTION_SHEET]
    assert q_ws.cell(row=2, column=en_idx).value == "Custom EN"
    assert q_ws.cell(row=2, column=fr_idx).value == "Custom FR"
    o_ws = wb[OPTIONS_SHEET]
    assert o_ws.cell(row=2, column=o_fr_idx).value == "Custom FR Opt"


def test_workbook_doctor_rejects_overlong_values(tmp_path: Path, monkeypatch) -> None:
    from qsync.translations import QUALTRICS_TRANSLATION_VALUE_MAX_CHARS

    payload = _survey_payload()
    payload["result"]["SurveyOptions"] = {
        "SurveyLanguage": "EN",
        "AvailableLanguages": ["EN", "FR"],
    }

    monkeypatch.setenv("QSYNC_ROOT", str(tmp_path))
    _write_cached_survey(tmp_path, "SV_TEST", payload)

    workbook_path = tmp_path / "workbook.xlsx"
    init_workbook_from_survey(
        "SV_TEST",
        payload,
        workbook_path,
        languages=["FR"],
    )

    wb = load_workbook(workbook_path)
    ws = wb[QUESTION_SHEET]
    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    col_idx = headers.index("text_fr") + 1
    ws.cell(row=2, column=col_idx).value = "x" * (
        QUALTRICS_TRANSLATION_VALUE_MAX_CHARS + 1
    )
    wb.save(workbook_path)

    report = run_translation_doctor(
        "SV_TEST",
        ["FR"],
        base_language="EN",
        workbook_path=workbook_path,
    )

    assert any("length limit" in err.lower() for err in report.errors)
