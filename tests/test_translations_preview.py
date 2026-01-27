from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from qsync.dimensions.translations_core import preview_translations
from qsync.excel_io import init_workbook_from_survey, QUESTION_SHEET, OPTIONS_SHEET, SUBITEMS_SHEET
from qsync.workbook_resolver import WorkbookResolver


def _survey_payload() -> dict:
    return {
        "result": {
            "SurveyID": "SV_TEST",
            "SurveyOptions": {
                "SurveyLanguage": "EN",
                "SurveyTitle": "TestSurvey",
                "AvailableLanguages": {"EN": True, "FR": True},
            },
            "SurveyFlow": {"Flow": [{"Type": "Standard", "ID": "BL_1"}]},
            "Blocks": {
                "BL_1": {
                    "Type": "Default",
                    "Description": "Block 1",
                    "BlockElements": [
                        {"Type": "Question", "QuestionID": "QID1"},
                        {"Type": "Question", "QuestionID": "QID2"},
                    ],
                }
            },
            "Questions": {
                "QID1": {
                    "QuestionType": "Matrix",
                    "Selector": "Likert",
                    "QuestionText": "Base Q1",
                    "DataExportTag": "q1",
                    "Choices": {"1": {"Display": "Row 1"}},
                    "ChoiceOrder": ["1"],
                    "Answers": {"1": {"Display": "Col 1"}},
                    "AnswerOrder": ["1"],
                },
                "QID2": {
                    "QuestionType": "MC",
                    "Selector": "SAVR",
                    "QuestionText": "Base Q2",
                    "DataExportTag": "q2",
                    "Choices": {"1": {"Display": "Choice 1"}},
                    "ChoiceOrder": ["1"],
                },
            },
        }
    }


def _write_inventory(tmp_path: Path) -> None:
    surveys_dir = tmp_path / "surveys"
    surveys_dir.mkdir(parents=True, exist_ok=True)
    inventory = surveys_dir / "inventory.csv"
    inventory.write_text(
        "id,name,lastModified\nSV_TEST,TestSurvey,2026-01-23T12:00:00Z\n",
        encoding="utf-8",
    )


def _write_cached_survey(tmp_path: Path, payload: dict) -> None:
    surveys_dir = tmp_path / "surveys"
    surveys_dir.mkdir(parents=True, exist_ok=True)
    path = surveys_dir / "TestSurvey__SV_TEST.json"
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def _write_workbook(tmp_path: Path, payload: dict) -> Path:
    resolver = WorkbookResolver()
    workbook_path = resolver.default_path("SV_TEST")
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    init_workbook_from_survey("SV_TEST", payload, workbook_path, languages=["FR"])
    return workbook_path


def test_preview_translations_no_changes(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setenv("QSYNC_ROOT", str(tmp_path))
    payload = _survey_payload()
    _write_inventory(tmp_path)
    _write_cached_survey(tmp_path, payload)
    _write_workbook(tmp_path, payload)

    lines = preview_translations("SV_TEST", ["FR"])
    assert lines == ["No differences between Excel and cached survey."]


def test_preview_translations_one_qid_change(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setenv("QSYNC_ROOT", str(tmp_path))
    payload = _survey_payload()
    _write_inventory(tmp_path)
    _write_cached_survey(tmp_path, payload)
    workbook_path = _write_workbook(tmp_path, payload)

    wb = load_workbook(workbook_path)
    q_ws = wb[QUESTION_SHEET]
    q_headers = [cell.value for cell in next(q_ws.iter_rows(max_row=1))]
    qid_idx = q_headers.index("QID") + 1
    text_idx = q_headers.index("Text_fr_MD") + 1
    for row in range(2, q_ws.max_row + 1):
        if str(q_ws.cell(row=row, column=qid_idx).value or "").strip() == "QID1":
            q_ws.cell(row=row, column=text_idx).value = "Bonjour"
            break
    wb.save(workbook_path)

    lines = preview_translations("SV_TEST", ["FR"])
    assert any(line.startswith("- QID1:") and "FR=1" in line for line in lines)


def test_preview_translations_multiple_sheets(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setenv("QSYNC_ROOT", str(tmp_path))
    payload = _survey_payload()
    _write_inventory(tmp_path)
    _write_cached_survey(tmp_path, payload)
    workbook_path = _write_workbook(tmp_path, payload)

    wb = load_workbook(workbook_path)

    q_ws = wb[QUESTION_SHEET]
    q_headers = [cell.value for cell in next(q_ws.iter_rows(max_row=1))]
    qid_idx = q_headers.index("QID") + 1
    text_idx = q_headers.index("Text_fr_MD") + 1
    for row in range(2, q_ws.max_row + 1):
        if str(q_ws.cell(row=row, column=qid_idx).value or "").strip() == "QID2":
            q_ws.cell(row=row, column=text_idx).value = "Bonjour 2"
            break

    o_ws = wb[OPTIONS_SHEET]
    o_headers = [cell.value for cell in next(o_ws.iter_rows(max_row=1))]
    o_qid_idx = o_headers.index("QID") + 1
    o_label_idx = o_headers.index("Label_fr_MD") + 1
    for row in range(2, o_ws.max_row + 1):
        if str(o_ws.cell(row=row, column=o_qid_idx).value or "").strip() == "QID1":
            o_ws.cell(row=row, column=o_label_idx).value = "FR option"
            break

    s_ws = wb[SUBITEMS_SHEET]
    s_headers = [cell.value for cell in next(s_ws.iter_rows(max_row=1))]
    s_qid_idx = s_headers.index("QID") + 1
    s_label_idx = s_headers.index("Label_fr_MD") + 1
    for row in range(2, s_ws.max_row + 1):
        if str(s_ws.cell(row=row, column=s_qid_idx).value or "").strip() == "QID1":
            s_ws.cell(row=row, column=s_label_idx).value = "FR subitem"
            break

    wb.save(workbook_path)

    lines = preview_translations("SV_TEST", ["FR"])
    assert any(line.startswith("- QID1:") and "FR=2" in line for line in lines)
    assert any(line.startswith("- QID2:") and "FR=1" in line for line in lines)


def test_preview_translations_missing_language_blocks(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setenv("QSYNC_ROOT", str(tmp_path))
    payload = _survey_payload()
    _write_inventory(tmp_path)
    _write_cached_survey(tmp_path, payload)
    workbook_path = _write_workbook(tmp_path, payload)

    wb = load_workbook(workbook_path)
    q_ws = wb[QUESTION_SHEET]
    q_headers = [cell.value for cell in next(q_ws.iter_rows(max_row=1))]
    qid_idx = q_headers.index("QID") + 1
    text_idx = q_headers.index("Text_fr_MD") + 1
    for row in range(2, q_ws.max_row + 1):
        if str(q_ws.cell(row=row, column=qid_idx).value or "").strip() == "QID2":
            q_ws.cell(row=row, column=text_idx).value = "Bonjour missing"
            break
    wb.save(workbook_path)

    lines = preview_translations("SV_TEST", ["FR"])
    assert any(line.startswith("- QID2:") and "FR=1" in line for line in lines)
