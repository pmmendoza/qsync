"""Integration-style test for embedded data preview/apply workflow."""

from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from qsync import excel_io
from qsync.sync_core import apply_changes, preview_changes


def _payload_with_debug(value: str) -> dict:
    return {
        "result": {
            "SurveyFlow": {
                "Flow": [
                    {
                        "Type": "EmbeddedData",
                        "FlowID": "FL_3",
                        "EmbeddedData": [
                            {
                                "Field": "DEBUG",
                                "Type": "Custom",
                                "Value": value,
                                "Description": "DEBUG",
                                "DataVisibility": [],
                                "AnalyzeText": False,
                            }
                        ],
                    }
                ]
            },
            "Questions": {},
            "Blocks": {},
        }
    }


class EmbeddedDataApplyTests(unittest.TestCase):
    def test_preview_apply_embedded_data(self) -> None:
        payload = _payload_with_debug("F")
        with tempfile.TemporaryDirectory(prefix="qsync_test_") as tmpdir:
            root = Path(tmpdir)
            surveys_dir = root / "surveys"
            backups_dir = surveys_dir / "backups"
            excel_dir = root / "excel"
            surveys_dir.mkdir(parents=True, exist_ok=True)
            backups_dir.mkdir(parents=True, exist_ok=True)
            excel_dir.mkdir(parents=True, exist_ok=True)

            survey_id = "SV_TEST"
            cached_path = surveys_dir / f"TEST__{survey_id}.json"
            backup_path = backups_dir / f"TEST__{survey_id}.json"
            cached_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
            backup_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")

            xlsx_path = excel_dir / f"{survey_id}.xlsx"
            excel_io.init_workbook_from_survey(survey_id, payload, xlsx_path)

            wb = load_workbook(xlsx_path)
            ws = wb[excel_io.EMBEDDED_DATA_SHEET]
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            idx = {name: i for i, name in enumerate(headers)}
            field_idx = idx["Field"]
            value_idx = idx["Value"]

            for row in ws.iter_rows(min_row=2, values_only=False):
                field = str(row[field_idx].value or "").strip()
                if field == "DEBUG":
                    row[value_idx].value = "T"
                    break
            wb.save(xlsx_path)

            with patch("qsync.qualtrics_client._workspace_root", return_value=root):
                preview = preview_changes(survey_id, xlsx_path)
                self.assertEqual(len(preview), 1)
                self.assertEqual(preview[0].kind, "embedded")

                result = apply_changes(survey_id, xlsx_path)
                self.assertEqual(result.qids, [])
                self.assertEqual(len(result.embedded_fields), 1)

            updated = json.loads(cached_path.read_text(encoding="utf-8"))
            embedded = updated["result"]["SurveyFlow"]["Flow"][0]["EmbeddedData"][0]
            self.assertEqual(embedded["Value"], "T")
