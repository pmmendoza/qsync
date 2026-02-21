from __future__ import annotations

import json
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

import pytest
from openpyxl import load_workbook

from qsync import excel_io
from qsync.dimensions import edf as edf_dimension
from qsync.dimensions import items as items_dimension
from qsync.dimensions.types import DimensionChanges
from qsync.sync_orchestrator import render_cell
from qsync.workbook_resolver import WorkbookResolver


def _payload_with_embedded() -> dict:
    return {
        "result": {
            "SurveyFlow": {
                "Flow": [
                    {"Type": "Standard", "ID": "BL_1"},
                    {
                        "Type": "EmbeddedData",
                        "FlowID": "FL_1",
                        "EmbeddedData": [
                            {
                                "Field": "DEBUG",
                                "Type": "Custom",
                                "Value": "F",
                                "Description": "DEBUG",
                                "DataVisibility": [],
                                "AnalyzeText": False,
                            }
                        ],
                    },
                ]
            },
            "Questions": {
                "QID1": {
                    "QuestionText": "Old text",
                    "QuestionType": "TE",
                    "Selector": "SL",
                    "SubSelector": "TX",
                    "DataExportTag": "Q1",
                    "QuestionDescription": "Q1",
                    "Configuration": {},
                    "Choices": {},
                    "Answers": {},
                }
            },
            "Blocks": {
                "BL_1": {
                    "Type": "Standard",
                    "ID": "BL_1",
                    "Description": "Block 1",
                    "BlockElements": [{"Type": "Question", "QuestionID": "QID1"}],
                }
            },
        }
    }


def _write_cached_survey(root: Path, survey_id: str, payload: dict) -> None:
    surveys_dir = root / "surveys"
    backups_dir = surveys_dir / "backups"
    surveys_dir.mkdir(parents=True, exist_ok=True)
    backups_dir.mkdir(parents=True, exist_ok=True)
    cached_path = surveys_dir / f"TEST__{survey_id}.json"
    backup_path = backups_dir / f"TEST__{survey_id}.json"
    cached_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    backup_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def _update_embedded_value(xlsx_path: Path, field: str, value: str) -> None:
    wb = load_workbook(xlsx_path)
    ws = wb[excel_io.EMBEDDED_DATA_SHEET]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: i for i, name in enumerate(headers)}
    field_idx = idx["Field"]
    value_idx = idx["Value"]
    for row in ws.iter_rows(min_row=2, values_only=False):
        if str(row[field_idx].value or "").strip() == field:
            row[value_idx].value = value
            break
    wb.save(xlsx_path)


def _remove_embedded_row(xlsx_path: Path, field: str) -> None:
    wb = load_workbook(xlsx_path)
    ws = wb[excel_io.EMBEDDED_DATA_SHEET]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: i for i, name in enumerate(headers)}
    field_idx = idx["Field"]
    for row in ws.iter_rows(min_row=2, values_only=False):
        if str(row[field_idx].value or "").strip() == field:
            ws.delete_rows(row[0].row, 1)
            break
    wb.save(xlsx_path)


def _find_base_col(headers, prefix):
    if prefix == "Text":
        lang_map = excel_io._question_text_lang_columns_from_headers(headers)
        if "EN" in lang_map:
            return headers.index(lang_map["EN"][0])
        if lang_map:
            return headers.index(next(iter(lang_map.values()))[0])
        if "text_en" in headers:
            return headers.index("text_en")
    for h in headers:
        if str(h).startswith(f"{prefix}_") and str(h).endswith("_MD"):
            return headers.index(h)
    return headers.index(f"{prefix}_en_MD")


def _update_question_text(xlsx_path: Path, qid: str, text: str) -> None:
    wb = load_workbook(xlsx_path)
    ws = wb[excel_io.QUESTION_SHEET]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: i for i, name in enumerate(headers)}
    qid_idx = idx["QID"]
    text_idx = _find_base_col(headers, "Text")
    for row in ws.iter_rows(min_row=2, values_only=False):
        if str(row[qid_idx].value or "").strip() == qid:
            row[text_idx].value = text
            break
    wb.save(xlsx_path)


def _update_question_translation_text(
    xlsx_path: Path, qid: str, language: str, text: str
) -> None:
    wb = load_workbook(xlsx_path)
    ws = wb[excel_io.QUESTION_SHEET]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: i for i, name in enumerate(headers)}
    qid_idx = idx["QID"]
    col_name = f"text_{str(language).lower()}"
    if col_name not in idx:
        raise KeyError(f"Missing translation column: {col_name}")
    text_idx = idx[col_name]
    for row in ws.iter_rows(min_row=2, values_only=False):
        if str(row[qid_idx].value or "").strip() == qid:
            row[text_idx].value = text
            break
    wb.save(xlsx_path)


def _read_question_text(xlsx_path: Path, qid: str) -> str:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[excel_io.QUESTION_SHEET]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: i for i, name in enumerate(headers)}
    qid_idx = idx["QID"]
    text_idx = _find_base_col(headers, "Text")
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[qid_idx] or "").strip() == qid:
            return str(row[text_idx] or "")
    return ""


def _read_question_translation_text(
    xlsx_path: Path, qid: str, language: str
) -> str:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[excel_io.QUESTION_SHEET]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: i for i, name in enumerate(headers)}
    qid_idx = idx["QID"]
    col_name = f"text_{str(language).lower()}"
    if col_name not in idx:
        raise KeyError(f"Missing translation column: {col_name}")
    text_idx = idx[col_name]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[qid_idx] or "").strip() == qid:
            return str(row[text_idx] or "")
    return ""


def test_edf_only_change_does_not_mark_items(tmp_path: Path, monkeypatch) -> None:
    payload = _payload_with_embedded()
    survey_id = "SV_TEST"
    _write_cached_survey(tmp_path, survey_id, payload)

    excel_dir = tmp_path / "excel"
    excel_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = WorkbookResolver(root=tmp_path).resolve(survey_id)
    excel_io.init_workbook_from_survey(survey_id, payload, xlsx_path)

    _update_embedded_value(xlsx_path, "DEBUG", "T")

    monkeypatch.setenv("QSYNC_ROOT", str(tmp_path))
    items_changes = items_dimension.detect_changes(survey_id)
    edf_changes = edf_dimension.detect_changes(survey_id)

    assert items_changes.has_changes is False
    assert edf_changes.has_changes is True


def test_render_cell_dual_badge() -> None:
    status = DimensionChanges(
        dimension="edf",
        has_changes=True,
        change_summary="",
        affected_qids=set(),
        warning_detail="warn",
        status_kind="unstaged",
        edit_count=3,
    )
    assert render_cell(status) == "⚡ 3 ⚠"


def test_repair_workbook_repairs_missing_rows_and_preserves_question_cells(
    tmp_path: Path, monkeypatch
) -> None:
    payload = _payload_with_embedded()
    survey_id = "SV_REPAIR"
    _write_cached_survey(tmp_path, survey_id, payload)

    xlsx_path = WorkbookResolver(root=tmp_path).resolve(survey_id)
    excel_io.init_workbook_from_survey(survey_id, payload, xlsx_path)
    _remove_embedded_row(xlsx_path, "DEBUG")
    _update_question_text(xlsx_path, "QID1", "Keep this local edit")

    monkeypatch.setenv("QSYNC_ROOT", str(tmp_path))
    with patch("qsync.qualtrics_client._workspace_root", return_value=tmp_path):
        report = edf_dimension.repair_workbook(
            survey_id,
            xlsx_path=xlsx_path,
            dry_run=False,
            refresh_cache=False,
            retain_backups=3,
        )

    fields = [row.field for row in excel_io.load_embedded_data_from_workbook(xlsx_path)]
    assert "DEBUG" in fields
    assert _read_question_text(xlsx_path, "QID1") == "Keep this local edit"
    assert report.changed is True
    assert report.rows_added >= 1
    assert report.backup_path is not None
    assert report.backup_path.exists()


def test_repair_workbook_dry_run_does_not_modify_file(
    tmp_path: Path, monkeypatch
) -> None:
    payload = _payload_with_embedded()
    survey_id = "SV_DRYRUN"
    _write_cached_survey(tmp_path, survey_id, payload)

    xlsx_path = WorkbookResolver(root=tmp_path).resolve(survey_id)
    excel_io.init_workbook_from_survey(survey_id, payload, xlsx_path)
    _remove_embedded_row(xlsx_path, "DEBUG")

    monkeypatch.setenv("QSYNC_ROOT", str(tmp_path))
    with patch("qsync.qualtrics_client._workspace_root", return_value=tmp_path):
        report = edf_dimension.repair_workbook(
            survey_id,
            xlsx_path=xlsx_path,
            dry_run=True,
            refresh_cache=False,
            retain_backups=3,
        )

    fields = [row.field for row in excel_io.load_embedded_data_from_workbook(xlsx_path)]
    assert "DEBUG" not in fields
    assert report.changed is True
    assert report.backup_path is None


def test_repair_workbook_preserves_translation_cells(
    tmp_path: Path, monkeypatch
) -> None:
    payload = _payload_with_embedded()
    payload["result"]["SurveyOptions"] = {
        "SurveyLanguage": "EN",
        "AvailableLanguages": {"EN": True, "FR": True},
    }
    payload["result"]["Questions"]["QID1"]["Language"] = {
        "FR": {"QuestionText": "Texte d'origine"}
    }
    survey_id = "SV_REPAIR_TRANSLATIONS"
    _write_cached_survey(tmp_path, survey_id, payload)

    xlsx_path = WorkbookResolver(root=tmp_path).resolve(survey_id)
    excel_io.init_workbook_from_survey(survey_id, payload, xlsx_path, languages=["FR"])
    _remove_embedded_row(xlsx_path, "DEBUG")
    _update_question_translation_text(
        xlsx_path, "QID1", "FR", "Conserver cette traduction locale"
    )

    monkeypatch.setenv("QSYNC_ROOT", str(tmp_path))
    with patch("qsync.qualtrics_client._workspace_root", return_value=tmp_path):
        report = edf_dimension.repair_workbook(
            survey_id,
            xlsx_path=xlsx_path,
            dry_run=False,
            refresh_cache=False,
            retain_backups=3,
        )

    assert report.changed is True
    assert (
        _read_question_translation_text(xlsx_path, "QID1", "FR")
        == "Conserver cette traduction locale"
    )


def test_repair_workbook_dry_run_is_stable_across_repeated_runs(
    tmp_path: Path, monkeypatch
) -> None:
    payload = _payload_with_embedded()
    survey_id = "SV_DRYRUN_STABLE"
    _write_cached_survey(tmp_path, survey_id, payload)

    xlsx_path = WorkbookResolver(root=tmp_path).resolve(survey_id)
    excel_io.init_workbook_from_survey(survey_id, payload, xlsx_path)
    _remove_embedded_row(xlsx_path, "DEBUG")
    bytes_before = xlsx_path.read_bytes()

    monkeypatch.setenv("QSYNC_ROOT", str(tmp_path))
    with patch("qsync.qualtrics_client._workspace_root", return_value=tmp_path):
        report_a = edf_dimension.repair_workbook(
            survey_id,
            xlsx_path=xlsx_path,
            dry_run=True,
            refresh_cache=False,
            retain_backups=3,
        )
        report_b = edf_dimension.repair_workbook(
            survey_id,
            xlsx_path=xlsx_path,
            dry_run=True,
            refresh_cache=False,
            retain_backups=3,
        )

    bytes_after = xlsx_path.read_bytes()
    assert bytes_after == bytes_before
    assert report_a.changed is True and report_b.changed is True
    assert report_a.rows_before == report_b.rows_before
    assert report_a.rows_after == report_b.rows_after
    assert report_a.rows_added == report_b.rows_added
    assert report_a.rows_removed == report_b.rows_removed
    assert report_a.backup_path is None and report_b.backup_path is None


def test_repair_workbook_requires_cached_json_when_not_refreshing(
    tmp_path: Path, monkeypatch
) -> None:
    payload = _payload_with_embedded()
    survey_id = "SV_NOCACHE"

    xlsx_path = WorkbookResolver(root=tmp_path).resolve(survey_id)
    excel_io.init_workbook_from_survey(survey_id, payload, xlsx_path)

    monkeypatch.setenv("QSYNC_ROOT", str(tmp_path))
    with patch("qsync.qualtrics_client._workspace_root", return_value=tmp_path):
        with pytest.raises(RuntimeError, match="No cached survey JSON"):
            edf_dimension.repair_workbook(
                survey_id,
                xlsx_path=xlsx_path,
                dry_run=True,
                refresh_cache=False,
            )


def test_repair_workbook_write_failure_reports_clean_error(
    tmp_path: Path, monkeypatch
) -> None:
    payload = _payload_with_embedded()
    survey_id = "SV_LOCKED"
    _write_cached_survey(tmp_path, survey_id, payload)

    xlsx_path = WorkbookResolver(root=tmp_path).resolve(survey_id)
    excel_io.init_workbook_from_survey(survey_id, payload, xlsx_path)
    _remove_embedded_row(xlsx_path, "DEBUG")

    monkeypatch.setenv("QSYNC_ROOT", str(tmp_path))
    with (
        patch("qsync.qualtrics_client._workspace_root", return_value=tmp_path),
        patch(
            "qsync.dimensions.edf._create_workbook_backup",
            return_value=tmp_path / "noop.bak.xlsx",
        ),
        patch(
            "qsync.dimensions.edf.shutil.copy2", side_effect=PermissionError("locked")
        ),
    ):
        with pytest.raises(RuntimeError, match="Unable to write repaired workbook"):
            edf_dimension.repair_workbook(
                survey_id,
                xlsx_path=xlsx_path,
                dry_run=False,
                refresh_cache=False,
            )


def test_repair_workbook_refresh_cache_failure_bubbles(
    tmp_path: Path, monkeypatch
) -> None:
    payload = _payload_with_embedded()
    survey_id = "SV_REFRESH_FAIL"
    xlsx_path = WorkbookResolver(root=tmp_path).resolve(survey_id)
    excel_io.init_workbook_from_survey(survey_id, payload, xlsx_path)

    monkeypatch.setenv("QSYNC_ROOT", str(tmp_path))
    with patch(
        "qsync.dimensions.edf.refresh_survey_cache",
        side_effect=RuntimeError("refresh failed"),
    ):
        with pytest.raises(RuntimeError, match="refresh failed"):
            edf_dimension.repair_workbook(
                survey_id,
                xlsx_path=xlsx_path,
                dry_run=True,
                refresh_cache=True,
            )


def test_cli_items_repair_edf_passes_flags(tmp_path: Path, monkeypatch) -> None:
    import qsync.cli as cli

    workbook = tmp_path / "SV_TEST.xlsx"
    workbook.touch()

    health = SimpleNamespace(
        missing_fields=[],
        extra_fields=[],
        duplicate_fields=[],
        ambiguous_fields=[],
    )
    fake_report = edf_dimension.EdfRepairReport(
        survey_id="SV_TEST",
        workbook_path=workbook,
        dry_run=True,
        changed=False,
        rows_before=1,
        rows_after=1,
        rows_added=0,
        rows_removed=0,
        duplicate_rows_removed=0,
        unchanged_rows=1,
        extra_rows_preserved=0,
        backup_path=None,
        before_health=health,
        after_health=health,
    )

    captured: dict[str, object] = {}

    def _fake_repair(*args, **kwargs):
        captured["args"] = args
        captured["kwargs"] = kwargs
        return fake_report

    monkeypatch.setenv("QSYNC_ROOT", str(tmp_path))
    with (
        patch.object(cli.os, "chdir", lambda *_args, **_kwargs: None),
        patch("qsync.dimensions.edf.repair_workbook", side_effect=_fake_repair),
    ):
        cli._main_impl(
            [
                "items",
                "repair-edf",
                "--survey-id",
                "SV_TEST",
                "--xlsx",
                str(workbook),
                "--dry-run",
                "--retain-backups",
                "7",
            ]
        )

    assert captured["args"] == ("SV_TEST",)
    kwargs = captured["kwargs"]
    assert kwargs["dry_run"] is True
    assert kwargs["refresh_cache"] is False
    assert kwargs["retain_backups"] == 7
    assert kwargs["xlsx_path"] == workbook


def test_cli_items_repair_edf_dry_run_output_is_checkable(
    tmp_path: Path, monkeypatch, capsys
) -> None:
    import qsync.cli as cli

    workbook = tmp_path / "SV_TEST.xlsx"
    workbook.touch()

    before_health = SimpleNamespace(
        missing_fields=["DEBUG"],
        extra_fields=["LEGACY"],
        duplicate_fields=[],
        ambiguous_fields=[],
    )
    after_health = SimpleNamespace(
        missing_fields=[],
        extra_fields=[],
        duplicate_fields=[],
        ambiguous_fields=[],
    )
    fake_report = edf_dimension.EdfRepairReport(
        survey_id="SV_TEST",
        workbook_path=workbook,
        dry_run=True,
        changed=True,
        rows_before=4,
        rows_after=5,
        rows_added=1,
        rows_removed=0,
        duplicate_rows_removed=0,
        unchanged_rows=4,
        extra_rows_preserved=0,
        backup_path=None,
        before_health=before_health,
        after_health=after_health,
    )

    monkeypatch.setenv("QSYNC_ROOT", str(tmp_path))
    with (
        patch.object(cli.os, "chdir", lambda *_args, **_kwargs: None),
        patch("qsync.dimensions.edf.repair_workbook", return_value=fake_report),
    ):
        cli._main_impl(
            [
                "items",
                "repair-edf",
                "--survey-id",
                "SV_TEST",
                "--xlsx",
                str(workbook),
                "--dry-run",
            ]
        )

    out = capsys.readouterr().out
    assert "Rows: 4 -> 5 (+1/-0, duplicates removed=0, unchanged=4)" in out
    assert "Issues: 2 -> 0 (extra rows preserved=0)" in out
    assert "Dry run complete; no workbook changes written." in out
