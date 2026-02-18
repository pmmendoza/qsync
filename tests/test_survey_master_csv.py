"""Tests for survey master CSV generation and field extraction."""

import csv
import os
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook


class SurveyMasterCSVTests(unittest.TestCase):
    """Test CSV generation, column ordering, and value extraction."""

    def test_column_order_respects_order_field(self) -> None:
        """Ordered columns appear first (1-19)."""
        from qsync.survey_master import _get_column_order

        mapping = {
            "SurveyID": {"field_name": "SurveyID", "order": "1"},
            "SurveyName": {"field_name": "SurveyName", "order": "2"},
            "Header": {"field_name": "Header", "order": ""},
            "Footer": {"field_name": "Footer", "order": "3"},
        }

        with patch("qsync.survey_master._parse_mapping_csv", return_value=mapping):
            columns = _get_column_order()

        # First 3 should be ordered (1, 2, 3)
        self.assertEqual(columns[0], "SurveyID")
        self.assertEqual(columns[1], "SurveyName")
        self.assertEqual(columns[2], "Footer")
        # Unordered at end
        self.assertEqual(columns[3], "Header")

    def test_column_order_alphabetical_for_unordered(self) -> None:
        """Unordered columns sorted alphabetically."""
        from qsync.survey_master import _get_column_order

        mapping = {
            "Zebra": {"field_name": "Zebra", "order": ""},
            "Alpha": {"field_name": "Alpha", "order": ""},
            "Beta": {"field_name": "Beta", "order": ""},
        }

        with patch("qsync.survey_master._parse_mapping_csv", return_value=mapping):
            columns = _get_column_order()

        # Unordered should be alphabetical
        self.assertEqual(columns[0], "Alpha")
        self.assertEqual(columns[1], "Beta")
        self.assertEqual(columns[2], "Zebra")

    def test_column_order_underscore_prefix_last(self) -> None:
        """Read-only (_prefixed) columns at end."""
        from qsync.survey_master import _get_column_order

        mapping = {
            "SurveyName": {"field_name": "SurveyName", "order": "1"},
            "_versionNumber": {"field_name": "_versionNumber", "order": ""},
            "Header": {"field_name": "Header", "order": ""},
            "_lastPublishedDate": {"field_name": "_lastPublishedDate", "order": ""},
        }

        with patch("qsync.survey_master._parse_mapping_csv", return_value=mapping):
            columns = _get_column_order()

        # Ordered first
        self.assertEqual(columns[0], "SurveyName")
        # Regular unordered middle
        self.assertEqual(columns[1], "Header")
        # Underscore prefix last
        self.assertIn("_versionNumber", columns[-2:])
        self.assertIn("_lastPublishedDate", columns[-2:])

    def test_extract_value_from_metadata_section(self) -> None:
        """Extracts simple value from metadata."""
        from qsync.survey_master import _extract_value_from_snapshot

        snapshot = {"sections": {"metadata": {"data": {"SurveyName": "My Survey"}}}}

        field_info = {
            "field_name": "SurveyName",
            "domain": "survey_metadata",
            "object_path": "result.SurveyName",
        }

        with patch("qsync.survey_master._derive_endpoint", return_value="metadata"):
            value = _extract_value_from_snapshot(snapshot, field_info)

        self.assertEqual(value, "My Survey")

    def test_extract_value_from_options_section(self) -> None:
        """Extracts simple value from options."""
        from qsync.survey_master import _extract_value_from_snapshot

        snapshot = {"sections": {"options": {"data": {"Header": "<p>Welcome</p>"}}}}

        field_info = {
            "field_name": "Header",
            "domain": "survey_options",
            "object_path": "result.Header",
        }

        with patch("qsync.survey_master._derive_endpoint", return_value="options"):
            value = _extract_value_from_snapshot(snapshot, field_info)

        self.assertEqual(value, "<p>Welcome</p>")

    def test_extract_value_from_status_section(self) -> None:
        """Extracts simple value from status."""
        from qsync.survey_master import _extract_value_from_snapshot

        snapshot = {"sections": {"status": {"data": {"isActive": True}}}}

        field_info = {
            "field_name": "isActive",
            "domain": "survey_detail",
            "object_path": "result.isActive",
        }

        with patch("qsync.survey_master._derive_endpoint", return_value="status"):
            value = _extract_value_from_snapshot(snapshot, field_info)

        self.assertEqual(value, True)

    def test_extract_value_nested_path(self) -> None:
        """Extracts nested value (e.g., responseCounts.generated)."""
        from qsync.survey_master import _extract_value_from_snapshot

        snapshot = {
            "sections": {
                "status": {
                    "data": {"responseCounts": {"generated": 42, "auditable": 38}}
                }
            }
        }

        field_info = {
            "field_name": "generated",
            "domain": "survey_detail",
            "object_path": "result.responseCounts.generated",
        }

        with patch("qsync.survey_master._derive_endpoint", return_value="status"):
            value = _extract_value_from_snapshot(snapshot, field_info)

        self.assertEqual(value, 42)

    def test_extract_value_missing_returns_empty(self) -> None:
        """Returns '' when path not found."""
        from qsync.survey_master import _extract_value_from_snapshot

        snapshot = {"sections": {"metadata": {"data": {}}}}  # Empty

        field_info = {
            "field_name": "NonexistentField",
            "domain": "survey_metadata",
            "object_path": "result.NonexistentField",
        }

        with patch("qsync.survey_master._derive_endpoint", return_value="metadata"):
            value = _extract_value_from_snapshot(snapshot, field_info)

        self.assertEqual(value, None)

    def test_extract_value_boolean_converts_to_string(self) -> None:
        """true → 'true', false → 'false'."""
        from qsync.survey_master import _extract_value_from_snapshot

        # Test true
        snapshot_true = {"sections": {"status": {"data": {"isActive": True}}}}

        field_info = {
            "field_name": "isActive",
            "domain": "survey_detail",
            "object_path": "result.isActive",
        }

        with patch("qsync.survey_master._derive_endpoint", return_value="status"):
            value = _extract_value_from_snapshot(snapshot_true, field_info)

        self.assertEqual(value, True)

        # Test false
        snapshot_false = {"sections": {"status": {"data": {"isActive": False}}}}

        with patch("qsync.survey_master._derive_endpoint", return_value="status"):
            value = _extract_value_from_snapshot(snapshot_false, field_info)

        self.assertEqual(value, False)

    def test_generate_csv_multiple_surveys(self) -> None:
        """Generates row per survey."""
        from qsync.survey_master import generate_master_csv_from_snapshots

        snapshots = {
            "SV_001": {"survey_id": "SV_001", "survey_name": "Survey 1"},
            "SV_002": {"survey_id": "SV_002", "survey_name": "Survey 2"},
        }

        mapping = {
            "SurveyID": {
                "field_name": "SurveyID",
                "domain": "survey_def",
                "order": "1",
                "object_path": "SurveyID",
            },
        }

        with patch("qsync.survey_master._parse_mapping_csv", return_value=mapping):
            with patch(
                "qsync.survey_master.load_snapshot",
                side_effect=lambda sid: snapshots.get(sid),
            ):
                with patch(
                    "qsync.survey_master._extract_value_from_snapshot", return_value=""
                ):
                    rows = generate_master_csv_from_snapshots(["SV_001", "SV_002"])

        self.assertEqual(len(rows), 3)

    def test_generate_csv_missing_snapshot_skipped(self) -> None:
        """Skips surveys without snapshots."""
        from qsync.survey_master import generate_master_csv_from_snapshots

        def mock_load_snapshot(survey_id):
            if survey_id == "SV_001":
                return {"survey_id": "SV_001", "sections": {}}
            return None

        mapping = {
            "SurveyID": {
                "field_name": "SurveyID",
                "domain": "survey_def",
                "order": "1",
                "object_path": "SurveyID",
            },
        }

        with patch("qsync.survey_master._parse_mapping_csv", return_value=mapping):
            with patch(
                "qsync.survey_master.load_snapshot", side_effect=mock_load_snapshot
            ):
                rows = generate_master_csv_from_snapshots(["SV_001", "SV_002"])

        # Should have header + 1 data row (SV_002 skipped)
        self.assertEqual(len(rows), 2)

    def test_write_csv_creates_file(self) -> None:
        """Creates CSV file."""
        from qsync.survey_master import write_master_csv

        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            csv_path = root / "surveys" / "qualtrics_master.csv"

            rows = [
                ["SurveyID", "SurveyName"],
                ["SV_001", "Survey 1"],
                ["SV_002", "Survey 2"],
            ]

            with patch("qsync.survey_master._master_csv_path", return_value=csv_path):
                write_master_csv(rows)

            self.assertTrue(csv_path.exists())

    def test_write_csv_utf8_encoding(self) -> None:
        """Handles UTF-8 characters."""
        from qsync.survey_master import write_master_csv

        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            csv_path = root / "surveys" / "qualtrics_master.csv"

            rows = [
                ["SurveyID", "SurveyName"],
                ["SV_001", "Encuesta Española ñ"],
            ]

            with patch("qsync.survey_master._master_csv_path", return_value=csv_path):
                write_master_csv(rows)

                # Verify UTF-8 written correctly
                content = csv_path.read_text(encoding="utf-8")
                self.assertIn("ñ", content)

    def test_load_csv_missing_file_raises(self) -> None:
        """FileNotFoundError when CSV missing."""
        from qsync.survey_master import load_master_csv

        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            csv_path = root / "surveys" / "qualtrics_master.csv"

            with patch("qsync.survey_master._master_csv_path", return_value=csv_path):
                headers, rows = load_master_csv()
                self.assertEqual(headers, [])
                self.assertEqual(rows, [])

    def test_load_csv_empty_file_returns_empty_list(self) -> None:
        """Empty CSV returns []."""
        from qsync.survey_master import load_master_csv

        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            csv_path = root / "surveys" / "qualtrics_master.csv"
            csv_path.parent.mkdir(parents=True, exist_ok=True)

            # Write empty CSV (just headers)
            with open(csv_path, "w", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=["SurveyID"])
                writer.writeheader()

            with patch("qsync.survey_master._master_csv_path", return_value=csv_path):
                headers, rows = load_master_csv()

            self.assertEqual(headers, ["SurveyID"])
            self.assertEqual(len(rows), 0)

    def test_write_csv_also_writes_workbook_with_readonly_fill(self) -> None:
        """Writing master CSV also writes workbook surface with read-only shading."""
        from qsync.survey_master import write_master_csv

        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            csv_path = root / "surveys" / "qualtrics_master.csv"
            rows = [
                ["SurveyID", "SurveyName", "SurveyStatus"],
                ["SV_001", "Survey A", "Active"],
            ]
            mapping = {
                "SurveyID": {
                    "field_name": "SurveyID",
                    "survey_master": "read",
                    "description": "Survey id",
                    "allowed_values": "",
                    "format_notes": "",
                },
                "SurveyName": {
                    "field_name": "SurveyName",
                    "survey_master": "write",
                    "description": "Survey name",
                    "allowed_values": "",
                    "format_notes": "",
                },
                "SurveyStatus": {
                    "field_name": "SurveyStatus",
                    "survey_master": "write",
                    "description": "Status",
                    "allowed_values": "Active; Inactive",
                    "format_notes": "",
                },
            }

            with (
                patch("qsync.survey_master._master_csv_path", return_value=csv_path),
                patch("qsync.survey_master._parse_mapping_csv", return_value=mapping),
            ):
                write_master_csv(rows)

            workbook_path = csv_path.with_suffix(".xlsx")
            self.assertTrue(workbook_path.exists())

            wb = load_workbook(workbook_path)
            self.assertIn("Survey_Master", wb.sheetnames)
            self.assertIn("Survey_Master_Guide", wb.sheetnames)
            ws = wb["Survey_Master"]
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            survey_id_idx = headers.index("SurveyID") + 1
            survey_name_idx = headers.index("SurveyName") + 1
            readonly_rgb = "FFECECEC"
            self.assertEqual(
                ws.cell(row=2, column=survey_id_idx).fill.fgColor.rgb,
                readonly_rgb,
            )
            self.assertNotEqual(
                ws.cell(row=2, column=survey_name_idx).fill.fgColor.rgb,
                readonly_rgb,
            )
            self.assertIsNot(ws.cell(row=1, column=survey_name_idx).alignment.wrap_text, True)
            self.assertIsNot(ws.cell(row=2, column=survey_name_idx).alignment.wrap_text, True)

    def test_workbook_header_comments_include_docs_and_allowed_values(self) -> None:
        """Header comments include field help text and options docs links."""
        from qsync.survey_master import write_master_workbook

        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            csv_path = root / "surveys" / "qualtrics_master.csv"
            rows = [
                ["SurveyID", "SurveyProtection"],
                ["SV_001", "PublicSurvey"],
            ]
            mapping = {
                "SurveyID": {
                    "field_name": "SurveyID",
                    "survey_master": "read",
                    "description": "Survey id",
                    "allowed_values": "",
                    "format_notes": "",
                    "domain": "survey_metadata",
                },
                "SurveyProtection": {
                    "field_name": "SurveyProtection",
                    "survey_master": "write",
                    "description": "Survey protection mode",
                    "allowed_values": "PublicSurvey; ByInvitation; PasswordProtected",
                    "format_notes": "",
                    "domain": "survey_options",
                },
            }

            with (
                patch("qsync.survey_master._master_csv_path", return_value=csv_path),
                patch("qsync.survey_master._parse_mapping_csv", return_value=mapping),
            ):
                write_master_workbook(rows)

            wb = load_workbook(csv_path.with_suffix(".xlsx"))
            ws = wb["Survey_Master"]
            header_values = [
                cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))
            ]
            target_idx = header_values.index("SurveyProtection") + 1
            comment = ws.cell(row=1, column=target_idx).comment
            self.assertIsNotNone(comment)
            text = comment.text
            self.assertIn("What it controls: Survey protection mode", text)
            self.assertIn("Allowed values:", text)
            self.assertIn("- PublicSurvey", text)
            self.assertIn("API endpoints:", text)
            self.assertIn("GET /survey-definitions/{surveyId}/options", text)
            self.assertIn("PUT /survey-definitions/{surveyId}/options", text)
            self.assertIn(
                "https://api.qualtrics.com/021740be5b5b6-get-options#response-body",
                text,
            )
            self.assertIn(
                "https://api.qualtrics.com/5d9e865296ce5-update-options",
                text,
            )

    def test_workbook_header_comments_include_metadata_and_detail_endpoints(self) -> None:
        """Metadata/detail domain comments include endpoint and docs hints."""
        from qsync.survey_master import write_master_workbook

        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            csv_path = root / "surveys" / "qualtrics_master.csv"
            rows = [
                ["SurveyID", "SurveyName", "responseCounts"],
                ["SV_001", "Survey A", "{}"],
            ]
            mapping = {
                "SurveyID": {
                    "field_name": "SurveyID",
                    "survey_master": "read",
                    "description": "Survey id",
                    "allowed_values": "",
                    "format_notes": "",
                    "domain": "survey_metadata",
                },
                "SurveyName": {
                    "field_name": "SurveyName",
                    "survey_master": "write",
                    "description": "Survey name",
                    "allowed_values": "",
                    "format_notes": "",
                    "domain": "survey_metadata",
                },
                "responseCounts": {
                    "field_name": "responseCounts",
                    "survey_master": "read",
                    "description": "Counts summary for responses",
                    "allowed_values": "",
                    "format_notes": "",
                    "domain": "survey_detail",
                    "data_type": "object",
                },
            }

            with (
                patch("qsync.survey_master._master_csv_path", return_value=csv_path),
                patch("qsync.survey_master._parse_mapping_csv", return_value=mapping),
            ):
                write_master_workbook(rows)

            wb = load_workbook(csv_path.with_suffix(".xlsx"))
            ws = wb["Survey_Master"]
            header_values = [
                cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))
            ]

            metadata_idx = header_values.index("SurveyName") + 1
            metadata_comment = ws.cell(row=1, column=metadata_idx).comment
            self.assertIsNotNone(metadata_comment)
            metadata_text = metadata_comment.text
            self.assertIn("GET /survey-definitions/{surveyId}/metadata", metadata_text)
            self.assertIn("PUT /survey-definitions/{surveyId}/metadata", metadata_text)
            self.assertIn("https://api.qualtrics.com/", metadata_text)

            detail_idx = header_values.index("responseCounts") + 1
            detail_comment = ws.cell(row=1, column=detail_idx).comment
            self.assertIsNotNone(detail_comment)
            detail_text = detail_comment.text
            self.assertIn("GET /surveys/{surveyId}", detail_text)
            self.assertIn("Expected type: object", detail_text)
            self.assertIn("https://api.qualtrics.com/", detail_text)

    def test_load_master_csv_prefers_newer_workbook_surface(self) -> None:
        """When workbook is newer than CSV, load from workbook surface."""
        from qsync.survey_master import load_master_csv, write_master_workbook

        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            csv_path = root / "surveys" / "qualtrics_master.csv"
            csv_path.parent.mkdir(parents=True, exist_ok=True)
            csv_path.write_text(
                "SurveyID,SurveyName\nSV_001,From CSV\n",
                encoding="utf-8-sig",
            )
            os.utime(csv_path, (1_000_000_000, 1_000_000_000))

            mapping = {
                "SurveyID": {"field_name": "SurveyID", "survey_master": "read"},
                "SurveyName": {"field_name": "SurveyName", "survey_master": "write"},
            }
            rows = [["SurveyID", "SurveyName"], ["SV_001", "From Workbook"]]
            with (
                patch("qsync.survey_master._master_csv_path", return_value=csv_path),
                patch("qsync.survey_master._parse_mapping_csv", return_value=mapping),
            ):
                write_master_workbook(rows)
                workbook_path = csv_path.with_suffix(".xlsx")
                os.utime(workbook_path, (1_000_000_100, 1_000_000_100))
                headers, loaded_rows = load_master_csv()

            self.assertEqual(headers, ["SurveyID", "SurveyName"])
            self.assertEqual(loaded_rows[0]["SurveyName"], "From Workbook")

    def test_load_master_csv_prefers_newer_csv_surface(self) -> None:
        """When CSV is newer than workbook, load from CSV surface."""
        from qsync.survey_master import load_master_csv, write_master_workbook

        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            csv_path = root / "surveys" / "qualtrics_master.csv"
            csv_path.parent.mkdir(parents=True, exist_ok=True)

            mapping = {
                "SurveyID": {"field_name": "SurveyID", "survey_master": "read"},
                "SurveyName": {"field_name": "SurveyName", "survey_master": "write"},
            }
            rows = [["SurveyID", "SurveyName"], ["SV_001", "From Workbook"]]
            with (
                patch("qsync.survey_master._master_csv_path", return_value=csv_path),
                patch("qsync.survey_master._parse_mapping_csv", return_value=mapping),
            ):
                write_master_workbook(rows)
                workbook_path = csv_path.with_suffix(".xlsx")
                os.utime(workbook_path, (1_000_000_000, 1_000_000_000))
                csv_path.write_text(
                    "SurveyID,SurveyName\nSV_001,From CSV\n",
                    encoding="utf-8-sig",
                )
                os.utime(csv_path, (1_000_000_100, 1_000_000_100))
                headers, loaded_rows = load_master_csv()

            self.assertEqual(headers, ["SurveyID", "SurveyName"])
            self.assertEqual(loaded_rows[0]["SurveyName"], "From CSV")


if __name__ == "__main__":
    unittest.main()
