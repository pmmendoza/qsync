from __future__ import annotations

import json
from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook

from qsync.dimensions.translations_core import apply_translations, push_translations
from qsync.pending_stage import load_pending, TranslationsPendingPayload
from qsync.qualtrics_client import SurveyCache
from qsync.workbook_resolver import WorkbookResolver
from qsync.excel_io import init_workbook_from_survey, QUESTION_SHEET
from qsync.excel_io import SURVEY_METADATA_SHEET
from qsync.excel_io import SBS_COLUMNS_SHEET, SBS_COLUMN_ANSWERS_SHEET


def _survey_payload() -> dict:
    return {
        "result": {
            "SurveyID": "SV_TEST",
            "SurveyOptions": {
                "SurveyLanguage": "EN",
                "SurveyTitle": "TestSurvey",
                "AvailableLanguages": {"EN": True, "FR": True},
                "MetaDataTranslations": {
                    "FR": {"SurveyMetaDescription": "Description FR old"},
                },
            },
            "SurveyFlow": {"Flow": [{"Type": "Standard", "ID": "BL_1"}]},
            "Blocks": {
                "BL_1": {
                    "Type": "Default",
                    "Description": "Block 1",
                    "BlockElements": [{"Type": "Question", "QuestionID": "QID1"}],
                }
            },
            "Questions": {
                "QID1": {
                    "QuestionType": "MC",
                    "Selector": "SAVR",
                    "QuestionText": "Base question",
                    "DataExportTag": "q1",
                    "Choices": {"1": {"Display": "Choice 1"}},
                    "ChoiceOrder": ["1"],
                    "Answers": {"1": {"Display": "Answer 1"}},
                    "AnswerOrder": ["1"],
                }
            },
        }
    }


def _sbs_survey_payload() -> dict:
    return {
        "result": {
            "SurveyID": "SV_TEST",
            "SurveyOptions": {
                "SurveyLanguage": "EN",
                "SurveyTitle": "TestSurvey",
                "AvailableLanguages": {"EN": True, "FR": True},
            },
            "SurveyFlow": {"Flow": [{"Type": "Standard", "ID": "BL_1"}]},
            "Blocks": {
                "BL_1": {
                    "Type": "Default",
                    "Description": "Block 1",
                    "BlockElements": [{"Type": "Question", "QuestionID": "QID15"}],
                }
            },
            "Questions": {
                "QID15": {
                    "QuestionType": "SBS",
                    "Selector": "SBSMatrix",
                    "QuestionText": "Base SBS question",
                    "DataExportTag": "q15",
                    "Choices": {"1": {"Display": "Statement 1"}},
                    "ChoiceOrder": ["1"],
                    "AdditionalQuestions": {
                        "1": {
                            "QuestionText": "Definitely didn't happen",
                            "Answers": {
                                "1": {"Display": "Answer 1"},
                                "2": {"Display": "Answer 2"},
                            },
                            "AnswerOrder": ["1", "2"],
                        },
                        "2": {
                            "QuestionText": "Probably didn't happen",
                            "Answers": {
                                "1": {"Display": "Answer 1"},
                                "2": {"Display": "Answer 2"},
                                "3": {"Display": "Answer 3"},
                            },
                            "AnswerOrder": ["1", "2", "3"],
                        },
                    },
                    "Language": {
                        "FR": {
                            "AdditionalQuestions": {
                                "1": {
                                    "QuestionText": "Certainement pas arrivé old",
                                    "Answers": {
                                        "1": {"Display": "Réponse 1 old"},
                                        "2": {"Display": "Réponse 2 old"},
                                    },
                                },
                                "2": {
                                    "QuestionText": "Probablement pas arrivé old",
                                    "Answers": {
                                        "1": {"Display": "Réponse 1 col2 old"},
                                        "2": {"Display": "Réponse 2 col2 old"},
                                        "3": {"Display": "Réponse 3 col2 old"},
                                    },
                                },
                            }
                        }
                    },
                }
            },
        }
    }


def _write_inventory(tmp_path: Path) -> None:
    surveys_dir = tmp_path / "surveys"
    surveys_dir.mkdir(parents=True, exist_ok=True)
    inventory = surveys_dir / "inventory.csv"
    inventory.write_text(
        "id,name,lastModified\nSV_TEST,TestSurvey,2026-01-23T12:00:00Z\n",
        encoding="utf-8",
    )


def _write_cached_survey(tmp_path: Path, payload: dict) -> Path:
    surveys_dir = tmp_path / "surveys"
    surveys_dir.mkdir(parents=True, exist_ok=True)
    path = surveys_dir / "TestSurvey__SV_TEST.json"
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    return path


def _write_workbook(tmp_path: Path, payload: dict) -> Path:
    resolver = WorkbookResolver()
    workbook_path = resolver.default_path("SV_TEST")
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    init_workbook_from_survey("SV_TEST", payload, workbook_path, languages=["FR"])
    wb = load_workbook(workbook_path)
    ws = wb[QUESTION_SHEET]
    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    qid_idx = headers.index("QID") + 1
    text_header = next(
        (
            str(header)
            for header in headers
            if str(header or "").lower() == "text_fr"
            or (
                str(header or "").startswith("Text_fr_")
                and str(header or "").endswith("_MD")
            )
        ),
        None,
    )
    assert text_header is not None
    text_idx = headers.index(text_header) + 1
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row=row, column=qid_idx).value or "").strip() == "QID1":
            ws.cell(row=row, column=text_idx).value = "Bonjour"
            break
    wb.save(workbook_path)
    return workbook_path


def _write_metadata_workbook(tmp_path: Path, payload: dict) -> Path:
    resolver = WorkbookResolver()
    workbook_path = resolver.default_path("SV_TEST")
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    init_workbook_from_survey("SV_TEST", payload, workbook_path, languages=["FR"])
    wb = load_workbook(workbook_path)
    ws = wb[SURVEY_METADATA_SHEET]
    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    lang_idx = headers.index("Language") + 1
    desc_idx = headers.index("SurveyDescription_MD") + 1
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row=row, column=lang_idx).value or "").strip() == "FR":
            ws.cell(row=row, column=desc_idx).value = "Description FR new"
            break
    wb.save(workbook_path)
    return workbook_path


def _write_sbs_workbook(tmp_path: Path, payload: dict) -> Path:
    resolver = WorkbookResolver()
    workbook_path = resolver.default_path("SV_TEST")
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    init_workbook_from_survey("SV_TEST", payload, workbook_path, languages=["FR"])
    wb = load_workbook(workbook_path)

    ws_cols = wb[SBS_COLUMNS_SHEET]
    col_headers = [cell.value for cell in next(ws_cols.iter_rows(max_row=1))]
    col_qid_idx = col_headers.index("QID") + 1
    col_id_idx = col_headers.index("ColumnId") + 1
    col_text_idx = col_headers.index("Label_fr_MD") + 1
    for row in range(2, ws_cols.max_row + 1):
        if str(ws_cols.cell(row=row, column=col_qid_idx).value or "").strip() != "QID15":
            continue
        if str(ws_cols.cell(row=row, column=col_id_idx).value or "").strip() != "1":
            continue
        ws_cols.cell(row=row, column=col_text_idx).value = "Certainement pas arrivé new"
        break

    ws_ans = wb[SBS_COLUMN_ANSWERS_SHEET]
    ans_headers = [cell.value for cell in next(ws_ans.iter_rows(max_row=1))]
    ans_qid_idx = ans_headers.index("QID") + 1
    ans_col_idx = ans_headers.index("ColumnId") + 1
    ans_id_idx = ans_headers.index("AnswerId") + 1
    ans_text_idx = ans_headers.index("Label_fr_MD") + 1
    for row in range(2, ws_ans.max_row + 1):
        if str(ws_ans.cell(row=row, column=ans_qid_idx).value or "").strip() != "QID15":
            continue
        if str(ws_ans.cell(row=row, column=ans_col_idx).value or "").strip() != "2":
            continue
        if str(ws_ans.cell(row=row, column=ans_id_idx).value or "").strip() != "3":
            continue
        ws_ans.cell(row=row, column=ans_text_idx).value = "Réponse 3 col2 new"
        break

    wb.save(workbook_path)
    return workbook_path


def test_apply_translations_writes_pending_payload(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setenv("QSYNC_ROOT", str(tmp_path))
    payload = _survey_payload()
    _write_inventory(tmp_path)
    survey_path = _write_cached_survey(tmp_path, payload)
    _write_workbook(tmp_path, payload)

    from qsync import drift_check
    from qsync.dimensions import translations_core

    monkeypatch.setattr(
        drift_check,
        "check_drift",
        lambda *args, **kwargs: drift_check.DriftReport(
            has_drift=False,
            summary="",
            diff_lines=[],
            recommendation="",
            changed_count=0,
        ),
    )
    monkeypatch.setattr(translations_core, "ensure_backup", lambda survey_id: None)
    monkeypatch.setattr(
        translations_core,
        "_inventory_last_modified",
        lambda survey_id: "2026-01-23T12:00:00Z",
    )

    record = apply_translations(
        "SV_TEST",
        ["FR"],
        allow_drift=True,
        interactive=False,
    )
    assert record is not None
    assert isinstance(record.payload, TranslationsPendingPayload)
    assert record.payload.qids == ["QID1"]
    assert record.payload.languages == ["FR"]
    assert record.payload.staged_last_modified == "2026-01-23T12:00:00Z"

    pending = load_pending("SV_TEST", "translations")
    assert pending is not None
    assert pending.payload.qids == ["QID1"]
    assert pending.payload.languages == ["FR"]

    updated = json.loads(survey_path.read_text(encoding="utf-8"))
    # Staging should not mutate cached survey JSON.
    lang_block = (
        updated.get("result", {})
        .get("Questions", {})
        .get("QID1", {})
        .get("Language", {})
        .get("FR", {})
    )
    assert lang_block == {}


def test_apply_translations_allows_html_attributes(
    tmp_path: Path, monkeypatch
) -> None:
    monkeypatch.setenv("QSYNC_ROOT", str(tmp_path))
    payload = _survey_payload()
    _write_inventory(tmp_path)
    _write_cached_survey(tmp_path, payload)

    resolver = WorkbookResolver()
    workbook_path = resolver.default_path("SV_TEST")
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    init_workbook_from_survey("SV_TEST", payload, workbook_path, languages=["FR"])
    wb = load_workbook(workbook_path)
    ws = wb[QUESTION_SHEET]
    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    qid_idx = headers.index("QID") + 1
    text_header = next(
        (
            str(header)
            for header in headers
            if str(header or "").lower() == "text_fr"
            or (
                str(header or "").startswith("Text_fr_")
                and str(header or "").endswith("_MD")
            )
        ),
        None,
    )
    assert text_header is not None
    text_idx = headers.index(text_header) + 1
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row=row, column=qid_idx).value or "").strip() != "QID1":
            continue
        ws.cell(row=row, column=text_idx).value = (
            '<a href="#" onclick="return false">Bonjour</a>'
        )
        break
    wb.save(workbook_path)

    from qsync import drift_check
    from qsync.dimensions import translations_core

    monkeypatch.setattr(
        drift_check,
        "check_drift",
        lambda *args, **kwargs: drift_check.DriftReport(
            has_drift=False,
            summary="",
            diff_lines=[],
            recommendation="",
            changed_count=0,
        ),
    )
    monkeypatch.setattr(translations_core, "ensure_backup", lambda survey_id: None)
    monkeypatch.setattr(
        translations_core,
        "_inventory_last_modified",
        lambda survey_id: "2026-01-23T12:00:00Z",
    )

    record = apply_translations(
        "SV_TEST",
        ["FR"],
        allow_drift=True,
        interactive=False,
    )
    assert record is not None
    assert record.payload.qids == ["QID1"]


def test_push_translations_includes_language_blocks(
    tmp_path: Path, monkeypatch
) -> None:
    monkeypatch.setenv("QSYNC_ROOT", str(tmp_path))
    payload = _survey_payload()
    _write_inventory(tmp_path)
    _write_cached_survey(tmp_path, payload)
    _write_workbook(tmp_path, payload)

    from qsync import drift_check
    from qsync.dimensions import translations_core

    monkeypatch.setattr(
        drift_check,
        "check_drift",
        lambda *args, **kwargs: drift_check.DriftReport(
            has_drift=False,
            summary="",
            diff_lines=[],
            recommendation="",
            changed_count=0,
        ),
    )
    monkeypatch.setattr(translations_core, "ensure_backup", lambda survey_id: None)

    apply_translations(
        "SV_TEST",
        ["FR"],
        allow_drift=True,
        interactive=False,
    )

    captured: dict[str, object] = {}

    def _fake_push_questions(survey: SurveyCache, qids, context=None) -> None:
        captured["qids"] = list(qids)
        captured["question"] = survey.questions.get("QID1")

    monkeypatch.setattr(translations_core, "push_questions", _fake_push_questions)
    monkeypatch.setattr(
        translations_core,
        "enforce_push_safeguards",
        lambda *args, **kwargs: SimpleNamespace(warnings=[]),
    )
    monkeypatch.setattr(
        translations_core, "auto_publish_after_push", lambda *args, **kwargs: None
    )

    pushed = push_translations(
        survey_id="SV_TEST",
        languages=["FR"],
        allow_drift=True,
        interactive=False,
        publish=False,
        force_live=True,
        force_preview=True,
    )
    assert pushed == ["QID1"]
    assert captured.get("qids") == ["QID1"]
    question = captured.get("question") or {}
    lang_block = (
        (question.get("Language") or {}).get("FR")
        if isinstance(question.get("Language"), dict)
        else {}
    )
    assert lang_block.get("QuestionText") == "Bonjour"


def test_apply_translations_tracks_metadata_keys(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setenv("QSYNC_ROOT", str(tmp_path))
    payload = _survey_payload()
    _write_inventory(tmp_path)
    survey_path = _write_cached_survey(tmp_path, payload)
    _write_metadata_workbook(tmp_path, payload)

    from qsync import drift_check
    from qsync.dimensions import translations_core

    monkeypatch.setattr(
        drift_check,
        "check_drift",
        lambda *args, **kwargs: drift_check.DriftReport(
            has_drift=False,
            summary="",
            diff_lines=[],
            recommendation="",
            changed_count=0,
        ),
    )
    monkeypatch.setattr(translations_core, "ensure_backup", lambda survey_id: None)
    monkeypatch.setattr(
        translations_core,
        "_inventory_last_modified",
        lambda survey_id: "2026-01-23T12:00:00Z",
    )

    record = apply_translations(
        "SV_TEST",
        ["FR"],
        allow_drift=True,
        interactive=False,
    )
    assert record is not None
    assert isinstance(record.payload, TranslationsPendingPayload)
    assert record.payload.metadata_keys == ["SurveyMetaDescription"]

    updated = json.loads(survey_path.read_text(encoding="utf-8"))
    # Staging should not mutate cached survey JSON.
    meta = (
        updated.get("result", {})
        .get("SurveyOptions", {})
        .get("MetaDataTranslations", {})
        .get("FR", {})
    )
    assert meta.get("SurveyMetaDescription") == "Description FR old"


def test_apply_translations_tracks_sbs_changes_under_parent_qid(
    tmp_path: Path, monkeypatch
) -> None:
    monkeypatch.setenv("QSYNC_ROOT", str(tmp_path))
    payload = _sbs_survey_payload()
    _write_inventory(tmp_path)
    _write_cached_survey(tmp_path, payload)
    _write_sbs_workbook(tmp_path, payload)

    from qsync import drift_check
    from qsync.dimensions import translations_core

    monkeypatch.setattr(
        drift_check,
        "check_drift",
        lambda *args, **kwargs: drift_check.DriftReport(
            has_drift=False,
            summary="",
            diff_lines=[],
            recommendation="",
            changed_count=0,
        ),
    )
    monkeypatch.setattr(translations_core, "ensure_backup", lambda survey_id: None)
    monkeypatch.setattr(
        translations_core,
        "_inventory_last_modified",
        lambda survey_id: "2026-01-23T12:00:00Z",
    )

    record = apply_translations(
        "SV_TEST",
        ["FR"],
        allow_drift=True,
        interactive=False,
    )
    assert record is not None
    assert isinstance(record.payload, TranslationsPendingPayload)
    assert record.payload.qids == ["QID15"]
    assert any(
        str(change.get("qid") or "") == "QID15#1"
        and str(change.get("field") or "") == "QuestionText"
        for change in record.payload.changes
    )
    assert any(
        str(change.get("qid") or "") == "QID15#2"
        and str(change.get("field") or "") == "Answer"
        and str(change.get("item_id") or "") == "3"
        for change in record.payload.changes
    )


def test_push_translations_writes_sbs_language_blocks(
    tmp_path: Path, monkeypatch
) -> None:
    monkeypatch.setenv("QSYNC_ROOT", str(tmp_path))
    payload = _sbs_survey_payload()
    _write_inventory(tmp_path)
    _write_cached_survey(tmp_path, payload)
    _write_sbs_workbook(tmp_path, payload)

    from qsync import drift_check
    from qsync.dimensions import translations_core

    monkeypatch.setattr(
        drift_check,
        "check_drift",
        lambda *args, **kwargs: drift_check.DriftReport(
            has_drift=False,
            summary="",
            diff_lines=[],
            recommendation="",
            changed_count=0,
        ),
    )
    monkeypatch.setattr(translations_core, "ensure_backup", lambda survey_id: None)

    apply_translations(
        "SV_TEST",
        ["FR"],
        allow_drift=True,
        interactive=False,
    )

    captured: dict[str, object] = {}

    def _fake_push_questions(survey: SurveyCache, qids, context=None) -> None:
        captured["qids"] = list(qids)
        captured["question"] = survey.questions.get("QID15")

    monkeypatch.setattr(translations_core, "push_questions", _fake_push_questions)
    monkeypatch.setattr(
        translations_core,
        "enforce_push_safeguards",
        lambda *args, **kwargs: SimpleNamespace(warnings=[]),
    )
    monkeypatch.setattr(
        translations_core, "auto_publish_after_push", lambda *args, **kwargs: None
    )

    pushed = push_translations(
        survey_id="SV_TEST",
        languages=["FR"],
        allow_drift=True,
        interactive=False,
        publish=False,
        force_live=True,
        force_preview=True,
    )

    assert pushed == ["QID15"]
    assert captured.get("qids") == ["QID15"]
    question = captured.get("question") or {}
    add_lang = (
        (question.get("Language") or {}).get("FR", {}).get("AdditionalQuestions", {})
        if isinstance((question.get("Language") or {}), dict)
        else {}
    )
    assert add_lang.get("1", {}).get("QuestionText") == "Certainement pas arrivé new"
    assert (
        add_lang.get("2", {}).get("Answers", {}).get("3", {}).get("Display")
        == "Réponse 3 col2 new"
    )
