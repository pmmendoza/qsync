from __future__ import annotations

import json
from pathlib import Path
from unittest.mock import patch

import pytest
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from qsync import excel_io
from qsync.dimensions import items as items_dimension
from qsync.sync_core import apply_changes, preview_changes


READONLY_RGB = "FFECECEC"


def _fill_rgb(cell) -> str | None:
    color = getattr(cell.fill, "fgColor", None)
    if color is None:
        return None
    return getattr(color, "rgb", None)


def _has_data_validation_formula_for_column(
    ws, header_index: dict[str, int], header_name: str, expected_formula: str
) -> bool:
    col_idx = header_index.get(header_name)
    if not col_idx:
        return False
    col_letter = get_column_letter(col_idx)
    expected_range = f"{col_letter}2:{col_letter}{ws.max_row}"
    expected_single = f"{col_letter}2"
    data_validations = ws.data_validations.dataValidation
    for dv in data_validations:
        if str(getattr(dv, "formula1", "")).strip() != expected_formula:
            continue
        sqref = str(getattr(dv, "sqref", ""))
        if expected_range in sqref or expected_single in sqref:
            return True
    return False


def _question_settings_payload() -> dict:
    return {
        "result": {
            "SurveyID": "SV_TEST",
            "SurveyOptions": {
                "SurveyLanguage": "EN",
                "AvailableLanguages": {"EN": True},
            },
            "SurveyFlow": {"Flow": [{"Type": "Standard", "ID": "BL_1"}]},
            "Blocks": {
                "BL_1": {
                    "Type": "Standard",
                    "ID": "BL_1",
                    "Description": "Block 1",
                    "BlockElements": [{"Type": "Question", "QuestionID": "QID1"}],
                }
            },
            "Questions": {
                "QID1": {
                    "QuestionType": "TE",
                    "Selector": "SL",
                    "SubSelector": "TX",
                    "QuestionText": "Base text",
                    "DataExportTag": "Q1",
                    "QuestionDescription": "Q1",
                    "Validation": {
                        "Settings": {
                            "ForceResponse": "ON",
                            "Type": "None",
                            "MinChars": "5",
                        }
                    },
                }
            },
        }
    }


def _all_surfaces_payload() -> dict:
    return {
        "result": {
            "SurveyID": "SV_SURFACES",
            "SurveyName": "Survey Name",
            "SurveyDescription": "Survey Description",
            "SurveyMetaDescription": "Meta Description",
            "SurveyOptions": {
                "SurveyLanguage": "EN",
                "AvailableLanguages": {"EN": True},
            },
            "SurveyFlow": {
                "Flow": [
                    {"Type": "Standard", "ID": "BL_1"},
                    {
                        "Type": "EmbeddedData",
                        "FlowID": "FL_1",
                        "EmbeddedData": [
                            {
                                "Field": "DEBUG",
                                "Type": "Custom",
                                "Value": "F",
                                "Description": "DEBUG",
                                "DataVisibility": [],
                                "AnalyzeText": False,
                            }
                        ],
                    },
                ]
            },
            "Blocks": {
                "BL_1": {
                    "Type": "Standard",
                    "ID": "BL_1",
                    "Description": "Main",
                    "BlockElements": [
                        {"Type": "Question", "QuestionID": "QID1"},
                        {"Type": "Question", "QuestionID": "QID2"},
                        {"Type": "Question", "QuestionID": "QID3"},
                    ],
                }
            },
            "Questions": {
                "QID1": {
                    "QuestionType": "MC",
                    "Selector": "SAVR",
                    "QuestionText": "MC text",
                    "DataExportTag": "MC1",
                    "Choices": {
                        "1": {"Display": "Yes"},
                        "2": {"Display": "No"},
                    },
                },
                "QID2": {
                    "QuestionType": "Matrix",
                    "Selector": "Likert",
                    "QuestionText": "Matrix text",
                    "DataExportTag": "M1",
                    "Choices": {
                        "1": {"Display": "Row 1"},
                        "2": {"Display": "Row 2"},
                    },
                    "Answers": {
                        "1": {"Display": "Low"},
                        "2": {"Display": "High"},
                    },
                },
                "QID3": {
                    "QuestionType": "SBS",
                    "Selector": "SBSMatrix",
                    "QuestionText": "SBS text",
                    "DataExportTag": "SBS1",
                    "Choices": {"1": {"Display": "Statement 1"}},
                    "ChoiceOrder": ["1"],
                    "AdditionalQuestions": {
                        "1": {
                            "QuestionText": "Column 1",
                            "Answers": {
                                "1": {"Display": "A1"},
                                "2": {"Display": "A2"},
                            },
                            "AnswerOrder": ["1", "2"],
                        }
                    },
                },
            },
        }
    }


def _minimal_payload_no_embedded_or_sbs() -> dict:
    return {
        "result": {
            "SurveyID": "SV_MIN",
            "SurveyOptions": {
                "SurveyLanguage": "EN",
                "AvailableLanguages": {"EN": True},
            },
            "SurveyFlow": {"Flow": [{"Type": "Standard", "ID": "BL_1"}]},
            "Blocks": {
                "BL_1": {
                    "Type": "Standard",
                    "ID": "BL_1",
                    "Description": "Block 1",
                    "BlockElements": [{"Type": "Question", "QuestionID": "QID1"}],
                }
            },
            "Questions": {
                "QID1": {
                    "QuestionType": "TE",
                    "Selector": "SL",
                    "SubSelector": "TX",
                    "QuestionText": "Hello",
                    "DataExportTag": "Q1",
                }
            },
        }
    }


def _fr_base_payload() -> dict:
    return {
        "result": {
            "SurveyID": "SV_FR",
            "SurveyOptions": {
                "SurveyLanguage": "FR",
                "AvailableLanguages": {"FR": True, "EN": True},
            },
            "SurveyFlow": {"Flow": [{"Type": "Standard", "ID": "BL_1"}]},
            "Blocks": {
                "BL_1": {
                    "Type": "Standard",
                    "ID": "BL_1",
                    "Description": "Block 1",
                    "BlockElements": [{"Type": "Question", "QuestionID": "QID1"}],
                }
            },
            "Questions": {
                "QID1": {
                    "QuestionType": "TE",
                    "Selector": "SL",
                    "SubSelector": "TX",
                    "QuestionText": "Texte FR",
                    "DataExportTag": "Q1",
                    "QuestionDescription": "Q1",
                    "Language": {"EN": {"QuestionText": "Text EN"}},
                }
            },
        }
    }


def _write_cached(root: Path, survey_id: str, payload: dict) -> Path:
    surveys_dir = root / "surveys"
    backups_dir = surveys_dir / "backups"
    surveys_dir.mkdir(parents=True, exist_ok=True)
    backups_dir.mkdir(parents=True, exist_ok=True)
    cached_path = surveys_dir / f"TEST__{survey_id}.json"
    backup_path = backups_dir / f"TEST__{survey_id}.json"
    cached_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    backup_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    return cached_path


def test_questions_sheet_includes_config_column_and_required_highlight(tmp_path: Path) -> None:
    payload = _question_settings_payload()
    xlsx_path = tmp_path / "SV_TEST.xlsx"
    excel_io.init_workbook_from_survey("SV_TEST", payload, xlsx_path)

    wb = load_workbook(xlsx_path)
    ws = wb[excel_io.QUESTION_SHEET]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {str(name): i + 1 for i, name in enumerate(headers)}

    for required_col in (
        "RequiredResponse",
        "ForceResponseMode",
        "ValidationType",
        "ValidationSettingsJSON",
        "RandomizationType",
        "RandomizationSettingsJSON",
        "QuestionConfigJSON",
    ):
        assert required_col in idx
    assert "QuestionKey" not in idx

    qid_col = idx["QID"]
    row_idx = None
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(row=r, column=qid_col).value or "").strip() == "QID1":
            row_idx = r
            break
    assert row_idx is not None

    assert ws.cell(row=row_idx, column=idx["RequiredResponse"]).value is True
    config = json.loads(ws.cell(row=row_idx, column=idx["QuestionConfigJSON"]).value)
    assert config["Validation"] == {
        "ForceResponse": "ON",
        "MinChars": "5",
        "Type": "None",
    }
    assert config["Randomization"] == {"Type": "None"}

    assert (
        _fill_rgb(ws.cell(row=row_idx, column=idx["OptionsPreview"])) == READONLY_RGB
    )
    assert (
        _fill_rgb(ws.cell(row=row_idx, column=idx["QuestionConfigJSON"]))
        == READONLY_RGB
    )
    assert (
        _fill_rgb(ws.cell(row=row_idx, column=idx["ForceResponseMode"]))
        != READONLY_RGB
    )
    assert (
        _fill_rgb(ws.cell(row=row_idx, column=idx["RandomizationType"]))
        != READONLY_RGB
    )

    assert _has_data_validation_formula_for_column(
        ws, idx, "ForceResponseMode", '"OFF,ON,RequestResponse"'
    )
    assert _has_data_validation_formula_for_column(
        ws, idx, "ValidationType", '"None,MinChoices,CustomValidation,ChoicesTotal"'
    )
    assert _has_data_validation_formula_for_column(
        ws, idx, "RandomizationType", '"None,All,Subset,Advanced"'
    )

    required_col_letter = get_column_letter(idx["RequiredResponse"])
    expected_formula = f"=${required_col_letter}2=TRUE"
    has_required_rule = False
    for rules in ws.conditional_formatting._cf_rules.values():  # type: ignore[attr-defined]
        for rule in rules:
            formulas = [str(f) for f in getattr(rule, "formula", [])]
            if expected_formula in formulas:
                has_required_rule = True
                break
        if has_required_rule:
            break
    assert has_required_rule


def test_questions_refresh_clears_stale_readonly_fill_for_force_response(
    tmp_path: Path,
) -> None:
    payload = _question_settings_payload()
    xlsx_path = tmp_path / "SV_TEST.xlsx"
    excel_io.init_workbook_from_survey("SV_TEST", payload, xlsx_path)

    wb = load_workbook(xlsx_path)
    ws = wb[excel_io.QUESTION_SHEET]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {str(name): i + 1 for i, name in enumerate(headers)}
    ws.cell(row=2, column=idx["ForceResponseMode"]).fill = PatternFill(
        fill_type="solid",
        fgColor=READONLY_RGB,
    )
    wb.save(xlsx_path)

    # Refresh formatting via workbook update.
    excel_io.init_workbook_from_survey("SV_TEST", payload, xlsx_path)

    wb = load_workbook(xlsx_path)
    ws = wb[excel_io.QUESTION_SHEET]
    assert _fill_rgb(ws.cell(row=2, column=idx["ForceResponseMode"])) != READONLY_RGB
    assert _fill_rgb(ws.cell(row=2, column=idx["RequiredResponse"])) == READONLY_RGB


def test_all_major_workbook_tables_mark_readonly_columns_gray(tmp_path: Path) -> None:
    payload = _all_surfaces_payload()
    xlsx_path = tmp_path / "SV_SURFACES.xlsx"
    excel_io.init_workbook_from_survey("SV_SURFACES", payload, xlsx_path)

    wb = load_workbook(xlsx_path)
    checks = [
        (excel_io.QUESTION_SHEET, "QID", "text_en"),
        (excel_io.OPTIONS_SHEET, "ChoiceId", "Label_en_MD"),
        (excel_io.SUBITEMS_SHEET, "AnswerId", "Label_en_MD"),
        (excel_io.SBS_COLUMNS_SHEET, "ColumnId", "Label_en_MD"),
        (excel_io.SBS_COLUMN_ANSWERS_SHEET, "ColumnId", "Label_en_MD"),
        (excel_io.SURVEY_METADATA_SHEET, "Language", "SurveyTitle_MD"),
        (excel_io.EMBEDDED_DATA_SHEET, "Field", "Value"),
    ]

    for sheet_name, readonly_col, editable_col in checks:
        ws = wb[sheet_name]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {str(name): i + 1 for i, name in enumerate(headers)}
        assert ws.max_row >= 2
        assert _fill_rgb(ws.cell(row=2, column=idx[readonly_col])) == READONLY_RGB
        assert _fill_rgb(ws.cell(row=2, column=idx[editable_col])) != READONLY_RGB


def test_subitems_self_heal_resets_system_columns_and_warns_for_orphans(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    payload = _all_surfaces_payload()
    survey_id = "SV_SURFACES"
    _write_cached(tmp_path, survey_id, payload)

    xlsx_path = tmp_path / "excel" / f"{survey_id}.xlsx"
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    excel_io.init_workbook_from_survey(survey_id, payload, xlsx_path)

    wb = load_workbook(xlsx_path)
    ws = wb[excel_io.SUBITEMS_SHEET]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {str(name): i for i, name in enumerate(headers)}

    target_row = None
    for r in range(2, ws.max_row + 1):
        qid = str(ws.cell(row=r, column=idx["QID"] + 1).value or "").strip()
        answer_id = str(ws.cell(row=r, column=idx["AnswerId"] + 1).value or "").strip()
        if qid == "QID2" and answer_id == "1":
            target_row = r
            break
    assert target_row is not None

    ws.cell(row=target_row, column=idx["SurveyID"] + 1).value = "SV_WRONG"
    ws.cell(row=target_row, column=idx["QuestionType"] + 1).value = "TE"
    ws.cell(row=target_row, column=idx["ExportTag"] + 1).value = "WRONG_TAG"

    orphan_row = [None] * len(headers)
    orphan_row[idx["SurveyID"]] = survey_id
    orphan_row[idx["QID"]] = "QID2"
    orphan_row[idx["AnswerId"]] = "999"
    orphan_row[idx["QuestionType"]] = "Matrix"
    orphan_row[idx["ExportTag"]] = "M1"
    if "Field" in idx:
        orphan_row[idx["Field"]] = "Answer"
    if "Label_en_MD" in idx:
        orphan_row[idx["Label_en_MD"]] = "orphan statement"
    ws.append(orphan_row)
    wb.save(xlsx_path)

    with patch("qsync.qualtrics_client._workspace_root", return_value=tmp_path):
        _ = preview_changes(
            survey_id,
            xlsx_path,
            check_drift=False,
            annotate_dirty=False,
        )

    out = capsys.readouterr().out
    assert "System column SurveyID changed in Subitems" in out
    assert "System column QuestionType changed in Subitems" in out
    assert "System column ExportTag changed in Subitems" in out
    assert "Subitems has 1 orphan row(s) with unknown AnswerId/LabelId" in out

    wb = load_workbook(xlsx_path)
    ws = wb[excel_io.SUBITEMS_SHEET]
    assert ws.cell(row=target_row, column=idx["SurveyID"] + 1).value == survey_id
    assert ws.cell(row=target_row, column=idx["QuestionType"] + 1).value == "Matrix"
    assert ws.cell(row=target_row, column=idx["ExportTag"] + 1).value == "M1"


def test_preview_and_apply_include_question_validation_settings(tmp_path: Path) -> None:
    payload = _question_settings_payload()
    survey_id = "SV_TEST"
    cached_path = _write_cached(tmp_path, survey_id, payload)

    xlsx_path = tmp_path / "excel" / f"{survey_id}.xlsx"
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    excel_io.init_workbook_from_survey(survey_id, payload, xlsx_path)

    wb = load_workbook(xlsx_path)
    ws = wb[excel_io.QUESTION_SHEET]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {str(name): i + 1 for i, name in enumerate(headers)}
    qid_col = idx["QID"]

    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row=row, column=qid_col).value or "").strip() == "QID1":
            ws.cell(row=row, column=idx["ForceResponseMode"]).value = "OFF"
            ws.cell(row=row, column=idx["ValidationType"]).value = "MinChoices"
            ws.cell(row=row, column=idx["ValidationSettingsJSON"]).value = (
                '{"MinChoices":"2"}'
            )
            ws.cell(row=row, column=idx["RandomizationType"]).value = "All"
            ws.cell(row=row, column=idx["RandomizationSettingsJSON"]).value = (
                '{"TotalRandSubset":"2"}'
            )
            break
    wb.save(xlsx_path)

    with patch("qsync.qualtrics_client._workspace_root", return_value=tmp_path):
        changes = preview_changes(
            survey_id,
            xlsx_path,
            check_drift=False,
            self_heal_system_columns=False,
            annotate_dirty=False,
        )
        setting_changes = [c for c in changes if c.kind == "question_setting"]
        randomization_changes = [
            c for c in changes if c.kind == "question_randomization"
        ]
        assert len(setting_changes) == 1
        assert len(randomization_changes) == 1

        result = apply_changes(
            survey_id,
            xlsx_path,
            allow_drift=True,
            interactive=False,
        )

    assert "QID1" in result.qids
    updated = json.loads(cached_path.read_text(encoding="utf-8"))
    settings = (
        updated.get("result", {})
        .get("Questions", {})
        .get("QID1", {})
        .get("Validation", {})
        .get("Settings", {})
    )
    assert settings.get("ForceResponse") == "OFF"
    assert settings.get("Type") == "MinChoices"
    assert settings.get("MinChoices") == "2"
    assert "MinChars" not in settings
    randomization = (
        updated.get("result", {}).get("Questions", {}).get("QID1", {}).get("Randomization")
    )
    assert randomization == {"Type": "All", "TotalRandSubset": "2"}


def test_questions_short_columns_are_center_aligned_and_html_text_is_italic(tmp_path: Path) -> None:
    payload = _question_settings_payload()
    xlsx_path = tmp_path / "SV_TEST.xlsx"
    excel_io.init_workbook_from_survey("SV_TEST", payload, xlsx_path)

    wb = load_workbook(xlsx_path)
    ws = wb[excel_io.QUESTION_SHEET]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {str(name): i + 1 for i, name in enumerate(headers)}

    qid_row = 2
    ws.cell(row=qid_row, column=idx["ishtml_en"]).value = True
    excel_io._format_questions_sheet(ws)
    wb.save(xlsx_path)

    wb = load_workbook(xlsx_path)
    ws = wb[excel_io.QUESTION_SHEET]
    assert ws.cell(row=1, column=idx["QID"]).alignment.horizontal == "center"
    assert ws.cell(row=2, column=idx["ValidationType"]).alignment.horizontal == "center"
    assert (
        ws.cell(row=2, column=idx["QuestionConfigJSON"]).alignment.horizontal == "left"
    )
    assert ws.cell(row=2, column=idx["ishtml_en"]).alignment.horizontal == "center"
    assert ws.cell(row=2, column=idx["text_en"]).font.italic is True


def test_init_hides_empty_sbs_and_embedded_sheets(tmp_path: Path) -> None:
    payload = _minimal_payload_no_embedded_or_sbs()
    xlsx_path = tmp_path / "SV_MIN.xlsx"
    excel_io.init_workbook_from_survey("SV_MIN", payload, xlsx_path)

    wb = load_workbook(xlsx_path)
    assert wb[excel_io.SBS_COLUMNS_SHEET].sheet_state == "hidden"
    assert wb[excel_io.SBS_COLUMN_ANSWERS_SHEET].sheet_state == "hidden"
    assert wb[excel_io.EMBEDDED_DATA_SHEET].sheet_state == "hidden"


def test_refresh_unhides_sbs_and_embedded_sheets_when_rows_exist(tmp_path: Path) -> None:
    xlsx_path = tmp_path / "SV_SURFACES.xlsx"
    excel_io.init_workbook_from_survey(
        "SV_SURFACES",
        _minimal_payload_no_embedded_or_sbs(),
        xlsx_path,
    )
    excel_io.init_workbook_from_survey("SV_SURFACES", _all_surfaces_payload(), xlsx_path)

    wb = load_workbook(xlsx_path)
    assert wb[excel_io.SBS_COLUMNS_SHEET].sheet_state == "visible"
    assert wb[excel_io.SBS_COLUMN_ANSWERS_SHEET].sheet_state == "visible"
    assert wb[excel_io.EMBEDDED_DATA_SHEET].sheet_state == "visible"

    sbs_ws = wb[excel_io.SBS_COLUMNS_SHEET]
    sbs_headers = [cell.value for cell in next(sbs_ws.iter_rows(min_row=1, max_row=1))]
    sbs_idx = {str(name): i + 1 for i, name in enumerate(sbs_headers)}
    assert sbs_ws.cell(row=1, column=sbs_idx["QID"]).alignment.horizontal == "center"
    assert (
        sbs_ws.cell(row=2, column=sbs_idx["Label_en_IsHTML"]).alignment.horizontal
        == "center"
    )


def test_preview_and_detect_changes_handle_non_en_base_without_en_text_column(
    tmp_path: Path,
) -> None:
    payload = _fr_base_payload()
    survey_id = "SV_FR"
    _write_cached(tmp_path, survey_id, payload)

    xlsx_path = tmp_path / "excel" / f"{survey_id}.xlsx"
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    excel_io.init_workbook_from_survey(
        survey_id,
        payload,
        xlsx_path,
        languages=["FR"],
    )

    with patch("qsync.qualtrics_client._workspace_root", return_value=tmp_path):
        changes = preview_changes(
            survey_id,
            xlsx_path,
            check_drift=False,
            annotate_dirty=False,
            self_heal_system_columns=False,
        )
        detected = items_dimension.detect_changes(survey_id)

    assert changes == []
    assert detected.has_changes is False


def test_load_questions_prefers_declared_base_language_column(tmp_path: Path) -> None:
    payload = _fr_base_payload()
    xlsx_path = tmp_path / "SV_FR.xlsx"
    excel_io.init_workbook_from_survey(
        "SV_FR",
        payload,
        xlsx_path,
        languages=["FR", "EN"],
    )

    wb = load_workbook(xlsx_path)
    ws = wb[excel_io.QUESTION_SHEET]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {str(name): i + 1 for i, name in enumerate(headers)}
    ws.cell(row=2, column=idx["text_fr"]).value = "BASE_FR"
    ws.cell(row=2, column=idx["text_en"]).value = "TRANSLATION_EN"
    wb.save(xlsx_path)

    row_fr = excel_io.load_questions_from_workbook(
        xlsx_path,
        base_language="FR",
    )["QID1"]
    row_en = excel_io.load_questions_from_workbook(
        xlsx_path,
        base_language="EN",
    )["QID1"]
    row_default = excel_io.load_questions_from_workbook(xlsx_path)["QID1"]

    assert row_fr.text_en_md == "BASE_FR"
    assert row_en.text_en_md == "TRANSLATION_EN"
    assert row_default.text_en_md == "BASE_FR"


def test_load_questions_raises_actionable_error_when_text_columns_missing(
    tmp_path: Path,
) -> None:
    payload = _question_settings_payload()
    xlsx_path = tmp_path / "SV_TEST.xlsx"
    excel_io.init_workbook_from_survey("SV_TEST", payload, xlsx_path)

    wb = load_workbook(xlsx_path)
    ws = wb[excel_io.QUESTION_SHEET]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    for name in ("text_en", "ishtml_en"):
        ws.delete_cols(headers.index(name) + 1, 1)
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    wb.save(xlsx_path)

    with pytest.raises(ValueError, match="Run `qsync items pull --survey-id"):
        excel_io.load_questions_from_workbook(xlsx_path, base_language="EN")
