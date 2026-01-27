from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from qsync import excel_io
from qsync.sync_core import ERROR_ID_EMBEDDED_DANGEROUS_SKIPPED, apply_changes


def _payload_with_empty_default(field: str) -> dict:
    return {
        "result": {
            "SurveyFlow": {
                "Flow": [
                    {
                        "Type": "EmbeddedData",
                        "FlowID": "FL_3",
                        "EmbeddedData": [
                            {
                                "Field": field,
                                "Type": "Custom",
                                "Value": "",
                                "Description": field,
                                "DataVisibility": [],
                                "AnalyzeText": False,
                            }
                        ],
                    }
                ]
            },
            "Questions": {},
            "Blocks": {},
        }
    }


def test_dangerous_embedded_default_is_skipped_with_guidance(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch, capsys: pytest.CaptureFixture[str]
) -> None:
    root = tmp_path
    surveys_dir = root / "surveys"
    backups_dir = surveys_dir / "backups"
    excel_dir = root / "excel"
    surveys_dir.mkdir(parents=True, exist_ok=True)
    backups_dir.mkdir(parents=True, exist_ok=True)
    excel_dir.mkdir(parents=True, exist_ok=True)
    (root / "survey_js").mkdir(parents=True, exist_ok=True)

    survey_id = "SV_TEST"
    payload = _payload_with_empty_default("DEBUG")
    cached_path = surveys_dir / f"TEST__{survey_id}.json"
    backup_path = backups_dir / f"TEST__{survey_id}.json"
    cached_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    backup_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")

    xlsx_path = excel_dir / f"{survey_id}.xlsx"
    excel_io.init_workbook_from_survey(survey_id, payload, xlsx_path)

    wb = load_workbook(xlsx_path)
    ws = wb[excel_io.EMBEDDED_DATA_SHEET]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: i for i, name in enumerate(headers)}
    field_idx = idx["Field"]
    value_idx = idx["Value"]
    for row in ws.iter_rows(min_row=2, values_only=False):
        if str(row[field_idx].value or "").strip() == "DEBUG":
            row[value_idx].value = "T"
            break
    wb.save(xlsx_path)

    # Ensure logging routes into the temp workspace.
    monkeypatch.setenv("QSYNC_ROOT", str(root))

    result = apply_changes(survey_id, xlsx_path, allow_dangerous=False, embedded_only=True)
    assert result.embedded_fields == []

    out = capsys.readouterr().out
    assert ERROR_ID_EMBEDDED_DANGEROUS_SKIPPED in out
    assert "--allow-dangerous" in out

    # Value should remain unset.
    updated = json.loads(cached_path.read_text(encoding="utf-8"))
    embedded = updated["result"]["SurveyFlow"]["Flow"][0]["EmbeddedData"][0]
    assert embedded["Value"] == ""

    log_path = root / "logs" / "qualtrics_push.log"
    assert log_path.exists()
    last = json.loads(log_path.read_text(encoding="utf-8").strip().splitlines()[-1])
    assert last["error"]["error_id"] == ERROR_ID_EMBEDDED_DANGEROUS_SKIPPED

