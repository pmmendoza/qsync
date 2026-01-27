"""Advanced tests for embedded data: dynamic patterns, comments, and rollback."""

from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from qsync import excel_io
from qsync.sync_core import apply_changes


class DynamicPatternDetectionTests(unittest.TestCase):
    """Tests for dynamic JS pattern detection (concatenation and template literals)."""

    def test_dynamic_pattern_concatenation(self) -> None:
        """Test detection of dynamic fields using string concatenation."""
        payload = {
            "result": {
                "SurveyFlow": {"Flow": []},
                "Questions": {
                    "QID1": {
                        "QuestionJS": """
                        for (let i = 0; i < 5; i++) {
                            Qualtrics.SurveyEngine.setEmbeddedData('field_' + i, value);
                        }
                        """
                    }
                },
                "Blocks": {},
            }
        }
        rows = excel_io.build_embedded_data_rows("SV_TEST", payload)
        by_field = {row.field: row for row in rows}

        self.assertIn("field_*", by_field)
        self.assertEqual(by_field["field_*"].written_by_qids, "QID1")
        self.assertEqual(by_field["field_*"].ed_type, "JS-only")

    def test_dynamic_pattern_template_literal(self) -> None:
        """Test detection of dynamic fields using template literals."""
        payload = {
            "result": {
                "SurveyFlow": {"Flow": []},
                "Questions": {
                    "QID2": {
                        "QuestionJS": """
                        for (let i = 0; i < 5; i++) {
                            Qualtrics.SurveyEngine.setEmbeddedData(`item_${i}_label`, label);
                        }
                        """
                    }
                },
                "Blocks": {},
            }
        }
        rows = excel_io.build_embedded_data_rows("SV_TEST", payload)
        by_field = {row.field: row for row in rows}

        self.assertIn("item_*", by_field)
        self.assertEqual(by_field["item_*"].written_by_qids, "QID2")

    def test_dynamic_pattern_multiple_prefixes(self) -> None:
        """Test that different prefixes create separate wildcard entries."""
        payload = {
            "result": {
                "SurveyFlow": {"Flow": []},
                "Questions": {
                    "QID1": {
                        "QuestionJS": """
                        setEmbeddedData('alpha_' + x, 1);
                        setEmbeddedData('beta_' + y, 2);
                        """
                    }
                },
                "Blocks": {},
            }
        }
        rows = excel_io.build_embedded_data_rows("SV_TEST", payload)
        by_field = {row.field: row for row in rows}

        self.assertIn("alpha_*", by_field)
        self.assertIn("beta_*", by_field)

    def test_static_and_dynamic_mixed(self) -> None:
        """Test that static and dynamic fields are both detected."""
        payload = {
            "result": {
                "SurveyFlow": {"Flow": []},
                "Questions": {
                    "QID1": {
                        "QuestionJS": """
                        setEmbeddedData('static_field', 'value');
                        setEmbeddedData('dynamic_' + i, value);
                        """
                    }
                },
                "Blocks": {},
            }
        }
        rows = excel_io.build_embedded_data_rows("SV_TEST", payload)
        by_field = {row.field: row for row in rows}

        self.assertIn("static_field", by_field)
        self.assertIn("dynamic_*", by_field)


class CommentStrippingTests(unittest.TestCase):
    """Tests for comment stripping in JS analysis."""

    def test_single_line_comments_ignored(self) -> None:
        """Test that single-line comments are ignored."""
        payload = {
            "result": {
                "SurveyFlow": {"Flow": []},
                "Questions": {
                    "QID1": {
                        "QuestionJS": """
                        // setEmbeddedData('commented_out', 1);
                        setEmbeddedData('active_field', 2);
                        """
                    }
                },
                "Blocks": {},
            }
        }
        rows = excel_io.build_embedded_data_rows("SV_TEST", payload)
        fields = [row.field for row in rows]

        self.assertIn("active_field", fields)
        self.assertNotIn("commented_out", fields)

    def test_block_comments_ignored(self) -> None:
        """Test that block comments are ignored."""
        payload = {
            "result": {
                "SurveyFlow": {"Flow": []},
                "Questions": {
                    "QID1": {
                        "QuestionJS": """
                        /*
                        setEmbeddedData('block_commented', 1);
                        */
                        setEmbeddedData('active_field', 2);
                        """
                    }
                },
                "Blocks": {},
            }
        }
        rows = excel_io.build_embedded_data_rows("SV_TEST", payload)
        fields = [row.field for row in rows]

        self.assertIn("active_field", fields)
        self.assertNotIn("block_commented", fields)

    def test_inline_block_comments_ignored(self) -> None:
        """Test that inline block comments are ignored."""
        payload = {
            "result": {
                "SurveyFlow": {"Flow": []},
                "Questions": {
                    "QID1": {
                        "QuestionJS": """
                        /* inline */ setEmbeddedData('after_comment', 1);
                        setEmbeddedData(/* mid */ 'mid_comment', 2);
                        """
                    }
                },
                "Blocks": {},
            }
        }
        rows = excel_io.build_embedded_data_rows("SV_TEST", payload)
        fields = [row.field for row in rows]

        self.assertIn("after_comment", fields)
        self.assertIn("mid_comment", fields)


class ApplyRollbackTests(unittest.TestCase):
    """Tests for atomic rollback on apply failure."""

    def test_apply_rollback_on_exception(self) -> None:
        """Test that cache is restored when apply fails."""
        payload = {
            "result": {
                "SurveyFlow": {
                    "Flow": [
                        {
                            "Type": "EmbeddedData",
                            "FlowID": "FL_1",
                            "EmbeddedData": [
                                {
                                    "Field": "test_field",
                                    "Type": "Custom",
                                    "Value": "original",
                                    "Description": "test_field",
                                    "DataVisibility": [],
                                    "AnalyzeText": False,
                                }
                            ],
                        }
                    ]
                },
                "Questions": {},
                "Blocks": {},
            }
        }

        with tempfile.TemporaryDirectory(prefix="qsync_test_") as tmpdir:
            root = Path(tmpdir)
            surveys_dir = root / "surveys"
            backups_dir = surveys_dir / "backups"
            excel_dir = root / "excel"
            surveys_dir.mkdir(parents=True, exist_ok=True)
            backups_dir.mkdir(parents=True, exist_ok=True)
            excel_dir.mkdir(parents=True, exist_ok=True)

            survey_id = "SV_TEST"
            cached_path = surveys_dir / f"TEST__{survey_id}.json"
            backup_path = backups_dir / f"TEST__{survey_id}.json"
            original_json = json.dumps(payload, indent=2)
            cached_path.write_text(original_json, encoding="utf-8")
            backup_path.write_text(original_json, encoding="utf-8")

            xlsx_path = excel_dir / f"{survey_id}.xlsx"
            excel_io.init_workbook_from_survey(survey_id, payload, xlsx_path)

            # Corrupt the Excel file to trigger an exception
            wb = load_workbook(xlsx_path)
            ws = wb[excel_io.EMBEDDED_DATA_SHEET]
            # Delete the Value column header to cause an error
            for cell in ws[1]:
                if cell.value == "Value":
                    cell.value = "CORRUPTED_HEADER"
                    break
            wb.save(xlsx_path)

            # Apply should fail and rollback
            with patch("qsync.qualtrics_client._workspace_root", return_value=root):
                with self.assertRaises(Exception):
                    apply_changes(survey_id, xlsx_path)

            # Verify that the cache was restored to original state
            restored_json = cached_path.read_text(encoding="utf-8")
            self.assertEqual(restored_json, original_json)

    def test_apply_success_modifies_cache(self) -> None:
        """Test that successful apply modifies the cache."""
        payload = {
            "result": {
                "SurveyFlow": {
                    "Flow": [
                        {
                            "Type": "EmbeddedData",
                            "FlowID": "FL_1",
                            "EmbeddedData": [
                                {
                                    "Field": "test_field",
                                    "Type": "Custom",
                                    "Value": "original",
                                    "Description": "test_field",
                                    "DataVisibility": [],
                                    "AnalyzeText": False,
                                }
                            ],
                        }
                    ]
                },
                "Questions": {},
                "Blocks": {},
            }
        }

        with tempfile.TemporaryDirectory(prefix="qsync_test_") as tmpdir:
            root = Path(tmpdir)
            surveys_dir = root / "surveys"
            backups_dir = surveys_dir / "backups"
            excel_dir = root / "excel"
            surveys_dir.mkdir(parents=True, exist_ok=True)
            backups_dir.mkdir(parents=True, exist_ok=True)
            excel_dir.mkdir(parents=True, exist_ok=True)

            survey_id = "SV_TEST"
            cached_path = surveys_dir / f"TEST__{survey_id}.json"
            backup_path = backups_dir / f"TEST__{survey_id}.json"
            original_json = json.dumps(payload, indent=2)
            cached_path.write_text(original_json, encoding="utf-8")
            backup_path.write_text(original_json, encoding="utf-8")

            xlsx_path = excel_dir / f"{survey_id}.xlsx"
            excel_io.init_workbook_from_survey(survey_id, payload, xlsx_path)

            # Edit the value
            wb = load_workbook(xlsx_path)
            ws = wb[excel_io.EMBEDDED_DATA_SHEET]
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            idx = {name: i for i, name in enumerate(headers)}
            field_idx = idx["Field"]
            value_idx = idx["Value"]

            for row in ws.iter_rows(min_row=2, values_only=False):
                field = str(row[field_idx].value or "").strip()
                if field == "test_field":
                    row[value_idx].value = "modified"
                    break
            wb.save(xlsx_path)

            # Apply should succeed
            with patch("qsync.qualtrics_client._workspace_root", return_value=root):
                result = apply_changes(survey_id, xlsx_path)

            self.assertEqual(len(result.embedded_fields), 1)

            # Verify cache was modified
            modified_json = cached_path.read_text(encoding="utf-8")
            self.assertNotEqual(modified_json, original_json)

            modified_payload = json.loads(modified_json)
            embedded_value = modified_payload["result"]["SurveyFlow"]["Flow"][0][
                "EmbeddedData"
            ][0]["Value"]
            self.assertEqual(embedded_value, "modified")


if __name__ == "__main__":
    unittest.main()
