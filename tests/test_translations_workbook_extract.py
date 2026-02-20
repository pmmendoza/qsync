from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from qsync.errors import QsyncValidationError
from qsync.excel_io import (
    OPTIONS_SHEET,
    QUESTION_SHEET,
    SBS_COLUMNS_SHEET,
    SBS_COLUMN_ANSWERS_SHEET,
    SUBITEMS_SHEET,
    SURVEY_METADATA_SHEET,
    init_workbook_from_survey,
)
from qsync.dimensions.translations_workbook_extract import (
    diff_workbook_vs_cache,
    extract_workbook_values,
)


def _survey_payload() -> dict:
    return {
        "result": {
            "SurveyID": "SV_TEST",
            "SurveyDescription": "Base description",
            "SurveyOptions": {
                "SurveyLanguage": "EN",
                "AvailableLanguages": ["EN", "FR"],
                "MetaDataTranslations": {
                    "FR": {
                        "SurveyTitle": "Titre FR old",
                        "SurveyMetaDescription": "Description FR old",
                    }
                },
            },
            "SurveyFlow": {"Flow": [{"Type": "Standard", "ID": "BL_1"}]},
            "Blocks": {
                "BL_1": {
                    "Type": "Default",
                    "Description": "Block 1",
                    "BlockElements": [{"Type": "Question", "QuestionID": "QID1"}],
                }
            },
            "Questions": {
                "QID1": {
                    "QuestionType": "Slider",
                    "Selector": "SL",
                    "QuestionText": "Base question",
                    "DataExportTag": "q1",
                    "Choices": {"1": {"Display": "Choice 1"}},
                    "ChoiceOrder": ["1"],
                    "Answers": {
                        "1": {"Display": "Answer 1"},
                        "2": {"Display": "Answer 2"},
                    },
                    "AnswerOrder": ["1", "2"],
                    "Labels": {
                        "1": {"Display": "Low"},
                        "2": {"Display": "High"},
                    },
                    "Language": {
                        "FR": {
                            "QuestionText": "Question FR old",
                            "Choices": {"1": {"Display": "Choix old"}},
                            "Answers": {"1": {"Display": "Réponse old"}},
                            "Labels": {"1": {"Display": "Bas old"}},
                        }
                    },
                }
            },
        }
    }


def _sbs_survey_payload() -> dict:
    return {
        "result": {
            "SurveyID": "SV_TEST",
            "SurveyOptions": {
                "SurveyLanguage": "EN",
                "AvailableLanguages": ["EN", "FR"],
            },
            "SurveyFlow": {"Flow": [{"Type": "Standard", "ID": "BL_1"}]},
            "Blocks": {
                "BL_1": {
                    "Type": "Default",
                    "Description": "Block 1",
                    "BlockElements": [{"Type": "Question", "QuestionID": "QID15"}],
                }
            },
            "Questions": {
                "QID15": {
                    "QuestionType": "SBS",
                    "Selector": "SBSMatrix",
                    "QuestionText": "Base SBS question",
                    "DataExportTag": "q15",
                    "Choices": {"1": {"Display": "Statement 1"}},
                    "ChoiceOrder": ["1"],
                    "AdditionalQuestions": {
                        "1": {
                            "QuestionText": "Definitely didn't happen",
                            "Answers": {
                                "1": {"Display": "Answer 1"},
                                "2": {"Display": "Answer 2"},
                            },
                            "AnswerOrder": ["1", "2"],
                        },
                        "2": {
                            "QuestionText": "Probably didn't happen",
                            "Answers": {
                                "1": {"Display": "Answer 1"},
                                "2": {"Display": "Answer 2"},
                                "3": {"Display": "Answer 3"},
                            },
                            "AnswerOrder": ["1", "2", "3"],
                        },
                    },
                    "Language": {
                        "FR": {
                            "QuestionText": "Question SBS FR old",
                            "Choices": {"1": {"Display": "Déclaration old"}},
                            "AdditionalQuestions": {
                                "1": {
                                    "QuestionText": "Certainement pas arrivé old",
                                    "Answers": {
                                        "1": {"Display": "Réponse 1 old"},
                                        "2": {"Display": "Réponse 2 old"},
                                    },
                                },
                                "2": {
                                    "QuestionText": "Probablement pas arrivé old",
                                    "Answers": {
                                        "1": {"Display": "Réponse 1 col2 old"},
                                        "2": {"Display": "Réponse 2 col2 old"},
                                        "3": {"Display": "Réponse 3 col2 old"},
                                    },
                                },
                            },
                        }
                    },
                }
            },
        }
    }


def _find_row(ws, qid: str) -> int:
    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    qid_idx = headers.index("QID") + 1
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row=row, column=qid_idx).value or "").strip() == qid:
            return row
    raise AssertionError(f"Row for {qid} not found")


def test_diff_workbook_vs_cache_extracts_changes(tmp_path: Path) -> None:
    payload = _survey_payload()
    workbook_path = tmp_path / "workbook.xlsx"
    init_workbook_from_survey("SV_TEST", payload, workbook_path, languages=["FR"])

    wb = load_workbook(workbook_path)

    q_ws = wb[QUESTION_SHEET]
    q_headers = [cell.value for cell in next(q_ws.iter_rows(max_row=1))]
    q_row = _find_row(q_ws, "QID1")
    q_text_idx = q_headers.index("text_fr") + 1
    q_ws.cell(row=q_row, column=q_text_idx).value = "Question FR new"

    o_ws = wb[OPTIONS_SHEET]
    o_headers = [cell.value for cell in next(o_ws.iter_rows(max_row=1))]
    o_qid_idx = o_headers.index("QID") + 1
    o_choice_idx = o_headers.index("ChoiceId") + 1
    o_label_idx = o_headers.index("Label_fr_MD") + 1
    for row in range(2, o_ws.max_row + 1):
        if str(o_ws.cell(row=row, column=o_qid_idx).value or "").strip() != "QID1":
            continue
        if str(o_ws.cell(row=row, column=o_choice_idx).value or "").strip() != "1":
            continue
        o_ws.cell(row=row, column=o_label_idx).value = "Choix new"

    s_ws = wb[SUBITEMS_SHEET]
    s_headers = [cell.value for cell in next(s_ws.iter_rows(max_row=1))]
    s_qid_idx = s_headers.index("QID") + 1
    s_answer_idx = s_headers.index("AnswerId") + 1
    s_field_idx = s_headers.index("Field") + 1
    s_label_idx = s_headers.index("Label_fr_MD") + 1
    for row in range(2, s_ws.max_row + 1):
        qid = str(s_ws.cell(row=row, column=s_qid_idx).value or "").strip()
        answer_id = str(s_ws.cell(row=row, column=s_answer_idx).value or "").strip()
        field = str(s_ws.cell(row=row, column=s_field_idx).value or "").strip()
        if qid != "QID1":
            continue
        if field == "Answer" and answer_id == "1":
            s_ws.cell(row=row, column=s_label_idx).value = "Réponse new"
        if field == "Label" and answer_id == "1":
            s_ws.cell(row=row, column=s_label_idx).value = "Bas new"

    wb.save(workbook_path)

    changes = diff_workbook_vs_cache(payload, workbook_path, ["FR"])
    assert len(changes) == 4
    assert any(change.field == "QuestionText" for change in changes)
    assert any(change.field == "Choice" for change in changes)
    assert any(change.field == "Answer" for change in changes)
    assert any(change.field == "Label" for change in changes)

    label_change = next(change for change in changes if change.field == "Label")
    assert label_change.item_id == "1"
    assert label_change.old_value == "Bas old"


def test_extract_workbook_values_rejects_duplicate_subitems(tmp_path: Path) -> None:
    payload = _survey_payload()
    workbook_path = tmp_path / "workbook.xlsx"
    init_workbook_from_survey("SV_TEST", payload, workbook_path, languages=["FR"])

    wb = load_workbook(workbook_path)
    ws = wb[SUBITEMS_SHEET]
    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    qid_idx = headers.index("QID") + 1
    answer_idx = headers.index("AnswerId") + 1
    field_idx = headers.index("Field") + 1

    duplicate_values = None
    for row in range(2, ws.max_row + 1):
        qid = str(ws.cell(row=row, column=qid_idx).value or "").strip()
        answer_id = str(ws.cell(row=row, column=answer_idx).value or "").strip()
        field = str(ws.cell(row=row, column=field_idx).value or "").strip()
        if qid == "QID1" and answer_id == "1" and field == "Answer":
            duplicate_values = [cell.value for cell in ws[row]]
            break

    assert duplicate_values is not None
    ws.append(duplicate_values)
    wb.save(workbook_path)

    with pytest.raises(QsyncValidationError):
        extract_workbook_values(workbook_path, ["FR"])


def test_extract_workbook_values_skips_missing_language_columns(tmp_path: Path) -> None:
    payload = _survey_payload()
    workbook_path = tmp_path / "workbook.xlsx"
    init_workbook_from_survey("SV_TEST", payload, workbook_path, languages=None)

    values = extract_workbook_values(workbook_path, ["FR"])
    assert values
    assert all(value.field == "Metadata" for value in values)


def test_diff_workbook_vs_cache_includes_metadata(tmp_path: Path) -> None:
    payload = _survey_payload()
    workbook_path = tmp_path / "workbook.xlsx"
    init_workbook_from_survey("SV_TEST", payload, workbook_path, languages=["FR"])

    wb = load_workbook(workbook_path)
    ws = wb[SURVEY_METADATA_SHEET]
    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    lang_idx = headers.index("Language") + 1
    desc_idx = headers.index("SurveyDescription_MD") + 1

    for row in range(2, ws.max_row + 1):
        lang = str(ws.cell(row=row, column=lang_idx).value or "").strip()
        if lang == "FR":
            ws.cell(row=row, column=desc_idx).value = "Description FR new"
            break

    wb.save(workbook_path)

    changes = diff_workbook_vs_cache(payload, workbook_path, ["FR"])
    assert any(change.field == "Metadata" for change in changes)
    change = next(change for change in changes if change.field == "Metadata")
    assert change.item_id == "SurveyDescription"
    assert change.old_value == "Description FR old"


def test_diff_workbook_vs_cache_ignores_markdown_roundtrip_noise(tmp_path: Path) -> None:
    payload = _survey_payload()
    # Simulate the real-world case where upstream HTML contains an empty `<strong> </strong>`
    # and a list followed by a `<br>`, which historically produced phantom diffs.
    payload["result"]["Questions"]["QID1"]["Language"]["FR"]["QuestionText"] = (
        "<ul><li>One</li><li>Two</li></ul><br><strong>After</strong> <strong> </strong>"
    )
    workbook_path = tmp_path / "workbook.xlsx"
    init_workbook_from_survey("SV_TEST", payload, workbook_path, languages=["FR"])

    # Force a "legacy" workbook cell representation:
    # - Blank line after list (list newline + `<br>` newline)
    # - Extra `****` from empty `<strong> </strong>`
    wb = load_workbook(workbook_path)
    q_ws = wb[QUESTION_SHEET]
    q_headers = [cell.value for cell in next(q_ws.iter_rows(max_row=1))]
    q_row = _find_row(q_ws, "QID1")
    q_text_idx = q_headers.index("text_fr") + 1
    q_ws.cell(row=q_row, column=q_text_idx).value = "- One\n- Two\n\n**After******"
    wb.save(workbook_path)

    changes = diff_workbook_vs_cache(payload, workbook_path, ["FR"])
    assert changes == []


def test_diff_workbook_vs_cache_ignores_markdown_style_only_noise(
    tmp_path: Path,
) -> None:
    payload = _survey_payload()
    payload["result"]["Questions"]["QID1"]["Language"]["FR"]["QuestionText"] = (
        "Wat is het <strong>hoogste opleidingsniveau dat je succesvol hebt afgerond</strong>? "
        "<em>(Als je nu studeert, kies dan het hoogste niveau dat je al hebt afgerond.)</em>"
    )
    workbook_path = tmp_path / "workbook.xlsx"
    init_workbook_from_survey("SV_TEST", payload, workbook_path, languages=["FR"])

    wb = load_workbook(workbook_path)
    q_ws = wb[QUESTION_SHEET]
    q_headers = [cell.value for cell in next(q_ws.iter_rows(max_row=1))]
    q_row = _find_row(q_ws, "QID1")

    text_header = next(
        (
            str(header)
            for header in q_headers
            if str(header or "").lower() == "text_fr"
            or (
                str(header or "").startswith("Text_fr_")
                and str(header or "").endswith("_MD")
            )
        ),
        None,
    )
    assert text_header is not None
    q_text_idx = q_headers.index(text_header) + 1
    q_ws.cell(row=q_row, column=q_text_idx).value = (
        "Wat is het **hoogste opleidingsniveau dat je succesvol hebt afgerond**? "
        "*(Als je nu studeert, kies dan het hoogste niveau dat je al hebt afgerond.)*"
    )
    wb.save(workbook_path)

    changes = diff_workbook_vs_cache(payload, workbook_path, ["FR"])
    assert changes == []


def test_diff_workbook_vs_cache_extracts_sbs_column_changes(tmp_path: Path) -> None:
    payload = _sbs_survey_payload()
    workbook_path = tmp_path / "workbook_sbs.xlsx"
    init_workbook_from_survey("SV_TEST", payload, workbook_path, languages=["FR"])

    wb = load_workbook(workbook_path)

    col_ws = wb[SBS_COLUMNS_SHEET]
    col_headers = [cell.value for cell in next(col_ws.iter_rows(max_row=1))]
    col_qid_idx = col_headers.index("QID") + 1
    col_id_idx = col_headers.index("ColumnId") + 1
    col_label_idx = col_headers.index("Label_fr_MD") + 1
    for row in range(2, col_ws.max_row + 1):
        if str(col_ws.cell(row=row, column=col_qid_idx).value or "").strip() != "QID15":
            continue
        if str(col_ws.cell(row=row, column=col_id_idx).value or "").strip() != "1":
            continue
        col_ws.cell(row=row, column=col_label_idx).value = "Certainement pas arrivé new"
        break

    ans_ws = wb[SBS_COLUMN_ANSWERS_SHEET]
    ans_headers = [cell.value for cell in next(ans_ws.iter_rows(max_row=1))]
    ans_qid_idx = ans_headers.index("QID") + 1
    ans_col_idx = ans_headers.index("ColumnId") + 1
    ans_id_idx = ans_headers.index("AnswerId") + 1
    ans_label_idx = ans_headers.index("Label_fr_MD") + 1
    for row in range(2, ans_ws.max_row + 1):
        if str(ans_ws.cell(row=row, column=ans_qid_idx).value or "").strip() != "QID15":
            continue
        if str(ans_ws.cell(row=row, column=ans_col_idx).value or "").strip() != "2":
            continue
        if str(ans_ws.cell(row=row, column=ans_id_idx).value or "").strip() != "3":
            continue
        ans_ws.cell(row=row, column=ans_label_idx).value = "Réponse 3 col2 new"
        break

    wb.save(workbook_path)

    changes = diff_workbook_vs_cache(payload, workbook_path, ["FR"])
    assert any(
        change.qid == "QID15#1" and change.field == "QuestionText" for change in changes
    )
    assert any(
        change.qid == "QID15#2"
        and change.field == "Answer"
        and change.item_id == "3"
        for change in changes
    )
