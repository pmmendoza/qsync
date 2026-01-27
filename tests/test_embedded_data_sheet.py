"""Tests for Embedded_Data worksheet creation."""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from qsync import excel_io


def _payload_with_recipient() -> dict:
    return {
        "result": {
            "SurveyFlow": {
                "Flow": [
                    {
                        "Type": "EmbeddedData",
                        "FlowID": "FL_9",
                        "EmbeddedData": [
                            {
                                "Field": "DEBUG",
                                "Type": "Custom",
                                "Value": "F",
                                "Description": "DEBUG",
                                "DataVisibility": [],
                                "AnalyzeText": False,
                            },
                            {
                                "Field": "PROLIFIC_PID",
                                "Type": "Recipient",
                                "Value": None,
                                "Description": "PROLIFIC_PID",
                                "DataVisibility": [],
                                "AnalyzeText": False,
                            },
                        ],
                    }
                ]
            },
            "Questions": {},
            "Blocks": {},
        }
    }


class EmbeddedDataSheetTests(unittest.TestCase):
    def test_embedded_data_sheet_contains_placeholder_value(self) -> None:
        payload = _payload_with_recipient()
        with tempfile.TemporaryDirectory(prefix="qsync_test_") as tmpdir:
            xlsx_path = Path(tmpdir) / "SV_TEST.xlsx"
            excel_io.init_workbook_from_survey("SV_TEST", payload, xlsx_path)

            wb = load_workbook(xlsx_path, data_only=True)
            self.assertIn(excel_io.EMBEDDED_DATA_SHEET, wb.sheetnames)
            ws = wb[excel_io.EMBEDDED_DATA_SHEET]
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            idx = {name: i for i, name in enumerate(headers)}
            field_idx = idx["Field"]
            value_idx = idx["Value"]

            values = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                field = str(row[field_idx] or "").strip()
                if field:
                    values[field] = row[value_idx]

            self.assertEqual(values.get("DEBUG"), "F")
            self.assertEqual(values.get("PROLIFIC_PID"), excel_io.EMBEDDED_EMPTY_VALUE)
