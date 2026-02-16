from __future__ import annotations

from dataclasses import dataclass
import html
from pathlib import Path
from typing import Any, Iterable, Literal

from openpyxl import load_workbook

from ..errors import QsyncValidationError
from ..excel_io import (
    OPTIONS_SHEET,
    QUESTION_SHEET,
    SBS_COLUMNS_SHEET,
    SBS_COLUMN_ANSWERS_SHEET,
    SUBITEMS_SHEET,
    SURVEY_METADATA_SHEET,
    _iter_sheet_rows,
    _language_from_suffix,
    _language_suffix,
)
from ..markdown_codec import html_to_md, md_to_html, normalize_markdown_for_compare, normalize_text
from ..scope_filter import ScopeFilter
from ..translations_utils import normalize_language_code, normalize_language_list
from .translations_language_blocks import (
    get_base_language as get_base_language_from_options,
    read_answer_display,
    read_choice_display,
    read_label_display,
    read_question_text,
    read_sbs_column_answer_display,
    read_sbs_column_question_text,
)

TranslationField = Literal["QuestionText", "Choice", "Answer", "Label", "Metadata"]
TranslationKey = tuple[str, TranslationField, str | None]

SURVEY_METADATA_QID = "SurveyMetadata"


@dataclass(frozen=True)
class WorkbookTranslationValue:
    qid: str
    language: str
    field: TranslationField
    item_id: str | None
    text: str
    is_html: bool

    @property
    def key(self) -> TranslationKey:
        return (self.qid, self.field, self.item_id)

    @property
    def html_value(self) -> str:
        if self.is_html:
            return normalize_text(self.text)
        return md_to_html(self.text)


@dataclass(frozen=True)
class TranslationChange:
    qid: str
    language: str
    field: TranslationField
    item_id: str | None
    old_value: str
    new_value: str


def resolve_languages_from_workbook(wb) -> list[str]:
    languages: list[str] = []
    seen: set[str] = set()
    for sheet_name, prefix in (
        (QUESTION_SHEET, "Text"),
        (OPTIONS_SHEET, "Label"),
        (SUBITEMS_SHEET, "Label"),
        (SBS_COLUMNS_SHEET, "Label"),
        (SBS_COLUMN_ANSWERS_SHEET, "Label"),
    ):
        if sheet_name not in wb.sheetnames:
            continue
        headers, _ = _iter_sheet_rows(wb[sheet_name])
        for name in headers:
            header = str(name or "")
            if not header.startswith(f"{prefix}_") or not header.endswith("_MD"):
                continue
            suffix = header[len(prefix) + 1 : -len("_MD")]
            code = _language_from_suffix(suffix)
            if not code or code in seen:
                continue
            seen.add(code)
            languages.append(code)
    if SURVEY_METADATA_SHEET in wb.sheetnames:
        headers, data_rows = _iter_sheet_rows(wb[SURVEY_METADATA_SHEET])
        if headers and "Language" in headers:
            lang_idx = headers.index("Language")
            for row in data_rows:
                raw = row[lang_idx].value if lang_idx < len(row) else None
                code = normalize_language_code(str(raw or ""))
                if not code or code in seen:
                    continue
                seen.add(code)
                languages.append(code)
    return normalize_language_list(languages)


def _normalize_translation_compare(value: str | None) -> str:
    text = html.unescape(str(value or ""))
    text = text.replace("\u00a0", " ")
    return normalize_text(text)


def _normalize_field(value: str | None) -> str:
    raw = str(value or "").strip()
    if not raw:
        return "Answer"
    lowered = raw.lower()
    if lowered == "answer":
        return "Answer"
    if lowered == "label":
        return "Label"
    raise QsyncValidationError(
        error_id="QSYNC-TRANS-INVALID-FIELD",
        problem=f"Invalid Field value '{raw}' in Subitems sheet.",
        why="Field must be 'Answer' or 'Label' to disambiguate subitems.",
        impact="Translations cannot be staged safely.",
        action="Set Field to Answer or Label and retry.",
    )


def _make_sbs_scoped_qid(qid: str, column_id: str) -> str:
    qid_s = str(qid or "").strip()
    column_s = str(column_id or "").strip()
    if not qid_s or not column_s:
        return ""
    return f"{qid_s}#{column_s}"


def _split_sbs_scoped_qid(qid: str) -> tuple[str, str] | None:
    raw = str(qid or "").strip()
    if "#" not in raw:
        return None
    base_qid, column_id = raw.split("#", 1)
    base_qid = base_qid.strip()
    column_id = column_id.strip()
    if not base_qid or not column_id:
        return None
    return base_qid, column_id


def _extract_question_values(
    wb,
    language: str,
    *,
    scope_qids: set[str] | None,
) -> list[WorkbookTranslationValue]:
    if QUESTION_SHEET not in wb.sheetnames:
        return []
    ws = wb[QUESTION_SHEET]
    headers, data_rows = _iter_sheet_rows(ws)
    if not headers or "QID" not in headers:
        return []
    qid_idx = headers.index("QID")
    suffix = _language_suffix(language)
    if not suffix:
        return []
    text_col = f"Text_{suffix}_MD"
    html_col = f"Text_{suffix}_IsHTML"
    if text_col not in headers:
        return []
    text_idx = headers.index(text_col)
    html_idx = headers.index(html_col) if html_col in headers else None

    values: list[WorkbookTranslationValue] = []
    for row in data_rows:
        qid_val = row[qid_idx].value
        qid = str(qid_val or "").strip()
        if not qid:
            continue
        if scope_qids is not None and qid not in scope_qids:
            continue
        cell = row[text_idx]
        raw = cell.value if cell is not None else None
        if raw is None or str(raw).strip() == "":
            continue
        is_html = False
        if html_idx is not None:
            html_cell = row[html_idx]
            is_html = bool(html_cell.value) if html_cell is not None else False
        text = normalize_text(str(raw)) if is_html else str(raw)
        values.append(
            WorkbookTranslationValue(
                qid=qid,
                language=language,
                field="QuestionText",
                item_id=None,
                text=text,
                is_html=is_html,
            )
        )
    return values


def _extract_option_values(
    wb,
    language: str,
    *,
    scope_qids: set[str] | None,
    question_rows: dict[str, Any] | None = None,
) -> list[WorkbookTranslationValue]:
    if OPTIONS_SHEET not in wb.sheetnames:
        return []
    ws = wb[OPTIONS_SHEET]
    headers, data_rows = _iter_sheet_rows(ws)
    if not headers or "QID" not in headers or "ChoiceId" not in headers:
        return []
    qid_idx = headers.index("QID")
    choice_idx = headers.index("ChoiceId")
    suffix = _language_suffix(language)
    if not suffix:
        return []
    text_col = f"Label_{suffix}_MD"
    html_col = f"Label_{suffix}_IsHTML"
    if text_col not in headers:
        return []
    text_idx = headers.index(text_col)
    html_idx = headers.index(html_col) if html_col in headers else None

    values: list[WorkbookTranslationValue] = []
    for row in data_rows:
        qid_val = row[qid_idx].value
        choice_val = row[choice_idx].value
        qid = str(qid_val or "").strip()
        choice_id = str(choice_val or "").strip()
        if not qid or not choice_id:
            continue
        if scope_qids is not None and qid not in scope_qids:
            continue
        qtype = ""
        if question_rows and qid in question_rows:
            qtype = str(getattr(question_rows[qid], "question_type", "") or "")
        cell = row[text_idx]
        raw = cell.value if cell is not None else None
        if raw is None or str(raw).strip() == "":
            continue
        is_html = False
        if html_idx is not None:
            html_cell = row[html_idx]
            is_html = bool(html_cell.value) if html_cell is not None else False
        text = normalize_text(str(raw)) if is_html else str(raw)
        field: TranslationField = "Choice"
        if qtype.lower() == "matrix":
            field = "Answer"
        values.append(
            WorkbookTranslationValue(
                qid=qid,
                language=language,
                field=field,
                item_id=choice_id,
                text=text,
                is_html=is_html,
            )
        )
    return values


def _extract_sbs_column_values(
    wb,
    language: str,
    *,
    scope_qids: set[str] | None,
) -> list[WorkbookTranslationValue]:
    if SBS_COLUMNS_SHEET not in wb.sheetnames:
        return []
    ws = wb[SBS_COLUMNS_SHEET]
    headers, data_rows = _iter_sheet_rows(ws)
    if not headers or "QID" not in headers or "ColumnId" not in headers:
        return []
    qid_idx = headers.index("QID")
    column_idx = headers.index("ColumnId")
    suffix = _language_suffix(language)
    if not suffix:
        return []
    text_col = f"Label_{suffix}_MD"
    html_col = f"Label_{suffix}_IsHTML"
    if text_col not in headers:
        return []
    text_idx = headers.index(text_col)
    html_idx = headers.index(html_col) if html_col in headers else None

    values: list[WorkbookTranslationValue] = []
    for row in data_rows:
        qid_val = row[qid_idx].value
        column_val = row[column_idx].value
        qid = str(qid_val or "").strip()
        column_id = str(column_val or "").strip()
        if not qid or not column_id:
            continue
        if scope_qids is not None and qid not in scope_qids:
            continue
        scoped_qid = _make_sbs_scoped_qid(qid, column_id)
        if not scoped_qid:
            continue
        cell = row[text_idx]
        raw = cell.value if cell is not None else None
        if raw is None or str(raw).strip() == "":
            continue
        is_html = False
        if html_idx is not None:
            html_cell = row[html_idx]
            is_html = bool(html_cell.value) if html_cell is not None else False
        text = normalize_text(str(raw)) if is_html else str(raw)
        values.append(
            WorkbookTranslationValue(
                qid=scoped_qid,
                language=language,
                field="QuestionText",
                item_id=None,
                text=text,
                is_html=is_html,
            )
        )
    return values


def _extract_sbs_column_answer_values(
    wb,
    language: str,
    *,
    scope_qids: set[str] | None,
) -> list[WorkbookTranslationValue]:
    if SBS_COLUMN_ANSWERS_SHEET not in wb.sheetnames:
        return []
    ws = wb[SBS_COLUMN_ANSWERS_SHEET]
    headers, data_rows = _iter_sheet_rows(ws)
    if (
        not headers
        or "QID" not in headers
        or "ColumnId" not in headers
        or "AnswerId" not in headers
    ):
        return []
    qid_idx = headers.index("QID")
    column_idx = headers.index("ColumnId")
    answer_idx = headers.index("AnswerId")
    suffix = _language_suffix(language)
    if not suffix:
        return []
    text_col = f"Label_{suffix}_MD"
    html_col = f"Label_{suffix}_IsHTML"
    if text_col not in headers:
        return []
    text_idx = headers.index(text_col)
    html_idx = headers.index(html_col) if html_col in headers else None

    values: list[WorkbookTranslationValue] = []
    for row in data_rows:
        qid_val = row[qid_idx].value
        column_val = row[column_idx].value
        answer_val = row[answer_idx].value
        qid = str(qid_val or "").strip()
        column_id = str(column_val or "").strip()
        answer_id = str(answer_val or "").strip()
        if not qid or not column_id or not answer_id:
            continue
        if scope_qids is not None and qid not in scope_qids:
            continue
        scoped_qid = _make_sbs_scoped_qid(qid, column_id)
        if not scoped_qid:
            continue
        cell = row[text_idx]
        raw = cell.value if cell is not None else None
        if raw is None or str(raw).strip() == "":
            continue
        is_html = False
        if html_idx is not None:
            html_cell = row[html_idx]
            is_html = bool(html_cell.value) if html_cell is not None else False
        text = normalize_text(str(raw)) if is_html else str(raw)
        values.append(
            WorkbookTranslationValue(
                qid=scoped_qid,
                language=language,
                field="Answer",
                item_id=answer_id,
                text=text,
                is_html=is_html,
            )
        )
    return values


def _extract_subitem_values(
    wb,
    language: str,
    *,
    scope_qids: set[str] | None,
    question_rows: dict[str, Any] | None = None,
) -> list[WorkbookTranslationValue]:
    if SUBITEMS_SHEET not in wb.sheetnames:
        return []
    ws = wb[SUBITEMS_SHEET]
    headers, data_rows = _iter_sheet_rows(ws)
    if not headers or "QID" not in headers or "AnswerId" not in headers:
        return []
    qid_idx = headers.index("QID")
    answer_idx = headers.index("AnswerId")
    field_idx = headers.index("Field") if "Field" in headers else None
    suffix = _language_suffix(language)
    if not suffix:
        return []
    text_col = f"Label_{suffix}_MD"
    html_col = f"Label_{suffix}_IsHTML"
    if text_col not in headers:
        return []
    text_idx = headers.index(text_col)
    html_idx = headers.index(html_col) if html_col in headers else None

    values: list[WorkbookTranslationValue] = []
    seen_keys: set[tuple[str, str, str]] = set()
    for row_idx, row in enumerate(data_rows, start=2):
        qid_val = row[qid_idx].value
        answer_val = row[answer_idx].value
        qid = str(qid_val or "").strip()
        answer_id = str(answer_val or "").strip()
        if not qid or not answer_id:
            continue
        if scope_qids is not None and qid not in scope_qids:
            continue
        field_val = row[field_idx].value if field_idx is not None else "Answer"
        field = _normalize_field(field_val)
        dedupe_key = (qid, field, answer_id)
        if dedupe_key in seen_keys:
            raise QsyncValidationError(
                error_id="QSYNC-TRANS-DUPLICATE-SUBITEM",
                problem=(
                    "Duplicate Subitems rows detected for "
                    f"QID={qid}, Field={field}, AnswerId={answer_id}."
                ),
                why="The Subitems sheet must be unique on (QID, Field, AnswerId).",
                impact="Translations cannot be staged safely.",
                action="Remove duplicate rows and retry.",
                context={"row": row_idx},
            )
        seen_keys.add(dedupe_key)
        cell = row[text_idx]
        raw = cell.value if cell is not None else None
        if raw is None or str(raw).strip() == "":
            continue
        is_html = False
        if html_idx is not None:
            html_cell = row[html_idx]
            is_html = bool(html_cell.value) if html_cell is not None else False
        text = normalize_text(str(raw)) if is_html else str(raw)
        effective_field = field
        if question_rows and qid in question_rows:
            qtype = str(getattr(question_rows[qid], "question_type", "") or "")
            if qtype.lower() == "matrix" and field == "Answer":
                effective_field = "Choice"
        values.append(
            WorkbookTranslationValue(
                qid=qid,
                language=language,
                field=effective_field,  # type: ignore[arg-type]
                item_id=answer_id,
                text=text,
                is_html=is_html,
            )
        )
    return values


def _metadata_keys_from_headers(headers: list[str]) -> list[str]:
    keys: list[str] = []
    for name in headers:
        header = str(name or "")
        if not header.endswith("_MD"):
            continue
        key = header[: -len("_MD")]
        if not key or key.lower() == "language":
            continue
        if key in keys:
            continue
        keys.append(key)
    return keys


def _extract_metadata_values(
    wb,
    language: str,
    *,
    include_metadata: bool,
) -> list[WorkbookTranslationValue]:
    if not include_metadata or SURVEY_METADATA_SHEET not in wb.sheetnames:
        return []
    ws = wb[SURVEY_METADATA_SHEET]
    headers, data_rows = _iter_sheet_rows(ws)
    if not headers or "Language" not in headers:
        return []
    lang_idx = headers.index("Language")
    keys = _metadata_keys_from_headers(headers)
    if not keys:
        return []
    md_indices = {
        key: headers.index(f"{key}_MD") for key in keys if f"{key}_MD" in headers
    }
    html_indices = {
        key: headers.index(f"{key}_IsHTML")
        for key in keys
        if f"{key}_IsHTML" in headers
    }

    values: list[WorkbookTranslationValue] = []
    lang_norm = normalize_language_code(language)
    for row in data_rows:
        raw_lang = row[lang_idx].value if lang_idx < len(row) else None
        row_lang = normalize_language_code(str(raw_lang or ""))
        if not row_lang or row_lang != lang_norm:
            continue
        for key in keys:
            md_idx = md_indices.get(key)
            if md_idx is None:
                continue
            cell = row[md_idx]
            raw = cell.value if cell is not None else None
            if raw is None or str(raw).strip() == "":
                continue
            is_html = False
            html_idx = html_indices.get(key)
            if html_idx is not None:
                html_cell = row[html_idx]
                is_html = bool(html_cell.value) if html_cell is not None else False
            text = normalize_text(str(raw)) if is_html else str(raw)
            values.append(
                WorkbookTranslationValue(
                    qid=SURVEY_METADATA_QID,
                    language=language,
                    field="Metadata",
                    item_id=key,
                    text=text,
                    is_html=is_html,
                )
            )
    return values


def extract_workbook_values(
    workbook_path: Path,
    languages: Iterable[str],
    *,
    scope: ScopeFilter | None = None,
    question_rows: dict[str, Any] | None = None,
) -> list[WorkbookTranslationValue]:
    wb = load_workbook(workbook_path, data_only=True)
    langs = normalize_language_list(languages)
    scope_qids: set[str] | None = None
    if scope and question_rows:
        scope_qids = {
            qid
            for qid, row in question_rows.items()
            if scope.matches(
                qid=qid, tags=[row.data_export_tag] if row.data_export_tag else None
            )
        }

    values: list[WorkbookTranslationValue] = []
    for lang in langs:
        values.extend(_extract_question_values(wb, lang, scope_qids=scope_qids))
        values.extend(_extract_sbs_column_values(wb, lang, scope_qids=scope_qids))
        values.extend(
            _extract_option_values(
                wb,
                lang,
                scope_qids=scope_qids,
                question_rows=question_rows,
            )
        )
        values.extend(
            _extract_sbs_column_answer_values(wb, lang, scope_qids=scope_qids)
        )
        values.extend(
            _extract_subitem_values(
                wb,
                lang,
                scope_qids=scope_qids,
                question_rows=question_rows,
            )
        )
        values.extend(
            _extract_metadata_values(wb, lang, include_metadata=scope is None)
        )
    return values


def diff_workbook_vs_cache(
    survey_payload: dict,
    workbook_path: Path,
    languages: Iterable[str],
    *,
    scope: ScopeFilter | None = None,
    question_rows: dict[str, Any] | None = None,
) -> list[TranslationChange]:
    values = extract_workbook_values(
        workbook_path,
        languages,
        scope=scope,
        question_rows=question_rows,
    )
    questions = (survey_payload.get("result") or {}).get("Questions") or {}
    changes: list[TranslationChange] = []
    for value in values:
        sbs_scope = _split_sbs_scoped_qid(value.qid)
        if value.field == "Metadata":
            old = _read_metadata_value(
                survey_payload, value.language, value.item_id or ""
            )
        elif sbs_scope is not None:
            qid, column_id = sbs_scope
            question = questions.get(qid)
            if not question:
                continue
            if value.field == "QuestionText":
                old = read_sbs_column_question_text(question, value.language, column_id)
            elif value.field == "Answer":
                old = read_sbs_column_answer_display(
                    question,
                    value.language,
                    column_id,
                    value.item_id or "",
                )
            else:
                continue
        else:
            question = questions.get(value.qid)
            if not question:
                continue
            if value.field == "QuestionText":
                old = read_question_text(question, value.language)
            elif value.field == "Choice":
                old = read_choice_display(question, value.language, value.item_id or "")
            elif value.field == "Answer":
                old = read_answer_display(question, value.language, value.item_id or "")
            else:
                old = read_label_display(question, value.language, value.item_id or "")
        if value.is_html:
            old_text = _normalize_translation_compare(old)
            new_text = _normalize_translation_compare(value.html_value)
            if old_text == new_text:
                continue
        else:
            md_old = normalize_markdown_for_compare(html_to_md(str(old or "")))
            md_new = normalize_markdown_for_compare(value.text)
            if md_old == md_new:
                continue
        changes.append(
            TranslationChange(
                qid=value.qid,
                language=normalize_language_code(value.language),
                field=value.field,
                item_id=value.item_id,
                old_value=str(old or ""),
                new_value=str(value.html_value or ""),
            )
        )
    return changes


def build_workbook_value_map(
    workbook_path: Path,
    language: str,
    *,
    scope: ScopeFilter | None = None,
    question_rows: dict[str, Any] | None = None,
) -> dict[TranslationKey, str]:
    values = extract_workbook_values(
        workbook_path,
        [language],
        scope=scope,
        question_rows=question_rows,
    )
    return {value.key: value.html_value for value in values}


def build_base_value_map_for_keys(
    survey_payload: dict,
    keys: Iterable[TranslationKey],
) -> dict[TranslationKey, str]:
    questions = (survey_payload.get("result") or {}).get("Questions") or {}
    result = survey_payload.get("result") or {}
    base_map: dict[TranslationKey, str] = {}
    for qid, field, item_id in keys:
        sbs_scope = _split_sbs_scoped_qid(qid)
        if field == "Metadata":
            value = _base_metadata_value(result, str(item_id or ""))
        elif sbs_scope is not None:
            base_qid, column_id = sbs_scope
            question = questions.get(base_qid)
            if not question:
                continue
            additional = question.get("AdditionalQuestions") or {}
            if not isinstance(additional, dict):
                additional = {}
            column = additional.get(str(column_id))
            if not isinstance(column, dict):
                continue
            if field == "QuestionText":
                value = column.get("QuestionText")
            elif field == "Answer":
                answers = column.get("Answers") or {}
                if not isinstance(answers, dict):
                    answers = {}
                value = (answers.get(str(item_id)) or {}).get("Display")
            else:
                continue
        else:
            question = questions.get(qid)
            if not question:
                continue
            if field == "QuestionText":
                value = question.get("QuestionText")
            elif field == "Choice":
                value = (
                    (question.get("Choices") or {}).get(str(item_id), {}).get("Display")
                )
            elif field == "Answer":
                value = (
                    (question.get("Answers") or {}).get(str(item_id), {}).get("Display")
                )
            else:
                value = (
                    (question.get("Labels") or {}).get(str(item_id), {}).get("Display")
                )
        base_map[(qid, field, item_id)] = str(value) if value is not None else ""
    return base_map


def _base_metadata_value(result: dict, key: str) -> str:
    if not key:
        return ""
    if key in result:
        return str(result.get(key) or "")
    options = result.get("SurveyOptions") or {}
    if isinstance(options, dict) and key in options:
        return str(options.get(key) or "")
    return ""


def _read_metadata_value(payload: dict, language: str, key: str) -> str:
    result = payload.get("result") or {}
    options = result.get("SurveyOptions") or {}
    base_lang = get_base_language_from_options(payload)
    lang = normalize_language_code(language)
    if base_lang and lang == base_lang:
        return _base_metadata_value(result, key)
    meta = options.get("MetaDataTranslations")
    if isinstance(meta, dict):
        entry = meta.get(lang) or meta.get(lang.lower()) or meta.get(lang.upper())
        if isinstance(entry, dict):
            if key in entry:
                return str(entry.get(key) or "")
            if key == "SurveyDescription" and "SurveyMetaDescription" in entry:
                return str(entry.get("SurveyMetaDescription") or "")
    return ""
