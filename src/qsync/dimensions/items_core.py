"""Core survey↔Excel sync operations for qsync."""

from __future__ import annotations

from dataclasses import dataclass
import difflib
from pathlib import Path
from typing import Dict, List, Set, Tuple

from .. import excel_io
from ..markdown_codec import (
    normalize_markdown_for_compare,
    normalize_text,
    validate_html_fragment,
    html_to_md,
)
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from ..qualtrics_client import (
    SurveyCache,
    ensure_backup,
    load_cached_survey,
    refresh_survey_cache,
    push_questions,
    push_survey_flow,
)
from ..drift_check import check_drift as check_drift_fn, enforce_no_drift
from ..scope_filter import ScopeFilter
from ..push_safeguards import enforce_push_safeguards, SafeguardConfig
from ..auto_publish import auto_publish_after_push

ERROR_ID_EMBEDDED_DANGEROUS_SKIPPED = "QSYNC-EMBEDDED-DANGEROUS-001"


@dataclass
class PreviewChange:
    """One previewed change between cached survey HTML and Excel-specified wording."""

    kind: str  # "question", "option", or "subitem"
    qid: str
    old_html: str
    new_html: str
    choice_id: str | None = None
    answer_id: str | None = None
    diff_lines: List[str] | None = None
    data_export_tag: str | None = None
    field: str | None = None
    flow_id: str | None = None
    is_dangerous: bool = False


@dataclass
class ApplyResult:
    """Result summary from apply_changes."""

    qids: list[str]
    embedded_fields: list[dict[str, str]]


@dataclass
class EmbeddedDataHealth:
    """Health status for Embedded_Data rows in a workbook."""

    is_valid: bool
    missing_fields: list[str]
    extra_fields: list[str]
    duplicate_fields: list[str]
    ambiguous_fields: list[str]


def _display_to_str(obj: dict) -> str:
    """Convert a Qualtrics Display field to a stable string.

    Treats None as empty, but preserves numeric values such as 0.
    """

    if obj is None:
        return ""
    val = obj.get("Display")
    if val is None:
        return ""
    return str(val)


def _diff_lines(old_html: str, new_html: str, context: str | None = None) -> List[str]:
    """Generate unified diff lines between old and new content.

    Args:
        old_html: Original content from cache
        new_html: New content from Excel
        context: Optional context string to add to fromfile/tofile labels (e.g., field name for EDFs)

    Returns:
        List of unified diff lines
    """
    fromfile = "cached"
    tofile = "excel"
    if context:
        fromfile = f"cached ({context})"
        tofile = f"excel ({context})"

    return list(
        difflib.unified_diff(
            (old_html or "").splitlines(),
            (new_html or "").splitlines(),
            fromfile=fromfile,
            tofile=tofile,
            lineterm="",
        )
    )


def _normalize_embedded_value(value: str | None) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text == excel_io.EMBEDDED_EMPTY_VALUE:
        return None
    return text


def _display_embedded_value(value: str | None) -> str:
    return value if value is not None else excel_io.EMBEDDED_EMPTY_VALUE


def _format_field_list(fields: List[str], *, limit: int = 6) -> str:
    if not fields:
        return ""
    deduped = sorted({str(field).strip() for field in fields if str(field).strip()})
    if len(deduped) <= limit:
        return ", ".join(deduped)
    overflow = len(deduped) - limit
    return f"{', '.join(deduped[:limit])}, +{overflow} more"


def check_embedded_data_health(
    survey_id: str, survey_payload: dict, workbook_path: Path
) -> EmbeddedDataHealth:
    """Check if Embedded_Data rows align with expected SurveyFlow rows.

    Returns:
        EmbeddedDataHealth with missing/extra/duplicate/ambiguous field lists.
    """

    expected_rows = excel_io.build_embedded_data_rows(survey_id, survey_payload)
    excel_rows = excel_io.load_embedded_data_from_workbook(workbook_path)

    expected_by_key: Dict[Tuple[str, str], excel_io.EmbeddedDataRow] = {}
    expected_by_field: Dict[str, List[excel_io.EmbeddedDataRow]] = {}
    for row in expected_rows:
        key = (row.flow_id or "", row.field)
        expected_by_key[key] = row
        expected_by_field.setdefault(row.field, []).append(row)

    mapped_excel: Dict[Tuple[str, str], excel_io.EmbeddedDataRow] = {}
    extra_fields: List[str] = []
    duplicate_fields: List[str] = []
    ambiguous_fields: List[str] = []

    for row in excel_rows:
        flow_id = row.flow_id or ""
        field = row.field
        key: Tuple[str, str] | None = None
        if flow_id:
            key = (flow_id, field)
            if key not in expected_by_key:
                extra_fields.append(field)
                continue
        else:
            matches = expected_by_field.get(field, [])
            if not matches:
                extra_fields.append(field)
                continue
            if len(matches) > 1:
                ambiguous_fields.append(field)
                continue
            key = (matches[0].flow_id or "", field)
        if key in mapped_excel:
            duplicate_fields.append(field)
            continue
        mapped_excel[key] = row

    missing_keys = set(expected_by_key) - set(mapped_excel)
    missing_fields = sorted({field for _, field in missing_keys})

    is_valid = not (
        missing_fields or extra_fields or duplicate_fields or ambiguous_fields
    )
    return EmbeddedDataHealth(
        is_valid=is_valid,
        missing_fields=missing_fields,
        extra_fields=sorted({f for f in extra_fields if str(f).strip()}),
        duplicate_fields=sorted({f for f in duplicate_fields if str(f).strip()}),
        ambiguous_fields=sorted({f for f in ambiguous_fields if str(f).strip()}),
    )


def format_embedded_data_health_warning(
    health: EmbeddedDataHealth, *, survey_id: str
) -> str:
    if health.is_valid:
        return ""
    issues: List[str] = []
    if health.missing_fields:
        issues.append(f"missing fields: {_format_field_list(health.missing_fields)}")
    if health.extra_fields:
        issues.append(f"extra fields: {_format_field_list(health.extra_fields)}")
    if health.duplicate_fields:
        issues.append(
            f"duplicate fields: {_format_field_list(health.duplicate_fields)}"
        )
    if health.ambiguous_fields:
        issues.append(
            f"ambiguous fields (missing FlowID): {_format_field_list(health.ambiguous_fields)}"
        )
    issue_summary = "; ".join([issue for issue in issues if issue])
    return (
        "Embedded_Data worksheet is inconsistent with the cached survey "
        f"({issue_summary}). Repair: qsync items pull --survey-id {survey_id} "
        "(warning: may overwrite unstaged changes)"
    )


def _collect_embedded_data_changes(
    survey_id: str, survey_payload: dict, workbook_path: Path
) -> List[dict]:
    expected_rows = excel_io.build_embedded_data_rows(survey_id, survey_payload)
    excel_rows = excel_io.load_embedded_data_from_workbook(workbook_path)

    expected_by_key: Dict[Tuple[str, str], excel_io.EmbeddedDataRow] = {}
    expected_by_field: Dict[str, List[excel_io.EmbeddedDataRow]] = {}
    for row in expected_rows:
        key = (row.flow_id or "", row.field)
        expected_by_key[key] = row
        expected_by_field.setdefault(row.field, []).append(row)

    mapped_excel: Dict[Tuple[str, str], excel_io.EmbeddedDataRow] = {}
    extra_rows: List[excel_io.EmbeddedDataRow] = []

    for row in excel_rows:
        flow_id = row.flow_id or ""
        field = row.field
        key: Tuple[str, str] | None = None
        if flow_id:
            key = (flow_id, field)
            if key not in expected_by_key:
                extra_rows.append(row)
                continue
        else:
            matches = expected_by_field.get(field, [])
            if not matches:
                extra_rows.append(row)
                continue
            if len(matches) > 1:
                raise ValueError(
                    f"Embedded data field '{field}' appears multiple times in SurveyFlow; "
                    "re-run qsync init to capture FlowID values."
                )
            key = (matches[0].flow_id or "", field)
        if key in mapped_excel:
            raise ValueError(
                "Duplicate embedded data row for field "
                f"'{field}' in Embedded_Data sheet. "
                f"Run `qsync survey cleanup-embedded-data --survey-id {survey_id} "
                "--placeholder-only` if this is caused by placeholder duplicates."
            )
        mapped_excel[key] = row

    missing_keys = set(expected_by_key) - set(mapped_excel)
    if missing_keys:
        missing_fields = ", ".join(sorted({field for _, field in missing_keys}))
        raise ValueError(
            "Embedded_Data sheet is missing rows from SurveyFlow "
            f"(fields: {missing_fields}). Re-run qsync init."
        )
    if extra_rows:
        extra_fields = ", ".join(sorted({row.field for row in extra_rows}))
        raise ValueError(
            "Embedded_Data sheet contains unknown rows "
            f"(fields: {extra_fields}). Remove extra rows or re-run qsync init."
        )

    changes: List[dict] = []
    for key, expected in expected_by_key.items():
        excel_row = mapped_excel.get(key)
        if not excel_row:
            continue
        old_value = _normalize_embedded_value(expected.value)
        new_value = _normalize_embedded_value(excel_row.value)

        if old_value is not None and new_value is None:
            raise ValueError(
                f"Embedded data field '{expected.field}' has a default value; "
                "clearing defaults via Excel is not supported."
            )

        if old_value == new_value:
            continue

        changes.append(
            {
                "row": expected,
                "old_value": old_value,
                "new_value": new_value,
                "is_dangerous": old_value is None,
            }
        )

    return changes


def _index_embedded_flow_nodes(
    survey_payload: dict,
) -> Tuple[Dict[str, dict], List[dict]]:
    nodes = excel_io._iter_embedded_data_nodes(survey_payload)
    flow_id_map: Dict[str, dict] = {}
    ordered_nodes: List[dict] = []
    for _, node in nodes:
        ordered_nodes.append(node)
        flow_id = str(node.get("FlowID") or "").strip()
        if flow_id:
            flow_id_map[flow_id] = node
    return flow_id_map, ordered_nodes


def _find_embedded_entry(node: dict, field: str) -> dict | None:
    for entry in node.get("EmbeddedData", []) or []:
        if str(entry.get("Field") or "").strip() == field:
            return entry
    return None


def _next_flow_id(survey_payload: dict) -> str:
    ids: List[int] = []
    for _, node in excel_io._iter_embedded_data_nodes(survey_payload):
        flow_id = str(node.get("FlowID") or "")
        if flow_id.startswith("FL_"):
            try:
                ids.append(int(flow_id.split("_", 1)[1]))
            except ValueError:
                continue
    next_id = max(ids) + 1 if ids else 1
    return f"FL_{next_id}"


def _ensure_embedded_data_node(
    survey_payload: dict, *, ordered_nodes: List[dict]
) -> dict:
    if ordered_nodes:
        return ordered_nodes[0]
    survey_flow = survey_payload.get("result", {}).get("SurveyFlow", {})
    flow_list = survey_flow.get("Flow")
    if not isinstance(flow_list, list):
        raise ValueError("SurveyFlow.Flow is missing or not a list.")
    node = {
        "Type": "EmbeddedData",
        "FlowID": _next_flow_id(survey_payload),
        "EmbeddedData": [],
    }
    flow_list.insert(0, node)
    return node


def _make_embedded_entry(field: str, value: str, template: dict | None = None) -> dict:
    entry = {
        "Field": field,
        "Type": "Custom",
        "Value": value,
        "Description": field,
        "DataVisibility": [],
        "AnalyzeText": False,
    }
    if template:
        for key in ("DataVisibility", "AnalyzeText"):
            if key in template:
                entry[key] = template[key]
    return entry


def _find_embedded_field_in_flow(survey_payload: dict, field: str) -> list[dict]:
    matches: list[dict] = []
    for _, node in excel_io._iter_embedded_data_nodes(survey_payload):
        for entry in node.get("EmbeddedData", []) or []:
            if str(entry.get("Field") or "").strip() == field:
                matches.append(
                    {"flow_id": str(node.get("FlowID") or ""), "entry": entry}
                )
    return matches


def _find_js_field_match(
    survey_payload: dict, field: str
) -> tuple[str, list[str]] | None:
    js_map = excel_io._collect_js_embedded_data_fields(survey_payload)
    for candidate, qids in js_map.items():
        if candidate == field:
            return candidate, qids
        if candidate.endswith("*") and field.startswith(candidate[:-1]):
            return candidate, qids
    return None


def stage_add_embedded_field(
    survey_id: str,
    *,
    field: str,
    value: str | None = None,
    flow_id: str | None = None,
    dry_run: bool = False,
) -> dict[str, str]:
    """Stage an embedded data field addition in the cached SurveyFlow."""

    field_name = (field or "").strip()
    if not field_name:
        raise ValueError("Embedded field name must be non-empty.")

    survey: SurveyCache = load_cached_survey(survey_id)
    existing = _find_embedded_field_in_flow(survey.payload, field_name)
    if existing:
        flow_ids = ", ".join(sorted({entry.get("flow_id") or "" for entry in existing}))
        suffix = f" (FlowID(s)={flow_ids})" if flow_ids else ""
        raise ValueError(
            f"Embedded field '{field_name}' already exists in SurveyFlow{suffix}; aborting."
        )

    js_match = _find_js_field_match(survey.payload, field_name)
    if js_match:
        pattern, qids = js_match
        qid_list = ", ".join(sorted(qids)) if qids else "unknown QIDs"
        raise ValueError(
            f"Embedded field '{field_name}' is written by QuestionJS ({pattern}) in {qid_list}; "
            "aborting."
        )

    flow_id = (flow_id or "").strip() or None
    flow_id_map, ordered_nodes = _index_embedded_flow_nodes(survey.payload)
    if flow_id:
        node = flow_id_map.get(flow_id)
        if not node:
            raise ValueError(f"EmbeddedData FlowID '{flow_id}' not found; aborting.")
    else:
        node = _ensure_embedded_data_node(survey.payload, ordered_nodes=ordered_nodes)
    entries = node.get("EmbeddedData") or []
    template = entries[0] if entries else None
    default_value = "" if value is None else str(value)
    new_entry = _make_embedded_entry(field_name, default_value, template=template)
    if dry_run:
        return {"flow_id": str(node.get("FlowID") or ""), "field": field_name}
    entries.insert(0, new_entry)
    node["EmbeddedData"] = entries
    survey.save()
    return {"flow_id": str(node.get("FlowID") or ""), "field": field_name}


def stage_remove_embedded_field(
    survey_id: str,
    *,
    field: str,
    flow_id: str | None = None,
    dry_run: bool = False,
) -> list[dict[str, str]]:
    """Stage removal of an embedded data field from the cached SurveyFlow."""

    field_name = (field or "").strip()
    if not field_name:
        raise ValueError("Embedded field name must be non-empty.")

    survey: SurveyCache = load_cached_survey(survey_id)
    js_match = _find_js_field_match(survey.payload, field_name)
    if js_match:
        pattern, qids = js_match
        qid_list = ", ".join(sorted(qids)) if qids else "unknown QIDs"
        raise ValueError(
            f"Embedded field '{field_name}' is written by QuestionJS ({pattern}) in {qid_list}; "
            "aborting."
        )

    flow_id = (flow_id or "").strip() or None
    removed: list[dict[str, str]] = []
    if flow_id:
        flow_id_map, _ = _index_embedded_flow_nodes(survey.payload)
        node = flow_id_map.get(flow_id)
        if not node:
            raise ValueError(f"EmbeddedData FlowID '{flow_id}' not found; aborting.")
        entries = node.get("EmbeddedData") or []
        matching = [
            entry
            for entry in entries
            if str(entry.get("Field") or "").strip() == field_name
        ]
        if matching:
            removed.append({"flow_id": flow_id, "field": field_name})
        if not dry_run and matching:
            node["EmbeddedData"] = [
                entry
                for entry in entries
                if str(entry.get("Field") or "").strip() != field_name
            ]
    else:
        for _, node in excel_io._iter_embedded_data_nodes(survey.payload):
            entries = node.get("EmbeddedData") or []
            matching = [
                entry
                for entry in entries
                if str(entry.get("Field") or "").strip() == field_name
            ]
            if matching:
                removed.append(
                    {"flow_id": str(node.get("FlowID") or ""), "field": field_name}
                )
            if not dry_run and matching:
                node["EmbeddedData"] = [
                    entry
                    for entry in entries
                    if str(entry.get("Field") or "").strip() != field_name
                ]

    if removed and not dry_run:
        survey.save()

    return removed


def init_survey_to_excel(
    survey_id: str, xlsx_path: Path, *, languages: set[str] | list[str] | None = None
) -> None:
    """Initialize or refresh a survey workbook from the latest Qualtrics cache.

    This function refreshes the cached survey JSON from Qualtrics, then (re)builds
    the Excel workbook used for wording edits.

    Args:
        survey_id: Qualtrics survey ID (e.g., `SV_xxx`).
        xlsx_path: Target workbook path (typically under `excel/`).
        languages: Optional list of language codes to add as translation columns.
                  If None (default), auto-detects all enabled languages from Qualtrics.
                  If explicit list provided, creates only those columns.

    Raises:
        requests.HTTPError: If the Qualtrics API call fails while refreshing.

    Example:
        >>> from pathlib import Path
        >>> from qsync.sync_core import init_survey_to_excel
        >>> init_survey_to_excel("SV_xxx", Path("excel/SV_xxx-mylabel.xlsx"))
    """

    # Check for drift before refreshing
    drift_report = check_drift_fn(survey_id, dimension="items", interactive=True)
    if drift_report.has_drift:
        print("[qsync:items] WARNING: Drift detected between cache and API.")
        drift_report.display(interactive=True)

    # Refresh cache from Qualtrics and detect drift.
    survey, changed = refresh_survey_cache(survey_id)
    if changed:
        from ..survey_ref import format_survey_ref

        survey_ref = format_survey_ref(survey_id)
        print(
            f"[qsync:init] WARNING: Cached survey definition for {survey_ref} "
            "differs from the current online version. Excel may be out of sync. "
            "Next: re-run `qsync init --survey-id ...` and re-check diffs before editing/pushing."
        )

    # Auto-detect enabled languages if not explicitly specified
    if languages is None:
        from ..translations import list_enabled_languages

        print("[qsync:init] Auto-detecting enabled languages...")
        languages = list_enabled_languages(survey_id)
        if languages:
            langs_str = ", ".join(sorted(languages))
            print(f"[qsync:init] Creating columns for languages: {langs_str}")
        else:
            print("[qsync:init] No additional languages enabled (base language only)")

    excel_io.init_workbook_from_survey(
        survey_id,
        survey.payload,
        Path(xlsx_path),
        languages=list(languages) if languages else None,
    )
    print(f"[qsync:init] Updated workbook at {xlsx_path}")

    # Translation columns are populated from the cached survey definition only.


def _filter_qids_by_column(
    questions: Dict[str, excel_io.QuestionRow],
    filter_column: str | None,
    filter_value: str | None,
) -> Set[str]:
    if not filter_column:
        return set(questions.keys())
    val_norm = (filter_value or "TRUE").strip().lower()
    qids: Set[str] = set()
    for qid, row in questions.items():
        if filter_column == "InPre":
            col_val = row.in_pre
            if bool(col_val) == (val_norm in {"true", "1", "yes", "y"}):
                qids.add(qid)
        elif filter_column == "InPost":
            col_val = row.in_post
            if bool(col_val) == (val_norm in {"true", "1", "yes", "y"}):
                qids.add(qid)
        elif filter_column == "BlockName":
            # Match block names case-insensitively
            col_val = (row.block_name or "").strip().lower()
            if col_val and col_val == val_norm:
                qids.add(qid)
        else:
            # Unknown filter column for Phase 1: ignore.
            continue
    return qids


def _apply_include_filters(
    questions: Dict[str, excel_io.QuestionRow],
    qids: Set[str],
    include_qids: Set[str] | None,
    include_tags: Set[str] | None,
    *,
    context: str,
) -> Set[str]:
    if not qids:
        return set()
    scoped = set(qids)
    if include_qids:
        missing = [qid for qid in include_qids if qid not in questions]
        for qid in missing:
            print(
                f"[qsync:{context}] WARNING: QID {qid} not found in workbook; skipping. "
                "Next: run `qsync init --survey-id ...` to refresh the workbook, or remove the QID filter."
            )
        scoped &= set(include_qids)
    if include_tags:
        tag_map: Dict[str, str] = {}
        for row in questions.values():
            tag = (row.data_export_tag or "").strip()
            if tag:
                tag_map[tag.lower()] = row.qid
        resolved = set()
        for tag in include_tags:
            qid = tag_map.get((tag or "").strip().lower())
            if not qid:
                print(
                    f"[qsync:{context}] WARNING: DataExportTag '{tag}' not found; skipping. "
                    "Next: verify the tag in Qualtrics, then re-run `qsync init --survey-id ...` to refresh the workbook."
                )
                continue
            resolved.add(qid)
        if resolved:
            scoped &= resolved
        else:
            scoped.clear()
    return scoped


def _annotate_dirty_in_workbook(xlsx_path: Path, changes: List[PreviewChange]) -> None:
    """Mark dirty rows/cells in the Excel workbook based on preview changes.

    - Adds/updates a `Dirty` column on Questions/Options sheets.
    - Highlights the edited field (Text_en_MD or Label_en_MD).
    """

    if not changes:
        return

    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        return

    wb = load_workbook(xlsx_path)

    # Helper to ensure Dirty column exists and return index
    def ensure_dirty_column(ws_name: str) -> Tuple[object, Dict[str, int], int]:
        ws = wb[ws_name]
        headers, _ = excel_io._iter_sheet_rows(ws)
        if not headers:
            return ws, {}, -1
        if "Dirty" not in headers:
            headers.append("Dirty")
            ws.cell(row=1, column=len(headers), value="Dirty")
        # rebuild header/index
        header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=False))
        headers = [c.value or "" for c in header_cells]
        col_idx = headers.index("Dirty") + 1
        idx_map = {name: i for i, name in enumerate(headers)}
        return ws, idx_map, col_idx

    # Clear existing Dirty flags
    if excel_io.QUESTION_SHEET in wb.sheetnames:
        ws_q, idx_q, dirty_col_q = ensure_dirty_column(excel_io.QUESTION_SHEET)
        if dirty_col_q > 0:
            for row in ws_q.iter_rows(min_row=2):
                row[dirty_col_q - 1].value = ""
    else:
        ws_q = None
        idx_q = {}
        dirty_col_q = -1

    if excel_io.OPTIONS_SHEET in wb.sheetnames:
        ws_o, idx_o, dirty_col_o = ensure_dirty_column(excel_io.OPTIONS_SHEET)
        if dirty_col_o > 0:
            for row in ws_o.iter_rows(min_row=2):
                row[dirty_col_o - 1].value = ""
    else:
        ws_o = None
        idx_o = {}
        dirty_col_o = -1

    if excel_io.EMBEDDED_DATA_SHEET in wb.sheetnames:
        ws_e, idx_e, dirty_col_e = ensure_dirty_column(excel_io.EMBEDDED_DATA_SHEET)
        if dirty_col_e > 0:
            for row in ws_e.iter_rows(min_row=2):
                row[dirty_col_e - 1].value = ""
    else:
        ws_e = None
        idx_e = {}
        dirty_col_e = -1

    # Apply new Dirty markers
    for change in changes:
        if change.kind == "question" and ws_q is not None and dirty_col_q > 0:
            qid_idx = idx_q.get("QID")
            text_idx = idx_q.get("Text_en_MD")
            for row in ws_q.iter_rows(min_row=2, values_only=False):
                if qid_idx is not None and row[qid_idx].value == change.qid:
                    row[dirty_col_q - 1].value = "Y"
                    break
        elif change.kind == "option" and ws_o is not None and dirty_col_o > 0:
            qid_idx = idx_o.get("QID")
            choice_idx = idx_o.get("ChoiceId")
            for row in ws_o.iter_rows(min_row=2, values_only=False):
                if (
                    qid_idx is not None
                    and choice_idx is not None
                    and row[qid_idx].value == change.qid
                    and str(row[choice_idx].value) == str(change.choice_id)
                ):
                    row[dirty_col_o - 1].value = "Y"
                    break
        elif change.kind == "embedded" and ws_e is not None and dirty_col_e > 0:
            field_idx = idx_e.get("Field")
            flow_idx = idx_e.get("FlowID")
            field = change.field or change.qid
            flow_id = change.flow_id or ""
            for row in ws_e.iter_rows(min_row=2, values_only=False):
                if field_idx is None:
                    continue
                row_field = str(row[field_idx].value or "").strip()
                if row_field != field:
                    continue
                if flow_idx is not None and flow_id:
                    row_flow = str(row[flow_idx].value or "").strip()
                    if row_flow != flow_id:
                        continue
                row[dirty_col_e - 1].value = "Y"
                break

    # After marking, (re)attach conditional formatting for Dirty flags
    if ws_q is not None and idx_q and dirty_col_q > 0 and "Text_en_MD" in idx_q:
        headers, _ = excel_io._iter_sheet_rows(ws_q)
        max_row_q = ws_q.max_row
        if "Dirty" in headers and "Text_en_MD" in headers and max_row_q >= 2:
            dirty_idx = headers.index("Dirty") + 1
            text_idx = headers.index("Text_en_MD") + 1
            dirty_col_letter = get_column_letter(dirty_idx)
            text_col_letter = get_column_letter(text_idx)
            formula = f'=${dirty_col_letter}2="Y"'
            from openpyxl.formatting.rule import FormulaRule

            rule = FormulaRule(formula=[formula], fill=excel_io._DIRTY_FILL)
            ws_q.conditional_formatting.add(
                f"{text_col_letter}2:{text_col_letter}{max_row_q}", rule
            )

    if ws_o is not None and idx_o and dirty_col_o > 0 and "Label_en_MD" in idx_o:
        headers_o, _ = excel_io._iter_sheet_rows(ws_o)
        max_row_o = ws_o.max_row
        if "Dirty" in headers_o and "Label_en_MD" in headers_o and max_row_o >= 2:
            dirty_idx = headers_o.index("Dirty") + 1
            text_idx = headers_o.index("Label_en_MD") + 1
            dirty_col_letter = get_column_letter(dirty_idx)
            text_col_letter = get_column_letter(text_idx)
            formula = f'=${dirty_col_letter}2="Y"'
            from openpyxl.formatting.rule import FormulaRule

            rule = FormulaRule(formula=[formula], fill=excel_io._DIRTY_FILL)
            ws_o.conditional_formatting.add(
                f"{text_col_letter}2:{text_col_letter}{max_row_o}", rule
            )

    if ws_e is not None and idx_e and dirty_col_e > 0 and "Value" in idx_e:
        headers_e, _ = excel_io._iter_sheet_rows(ws_e)
        max_row_e = ws_e.max_row
        if "Dirty" in headers_e and "Value" in headers_e and max_row_e >= 2:
            dirty_idx = headers_e.index("Dirty") + 1
            value_idx = headers_e.index("Value") + 1
            dirty_col_letter = get_column_letter(dirty_idx)
            value_col_letter = get_column_letter(value_idx)
            formula = f'=${dirty_col_letter}2="Y"'
            from openpyxl.formatting.rule import FormulaRule

            rule = FormulaRule(formula=[formula], fill=excel_io._DIRTY_FILL)
            ws_e.conditional_formatting.add(
                f"{value_col_letter}2:{value_col_letter}{max_row_e}", rule
            )

    wb.save(xlsx_path)


def _self_heal_system_columns(survey: SurveyCache, xlsx_path: Path) -> None:
    """Check core system columns against the survey JSON and self-heal if needed.

    - For Questions: SurveyID, QuestionType, DataExportTag.
    - For Options:  SurveyID, QuestionType, Code.

    When mismatches are found, print a warning and reset the Excel cell to the
    Qualtrics value. If a QID or ChoiceId does not exist in the survey JSON,
    warn and ignore that row.
    """

    from openpyxl import load_workbook

    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        return

    wb = load_workbook(xlsx_path)
    questions = survey.questions

    # Questions sheet
    if excel_io.QUESTION_SHEET in wb.sheetnames:
        ws_q = wb[excel_io.QUESTION_SHEET]
        headers, data_rows = excel_io._iter_sheet_rows(ws_q)
        idx = {name: i for i, name in enumerate(headers)}

        for row in ws_q.iter_rows(min_row=2, values_only=False):
            qid_cell = row[idx.get("QID")] if "QID" in idx else None
            if qid_cell is None or qid_cell.value is None:
                continue
            qid = str(qid_cell.value).strip()
            q_json = questions.get(qid)
            if not q_json:
                print(
                    f"[qsync] WARNING: Questions row {qid_cell.row} refers to unknown QID={qid}; "
                    "row will be ignored for syncing. Next: re-run `qsync init --survey-id ...` to refresh the workbook."
                )
                continue

            expected_survey_id = survey.survey_id
            expected_qtype = q_json.get("QuestionType") or ""
            expected_tag = q_json.get("DataExportTag") or ""

            # SurveyID
            if "SurveyID" in idx:
                cell = row[idx["SurveyID"]]
                if str(cell.value or "").strip() != expected_survey_id:
                    print(
                        f"[qsync] WARNING: System column SurveyID changed in Questions row "
                        f"{cell.row} (QID={qid}); resetting to Qualtrics value. "
                        "Next: avoid editing system columns in Excel."
                    )
                    cell.value = expected_survey_id

            # QuestionType
            if "QuestionType" in idx:
                cell = row[idx["QuestionType"]]
                if str(cell.value or "").strip() != expected_qtype:
                    print(
                        f"[qsync] WARNING: System column QuestionType changed in Questions row "
                        f"{cell.row} (QID={qid}); resetting to Qualtrics value. "
                        "Next: avoid editing system columns in Excel."
                    )
                    cell.value = expected_qtype

            # DataExportTag
            if "DataExportTag" in idx:
                cell = row[idx["DataExportTag"]]
                if str(cell.value or "").strip() != expected_tag:
                    print(
                        f"[qsync] WARNING: System column DataExportTag changed in Questions row "
                        f"{cell.row} (QID={qid}); resetting to Qualtrics value. "
                        "Next: avoid editing system columns in Excel."
                    )
                    cell.value = expected_tag

    # Options sheet
    if excel_io.OPTIONS_SHEET in wb.sheetnames:
        ws_o = wb[excel_io.OPTIONS_SHEET]
        headers, data_rows = excel_io._iter_sheet_rows(ws_o)
        idx = {name: i for i, name in enumerate(headers)}

        for row in ws_o.iter_rows(min_row=2, values_only=False):
            qid_cell = row[idx.get("QID")] if "QID" in idx else None
            choice_cell = row[idx.get("ChoiceId")] if "ChoiceId" in idx else None
            if qid_cell is None or choice_cell is None:
                continue
            qid = str(qid_cell.value or "").strip()
            choice_id = str(choice_cell.value or "").strip()
            q_json = questions.get(qid)
            if not q_json:
                print(
                    f"[qsync] WARNING: Options row {qid_cell.row} refers to unknown QID={qid}; "
                    "row will be ignored for syncing. Next: re-run `qsync init --survey-id ...` to refresh the workbook."
                )
                continue

            qtype = q_json.get("QuestionType") or ""
            if qtype == "Matrix":
                container = q_json.get("Answers") or {}
            else:
                container = q_json.get("Choices") or {}
            choice = container.get(choice_id)
            if not choice:
                print(
                    f"[qsync] WARNING: Options row {choice_cell.row} refers to unknown "
                    f"choiceId={choice_id} for QID={qid}; row will be ignored for syncing. "
                    "Next: re-run `qsync init --survey-id ...` to refresh the workbook."
                )
                continue

            expected_survey_id = survey.survey_id
            expected_qtype = q_json.get("QuestionType") or ""
            expected_code = choice.get("Recode")

            # SurveyID
            if "SurveyID" in idx:
                cell = row[idx["SurveyID"]]
                if str(cell.value or "").strip() != expected_survey_id:
                    print(
                        f"[qsync] WARNING: System column SurveyID changed in Options row "
                        f"{cell.row} (QID={qid}, ChoiceId={choice_id}); resetting. "
                        "Next: avoid editing system columns in Excel."
                    )
                    cell.value = expected_survey_id

            # QuestionType
            if "QuestionType" in idx:
                cell = row[idx["QuestionType"]]
                if str(cell.value or "").strip() != expected_qtype:
                    print(
                        f"[qsync] WARNING: System column QuestionType changed in Options row "
                        f"{cell.row} (QID={qid}, ChoiceId={choice_id}); resetting. "
                        "Next: avoid editing system columns in Excel."
                    )
                    cell.value = expected_qtype

            # Code
            if "Code" in idx:
                cell = row[idx["Code"]]
                excel_code = str(cell.value or "").strip()
                expected_code_str = "" if expected_code is None else str(expected_code)
                if excel_code != expected_code_str:
                    print(
                        f"[qsync] WARNING: System column Code changed in Options row "
                        f"{cell.row} (QID={qid}, ChoiceId={choice_id}); resetting. "
                        "Next: avoid editing system columns in Excel."
                    )
                    cell.value = expected_code

    # Embedded Data sheet
    if excel_io.EMBEDDED_DATA_SHEET in wb.sheetnames:
        ws_e = wb[excel_io.EMBEDDED_DATA_SHEET]
        headers, data_rows = excel_io._iter_sheet_rows(ws_e)
        idx = {name: i for i, name in enumerate(headers)}

        expected_rows = excel_io.build_embedded_data_rows(
            survey.survey_id, survey.payload
        )
        expected_by_key = {(row.flow_id or "", row.field): row for row in expected_rows}
        expected_by_field: Dict[str, List[excel_io.EmbeddedDataRow]] = {}
        for row in expected_rows:
            expected_by_field.setdefault(row.field, []).append(row)

        for row in ws_e.iter_rows(min_row=2, values_only=False):
            field_cell = row[idx.get("Field")] if "Field" in idx else None
            if field_cell is None or field_cell.value is None:
                continue
            field = str(field_cell.value).strip()
            if not field:
                continue
            flow_cell = row[idx.get("FlowID")] if "FlowID" in idx else None
            flow_id = str(flow_cell.value or "").strip() if flow_cell else ""

            expected = None
            if flow_id:
                expected = expected_by_key.get((flow_id, field))
            else:
                matches = expected_by_field.get(field, [])
                if len(matches) == 1:
                    expected = matches[0]
                    if flow_cell is not None and expected.flow_id:
                        print(
                            f"[qsync] WARNING: Embedded_Data row {field_cell.row} missing FlowID; "
                            f"resetting to {expected.flow_id}. Next: avoid clearing FlowID; re-run `qsync init --survey-id ...` if the sheet is corrupted."
                        )
                        flow_cell.value = expected.flow_id
                elif len(matches) > 1:
                    print(
                        f"[qsync] WARNING: Embedded_Data row {field_cell.row} has ambiguous field "
                        f"'{field}'; re-run qsync init to restore FlowID values. "
                        "Next: run `qsync init --survey-id ...` and use FlowID to disambiguate."
                    )
                    continue

            if not expected:
                print(
                    f"[qsync] WARNING: Embedded_Data row {field_cell.row} refers to "
                    f"unknown field '{field}'; row will be ignored for syncing. "
                    "Next: re-run `qsync init --survey-id ...` to refresh Embedded_Data rows."
                )
                continue

            def _reset_if_changed(col_name: str, expected_value: str | int | None):
                if col_name not in idx:
                    return
                cell = row[idx[col_name]]
                current_raw = cell.value
                current = "" if current_raw is None else str(current_raw).strip()
                expected_str = "" if expected_value is None else str(expected_value)
                if col_name == "FlowOrder":
                    try:
                        current_num = int(float(current))
                    except (TypeError, ValueError):
                        current_num = None
                    try:
                        expected_num = int(float(expected_str))
                    except (TypeError, ValueError):
                        expected_num = None
                    if (
                        current_num is not None
                        and expected_num is not None
                        and current_num == expected_num
                    ):
                        return
                if current != expected_str:
                    print(
                        f"[qsync] WARNING: System column {col_name} changed in "
                        f"Embedded_Data row {cell.row} (Field={field}); resetting. "
                        "Next: avoid editing system columns in Excel."
                    )
                    cell.value = expected_value

            _reset_if_changed("SurveyID", survey.survey_id)
            _reset_if_changed("FlowID", expected.flow_id or "")
            _reset_if_changed("FlowOrder", expected.flow_order)
            _reset_if_changed("Field", expected.field)
            _reset_if_changed("Type", expected.ed_type)
            _reset_if_changed("WrittenByQIDs", expected.written_by_qids or "")

    wb.save(xlsx_path)


def _sort_changes_by_flow_order(
    survey: SurveyCache, changes: List[PreviewChange]
) -> List[PreviewChange]:
    """Sort changes by survey flow order instead of QID number.

    Args:
        survey: Cached survey data
        changes: List of changes to sort

    Returns:
        Sorted list of changes
    """
    # Build QID to flow order mapping
    qid_to_order = {}
    flow_elements = (
        survey.payload.get("result", {}).get("SurveyFlow", {}).get("Flow", [])
    )

    def extract_qids(element, order_counter=[0]):
        if isinstance(element, dict):
            if element.get("Type") == "Standard" and "ID" in element:
                qid_to_order[element["ID"]] = order_counter[0]
                order_counter[0] += 1
            if "Flow" in element:
                for child in element["Flow"]:
                    extract_qids(child, order_counter)

    for elem in flow_elements:
        extract_qids(elem)

    # Sort changes: first by flow order (if available), then by QID
    def sort_key(change):
        qid = change.qid
        flow_order = qid_to_order.get(qid, 999999)  # Unknown QIDs go to end
        return (flow_order, qid)  # Secondary sort by QID for stability

    return sorted(changes, key=sort_key)


def preview_changes(
    survey_id: str,
    xlsx_path: Path,
    filter_column: str | None = None,
    filter_value: str | None = None,
    include_qids: Set[str] | None = None,
    include_tags: Set[str] | None = None,
    embedded_only: bool = False,
    skip_embedded: bool = False,
    scope_expr: str | None = None,
    check_drift: bool = True,
    self_heal_system_columns: bool = True,
    annotate_dirty: bool = True,
) -> List[PreviewChange]:
    """Compute diffs between an Excel workbook and the cached survey JSON.

    This is the engine behind `qsync preview`.

    Notes:
    - Does not write to Qualtrics.
    - Self-heals system columns in the workbook (SurveyID/QuestionType/etc.) to
      reduce accidental drift.
    - Updates the workbook's `Dirty` markers so edited rows are easier to spot.

    Args:
        survey_id: Qualtrics survey ID (e.g., `SV_xxx`).
        xlsx_path: Path to the Excel workbook for the survey.
        filter_column: Optional column on the Questions sheet used to scope QIDs
            (e.g., `InPre`, `InPost`, `BlockName`).
        filter_value: Optional value for the filter column (defaults to `"TRUE"`).
        include_qids: Optional explicit set of QIDs to include.
        include_tags: Optional explicit set of DataExportTags to include.
        skip_embedded: If True, do not read or validate Embedded_Data rows.
        scope_expr: Optional scope filter expression (e.g., 'qid:Q1 OR tag:baseline').

    Returns:
        List of proposed changes (question text, option labels, subitem labels).

    Example:
        >>> from pathlib import Path
        >>> from qsync.sync_core import preview_changes
        >>> changes = preview_changes("SV_xxx", Path("excel/SV_xxx-mylabel.xlsx"))
        >>> len(changes) >= 0
        True
    """

    # Check for drift before preview
    if check_drift:
        drift_report = check_drift_fn(survey_id, dimension="items", interactive=True)
        if drift_report.has_drift:
            drift_report.display(interactive=False)

    survey: SurveyCache = load_cached_survey(survey_id)
    # Self-heal system columns (SurveyID, QuestionType, DataExportTag, Code) if needed
    if self_heal_system_columns:
        _self_heal_system_columns(survey, Path(xlsx_path))
    questions_excel = excel_io.load_questions_from_workbook(Path(xlsx_path))
    options_excel = excel_io.load_options_from_workbook(Path(xlsx_path))
    subitems_excel = excel_io.load_subitems_from_workbook(Path(xlsx_path))

    # Apply scope filtering
    if scope_expr:
        scope_filter = ScopeFilter.parse(scope_expr)
        in_scope_qids = {
            qid
            for qid in questions_excel.keys()
            if scope_filter.matches(
                qid=qid,
                tags=(
                    [questions_excel[qid].data_export_tag]
                    if questions_excel[qid].data_export_tag
                    else None
                ),
            )
        }
    else:
        in_scope_qids = _filter_qids_by_column(
            questions_excel, filter_column, filter_value
        )
        in_scope_qids = _apply_include_filters(
            questions_excel,
            in_scope_qids,
            include_qids,
            include_tags,
            context="preview",
        )
    changes: List[PreviewChange] = []

    # Question text changes
    for qid in sorted(in_scope_qids):
        q_row = questions_excel.get(qid)
        if not q_row or q_row.externally_managed_by:
            continue
        q_json = survey.questions.get(qid)
        if not q_json:
            continue
        question_text_json = q_json.get("QuestionText") or ""
        # For non-HTML questions, compare at the Markdown level.
        if not q_row.text_en_is_html:
            md_old = normalize_markdown_for_compare(html_to_md(question_text_json))
            md_new = normalize_markdown_for_compare(q_row.text_en_md or "")
            if md_old == md_new:
                continue
        old_html = normalize_text(question_text_json)
        new_html = normalize_text(excel_io.question_row_to_html(q_row))
        if old_html != new_html:
            tag = (q_json.get("DataExportTag") or "").strip() or None
            changes.append(
                PreviewChange(
                    kind="question",
                    qid=qid,
                    old_html=old_html,
                    new_html=new_html,
                    diff_lines=_diff_lines(old_html, new_html),
                    data_export_tag=tag,
                )
            )

    # Option label changes
    for (qid, choice_id), opt_row in options_excel.items():
        if qid not in in_scope_qids:
            continue
        q_json = survey.questions.get(qid)
        if not q_json:
            continue
        # Options for externally managed questions (e.g. dynamic newsmem batteries)
        # are owned by scripts and should not be driven from Excel.
        tag = (q_json.get("DataExportTag") or "").strip()
        if tag in excel_io.EXTERNALLY_MANAGED_TAGS:
            continue
        if opt_row.externally_managed_by:
            continue
        qtype = (opt_row.question_type or q_json.get("QuestionType") or "").strip()
        if qtype == "Matrix":
            container = q_json.get("Answers") or {}
        else:
            container = q_json.get("Choices") or {}
        choice = container.get(choice_id)
        if not choice:
            continue
        display_json_str = _display_to_str(choice)
        # For non-HTML labels, compare at the Markdown level.
        if not opt_row.label_en_is_html:
            md_old = normalize_markdown_for_compare(html_to_md(display_json_str))
            md_new = normalize_markdown_for_compare(opt_row.label_en_md or "")
            if md_old == md_new:
                continue
        old_html = normalize_text(display_json_str)
        new_html = normalize_text(excel_io.option_row_to_html(opt_row))
        if old_html != new_html:
            tag = (q_json.get("DataExportTag") or "").strip() or None
            changes.append(
                PreviewChange(
                    kind="option",
                    qid=qid,
                    choice_id=choice_id,
                    old_html=old_html,
                    new_html=new_html,
                    diff_lines=_diff_lines(old_html, new_html),
                    data_export_tag=tag,
                )
            )

    # Subitem label changes (matrix statements, sliders, etc.)
    for (qid, answer_id), sub_row in subitems_excel.items():
        if qid not in in_scope_qids:
            continue
        q_json = survey.questions.get(qid)
        if not q_json:
            continue
        # Subitems for externally managed questions (e.g. recognition, salience,
        # cued recall) are script-owned and should not be driven from Excel.
        tag = (q_json.get("DataExportTag") or "").strip()
        if tag in excel_io.EXTERNALLY_MANAGED_TAGS:
            continue
        qtype = (sub_row.question_type or q_json.get("QuestionType") or "").strip()
        if qtype == "Matrix":
            container = q_json.get("Choices") or {}
        else:
            container = q_json.get("Answers") or {}
        answer = container.get(answer_id)
        if not answer:
            continue
        display_json_str = _display_to_str(answer)
        # For non-HTML subitems, compare at the Markdown level.
        if not sub_row.label_en_is_html:
            md_old = normalize_markdown_for_compare(html_to_md(display_json_str))
            md_new = normalize_markdown_for_compare(sub_row.label_en_md or "")
            if md_old == md_new:
                continue
        old_html = normalize_text(display_json_str)
        new_html = normalize_text(excel_io.subitem_row_to_html(sub_row))
        if old_html != new_html:
            tag = (q_json.get("DataExportTag") or "").strip() or None
            changes.append(
                PreviewChange(
                    kind="subitem",
                    qid=qid,
                    answer_id=answer_id,
                    old_html=old_html,
                    new_html=new_html,
                    diff_lines=_diff_lines(old_html, new_html),
                    data_export_tag=tag,
                )
            )

    # Embedded data default changes
    if not skip_embedded:
        embedded_changes = _collect_embedded_data_changes(
            survey_id, survey.payload, Path(xlsx_path)
        )
        for change in embedded_changes:
            row = change["row"]
            old_display = _display_embedded_value(change["old_value"])
            new_display = _display_embedded_value(change["new_value"])
            # Include field name in diff context for better readability
            field_context = f"Field: {row.field}"
            changes.append(
                PreviewChange(
                    kind="embedded",
                    qid=row.field,
                    field=row.field,
                    flow_id=row.flow_id,
                    old_html=old_display,
                    new_html=new_display,
                    diff_lines=_diff_lines(
                        old_display, new_display, context=field_context
                    ),
                    is_dangerous=bool(change.get("is_dangerous")),
                )
            )

    if embedded_only:
        changes = [change for change in changes if change.kind == "embedded"]

    # Sort changes by survey flow order instead of QID number
    changes = _sort_changes_by_flow_order(survey, changes)

    # Update Dirty annotations in Excel so users can see changed fields.
    if annotate_dirty:
        _annotate_dirty_in_workbook(Path(xlsx_path), changes)

    return changes


def apply_changes(
    survey_id: str,
    xlsx_path: Path,
    filter_column: str | None = None,
    filter_value: str | None = None,
    include_qids: Set[str] | None = None,
    include_tags: Set[str] | None = None,
    allow_dangerous: bool = False,
    embedded_only: bool = False,
    skip_embedded: bool = False,
    scope_expr: str | None = None,
    allow_drift: bool = False,
    interactive: bool = True,
) -> ApplyResult:
    """Stage Excel diffs into the cached survey JSON (no Qualtrics writes).

    This is the engine behind `qsync apply`.

    Notes:
    - Creates a local backup of the cached survey JSON before modifying it.
    - Writes only to the local cache (`surveys/<SurveyID>.json`), not to Qualtrics.

    Args:
        survey_id: Qualtrics survey ID (e.g., `SV_xxx`).
        xlsx_path: Path to the Excel workbook for the survey.
        filter_column: Optional column on the Questions sheet used to scope QIDs.
        filter_value: Optional value for the filter column (defaults to `"TRUE"`).
        include_qids: Optional explicit set of QIDs to include.
        include_tags: Optional explicit set of DataExportTags to include.
        allow_dangerous: Allow embedded data edits for fields without defaults.
        skip_embedded: Skip staging embedded data changes entirely.
        scope_expr: Optional scope filter expression (e.g., 'qid:Q1 OR tag:baseline').

    Returns:
        ApplyResult with staged QIDs and embedded data fields.

    Example:
        >>> from pathlib import Path
        >>> from qsync.sync_core import apply_changes
        >>> result = apply_changes("SV_xxx", Path("excel/SV_xxx-mylabel.xlsx"))
        >>> isinstance(result.qids, list)
        True
    """

    enforce_no_drift(
        survey_id=survey_id,
        dimension="items",
        allow_drift=allow_drift,
        interactive=interactive,
    )

    ensure_backup(survey_id)

    workbook_path = Path(xlsx_path)
    survey: SurveyCache = load_cached_survey(survey_id)
    original_payload_text = survey.path.read_text(encoding="utf-8")

    try:
        questions_excel = excel_io.load_questions_from_workbook(workbook_path)
        options_excel = excel_io.load_options_from_workbook(workbook_path)
        subitems_excel = excel_io.load_subitems_from_workbook(workbook_path)

        # Apply scope filtering
        if scope_expr:
            scope_filter = ScopeFilter.parse(scope_expr)
            in_scope_qids = {
                qid
                for qid in questions_excel.keys()
                if scope_filter.matches(
                    qid=qid,
                    tags=(
                        [questions_excel[qid].data_export_tag]
                        if questions_excel[qid].data_export_tag
                        else None
                    ),
                )
            }
        else:
            in_scope_qids = _filter_qids_by_column(
                questions_excel, filter_column, filter_value
            )
            in_scope_qids = _apply_include_filters(
                questions_excel,
                in_scope_qids,
                include_qids,
                include_tags,
                context="apply",
            )

        changed_qids: Set[str] = set()

        if not embedded_only:
            # Apply question text changes into the survey payload
            for qid in sorted(in_scope_qids):
                q_row = questions_excel.get(qid)
                if not q_row or q_row.externally_managed_by:
                    continue
                q_json = survey.questions.get(qid)
                if not q_json:
                    continue
                question_text_json = q_json.get("QuestionText") or ""
                if not q_row.text_en_is_html:
                    md_old = normalize_markdown_for_compare(
                        html_to_md(question_text_json)
                    )
                    md_new = normalize_markdown_for_compare(q_row.text_en_md or "")
                    if md_old == md_new:
                        continue
                old_html = normalize_text(question_text_json)
                new_html = normalize_text(excel_io.question_row_to_html(q_row))
                if old_html != new_html:
                    # Lightweight HTML validation for raw HTML cells
                    if q_row.text_en_is_html:
                        errors = validate_html_fragment(new_html)
                        if errors:
                            print(
                                f"[qsync:apply] WARNING: Potential HTML issues in question {qid}:"
                            )
                            for err in errors:
                                print(f"  - {err}")
                            print(
                                "  Next: fix invalid HTML in the workbook cell (or switch to markdown mode) before pushing."
                            )
                    q_json["QuestionText"] = new_html
                    # Keep QuestionText_Unsafe in sync where present.
                    if "QuestionText_Unsafe" in q_json:
                        q_json["QuestionText_Unsafe"] = new_html
                    changed_qids.add(qid)

            # Apply option label changes into the survey payload
            for (qid, choice_id), opt_row in options_excel.items():
                if qid not in in_scope_qids:
                    continue
                q_json = survey.questions.get(qid)
                if not q_json:
                    continue
                tag = (q_json.get("DataExportTag") or "").strip()
                if tag in excel_io.EXTERNALLY_MANAGED_TAGS:
                    continue
                if opt_row.externally_managed_by:
                    continue
                qtype = (
                    opt_row.question_type or q_json.get("QuestionType") or ""
                ).strip()
                if qtype == "Matrix":
                    container = q_json.get("Answers") or {}
                else:
                    container = q_json.get("Choices") or {}
                choice = container.get(choice_id)
                if not choice:
                    continue
                display_json_str = _display_to_str(choice)
                if not opt_row.label_en_is_html:
                    md_old = normalize_markdown_for_compare(
                        html_to_md(display_json_str)
                    )
                    md_new = normalize_markdown_for_compare(opt_row.label_en_md or "")
                    if md_old == md_new:
                        continue
                old_html = normalize_text(display_json_str)
                new_html = normalize_text(excel_io.option_row_to_html(opt_row))
                if old_html != new_html:
                    if opt_row.label_en_is_html:
                        errors = validate_html_fragment(new_html)
                        if errors:
                            print(
                                f"[qsync:apply] WARNING: Potential HTML issues in option "
                                f"{qid}/{choice_id}:"
                            )
                            for err in errors:
                                print(f"  - {err}")
                            print(
                                "  Next: fix invalid HTML in the workbook cell (or switch to markdown mode) before pushing."
                            )
                    choice["Display"] = new_html
                    if "Display_Unsafe" in choice:
                        choice["Display_Unsafe"] = new_html
                    changed_qids.add(qid)

            # Apply subitem text changes into the survey payload
            for (qid, answer_id), sub_row in subitems_excel.items():
                if qid not in in_scope_qids:
                    continue
                q_json = survey.questions.get(qid)
                if not q_json:
                    continue
                tag = (q_json.get("DataExportTag") or "").strip()
                if tag in excel_io.EXTERNALLY_MANAGED_TAGS:
                    continue
                qtype = (
                    sub_row.question_type or q_json.get("QuestionType") or ""
                ).strip()
                if qtype == "Matrix":
                    container = q_json.get("Choices") or {}
                else:
                    container = q_json.get("Answers") or {}
                answer = container.get(answer_id)
                if not answer:
                    continue
                display_json_str = _display_to_str(answer)
                if not sub_row.label_en_is_html:
                    md_old = normalize_markdown_for_compare(
                        html_to_md(display_json_str)
                    )
                    md_new = normalize_markdown_for_compare(sub_row.label_en_md or "")
                    if md_old == md_new:
                        continue
                old_html = normalize_text(display_json_str)
                new_html = normalize_text(excel_io.subitem_row_to_html(sub_row))
                if old_html != new_html:
                    if sub_row.label_en_is_html:
                        errors = validate_html_fragment(new_html)
                        if errors:
                            print(
                                f"[qsync:apply] WARNING: Potential HTML issues in subitem "
                                f"{qid}/{answer_id}:"
                            )
                            for err in errors:
                                print(f"  - {err}")
                            print(
                                "  Next: fix invalid HTML in the workbook cell (or switch to markdown mode) before pushing."
                            )
                    answer["Display"] = new_html
                    if "Display_Unsafe" in answer:
                        answer["Display_Unsafe"] = new_html
                    changed_qids.add(qid)

        embedded_applied: List[dict[str, str]] = []
        embedded_skipped: List[dict] = []

        if not skip_embedded:
            embedded_changes = _collect_embedded_data_changes(
                survey_id, survey.payload, workbook_path
            )

            if embedded_changes:
                flow_id_map, ordered_nodes = _index_embedded_flow_nodes(survey.payload)
                for change in embedded_changes:
                    if change.get("is_dangerous") and not allow_dangerous:
                        embedded_skipped.append(change)
                        continue
                    row = change["row"]
                    new_value = change["new_value"]
                    if new_value is None:
                        continue
                    if row.flow_id:
                        node = flow_id_map.get(row.flow_id)
                        if not node:
                            raise ValueError(
                                f"Embedded data FlowID {row.flow_id} not found; re-run qsync init."
                            )
                        entry = _find_embedded_entry(node, row.field)
                        if not entry:
                            raise ValueError(
                                f"Embedded data field '{row.field}' not found in FlowID {row.flow_id}."
                            )
                        if entry.get("Type") == "Recipient":
                            print(
                                f"[qsync:apply] WARNING: Embedded data field '{row.field}' is "
                                "Recipient type; setting default value without changing Type. "
                                "Next: confirm this field is intended to remain Recipient type in SurveyFlow."
                            )
                        entry["Value"] = new_value
                        embedded_applied.append(
                            {"flow_id": row.flow_id or "", "field": row.field}
                        )
                    else:
                        node = _ensure_embedded_data_node(
                            survey.payload, ordered_nodes=ordered_nodes
                        )
                        entry = _find_embedded_entry(node, row.field)
                        if entry:
                            entry["Value"] = new_value
                        else:
                            template = None
                            existing = node.get("EmbeddedData") or []
                            if existing:
                                template = existing[0]
                            node.setdefault("EmbeddedData", []).append(
                                _make_embedded_entry(
                                    row.field, new_value, template=template
                                )
                            )
                        embedded_applied.append(
                            {
                                "flow_id": str(node.get("FlowID") or ""),
                                "field": row.field,
                            }
                        )

        if embedded_skipped:
            skipped_fields = ", ".join(
                sorted({c["row"].field for c in embedded_skipped})
            )
            message = (
                f"[qsync:apply] WARNING: {ERROR_ID_EMBEDDED_DANGEROUS_SKIPPED}: "
                "Dangerous embedded data default changes were skipped "
                f"(fields: {skipped_fields}). Re-run with --allow-dangerous to apply.\n"
                "  Why this is flagged as dangerous:\n"
                "    - These fields currently have no default value in SurveyFlow.\n"
                "    - Adding defaults can change logic that depends on a field being empty/unset.\n"
                "  Impact:\n"
                "    - May alter branching/display logic and downstream exports/analysis.\n"
                "  How to proceed safely:\n"
                "    1. Verify no survey logic depends on these fields being initially empty.\n"
                "    2. Review Branch Logic / Display Logic that references these fields.\n"
                "    3. Re-run with --allow-dangerous if the change is intentional.\n"
            )
            print(message)

            from ..push_logger import log_push_event

            # Derive a stable workspace root even when QSYNC_ROOT isn't set.
            root = survey.path.parent.parent
            log_push_event(
                action="qsync.apply.embedded_dangerous_skipped",
                method="LOCAL",
                path="sync_core.apply_changes",
                survey_id=survey_id,
                status=None,
                error={
                    "error_id": ERROR_ID_EMBEDDED_DANGEROUS_SKIPPED,
                    "message": "Dangerous embedded data default changes were skipped",
                    "fields": sorted({c["row"].field for c in embedded_skipped}),
                },
                root=root,
            )

        if not changed_qids and not embedded_applied:
            print("[qsync:apply] No changes to stage.")
            return ApplyResult(qids=[], embedded_fields=[])

        sorted_qids = sorted(changed_qids)
        survey.save()
        from ..survey_ref import format_survey_ref

        survey_ref = format_survey_ref(survey_id)
        if embedded_applied and sorted_qids:
            print(
                f"[qsync:apply] Staged {len(sorted_qids)} question(s) and "
                f"{len(embedded_applied)} embedded field(s) in survey {survey_ref}. "
                "Run 'qsync push' to upload them to Qualtrics."
            )
        elif embedded_applied:
            print(
                f"[qsync:apply] Staged {len(embedded_applied)} embedded field(s) in survey "
                f"{survey_ref}. Run 'qsync push' to upload them to Qualtrics."
            )
        else:
            print(
                f"[qsync:apply] Staged {len(sorted_qids)} question(s) in survey {survey_ref}. "
                "Run 'qsync push' to upload them to Qualtrics."
            )
        return ApplyResult(qids=sorted_qids, embedded_fields=embedded_applied)
    except Exception:
        survey.path.write_text(original_payload_text, encoding="utf-8")
        raise


def _apply_pending_item_changes(
    survey_payload: dict,
    *,
    pending_changes: list[dict[str, object]] | None,
    embedded_changes: list[dict[str, object]] | None,
    allow_dangerous: bool = False,
) -> tuple[set[str], list[dict[str, str]]]:
    changed_qids: set[str] = set()
    embedded_applied: list[dict[str, str]] = []
    questions = (survey_payload.get("result") or {}).get("Questions") or {}

    for change in pending_changes or []:
        kind = str(change.get("kind") or "")
        qid = str(change.get("qid") or "").strip()
        if not qid:
            continue
        question = questions.get(qid)
        if not question:
            continue
        if kind == "question":
            new_html = str(change.get("new_html") or "")
            question["QuestionText"] = new_html
            if "QuestionText_Unsafe" in question:
                question["QuestionText_Unsafe"] = new_html
            changed_qids.add(qid)
            continue
        qtype = (question.get("QuestionType") or "").strip()
        if kind == "option":
            choice_id = str(change.get("choice_id") or "").strip()
            if not choice_id:
                continue
            container = (
                question.get("Answers") or {}
                if qtype == "Matrix"
                else question.get("Choices") or {}
            )
            choice = container.get(choice_id)
            if not choice:
                continue
            new_html = str(change.get("new_html") or "")
            choice["Display"] = new_html
            if "Display_Unsafe" in choice:
                choice["Display_Unsafe"] = new_html
            changed_qids.add(qid)
            continue
        if kind == "subitem":
            answer_id = str(change.get("answer_id") or "").strip()
            if not answer_id:
                continue
            container = (
                question.get("Choices") or {}
                if qtype == "Matrix"
                else question.get("Answers") or {}
            )
            answer = container.get(answer_id)
            if not answer:
                continue
            new_html = str(change.get("new_html") or "")
            answer["Display"] = new_html
            if "Display_Unsafe" in answer:
                answer["Display_Unsafe"] = new_html
            changed_qids.add(qid)

    if embedded_changes:
        flow_id_map, ordered_nodes = _index_embedded_flow_nodes(survey_payload)
        for change in embedded_changes:
            if change.get("is_dangerous") and not allow_dangerous:
                continue
            field = str(change.get("field") or "").strip()
            if not field:
                continue
            flow_id = str(change.get("flow_id") or "").strip()
            if "new_value" not in change or change.get("new_value") is None:
                embedded_applied.append({"flow_id": flow_id, "field": field})
                continue
            new_value = change.get("new_value")
            if flow_id:
                node = flow_id_map.get(flow_id)
                if not node:
                    continue
                entry = _find_embedded_entry(node, field)
                if not entry:
                    continue
                entry["Value"] = new_value
                embedded_applied.append({"flow_id": flow_id, "field": field})
            else:
                node = _ensure_embedded_data_node(
                    survey_payload, ordered_nodes=ordered_nodes
                )
                entry = _find_embedded_entry(node, field)
                if entry:
                    entry["Value"] = new_value
                else:
                    template = None
                    existing = node.get("EmbeddedData") or []
                    if existing:
                        template = existing[0]
                    node.setdefault("EmbeddedData", []).append(
                        _make_embedded_entry(field, new_value, template=template)
                    )
                embedded_applied.append(
                    {"flow_id": str(node.get("FlowID") or ""), "field": field}
                )

    return changed_qids, embedded_applied


def push_staged_changes(
    survey_id: str,
    qids: list[str],
    *,
    embedded_fields: list[dict[str, object]] | None = None,
    pending_changes: list[dict[str, object]] | None = None,
    workbook: str | None = None,
    filter_column: str | None = None,
    filter_value: str | None = None,
    publish: bool = True,
    publish_description: str | None = None,
    force_live: bool = False,
    force_preview: bool = False,
    interactive: bool = True,
    allow_drift: bool = False,
    skip_drift_check: bool = False,
) -> None:
    """Push staged question and SurveyFlow changes to Qualtrics.

    Args:
        survey_id: Qualtrics survey ID
        qids: List of question IDs to push
        embedded_fields: Optional list of embedded data field updates
        pending_changes: Optional staged item change list (pending vs cache)
        workbook: Optional workbook path for publish description
        filter_column: Optional filter column used
        filter_value: Optional filter value used
        publish: Whether to publish after push
        publish_description: Optional custom publish description
        force_live: Allow push despite live responses
        force_preview: Allow push with preview/test responses
        interactive: Whether to use interactive prompts
    """

    if not skip_drift_check:
        enforce_no_drift(
            survey_id=survey_id,
            dimension="items",
            allow_drift=allow_drift,
            interactive=interactive,
        )

    embedded_fields = embedded_fields or []
    if not qids and not embedded_fields:
        from ..survey_ref import format_survey_ref

        print(
            f"[qsync:push] No staged changes found for {format_survey_ref(survey_id)}."
        )
        return

    # Enforce push safeguards
    from ..survey_ref import format_survey_ref

    survey_ref = format_survey_ref(survey_id)
    config = SafeguardConfig(
        survey_id=survey_id,
        dimension="items",
        force_live=force_live,
        force_preview=force_preview,
        auto_yes=not interactive,
    )
    safeguard_result = enforce_push_safeguards(config)
    if safeguard_result.blocked:
        raise SystemExit(f"[qsync:items] Push blocked: {safeguard_result.message}")
    if safeguard_result.warnings:
        for warning in safeguard_result.warnings:
            print(f"[qsync:items] WARNING: {warning}")

    ensure_backup(survey_id)
    survey: SurveyCache = load_cached_survey(survey_id)
    changed_qids, embedded_applied = _apply_pending_item_changes(
        survey.payload,
        pending_changes=pending_changes,
        embedded_changes=embedded_fields,
        allow_dangerous=False,
    )
    sorted_qids = sorted(changed_qids or qids)
    embedded_fields = embedded_applied or embedded_fields or []

    push_context = {
        "origin": "qsync.push_staged_changes",
        "workbook": workbook,
        "filter": {"column": filter_column, "value": filter_value},
        "changed_qids": sorted_qids,
        "changed_count": len(sorted_qids),
        "embedded_fields": embedded_fields,
        "embedded_count": len(embedded_fields),
    }

    if sorted_qids:
        push_questions(survey, sorted_qids, context=push_context)
    if embedded_fields:
        push_survey_flow(survey, context=push_context)

    # Use auto-publish module
    if publish:
        auto_publish_after_push(
            survey_id=survey_id,
            dimension="items",
            changed_qids=sorted_qids,
            count=len(sorted_qids),
            skip_publish=False,
            auto_yes=not interactive,
        )
    else:
        # No publish, just report upload
        if sorted_qids and embedded_fields:
            print(
                f"[qsync:push] Uploaded {len(sorted_qids)} question(s) and "
                f"{len(embedded_fields)} embedded field(s) for {survey_ref}."
            )
        elif embedded_fields:
            print(
                f"[qsync:push] Uploaded {len(embedded_fields)} embedded field(s) for {survey_ref}."
            )
        else:
            print(
                f"[qsync:push] Uploaded {len(sorted_qids)} question(s) for {survey_ref}."
            )
