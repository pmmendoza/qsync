"""Survey translations workflow helpers (pull/preview/apply/push/doctor/drift)."""

from __future__ import annotations

import difflib
import json
import re
import sys
from datetime import datetime, timezone
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from ..api_push import send_api_request
from ..config import get_client_config
from ..errors import QsyncValidationError
from .. import excel_io
from ..pending_stage import (
    PendingStagedChanges,
    TranslationsPendingPayload,
    load_pending,
    save_pending,
    clear_pending,
)
from ..terminal_output import info, success, warn, prompt_yes_no
from ..qualtrics_client import (
    ensure_backup,
    load_cached_survey,
    publish_survey_definition,
    SurveyCache,
    push_questions,
)
from ..drift_check import check_drift, enforce_no_drift
from ..push_safeguards import enforce_push_safeguards, SafeguardConfig
from ..auto_publish import auto_publish_after_push
from ..scope_filter import ScopeFilter
from ..workbook_resolver import WorkbookResolver
from openpyxl import load_workbook
from ..survey_inventory import load_inventory_record
from ..translation_snapshots import translation_key_snapshot_path
from ..translations_utils import normalize_language_code, normalize_language_list
from .translations_language_blocks import (
    get_base_language as get_base_language_from_options,
    list_enabled_languages as list_enabled_languages_from_options,
    write_answer_display,
    write_choice_display,
    write_label_display,
    write_question_text,
    write_sbs_column_answer_display,
    write_sbs_column_question_text,
)
from .translations_workbook_extract import (
    SURVEY_METADATA_QID,
    _normalize_field,
    build_base_value_map_for_keys,
    build_workbook_value_map,
    diff_workbook_vs_cache,
    resolve_languages_from_workbook,
)

import requests

# Explicitly export private functions needed by other qsync modules
__all__ = [
    "TranslationDiff",
    "TranslationDoctorReport",
    "list_enabled_languages",
    "ensure_languages",
    "set_languages",
    "fetch_base_language",
    "load_pending_languages",
    "resolve_languages_for_cli",
    "translation_key_snapshot_path",
    "preview_translations",
    "apply_translations",
    "push_translations",
    "run_translation_doctor",
    "drift_translations",
    "_check_large_deltas",
    "_check_placeholders",
    "_check_html_hazards",
    "_check_value_length_limit",
    "_coverage_stats",
    "_coverage_stats_with_allowed_empties",
    "_normalize_language_list",
]


_PLACEHOLDER_RE = re.compile(r"\$\{e://Field/[^}]+\}")
# Basic HTML safety check for translation content. We intentionally scope the `on...=`
# detection to *attribute-like* occurrences (word boundary), to avoid false positives
# inside URLs like `...expirationTimestamp=...`.
_HTML_HAZARD_RE = re.compile(
    r"(<\s*script|<\s*form|\bon[a-z0-9_-]+\s*=)", re.IGNORECASE
)

# Qualtrics API rejects individual translation values above this length.
# Observed error: QVAL_3 "Parameter <key> exceeds maximum length of 10000."
QUALTRICS_TRANSLATION_VALUE_MAX_CHARS = 10_000


@dataclass(frozen=True)
class TranslationDiff:
    missing_keys: set[str]
    extra_keys: set[str]
    changed_keys: set[str]

    @property
    def is_empty(self) -> bool:
        return not (self.missing_keys or self.extra_keys or self.changed_keys)


@dataclass(frozen=True)
class TranslationDoctorReport:
    errors: list[str]
    warnings: list[str]
    coverage: dict[str, dict[str, int]]

    @property
    def ok(self) -> bool:
        return not self.errors


def _normalize_language_list(languages: Iterable[str] | None) -> list[str]:
    return normalize_language_list(languages)


def _resolve_translation_languages(
    survey_id: str,
    *,
    explicit_languages: Sequence[str] | None = None,
    interactive: bool = True,
) -> list[str]:
    """
    Resolve which languages to pull/push using 5-tier precedence with interactive prompts.

    Precedence (highest to lowest):
    1. Explicit --languages flag
    2. Staged/pending translations record
    3. Workbook columns present (text_xx, Label_XX_MD)
    4. Enabled languages in Qualtrics (API)

    When sources disagree and interactive=True, prompts user for resolution.

    Args:
        survey_id: Survey ID
        explicit_languages: Explicit languages from CLI flag
        interactive: If True, prompt user when sources disagree

    Returns:
        Normalized list of language codes

    Raises:
        QsyncValidationError: If no languages found via any source
    """
    sources: dict[str, list[str]] = {}

    # Tier 1: Explicit CLI flag (highest priority)
    if explicit_languages:
        return _normalize_language_list(explicit_languages)

    # Tier 2: Staged/pending translations record
    pending = load_pending(survey_id, "translations")
    if pending and isinstance(pending.payload, TranslationsPendingPayload):
        if pending.payload.languages:
            sources["pending"] = _normalize_language_list(pending.payload.languages)

    # Tier 3: Workbook columns
    try:
        resolver = WorkbookResolver()
        workbook_path = resolver.default_path(survey_id)
        if workbook_path.exists():
            wb = load_workbook(workbook_path, data_only=True)
            workbook_langs = resolve_languages_from_workbook(wb)
            if workbook_langs:
                sources["workbook"] = _normalize_language_list(workbook_langs)
    except Exception as exc:
        warn("[qsync:translations]", f"Could not read workbook: {exc}")

    # Tier 4: Enabled languages from API
    try:
        api_langs = list_enabled_languages(survey_id)
        if api_langs:
            sources["api"] = _normalize_language_list(api_langs)
    except Exception as exc:
        warn("[qsync:translations]", f"Could not fetch API languages: {exc}")

    # No languages found
    if not sources:
        raise QsyncValidationError(
            error_id="QSYNC-TRANSLATIONS-LANG-001",
            problem="No translation languages found.",
            why="No languages specified via CLI, pending record, workbook, or API.",
            impact="Cannot proceed with translation operation.",
            action="Enable languages in Qualtrics or specify --languages explicitly.",
            context={"survey_id": survey_id},
        )

    # Single source: use it directly (deterministic)
    if len(sources) == 1:
        source_name, langs = list(sources.items())[0]
        info(
            "[qsync:translations]",
            f"Using languages from {source_name}: {', '.join(langs)}",
        )
        return langs

    # Multiple sources: check for agreement
    all_lang_sets = [set(langs) for langs in sources.values()]
    if all(lang_set == all_lang_sets[0] for lang_set in all_lang_sets):
        # All sources agree
        langs = list(sources.values())[0]
        info(
            "[qsync:translations]",
            f"All sources agree on languages: {', '.join(langs)}",
        )
        return langs

    # Sources disagree: use highest precedence (deterministic order)
    precedence = ["pending", "workbook", "api", "disk"]
    for source_name in precedence:
        if source_name in sources:
            langs = sources[source_name]

            if interactive:
                # Show proposed resolution with conflicts
                conflicts = []
                for other_name, other_langs in sources.items():
                    if other_name != source_name and set(other_langs) != set(langs):
                        conflicts.append(f"{other_name}: [{', '.join(other_langs)}]")

                if conflicts:
                    info("[qsync:translations]", "Language sources disagree:")
                    info(
                        "[qsync:translations]",
                        f"  Proposed: {source_name}: [{', '.join(langs)}]",
                    )
                    for conflict in conflicts:
                        info("[qsync:translations]", f"  Found:    {conflict}")

                    if prompt_yes_no(
                        f"Use {source_name} languages ({', '.join(langs)})?",
                        default=True,
                    ):
                        return langs
                    else:
                        raise QsyncValidationError(
                            error_id="QSYNC-TRANSLATIONS-LANG-002",
                            problem="User aborted language resolution.",
                            why="Sources disagreed and user declined proposed resolution.",
                            impact="Cannot proceed with translation operation.",
                            action="Specify --languages explicitly or resolve conflicts manually.",
                            context={"survey_id": survey_id, "sources": sources},
                        )

            # Non-interactive: use deterministic precedence
            info(
                "[qsync:translations]",
                f"Using languages from {source_name} (precedence): {', '.join(langs)}",
            )
            return langs

    # Fallback (should never reach here)
    raise QsyncValidationError(
        error_id="QSYNC-TRANSLATIONS-LANG-003",
        problem="No valid language source found.",
        why="All resolution tiers failed or returned empty.",
        impact="Cannot proceed with translation operation.",
        action="Specify --languages explicitly.",
        context={"survey_id": survey_id},
    )


def _resolve_stage_languages(
    survey_id: str,
    survey_payload: dict,
    workbook_path: Path,
    explicit_languages: Sequence[str] | None,
    *,
    allow_empty: bool = False,
    emit_warnings: bool = True,
) -> list[str]:
    base_language = get_base_language_from_options(survey_payload)
    if explicit_languages:
        langs = normalize_language_list(explicit_languages)
    else:
        enabled = list_enabled_languages_from_options(survey_payload)
        enabled_non_base = [
            lang
            for lang in normalize_language_list(enabled)
            if not base_language or lang != base_language
        ]
        langs = list(enabled_non_base)
        try:
            wb = load_workbook(workbook_path, data_only=True)
            workbook_langs = resolve_languages_from_workbook(wb)
            workbook_non_base = [
                lang
                for lang in normalize_language_list(workbook_langs)
                if not base_language or lang != base_language
            ]
            if workbook_non_base:
                stale = [lang for lang in workbook_non_base if lang not in set(langs)]
                if stale and emit_warnings:
                    warn(
                        "[qsync:translations]",
                        "Ignoring workbook-only translation columns not enabled online: "
                        + ", ".join(stale),
                    )
                missing = [lang for lang in langs if lang not in set(workbook_non_base)]
                if missing and emit_warnings:
                    warn(
                        "[qsync:translations]",
                        "Workbook is missing columns for enabled translation languages: "
                        + ", ".join(missing)
                        + ". Run `qsync items pull --survey-id "
                        + survey_id
                        + "` to refresh.",
                    )
        except Exception as exc:
            warn("[qsync:translations]", f"Could not read workbook languages: {exc}")
    if base_language and base_language in langs:
        warn(
            "[qsync:translations]",
            f"Ignoring base language {base_language} (managed via items workflow).",
        )
        langs = [lang for lang in langs if lang != base_language]
    if not langs:
        if allow_empty:
            return []
        raise QsyncValidationError(
            error_id="QSYNC-TRANSLATIONS-LANG-004",
            problem="No non-base translation languages resolved.",
            why="No translation columns were found in the workbook and no enabled non-base languages were available.",
            impact="Cannot stage translations.",
            action="Enable languages in Qualtrics or add translation columns to the workbook.",
            context={"survey_id": survey_id, "base_language": base_language},
        )
    return langs


def _load_json(path: Path) -> dict[str, Any]:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise QsyncValidationError(
            error_id="QSYNC-TRANSLATIONS-JSON-001",
            problem=f"Invalid JSON in {path}",
            why=str(exc),
            impact="Translation maps could not be loaded.",
            action="Fix the JSON file and retry.",
        ) from exc
    if not isinstance(payload, dict):
        raise QsyncValidationError(
            error_id="QSYNC-TRANSLATIONS-JSON-002",
            problem=f"Translation map at {path} is not a JSON object",
            why=f"Expected dict, found {type(payload).__name__}.",
            impact="Translation maps could not be loaded.",
            action="Ensure the translation file is a JSON object of key -> string.",
        )
    return payload


def _write_json(path: Path, payload: Mapping[str, Any], *, backup: bool) -> None:
    serialized = json.dumps(payload, indent=2, ensure_ascii=True, sort_keys=True) + "\n"
    if path.exists():
        existing = path.read_text(encoding="utf-8")
        if existing == serialized:
            return
        if backup:
            backup_path = path.with_suffix(path.suffix + ".bak")
            backup_path.write_text(existing, encoding="utf-8")
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(serialized, encoding="utf-8")


def list_enabled_languages(
    survey_id: str,
    *,
    base_url: str | None = None,
    headers: dict | None = None,
) -> list[str]:
    if base_url is None or headers is None:
        base_url, headers = get_client_config()
    resp = send_api_request(
        action="qsync.translations.languages.list",
        method="GET",
        base_url=base_url,
        headers=headers,
        path=f"surveys/{survey_id}/languages",
        survey_id=survey_id,
        log_event=False,
        timeout=30,
    )
    result = resp.json().get("result") or {}
    langs = result.get("AvailableLanguages") or result.get("languages") or []
    return _normalize_language_list(langs)


def set_languages(
    survey_id: str, languages: Sequence[str], *, dry_run: bool = False
) -> list[str]:
    target = _normalize_language_list(languages)
    if dry_run:
        return target
    base_url, headers = get_client_config()
    send_api_request(
        action="qsync.translations.languages.set",
        method="PUT",
        base_url=base_url,
        headers=headers,
        path=f"surveys/{survey_id}/languages",
        survey_id=survey_id,
        json={"AvailableLanguages": target},
        timeout=30,
    )
    return target


def snapshot_translation_keys(survey_id: str, language: str, *, label: str) -> Path:
    lang = normalize_language_code(language)
    from ..qualtrics_client import refresh_survey_cache
    from ..translation_export import build_translation_map_from_cache

    survey, _ = refresh_survey_cache(survey_id)
    base_lang = normalize_language_code(
        get_base_language_from_options(survey.payload) or ""
    )
    if not base_lang:
        base_lang = lang or "EN"
    payload = build_translation_map_from_cache(
        survey.payload,
        language=lang,
        base_language=base_lang,
    )
    snapshot = {
        "survey_id": survey_id,
        "language": lang,
        "label": label,
        "pulled_at": datetime.now(timezone.utc).isoformat(),
        "keys": sorted(payload.keys()),
        "count": len(payload),
    }
    path = translation_key_snapshot_path(survey_id, label, lang)
    _write_json(path, snapshot, backup=False)
    return path


def load_translation_key_snapshot(path: Path) -> dict[str, Any]:
    return _load_json(path)


def diff_translation_key_snapshots(
    before: Mapping[str, Any], after: Mapping[str, Any]
) -> TranslationDiff:
    before_keys = set(before.get("keys") or [])
    after_keys = set(after.get("keys") or [])
    missing = before_keys - after_keys
    extra = after_keys - before_keys
    return TranslationDiff(missing_keys=missing, extra_keys=extra, changed_keys=set())


def _survey_web_root(base_url: str) -> str:
    root = base_url.strip()
    if not root.startswith("http"):
        root = "https://" + root.lstrip("/")
    if "/API/v3" in root:
        root = root.split("/API/v3", 1)[0]
    return root.rstrip("/")


def _fetch_survey_name(survey_id: str) -> str:
    base_url, headers = get_client_config()
    resp = send_api_request(
        action="qsync.translations.survey.name",
        method="GET",
        base_url=base_url,
        headers=headers,
        path=f"surveys/{survey_id}",
        survey_id=survey_id,
        log_event=False,
        timeout=30,
    )
    return resp.json().get("result", {}).get("name", survey_id)


def _ensure_smoke_survey(survey_id: str, *, allow_non_smoke: bool) -> None:
    if allow_non_smoke:
        return
    name = _fetch_survey_name(survey_id)
    if "smoke" not in name.lower():
        raise QsyncValidationError(
            error_id="QSYNC-TRANSLATIONS-SMOKE-001",
            problem=f"Survey '{name}' does not look like a smoke test survey.",
            why="Key stability and publish checks should only run against smoke surveys.",
            impact="Operation aborted to avoid unintended survey changes.",
            action="Use a smoke survey (name contains 'smoke') or pass --allow-non-smoke.",
            context={"survey_id": survey_id, "survey_name": name},
        )


def _runtime_contains_marker(
    survey_id: str,
    language: str,
    marker: str,
    *,
    timeout: int = 30,
) -> bool:
    base_url, _ = get_client_config()
    web_root = _survey_web_root(base_url)
    lang = normalize_language_code(language)
    url = f"{web_root}/jfe/form/{survey_id}?Q_Language={lang}"
    resp = requests.get(url, timeout=timeout)
    if resp.status_code >= 400:
        warn(
            "[qsync:translations]",
            f"Runtime check failed ({resp.status_code}) for {url}",
        )
        return False
    return marker in resp.text


def run_publish_requirement_check(
    survey_id: str,
    language: str,
    *,
    key: str | None = None,
    marker: str | None = None,
    publish: bool = False,
    publish_description: str | None = None,
    restore: bool = True,
    publish_restore: bool = True,
    allow_non_smoke: bool = False,
) -> dict[str, Any]:
    _ensure_smoke_survey(survey_id, allow_non_smoke=allow_non_smoke)
    lang = normalize_language_code(language)
    base_lang = fetch_base_language(survey_id)
    if lang == base_lang:
        raise QsyncValidationError(
            error_id="QSYNC-TRANSLATIONS-PUBLISH-001",
            problem=f"Publish check requested for base language {lang}.",
            why="Base language edits are handled via the items workflow.",
            impact="Publish check cannot proceed.",
            action="Use a non-base language (e.g., FR/NL/CS).",
        )
    from ..qualtrics_client import refresh_survey_cache
    from ..translation_export import build_translation_map_from_cache

    survey, _ = refresh_survey_cache(survey_id)
    base_lang = normalize_language_code(
        get_base_language_from_options(survey.payload) or base_lang
    )
    original_map = build_translation_map_from_cache(
        survey.payload,
        language=lang,
        base_language=base_lang,
    )
    if not original_map:
        raise QsyncValidationError(
            error_id="QSYNC-TRANSLATIONS-PUBLISH-002",
            problem=f"No translation keys found for {survey_id}/{lang}.",
            why="Cached survey definition returned an empty translation map.",
            impact="Publish check cannot proceed.",
            action="Enable the language in Qualtrics and refresh the cache.",
        )

    selected_key = key
    if not selected_key:
        for candidate, value in original_map.items():
            if str(value or "").strip():
                selected_key = candidate
                break
    if not selected_key:
        selected_key = next(iter(original_map.keys()))

    marker = (
        marker
        or f"[qsync-publish-check-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}]"
    )
    original_value = str(original_map.get(selected_key, "") or "")
    new_value = (
        f"{original_value} {marker}".strip()
        if marker not in original_value
        else original_value
    )

    change = _change_for_translation_key(
        selected_key,
        lang,
        old_value=original_value,
        new_value=new_value,
    )
    if not change:
        raise QsyncValidationError(
            error_id="QSYNC-TRANSLATIONS-PUBLISH-003",
            problem=f"Unsupported translation key '{selected_key}'.",
            why="Publish check requires a QuestionText/Choice/Answer/Label or metadata key.",
            impact="Publish check cannot proceed.",
            action="Use a translation key like QID10_QuestionText or SurveyTitle.",
            context={"survey_id": survey_id, "language": lang, "key": selected_key},
        )

    ensure_backup(survey_id)
    qids, _, metadata_keys = _apply_translation_changes_to_payload(
        survey.payload, [change]
    )
    if qids:
        push_questions(
            survey,
            sorted(qids),
            context={
                "origin": "qsync.translations.publish_check",
                "language": lang,
                "key": selected_key,
            },
        )
    if metadata_keys:
        _push_survey_options_for_metadata(
            survey,
            sorted(metadata_keys),
            context={
                "origin": "qsync.translations.publish_check",
                "language": lang,
                "key": selected_key,
            },
        )

    pre_publish_visible = _runtime_contains_marker(survey_id, lang, marker)
    post_publish_visible = None
    if publish:
        desc = publish_description or f"qsync publish check ({lang})"
        publish_survey_definition(
            survey_id,
            description=desc,
            context={"origin": "qsync.translations.publish_check", "language": lang},
        )
        post_publish_visible = _runtime_contains_marker(survey_id, lang, marker)

    if restore:
        restore_change = dict(change)
        restore_change["new_value"] = original_value
        _apply_translation_changes_to_payload(survey.payload, [restore_change])
        if qids:
            push_questions(
                survey,
                sorted(qids),
                context={
                    "origin": "qsync.translations.publish_check.restore",
                    "language": lang,
                    "key": selected_key,
                },
            )
        if metadata_keys:
            _push_survey_options_for_metadata(
                survey,
                sorted(metadata_keys),
                context={
                    "origin": "qsync.translations.publish_check.restore",
                    "language": lang,
                    "key": selected_key,
                },
            )
        if publish and publish_restore:
            desc = publish_description or f"qsync publish check restore ({lang})"
            publish_survey_definition(
                survey_id,
                description=desc[:140],
                context={
                    "origin": "qsync.translations.publish_check.restore",
                    "language": lang,
                },
            )

    return {
        "survey_id": survey_id,
        "language": lang,
        "key": selected_key,
        "marker": marker,
        "pre_publish_visible": pre_publish_visible,
        "post_publish_visible": post_publish_visible,
    }


def _fetch_question_payload(survey_id: str, question_id: str) -> dict[str, Any]:
    base_url, headers = get_client_config()
    resp = send_api_request(
        action="qsync.translations.keycheck.fetch_question",
        method="GET",
        base_url=base_url,
        headers=headers,
        path=f"survey-definitions/{survey_id}/questions/{question_id}",
        survey_id=survey_id,
        log_event=False,
        timeout=30,
    )
    result = resp.json().get("result")
    if not isinstance(result, dict):
        raise QsyncValidationError(
            error_id="QSYNC-TRANSLATIONS-KEYCHECK-001",
            problem=f"Question {question_id} payload missing.",
            why="Qualtrics did not return a question result.",
            impact="Key stability check cannot proceed.",
            action="Verify the question ID and retry.",
            context={"survey_id": survey_id, "question_id": question_id},
        )
    return result


def _push_question_payload(
    survey_id: str, question_id: str, payload: dict[str, Any]
) -> None:
    base_url, headers = get_client_config()
    send_api_request(
        action="qsync.translations.keycheck.push_question",
        method="PUT",
        base_url=base_url,
        headers=headers,
        path=f"survey-definitions/{survey_id}/questions/{question_id}",
        survey_id=survey_id,
        json=payload,
        timeout=30,
    )


def _swap_choice_order(payload: dict[str, Any]) -> None:
    order = list(payload.get("ChoiceOrder") or [])
    if len(order) < 2:
        raise QsyncValidationError(
            error_id="QSYNC-TRANSLATIONS-KEYCHECK-002",
            problem="Question has fewer than two choices.",
            why="Cannot reorder choices without at least two entries.",
            impact="Key stability check aborted.",
            action="Select a question with at least two choices.",
        )
    order[0], order[1] = order[1], order[0]
    payload["ChoiceOrder"] = order


def _add_choice(payload: dict[str, Any], *, display: str) -> str:
    choices = payload.get("Choices")
    if not isinstance(choices, dict):
        raise QsyncValidationError(
            error_id="QSYNC-TRANSLATIONS-KEYCHECK-003",
            problem="Question has no Choices map.",
            why="Cannot add a choice to this question type.",
            impact="Key stability check aborted.",
            action="Select a multiple-choice style question.",
        )
    existing_ids = [int(c) for c in choices.keys() if str(c).isdigit()]
    next_id = payload.get("NextChoiceId")
    if isinstance(next_id, int):
        new_id = next_id
    elif existing_ids:
        new_id = max(existing_ids) + 1
    else:
        new_id = 1
    choices[str(new_id)] = {"Display": display}
    order = list(payload.get("ChoiceOrder") or [])
    if str(new_id) not in order:
        order.append(str(new_id))
    payload["ChoiceOrder"] = order
    payload["NextChoiceId"] = int(new_id) + 1
    return str(new_id)


def _remove_choice(payload: dict[str, Any], choice_id: str) -> None:
    choices = payload.get("Choices")
    if not isinstance(choices, dict):
        return
    choices.pop(choice_id, None)
    order = list(payload.get("ChoiceOrder") or [])
    payload["ChoiceOrder"] = [cid for cid in order if cid != choice_id]


def run_key_stability_check_publish(
    survey_id: str,
    language: str,
    *,
    label: str,
    publish_description: str | None = None,
    allow_non_smoke: bool = False,
) -> dict[str, Any]:
    _ensure_smoke_survey(survey_id, allow_non_smoke=allow_non_smoke)
    lang = normalize_language_code(language)
    before_path = snapshot_translation_keys(survey_id, lang, label=f"{label}_pre")
    desc = publish_description or f"qsync key check publish ({lang})"
    publish_survey_definition(
        survey_id,
        description=desc[:140],
        context={"origin": "qsync.translations.keycheck.publish", "language": lang},
    )
    after_path = snapshot_translation_keys(survey_id, lang, label=f"{label}_post")
    diff = diff_translation_key_snapshots(
        load_translation_key_snapshot(before_path),
        load_translation_key_snapshot(after_path),
    )
    return {
        "before": before_path,
        "after": after_path,
        "missing": sorted(diff.missing_keys),
        "extra": sorted(diff.extra_keys),
    }


def run_key_stability_check_reorder(
    survey_id: str,
    language: str,
    *,
    question_id: str,
    label: str,
    publish_description: str | None = None,
    allow_non_smoke: bool = False,
) -> dict[str, Any]:
    _ensure_smoke_survey(survey_id, allow_non_smoke=allow_non_smoke)
    lang = normalize_language_code(language)
    original = _fetch_question_payload(survey_id, question_id)
    working = json.loads(json.dumps(original))

    before_path = snapshot_translation_keys(survey_id, lang, label=f"{label}_pre")
    _swap_choice_order(working)
    _push_question_payload(survey_id, question_id, working)
    desc = publish_description or f"qsync key check reorder ({question_id})"
    publish_survey_definition(
        survey_id,
        description=desc[:140],
        context={
            "origin": "qsync.translations.keycheck.reorder",
            "question_id": question_id,
        },
    )
    after_path = snapshot_translation_keys(survey_id, lang, label=f"{label}_post")

    _push_question_payload(survey_id, question_id, original)
    publish_survey_definition(
        survey_id,
        description=(publish_description or f"qsync key check restore ({question_id})")[
            :140
        ],
        context={
            "origin": "qsync.translations.keycheck.reorder.restore",
            "question_id": question_id,
        },
    )

    diff = diff_translation_key_snapshots(
        load_translation_key_snapshot(before_path),
        load_translation_key_snapshot(after_path),
    )
    return {
        "before": before_path,
        "after": after_path,
        "missing": sorted(diff.missing_keys),
        "extra": sorted(diff.extra_keys),
    }


def run_key_stability_check_add_remove(
    survey_id: str,
    language: str,
    *,
    question_id: str,
    label: str,
    publish_description: str | None = None,
    allow_non_smoke: bool = False,
) -> dict[str, Any]:
    _ensure_smoke_survey(survey_id, allow_non_smoke=allow_non_smoke)
    lang = normalize_language_code(language)
    original = _fetch_question_payload(survey_id, question_id)
    working = json.loads(json.dumps(original))

    before_path = snapshot_translation_keys(survey_id, lang, label=f"{label}_pre")
    added_id = _add_choice(working, display="qsync key check choice")
    _push_question_payload(survey_id, question_id, working)
    desc = publish_description or f"qsync key check add ({question_id})"
    publish_survey_definition(
        survey_id,
        description=desc[:140],
        context={
            "origin": "qsync.translations.keycheck.add",
            "question_id": question_id,
        },
    )
    after_add_path = snapshot_translation_keys(
        survey_id, lang, label=f"{label}_post_add"
    )

    _remove_choice(working, added_id)
    _push_question_payload(survey_id, question_id, working)
    publish_survey_definition(
        survey_id,
        description=(publish_description or f"qsync key check remove ({question_id})")[
            :140
        ],
        context={
            "origin": "qsync.translations.keycheck.remove",
            "question_id": question_id,
        },
    )
    after_remove_path = snapshot_translation_keys(
        survey_id, lang, label=f"{label}_post_remove"
    )

    _push_question_payload(survey_id, question_id, original)
    publish_survey_definition(
        survey_id,
        description=(publish_description or f"qsync key check restore ({question_id})")[
            :140
        ],
        context={
            "origin": "qsync.translations.keycheck.restore",
            "question_id": question_id,
        },
    )

    diff_add = diff_translation_key_snapshots(
        load_translation_key_snapshot(before_path),
        load_translation_key_snapshot(after_add_path),
    )
    diff_remove = diff_translation_key_snapshots(
        load_translation_key_snapshot(after_add_path),
        load_translation_key_snapshot(after_remove_path),
    )
    return {
        "before": before_path,
        "after_add": after_add_path,
        "after_remove": after_remove_path,
        "missing_add": sorted(diff_add.missing_keys),
        "extra_add": sorted(diff_add.extra_keys),
        "missing_remove": sorted(diff_remove.missing_keys),
        "extra_remove": sorted(diff_remove.extra_keys),
    }


def ensure_languages(
    survey_id: str, languages: Sequence[str], *, dry_run: bool = False
) -> list[str]:
    existing = list_enabled_languages(survey_id)
    target = _normalize_language_list(existing + list(languages))
    if target == existing:
        return existing
    if dry_run:
        return target
    base_url, headers = get_client_config()
    send_api_request(
        action="qsync.translations.languages.ensure",
        method="PUT",
        base_url=base_url,
        headers=headers,
        path=f"surveys/{survey_id}/languages",
        survey_id=survey_id,
        json={"AvailableLanguages": target},
        timeout=30,
    )
    return target


def fetch_base_language(survey_id: str) -> str:
    base_url, headers = get_client_config()
    resp = send_api_request(
        action="qsync.translations.base_language",
        method="GET",
        base_url=base_url,
        headers=headers,
        path=f"survey-definitions/{survey_id}/options",
        survey_id=survey_id,
        log_event=False,
        timeout=30,
    )
    result = resp.json().get("result") or {}
    language = str(result.get("SurveyLanguage") or "").strip()
    if not language:
        raise QsyncValidationError(
            error_id="QSYNC-TRANSLATIONS-BASELANG-001",
            problem="SurveyLanguage missing from survey options.",
            why="Qualtrics did not return a base language.",
            impact="Cannot safely determine which language is writable via translations API.",
            action="Check survey options in Qualtrics and retry.",
            context={"survey_id": survey_id},
        )
    return normalize_language_code(language)


def _coverage_stats(payload: Mapping[str, Any]) -> dict[str, int]:
    total = len(payload)
    filled = 0
    empty = 0
    for value in payload.values():
        if isinstance(value, str) and value.strip():
            filled += 1
        else:
            empty += 1
    return {"total": total, "filled": filled, "empty": empty}


def _coverage_stats_with_allowed_empties(
    payload: Mapping[str, Any],
    *,
    allowed_empty_keys: set[str] | None,
) -> dict[str, int]:
    """
    Coverage stats where some keys are allowed to remain empty.

    This is useful for translation workflows where certain Qualtrics translation keys
    are intentionally blank in the base language (e.g. unused choice labels). In that
    case, we do not want to warn for other languages if they mirror the base.
    """

    if not allowed_empty_keys:
        return _coverage_stats(payload)

    total = len(payload)
    filled = 0
    empty = 0
    for key, value in payload.items():
        is_filled = isinstance(value, str) and value.strip()
        if is_filled:
            filled += 1
            continue
        if key in allowed_empty_keys:
            filled += 1
        else:
            empty += 1
    return {"total": total, "filled": filled, "empty": empty}


def _extract_placeholders(text: str) -> set[str]:
    return set(_PLACEHOLDER_RE.findall(text or ""))


def _check_placeholders(
    base_map: Mapping[str, Any] | None, target_map: Mapping[str, Any], language: str
) -> tuple[list[str], list[str]]:
    if not base_map:
        return [], []
    errors: list[str] = []
    warnings: list[str] = []
    for key, base_value in base_map.items():
        if not isinstance(base_value, str):
            continue
        base_tokens = _extract_placeholders(base_value)
        if not base_tokens:
            continue
        if key not in target_map:
            # Allow partial/local-overlay maps: if a key isn't present locally, we won't
            # be pushing it (we merge remote+local during push), so do not fail the check.
            continue
        target_value = target_map.get(key, "")
        target_tokens = (
            _extract_placeholders(target_value)
            if isinstance(target_value, str)
            else set()
        )
        if base_tokens != target_tokens:
            if base_tokens and not target_tokens:
                errors.append(
                    f"[{language}] {key}: missing placeholders {sorted(base_tokens)}"
                )
            else:
                errors.append(
                    f"[{language}] {key}: placeholder mismatch (base={sorted(base_tokens)}, target={sorted(target_tokens)})"
                )
    # Warn if target introduces placeholders not in base.
    for key, value in target_map.items():
        if not isinstance(value, str):
            continue
        target_tokens = _extract_placeholders(value)
        if not target_tokens:
            continue
        base_value = base_map.get(key, "") if base_map else ""
        base_tokens = (
            _extract_placeholders(base_value) if isinstance(base_value, str) else set()
        )
        extra = target_tokens - base_tokens
        if extra:
            warnings.append(
                f"[{language}] {key}: extra placeholders not in base {sorted(extra)}"
            )
    return errors, warnings


def _check_html_hazards(payload: Mapping[str, Any], language: str) -> list[str]:
    errors: list[str] = []
    for key, value in payload.items():
        if not isinstance(value, str):
            continue
        if _HTML_HAZARD_RE.search(value):
            errors.append(f"[{language}] {key}: HTML hazard detected")
    return errors


def _check_value_length_limit(
    payload: Mapping[str, Any],
    language: str,
    *,
    limit: int = QUALTRICS_TRANSLATION_VALUE_MAX_CHARS,
    max_samples: int = 20,
) -> list[str]:
    errors: list[str] = []
    for key, value in payload.items():
        if not isinstance(value, str):
            continue
        if len(value) <= limit:
            continue
        errors.append(
            f"[{language}] {key}: exceeds Qualtrics translation value length limit ({len(value)}>{limit})."
        )
        if len(errors) >= max_samples:
            break
    if errors and len(errors) >= max_samples:
        errors.append(
            f"[{language}] value length limit errors truncated (showing first {max_samples})."
        )
    return errors


def _approx_line_count(text: str) -> int:
    if not text:
        return 0
    br_count = len(re.findall(r"<br\\s*/?>", text, flags=re.IGNORECASE))
    newline_count = text.count("\n")
    return 1 + br_count + newline_count


def _check_large_deltas(
    base_map: Mapping[str, Any],
    target_map: Mapping[str, Any],
    language: str,
    *,
    min_chars: int = 20,
    ratio_low: float = 0.4,
    ratio_high: float = 2.5,
    line_ratio_low: float = 0.5,
    line_ratio_high: float = 2.5,
    max_samples: int = 20,
) -> list[str]:
    warnings: list[str] = []
    for key, base_value in base_map.items():
        if not isinstance(base_value, str):
            continue
        target_value = target_map.get(key)
        if not isinstance(target_value, str):
            continue
        base_text = base_value.strip()
        target_text = target_value.strip()
        if len(base_text) < min_chars:
            continue
        if not base_text or not target_text:
            continue
        ratio = len(target_text) / max(len(base_text), 1)
        base_lines = _approx_line_count(base_text)
        target_lines = _approx_line_count(target_text)
        line_ratio = target_lines / max(base_lines, 1)
        if (
            ratio < ratio_low
            or ratio > ratio_high
            or line_ratio < line_ratio_low
            or line_ratio > line_ratio_high
        ):
            warnings.append(
                f"[{language}] {key}: large delta (chars={len(base_text)}→{len(target_text)}, "
                f"ratio={ratio:.2f}, lines={base_lines}→{target_lines})"
            )
            if len(warnings) >= max_samples:
                break
    if warnings and len(warnings) >= max_samples:
        warnings.append(
            f"[{language}] large delta warnings truncated (showing first {max_samples})."
        )
    return warnings


def _collect_workbook_keys(wb, language: str) -> set[tuple[str, str, str | None]]:
    keys: set[tuple[str, str, str | None]] = set()
    suffix = excel_io._language_suffix(language)
    if not suffix:
        return keys

    if excel_io.QUESTION_SHEET in wb.sheetnames:
        ws = wb[excel_io.QUESTION_SHEET]
        headers, data_rows = excel_io._iter_sheet_rows(ws)
        question_lang_headers = (
            excel_io._question_text_lang_columns_from_headers(headers) if headers else {}
        )
        if headers and "QID" in headers and language in question_lang_headers:
            qid_idx = headers.index("QID")
            for row in data_rows:
                qid_val = row[qid_idx].value if qid_idx < len(row) else None
                qid = str(qid_val or "").strip()
                if qid:
                    keys.add((qid, "QuestionText", None))

    if excel_io.OPTIONS_SHEET in wb.sheetnames:
        ws = wb[excel_io.OPTIONS_SHEET]
        headers, data_rows = excel_io._iter_sheet_rows(ws)
        if (
            headers
            and "QID" in headers
            and "ChoiceId" in headers
            and f"Label_{suffix}_MD" in headers
        ):
            qid_idx = headers.index("QID")
            choice_idx = headers.index("ChoiceId")
            for row in data_rows:
                qid_val = row[qid_idx].value if qid_idx < len(row) else None
                choice_val = row[choice_idx].value if choice_idx < len(row) else None
                qid = str(qid_val or "").strip()
                choice_id = str(choice_val or "").strip()
                if qid and choice_id:
                    keys.add((qid, "Choice", choice_id))

    if excel_io.SUBITEMS_SHEET in wb.sheetnames:
        ws = wb[excel_io.SUBITEMS_SHEET]
        headers, data_rows = excel_io._iter_sheet_rows(ws)
        if (
            headers
            and "QID" in headers
            and "AnswerId" in headers
            and f"Label_{suffix}_MD" in headers
        ):
            qid_idx = headers.index("QID")
            answer_idx = headers.index("AnswerId")
            field_idx = headers.index("Field") if "Field" in headers else None
            for row in data_rows:
                qid_val = row[qid_idx].value if qid_idx < len(row) else None
                answer_val = row[answer_idx].value if answer_idx < len(row) else None
                qid = str(qid_val or "").strip()
                answer_id = str(answer_val or "").strip()
                if not qid or not answer_id:
                    continue
                field_val = row[field_idx].value if field_idx is not None else "Answer"
                field = _normalize_field(field_val)
                keys.add((qid, field, answer_id))

    if excel_io.SBS_COLUMNS_SHEET in wb.sheetnames:
        ws = wb[excel_io.SBS_COLUMNS_SHEET]
        headers, data_rows = excel_io._iter_sheet_rows(ws)
        if (
            headers
            and "QID" in headers
            and "ColumnId" in headers
            and f"Label_{suffix}_MD" in headers
        ):
            qid_idx = headers.index("QID")
            column_idx = headers.index("ColumnId")
            for row in data_rows:
                qid_val = row[qid_idx].value if qid_idx < len(row) else None
                column_val = row[column_idx].value if column_idx < len(row) else None
                qid = str(qid_val or "").strip()
                column_id = str(column_val or "").strip()
                scoped_qid = _make_sbs_scoped_qid(qid, column_id)
                if scoped_qid:
                    keys.add((scoped_qid, "QuestionText", None))

    if excel_io.SBS_COLUMN_ANSWERS_SHEET in wb.sheetnames:
        ws = wb[excel_io.SBS_COLUMN_ANSWERS_SHEET]
        headers, data_rows = excel_io._iter_sheet_rows(ws)
        if (
            headers
            and "QID" in headers
            and "ColumnId" in headers
            and "AnswerId" in headers
            and f"Label_{suffix}_MD" in headers
        ):
            qid_idx = headers.index("QID")
            column_idx = headers.index("ColumnId")
            answer_idx = headers.index("AnswerId")
            for row in data_rows:
                qid_val = row[qid_idx].value if qid_idx < len(row) else None
                column_val = row[column_idx].value if column_idx < len(row) else None
                answer_val = row[answer_idx].value if answer_idx < len(row) else None
                qid = str(qid_val or "").strip()
                column_id = str(column_val or "").strip()
                answer_id = str(answer_val or "").strip()
                scoped_qid = _make_sbs_scoped_qid(qid, column_id)
                if scoped_qid and answer_id:
                    keys.add((scoped_qid, "Answer", answer_id))

    if excel_io.SURVEY_METADATA_SHEET in wb.sheetnames:
        ws = wb[excel_io.SURVEY_METADATA_SHEET]
        headers, data_rows = excel_io._iter_sheet_rows(ws)
        if headers and "Language" in headers:
            lang_idx = headers.index("Language")
            target_lang = normalize_language_code(language)
            for row in data_rows:
                raw = row[lang_idx].value if lang_idx < len(row) else None
                lang = normalize_language_code(str(raw or ""))
                if not lang or lang != target_lang:
                    continue
                for key in excel_io.SURVEY_METADATA_KEYS:
                    keys.add((SURVEY_METADATA_QID, "Metadata", key))

    return keys


def _run_translation_doctor_from_workbook(
    survey_id: str,
    languages: Sequence[str],
    *,
    base_language: str | None = None,
    workbook_path: Path,
) -> TranslationDoctorReport:
    errors: list[str] = []
    warnings: list[str] = []
    coverage: dict[str, dict[str, int]] = {}

    if not workbook_path.exists():
        raise QsyncValidationError(
            error_id="QSYNC-TRANSLATIONS-DOCTOR-002",
            problem="Workbook not found for translation doctor.",
            why=f"Expected workbook at {workbook_path}.",
            impact="Cannot validate translations.",
            action="Run `qsync items pull --survey-id ...` to create the workbook.",
            context={"survey_id": survey_id, "workbook": str(workbook_path)},
        )

    survey = load_cached_survey(survey_id)
    wb = load_workbook(workbook_path, data_only=True)
    question_rows = excel_io.load_questions_from_workbook(workbook_path)

    base_lang = normalize_language_code(base_language or "")
    if not base_lang:
        base_lang = normalize_language_code(
            get_base_language_from_options(survey.payload) or ""
        )

    for lang in _normalize_language_list(languages):
        key_set = _collect_workbook_keys(wb, lang)
        if not key_set:
            errors.append(f"[{lang}] No translation columns found in workbook.")
            continue

        target_map = build_workbook_value_map(
            workbook_path,
            lang,
            scope=None,
            question_rows=question_rows,
        )
        full_map_str = {
            _translation_key_str(key): target_map.get(key, "") for key in key_set
        }

        base_map_str: dict[str, str] = {}
        allowed_empty_keys: set[str] = set()
        if base_lang:
            base_map = build_base_value_map_for_keys(survey.payload, key_set)
            base_map_str = {_translation_key_str(k): v for k, v in base_map.items()}
            allowed_empty_keys = {
                key
                for key, value in base_map_str.items()
                if not str(value or "").strip()
            }

        allowed = allowed_empty_keys if base_map_str and lang != base_lang else None
        if allowed:
            coverage[lang] = _coverage_stats_with_allowed_empties(
                full_map_str, allowed_empty_keys=allowed
            )
        else:
            coverage[lang] = _coverage_stats(full_map_str)

        if base_lang and lang == base_lang:
            pass
        elif coverage[lang]["empty"] > 0:
            warnings.append(
                f"[{lang}] Coverage incomplete: {coverage[lang]['filled']}/{coverage[lang]['total']} filled."
            )

        errors.extend(_check_html_hazards(full_map_str, lang))
        if lang != base_lang:
            errors.extend(_check_value_length_limit(full_map_str, lang))

        if base_map_str and lang != base_lang:
            ph_errors, ph_warnings = _check_placeholders(
                base_map_str, full_map_str, lang
            )
            errors.extend(ph_errors)
            warnings.extend(ph_warnings)
            warnings.extend(_check_large_deltas(base_map_str, full_map_str, lang))

    return TranslationDoctorReport(errors=errors, warnings=warnings, coverage=coverage)


def run_translation_doctor(
    survey_id: str,
    languages: Sequence[str],
    *,
    base_language: str | None = None,
    workbook_path: Path | None = None,
) -> TranslationDoctorReport:
    if workbook_path is None:
        raise QsyncValidationError(
            error_id="QSYNC-TRANSLATIONS-DOCTOR-003",
            problem="Workbook path is required for translation doctor.",
            why="Translation maps are no longer supported.",
            impact="Cannot validate translations.",
            action="Run `qsync items pull --survey-id ...` to create the workbook.",
            context={"survey_id": survey_id},
        )
    return _run_translation_doctor_from_workbook(
        survey_id,
        languages,
        base_language=base_language,
        workbook_path=workbook_path,
    )


def preview_translations(
    survey_id: str,
    languages: Sequence[str] | None,
    *,
    detailed: bool = False,
    scope: ScopeFilter | None = None,
) -> list[str]:
    from ..terminal_colors import colorize_unified_diff_lines

    resolver = WorkbookResolver()
    workbook_path = resolver.default_path(survey_id)
    if not workbook_path.exists():
        raise QsyncValidationError(
            error_id="QSYNC-TRANS-PREVIEW-001",
            problem="Workbook not found for translation preview.",
            why=f"Expected workbook at {workbook_path}.",
            impact="Cannot compare Excel translations to cached survey definition.",
            action="Run `qsync items pull --survey-id ...` to create the workbook.",
            context={"survey_id": survey_id, "workbook": str(workbook_path)},
        )

    survey = load_cached_survey(survey_id)
    question_rows = excel_io.load_questions_from_workbook(workbook_path)
    lang_list = _resolve_stage_languages(
        survey_id, survey.payload, workbook_path, languages
    )
    changes = diff_workbook_vs_cache(
        survey.payload,
        workbook_path,
        lang_list,
        scope=scope,
        question_rows=question_rows,
    )
    if not changes:
        return ["No differences between Excel and cached survey."]

    lines: list[str] = []
    by_qid: dict[str, dict[str, int]] = {}
    for change in changes:
        by_qid.setdefault(change.qid, {})
        by_qid[change.qid][change.language] = (
            by_qid[change.qid].get(change.language, 0) + 1
        )

    for qid in sorted(by_qid.keys()):
        lang_parts = [f"{lang}={count}" for lang, count in sorted(by_qid[qid].items())]
        lines.append(f"- {qid}: " + ", ".join(lang_parts))
        if not detailed:
            continue
        for change in [c for c in changes if c.qid == qid]:
            item = f"QID {change.qid} {change.field}"
            if change.item_id:
                item += f" {change.item_id}"
            lines.append(f"  [{change.language}] {item}")
            diff_lines = list(
                difflib.unified_diff(
                    str(change.old_value or "").splitlines(),
                    str(change.new_value or "").splitlines(),
                    fromfile="cache",
                    tofile="workbook",
                    lineterm="",
                )
            )
            if diff_lines:
                for line in colorize_unified_diff_lines(diff_lines):
                    lines.append(f"    {line}")
            else:
                lines.append(f"    - {change.old_value}")
                lines.append(f"    + {change.new_value}")
    return lines


def _translation_key_str(key: tuple[str, str, str | None]) -> str:
    qid, field, item_id = key
    if item_id:
        return f"{qid}:{field}:{item_id}"
    return f"{qid}:{field}"


def _normalize_metadata_key(key: str) -> str:
    if key == "SurveyDescription":
        return "SurveyMetaDescription"
    return key


def _split_sbs_scoped_qid(qid: str) -> tuple[str, str] | None:
    raw = str(qid or "").strip()
    if "#" not in raw:
        return None
    base_qid, column_id = raw.split("#", 1)
    base_qid = base_qid.strip()
    column_id = column_id.strip()
    if not base_qid or not column_id:
        return None
    return base_qid, column_id


def _canonical_question_qid(qid: str) -> str:
    scoped = _split_sbs_scoped_qid(qid)
    if scoped is None:
        return str(qid or "").strip()
    return scoped[0]


def _make_sbs_scoped_qid(qid: str, column_id: str) -> str:
    qid_s = str(qid or "").strip()
    column_s = str(column_id or "").strip()
    if not qid_s or not column_s:
        return ""
    return f"{qid_s}#{column_s}"


def _parse_translation_map_key(key: str) -> tuple[str, str, str | None] | None:
    raw = str(key or "").strip()
    if not raw:
        return None
    if raw in {"SurveyTitle", "SurveyDescription", "SurveyMetaDescription"}:
        if raw == "SurveyMetaDescription":
            return (SURVEY_METADATA_QID, "Metadata", "SurveyDescription")
        return (SURVEY_METADATA_QID, "Metadata", raw)
    if ":" in raw:
        parts = [part.strip() for part in raw.split(":") if part.strip()]
        if len(parts) >= 2:
            qid = parts[0]
            field = parts[1]
            item_id = parts[2] if len(parts) >= 3 else None
            return (qid, field, item_id)
    match = re.match(
        r"^(QID[^_]+)_(QuestionText|Choice\d+|Answer\d+|Label\d+)$",
        raw,
    )
    if not match:
        return None
    qid = match.group(1)
    suffix = match.group(2)
    if suffix == "QuestionText":
        return (qid, "QuestionText", None)
    for prefix in ("Choice", "Answer", "Label"):
        if suffix.startswith(prefix):
            item_id = suffix[len(prefix) :]
            return (qid, prefix, item_id or None)
    return None


def _change_for_translation_key(
    key: str,
    language: str,
    *,
    old_value: str,
    new_value: str,
) -> dict[str, object] | None:
    parsed = _parse_translation_map_key(key)
    if not parsed:
        return None
    qid, field, item_id = parsed
    return {
        "qid": qid,
        "language": language,
        "field": field,
        "item_id": item_id,
        "old_value": old_value,
        "new_value": new_value,
    }


def _apply_translation_changes_to_payload(
    payload: dict,
    changes: list[dict[str, object]],
) -> tuple[set[str], set[str], set[str]]:
    result = payload.get("result") or {}
    questions = result.get("Questions") or {}
    options = result.get("SurveyOptions")
    if not isinstance(options, dict):
        options = {}
        result["SurveyOptions"] = options
    base_lang = get_base_language_from_options(payload)
    meta_translations = options.get("MetaDataTranslations")
    if not isinstance(meta_translations, dict):
        meta_translations = {}
        options["MetaDataTranslations"] = meta_translations

    qids: set[str] = set()
    langs: set[str] = set()
    metadata_keys: set[str] = set()

    for change in changes:
        field = str(change.get("field") or "")
        qid = str(change.get("qid") or "")
        lang = normalize_language_code(str(change.get("language") or ""))
        item_id = str(change.get("item_id") or "")
        new_value = str(change.get("new_value") or "")

        if field == "Metadata":
            key_for_options = _normalize_metadata_key(item_id)
            if base_lang and lang == base_lang:
                if item_id:
                    result[item_id] = new_value
            else:
                if not key_for_options:
                    continue
                entry = meta_translations.get(lang)
                if not isinstance(entry, dict):
                    entry = {}
                    meta_translations[lang] = entry
                entry[key_for_options] = new_value
            if key_for_options:
                metadata_keys.add(key_for_options)
            langs.add(lang)
            continue

        sbs_scope = _split_sbs_scoped_qid(qid)
        base_qid = sbs_scope[0] if sbs_scope else qid
        question = questions.get(base_qid)
        if not question:
            continue
        if sbs_scope and field == "QuestionText":
            _base_qid, column_id = sbs_scope
            write_sbs_column_question_text(question, lang, column_id, new_value)
        elif sbs_scope and field == "Answer":
            _base_qid, column_id = sbs_scope
            if not item_id:
                continue
            write_sbs_column_answer_display(
                question,
                lang,
                column_id,
                item_id,
                new_value,
            )
        elif field == "QuestionText":
            write_question_text(question, lang, new_value)
        elif field == "Choice":
            write_choice_display(question, lang, item_id, new_value)
        elif field == "Answer":
            write_answer_display(question, lang, item_id, new_value)
        else:
            write_label_display(question, lang, item_id, new_value)
        qids.add(base_qid)
        langs.add(lang)

    return qids, langs, metadata_keys


def format_translation_changes(
    changes: list[dict[str, object]],
    *,
    detailed: bool = False,
) -> list[str]:
    from ..terminal_colors import colorize_unified_diff_lines

    if not changes:
        return ["No differences between staged translations and cache."]

    lines: list[str] = []
    by_qid: dict[str, dict[str, int]] = {}
    for change in changes:
        qid = str(change.get("qid") or "")
        lang = normalize_language_code(str(change.get("language") or ""))
        if not qid:
            continue
        by_qid.setdefault(qid, {})
        by_qid[qid][lang] = by_qid[qid].get(lang, 0) + 1

    for qid in sorted(by_qid.keys()):
        lang_parts = [f"{lang}={count}" for lang, count in sorted(by_qid[qid].items())]
        lines.append(f"- {qid}: " + ", ".join(lang_parts))
        if not detailed:
            continue
        for change in [c for c in changes if str(c.get("qid") or "") == qid]:
            item = f"QID {qid} {change.get('field')}"
            item_id = change.get("item_id")
            if item_id:
                item += f" {item_id}"
            lang = normalize_language_code(str(change.get("language") or ""))
            lines.append(f"  [{lang}] {item}")
            old_value = str(change.get("old_value") or "")
            new_value = str(change.get("new_value") or "")
            diff_lines = list(
                difflib.unified_diff(
                    old_value.splitlines(),
                    new_value.splitlines(),
                    fromfile="cache",
                    tofile="staged",
                    lineterm="",
                )
            )
            if diff_lines:
                for line in colorize_unified_diff_lines(diff_lines):
                    lines.append(f"    {line}")
            else:
                lines.append(f"    - {old_value}")
                lines.append(f"    + {new_value}")

    return lines


def _sanitize_metadata_translations(meta: dict[str, Any]) -> dict[str, Any]:
    cleaned: dict[str, Any] = {}
    for lang, entry in meta.items():
        if not isinstance(entry, dict):
            continue
        updated = dict(entry)
        if "SurveyDescription" in updated:
            if "SurveyMetaDescription" not in updated:
                updated["SurveyMetaDescription"] = updated["SurveyDescription"]
            updated.pop("SurveyDescription", None)
        cleaned[lang] = updated
    return cleaned


def _push_survey_options_for_metadata(
    survey: SurveyCache,
    metadata_keys: Sequence[str],
    *,
    context: dict[str, Any] | None = None,
) -> None:
    if not metadata_keys:
        return
    base_url, headers = get_client_config()
    resp = send_api_request(
        action="qsync.translations.push.options",
        method="GET",
        base_url=base_url,
        headers=headers,
        path=f"survey-definitions/{survey.survey_id}/options",
        survey_id=survey.survey_id,
        log_event=False,
        timeout=30,
    )
    current = resp.json().get("result") or {}
    merged = dict(current)

    cached_result = survey.payload.get("result") or {}
    cached_options = cached_result.get("SurveyOptions") or {}
    if isinstance(cached_options, dict) and "MetaDataTranslations" in cached_options:
        merged["MetaDataTranslations"] = _sanitize_metadata_translations(
            cached_options.get("MetaDataTranslations") or {}
        )

    send_api_request(
        action="qsync.translations.push.options",
        method="PUT",
        base_url=base_url,
        headers=headers,
        path=f"survey-definitions/{survey.survey_id}/options",
        survey_id=survey.survey_id,
        log_meta=(
            {
                "context": context,
                "metadata_keys": list(metadata_keys),
            }
            if context
            else {"metadata_keys": list(metadata_keys)}
        ),
        json=merged,
        timeout=30,
    )


def _inventory_last_modified(survey_id: str) -> str | None:
    record = load_inventory_record(survey_id) or {}
    return record.get("lastModified") or record.get("LastModified")


def apply_translations(
    survey_id: str,
    languages: Sequence[str] | None,
    *,
    scope: ScopeFilter | None = None,
    allow_drift: bool = False,
    interactive: bool = True,
) -> PendingStagedChanges | None:
    enforce_no_drift(
        survey_id=survey_id,
        dimension="translations",
        allow_drift=allow_drift,
        interactive=interactive,
    )

    resolver = WorkbookResolver()
    workbook_path = resolver.default_path(survey_id)
    if not workbook_path.exists():
        raise QsyncValidationError(
            error_id="QSYNC-TRANS-STAGE-001",
            problem="Workbook not found for translation staging.",
            why=f"Expected workbook at {workbook_path}.",
            impact="Cannot stage translations.",
            action="Run `qsync items pull --survey-id ...` to create the workbook.",
            context={"survey_id": survey_id, "workbook": str(workbook_path)},
        )

    survey = load_cached_survey(survey_id)
    question_rows = excel_io.load_questions_from_workbook(workbook_path)
    lang_list = _resolve_stage_languages(
        survey_id, survey.payload, workbook_path, languages
    )

    changes = diff_workbook_vs_cache(
        survey.payload,
        workbook_path,
        lang_list,
        scope=scope,
        question_rows=question_rows,
    )
    if not changes:
        clear_pending(survey_id, "translations")
        return None

    errors: list[str] = []
    warnings: list[str] = []
    keys_by_lang: dict[str, set[tuple[str, str, str | None]]] = {}
    for change in changes:
        lang = normalize_language_code(str(change.language or ""))
        if not lang:
            continue
        item_id = change.item_id if change.item_id is None else str(change.item_id)
        keys_by_lang.setdefault(lang, set()).add((change.qid, change.field, item_id))
    for lang in lang_list:
        normalized_lang = normalize_language_code(str(lang or ""))
        key_set = keys_by_lang.get(normalized_lang, set())
        if not key_set:
            continue
        target_map = build_workbook_value_map(
            workbook_path,
            normalized_lang,
            scope=scope,
            question_rows=question_rows,
        )
        if not target_map:
            continue
        target_map = {k: v for k, v in target_map.items() if k in key_set}
        if not target_map:
            continue
        base_map = build_base_value_map_for_keys(survey.payload, target_map.keys())
        base_map_str = {_translation_key_str(k): v for k, v in base_map.items()}
        target_map_str = {_translation_key_str(k): v for k, v in target_map.items()}
        errors.extend(_check_html_hazards(target_map_str, normalized_lang))
        errors.extend(_check_value_length_limit(target_map_str, normalized_lang))
        ph_errors, ph_warnings = _check_placeholders(
            base_map_str, target_map_str, normalized_lang
        )
        errors.extend(ph_errors)
        warnings.extend(ph_warnings)
        warnings.extend(
            _check_large_deltas(base_map_str, target_map_str, normalized_lang)
        )

    if errors:
        raise QsyncValidationError(
            error_id="QSYNC-TRANS-STAGE-002",
            problem="Translation validation failed.",
            why="One or more validation errors were detected in the workbook.",
            impact="Translations were not staged.",
            action="Fix the errors and retry.",
            context={"errors": errors[:20]},
        )
    for warning in warnings:
        warn("[qsync:translations]", warning)

    qids = sorted(
        {
            _canonical_question_qid(change.qid)
            for change in changes
            if change.field != "Metadata"
        }
    )
    langs = sorted({change.language for change in changes})
    metadata_keys = sorted(
        {
            _normalize_metadata_key(str(change.item_id))
            for change in changes
            if change.field == "Metadata" and change.item_id
        }
    )
    pending_changes = [
        {
            "qid": change.qid,
            "language": change.language,
            "field": change.field,
            "item_id": change.item_id,
            "old_value": change.old_value,
            "new_value": change.new_value,
        }
        for change in changes
    ]
    payload = TranslationsPendingPayload(
        qids=qids,
        languages=langs,
        metadata_keys=metadata_keys,
        staged_last_modified=_inventory_last_modified(survey_id),
        changes=pending_changes,
    )
    record = PendingStagedChanges(
        survey_id=survey_id,
        dimension="translations",
        payload=payload,
        schema_version=2,
    )
    save_pending(record)
    return record


def push_translations(
    survey_id: str,
    languages: Sequence[str] | None,
    *,
    scope: ScopeFilter | None = None,
    dry_run: bool = False,
    force_live: bool = False,
    force_preview: bool = False,
    interactive: bool = True,
    publish: bool = True,
    allow_drift: bool = False,
    prefer_pending: bool | None = None,
) -> list[str]:
    resolver = WorkbookResolver()
    workbook_path = resolver.default_path(survey_id)
    if not workbook_path.exists():
        raise QsyncValidationError(
            error_id="QSYNC-TRANS-PUSH-001",
            problem="Workbook not found for translation push.",
            why=f"Expected workbook at {workbook_path}.",
            impact="Cannot push translations.",
            action="Run `qsync items pull --survey-id ...` to create the workbook.",
            context={"survey_id": survey_id, "workbook": str(workbook_path)},
        )

    survey = load_cached_survey(survey_id)
    question_rows = excel_io.load_questions_from_workbook(workbook_path)
    lang_list = _resolve_stage_languages(
        survey_id, survey.payload, workbook_path, languages
    )

    if not allow_drift:
        enforce_no_drift(
            survey_id=survey_id,
            dimension="translations",
            allow_drift=False,
            interactive=interactive,
        )

    workbook_changes = diff_workbook_vs_cache(
        survey.payload,
        workbook_path,
        lang_list,
        scope=scope,
        question_rows=question_rows,
    )

    pending = load_pending(survey_id, "translations")
    if workbook_changes:
        if pending:
            decision = prefer_pending
            if decision is None:
                if interactive and sys.stdin.isatty():
                    from ..interactive_menu import select_from_list

                    choices = [
                        "Use staged changes (ignore Excel)",
                        "Restage from Excel (overwrite pending)",
                        "↩ Abort push",
                    ]
                    selection = select_from_list(
                        message="Excel differs from cache and staged changes exist. Which should be pushed?",
                        choices=choices,
                        default=choices[1],  # default is “restage” (legacy behavior)
                    )
                    if selection is None or selection.startswith("↩"):
                        decision = None
                    elif selection.startswith("Use staged"):
                        decision = True
                    else:
                        decision = False
                else:
                    decision = False
            if decision is True:
                warn(
                    "[qsync:translations]",
                    "Using staged changes and ignoring workbook differences.",
                )
            elif decision is False:
                print(
                    "[sync:translations] Excel differs from cache, re-staging from current Excel "
                    "(overriding stale staging)..."
                )
                print(
                    f"[sync:translations] Staging {len(workbook_changes)} change(s) from Excel..."
                )
                staged = apply_translations(
                    survey_id,
                    lang_list,
                    scope=scope,
                    allow_drift=allow_drift,
                    interactive=interactive,
                )
                if not staged:
                    print(
                        "[sync:translations] No stageable changes after staging; skipping."
                    )
                    return []
                pending = load_pending(survey_id, "translations")
            else:
                raise SystemExit("[qsync:translations] Aborted by user.")
        else:
            print(
                f"[sync:translations] Staging {len(workbook_changes)} change(s) from Excel..."
            )
            staged = apply_translations(
                survey_id,
                lang_list,
                scope=scope,
                allow_drift=allow_drift,
                interactive=interactive,
            )
            if not staged:
                print(
                    "[sync:translations] No stageable changes after staging; skipping."
                )
                return []
            pending = load_pending(survey_id, "translations")
    elif not pending:
        print("[sync:translations] No differences between Excel and cached survey.")
        return []

    if not pending or not isinstance(pending.payload, TranslationsPendingPayload):
        print("[sync:translations] No staged changes found.")
        return []

    if pending.schema_version < 2 or not getattr(pending.payload, "changes", None):
        # Attempt to rebuild pending changes for legacy payloads
        staged = apply_translations(
            survey_id,
            lang_list,
            scope=scope,
            allow_drift=allow_drift,
            interactive=interactive,
        )
        pending = load_pending(survey_id, "translations")
        if not pending or not isinstance(pending.payload, TranslationsPendingPayload):
            print("[sync:translations] No staged changes found.")
            return []

    qids = sorted(
        {
            _canonical_question_qid(qid)
            for qid in (pending.payload.qids or [])
            if qid
        }
    )
    langs = sorted({lang for lang in (pending.payload.languages or []) if lang})
    metadata_keys = sorted(
        {key for key in (pending.payload.metadata_keys or []) if key}
    )
    if not qids and not metadata_keys:
        print("[sync:translations] No staged changes to push.")
        return []

    current_last_modified = _inventory_last_modified(survey_id)
    if (
        pending.payload.staged_last_modified
        and current_last_modified
        and pending.payload.staged_last_modified != current_last_modified
    ):
        warn(
            "[qsync:translations]",
            "Survey lastModified changed since staging; consider refreshing the cache before pushing.",
        )

    if dry_run:
        info(
            "[qsync:translations]",
            f"Dry run: would push {len(qids)} question(s) and "
            f"{len(metadata_keys)} metadata key(s) for {survey_id}.",
        )
        return qids

    config = SafeguardConfig(
        survey_id=survey_id,
        dimension="translations",
        force_live=force_live,
        force_preview=force_preview,
        auto_yes=not interactive,
    )
    safeguard_result = enforce_push_safeguards(config)
    if safeguard_result.warnings:
        for warning in safeguard_result.warnings:
            warn("[qsync:translations]", warning)

    ensure_backup(survey_id)
    push_context = {
        "origin": "qsync.translations.push",
        "changed_qids": qids,
        "metadata_keys": metadata_keys,
        "changed_count": len(qids) + len(metadata_keys),
        "languages": langs,
    }
    _apply_translation_changes_to_payload(
        survey.payload,
        list(getattr(pending.payload, "changes", None) or []),
    )
    if qids:
        push_questions(survey, qids, context=push_context)
    if metadata_keys:
        _push_survey_options_for_metadata(survey, metadata_keys, context=push_context)

    if publish:
        auto_publish_after_push(
            survey_id=survey_id,
            dimension="translations",
            changed_qids=qids,
            count=len(qids) + len(metadata_keys),
            languages=langs,
            skip_publish=False,
            auto_yes=not interactive,
        )
    else:
        success(
            "[qsync:translations]",
            f"Uploaded {len(qids)} question(s) for {survey_id}.",
        )

    from ..qualtrics_client import refresh_survey_cache

    try:
        refresh_survey_cache(survey_id)
        clear_pending(survey_id, "translations")
    except Exception as exc:
        warn(
            "[qsync:translations]",
            f"Push succeeded but cache refresh failed: {exc}",
        )
    return qids


def drift_translations(
    survey_id: str,
    languages: Sequence[str],
    *,
    base_language: str | None = None,
) -> list[str]:
    report = check_drift(
        survey_id=survey_id,
        dimension="translations",
        interactive=False,
        context={"languages": list(languages)},
    )
    lines: list[str] = []
    if not report.has_drift:
        return ["No translation drift detected."]
    lines.append(report.summary)
    if report.context_lines:
        lines.extend(report.context_lines)
    if report.recommendation:
        lines.append(report.recommendation)
    return lines


def resolve_languages_for_cli(
    survey_id: str, languages: Sequence[str] | None
) -> list[str]:
    if languages:
        return _normalize_language_list(languages)
    return list_enabled_languages(survey_id)


def load_pending_languages(survey_id: str) -> list[str] | None:
    record = load_pending(survey_id, "translations")
    if not record or not isinstance(record.payload, TranslationsPendingPayload):
        return None
    return _normalize_language_list(record.payload.languages)


def ensure_pending_changes_record(
    survey_id: str,
    pending: PendingStagedChanges,
    *,
    scope: ScopeFilter | None = None,
    allow_drift: bool = False,
    interactive: bool = True,
) -> PendingStagedChanges | None:
    if not isinstance(pending.payload, TranslationsPendingPayload):
        return pending
    if pending.schema_version >= 2 and pending.payload.changes:
        return pending
    languages = pending.payload.languages or None
    staged = apply_translations(
        survey_id,
        languages,
        scope=scope,
        allow_drift=allow_drift,
        interactive=interactive,
    )
    if staged:
        return staged
    return load_pending(survey_id, "translations")
