from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from .. import excel_io
from ..config import resolve_root
from ..markdown_codec import (
    html_to_md,
    md_to_html,
    normalize_text,
    should_treat_as_html,
    validate_html_fragment,
)
from ..qualtrics_client import SurveyCache, load_cached_survey, refresh_survey_cache
from ..translations import list_enabled_languages
from ..survey_ref import format_survey_ref
from ..workbook_resolver import WorkbookResolver
from ..interactive_menu import (
    confirm,
    select_from_list,
    text_input,
    edit_text_in_editor,
)
from ..terminal_output import info, warn
from ..push_safeguards import SafeguardConfig, enforce_push_safeguards
from ..qualtrics_client import ensure_backup, push_questions
from ..auto_publish import auto_publish_after_push
from ..push_logger import log_push_event

EXTERNALLY_MANAGED_REGISTRY_REL = Path("surveys/externally_managed_items.csv")


@dataclass(frozen=True)
class ExternallyManagedEntry:
    qid: str | None
    data_export_tag: str | None
    owner: str


def _workspace_root() -> Path:
    return resolve_root(required=False) or Path.cwd()


def _registry_path() -> Path:
    return _workspace_root() / EXTERNALLY_MANAGED_REGISTRY_REL


def ensure_externally_managed_registry_exists() -> Path:
    """Ensure the externally-managed registry exists (seeded from code defaults)."""

    path = _registry_path()
    if path.exists():
        return path

    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f, fieldnames=["qid", "data_export_tag", "owner"], extrasaction="ignore"
        )
        writer.writeheader()
        for tag, owner in sorted(excel_io.EXTERNALLY_MANAGED_TAGS.items()):
            writer.writerow({"qid": "", "data_export_tag": tag, "owner": owner})
    return path


def load_externally_managed_registry() -> list[ExternallyManagedEntry]:
    """Load a workspace-local registry of externally-managed content.

    The registry is matched by either QID or DataExportTag.
    """

    path = ensure_externally_managed_registry_exists()
    entries: list[ExternallyManagedEntry] = []

    try:
        with path.open("r", newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                qid = (row.get("qid") or "").strip() or None
                tag = (row.get("data_export_tag") or "").strip() or None
                owner = (row.get("owner") or "").strip()
                if not owner:
                    continue
                entries.append(
                    ExternallyManagedEntry(qid=qid, data_export_tag=tag, owner=owner)
                )
    except Exception:
        # If the file is corrupted, fall back to code defaults.
        entries = []

    # Always seed with code defaults (tag-based) as a safety net.
    existing_tag_owners = {
        (e.data_export_tag, e.owner) for e in entries if e.data_export_tag
    }
    for tag, owner in sorted(excel_io.EXTERNALLY_MANAGED_TAGS.items()):
        if (tag, owner) not in existing_tag_owners:
            entries.append(
                ExternallyManagedEntry(qid=None, data_export_tag=tag, owner=owner)
            )

    return entries


def external_owner_for(*, qid: str | None, data_export_tag: str | None) -> str | None:
    qid_norm = (qid or "").strip()
    tag_norm = (data_export_tag or "").strip()
    if not qid_norm and not tag_norm:
        return None

    entries = load_externally_managed_registry()
    for e in entries:
        if e.qid and qid_norm and e.qid.strip() == qid_norm:
            return e.owner
        if e.data_export_tag and tag_norm and e.data_export_tag.strip() == tag_norm:
            return e.owner
    return None


def is_externally_managed(*, qid: str | None, data_export_tag: str | None) -> bool:
    return external_owner_for(qid=qid, data_export_tag=data_export_tag) is not None


def iter_active_qids_in_flow(survey: SurveyCache) -> Iterable[str]:
    """Yield in-flow QuestionIDs (QIDs) in survey-flow order (non-Trash blocks)."""

    result = survey.payload.get("result", {}) or {}
    blocks = result.get("Blocks", {}) or {}
    questions = survey.questions or {}
    flow_obj = result.get("SurveyFlow") or result.get("Flow") or {}

    ordered_block_ids: list[str] = []

    def _walk_flow(node: Any) -> None:
        if node is None:
            return
        if isinstance(node, list):
            for child in node:
                _walk_flow(child)
            return
        if not isinstance(node, dict):
            return

        node_type = str(node.get("Type") or "").strip()
        if node_type in {"Block", "Standard"} and node.get("ID"):
            bid = str(node.get("ID") or "").strip()
            if bid and bid not in ordered_block_ids:
                ordered_block_ids.append(bid)

        for key in ("Flow", "Then", "Else", "ElseFlow"):
            if key in node:
                _walk_flow(node.get(key))

    if isinstance(flow_obj, dict):
        _walk_flow(flow_obj.get("Flow"))
    elif isinstance(flow_obj, list):
        _walk_flow(flow_obj)

    trash_blocks = {
        str(bid).strip()
        for bid, payload in blocks.items()
        if str((payload or {}).get("Type") or "").strip() == "Trash"
    }

    seen_qids: set[str] = set()
    for bid in ordered_block_ids:
        if bid in trash_blocks:
            continue
        block = blocks.get(bid) or {}
        elements = block.get("BlockElements") or block.get("Elements") or []
        if not isinstance(elements, list):
            continue
        for elem in elements:
            if not isinstance(elem, dict):
                continue
            if str(elem.get("Type") or "").strip() not in {"", "Question"}:
                continue
            qid = str(elem.get("QuestionID") or "").strip()
            if not qid or qid in seen_qids:
                continue
            if qid not in questions:
                continue
            seen_qids.add(qid)
            yield qid


def iter_all_qids(survey: SurveyCache) -> list[str]:
    """Return all QIDs in the cached survey (sorted)."""
    return sorted(str(qid) for qid in survey.questions.keys())


def _text_preview(html: str, *, max_len: int = 70) -> str:
    raw = (html or "").strip()
    if not raw:
        return ""
    try:
        md = html_to_md(raw).replace("\n", " ").strip()
    except Exception:
        md = raw.replace("\n", " ").strip()
    md = " ".join(md.split())
    if len(md) <= max_len:
        return md
    return md[: max_len - 1].rstrip() + "…"


def summarize_structural_ops(
    structural_ops: list[dict[str, Any]],
) -> dict[str, dict[str, int]]:
    summary: dict[str, dict[str, int]] = {}
    op_map = {
        "choice_add": "add",
        "choice_edit": "edit",
        "choice_remove": "remove",
        "answer_add": "add",
        "answer_edit": "edit",
        "answer_remove": "remove",
        "question_text_edit": "edit",
        "sbs_column_add": "add",
        "sbs_column_edit": "edit",
        "sbs_column_remove": "remove",
        "sbs_column_answer_add": "add",
        "sbs_column_answer_edit": "edit",
        "sbs_column_answer_remove": "remove",
    }
    for op in structural_ops or []:
        qid = str(op.get("qid") or "").strip()
        if not qid:
            continue
        bucket = summary.setdefault(qid, {"add": 0, "edit": 0, "remove": 0, "other": 0})
        label = op_map.get(str(op.get("op") or ""), "other")
        bucket[label] = bucket.get(label, 0) + 1
    return summary


def _is_sbs_matrix_question(question: dict) -> bool:
    """Return True for Qualtrics SBS side-by-side matrix questions."""
    return (
        str(question.get("QuestionType") or "").strip() == "SBS"
        and str(question.get("Selector") or "").strip() == "SBSMatrix"
    )


def _is_matrix_question(question: dict) -> bool:
    return str(question.get("QuestionType") or "").strip() == "Matrix"


def _get_base_language(survey_payload: dict) -> str:
    """Return the base language code for workbook columns (best-effort)."""
    try:
        result = survey_payload.get("result", {})
        if not isinstance(result, dict):
            result = survey_payload
        options = result.get("SurveyOptions") or {}
        lang = str(options.get("SurveyLanguage") or "").strip()
        lang = lang.replace("-", "_").upper()
        return lang or "EN"
    except Exception:
        return "EN"


def _format_qid_label(
    *,
    survey: SurveyCache,
    qid: str,
    active_set: set[str],
    include_flow_status: bool,
) -> str:
    q = survey.questions.get(qid) or {}
    tag = (q.get("DataExportTag") or "").strip()
    preview = _text_preview(q.get("QuestionText") or "", max_len=55)
    suffix = []
    if tag:
        suffix.append(f"tag={tag}")
    qtype = (q.get("QuestionType") or "").strip()
    if qtype:
        suffix.append(qtype)
    if include_flow_status and qid not in active_set:
        suffix.append("not-in-flow")
    meta = f" [{', '.join(suffix)}]" if suffix else ""
    return f"{qid}{meta} {preview}".strip()


def inspect_question(
    *,
    survey_id: str,
    qid: str,
    refresh: bool = False,
) -> str:
    """Return a human-readable inspection string for one QID."""

    survey = (
        refresh_survey_cache(survey_id)[0] if refresh else load_cached_survey(survey_id)
    )
    question = survey.questions.get(qid)
    if not question:
        return (
            f"[qsync:items:inspect] QID {qid} not found in cached survey {survey_id}."
        )

    tag = (question.get("DataExportTag") or "").strip()
    qtype = (question.get("QuestionType") or "").strip()
    active_qids = list(iter_active_qids_in_flow(survey))
    active_set = set(active_qids)
    is_active = qid in active_set
    owner = external_owner_for(qid=qid, data_export_tag=tag)

    lines: list[str] = []
    lines.append(f"[qsync:items:inspect] survey_id={survey_id} qid={qid}")
    lines.append(f"- Active in flow: {'yes' if is_active else 'no'}")
    if is_active:
        lines.append(
            f"- Flow position: {active_qids.index(qid) + 1} / {len(active_qids)}"
        )
    lines.append(f"- QuestionType: {qtype or '(missing)'}")
    lines.append(f"- DataExportTag: {tag or '(missing)'}")
    if owner:
        lines.append(f"- Externally managed: yes (owner={owner})")
    else:
        lines.append("- Externally managed: no")
    qt = _text_preview(question.get("QuestionText") or "")
    if qt:
        lines.append(f"- QuestionText: {qt}")

    def _render_block(
        name: str, mapping: dict | None, order_key: str | None, next_key: str | None
    ) -> None:
        if not mapping:
            return
        lines.append(f"- {name}: {len(mapping)}")
        if order_key and question.get(order_key) is not None:
            lines.append(f"  - {order_key}: {len(list(question.get(order_key) or []))}")
        if next_key and question.get(next_key) is not None:
            lines.append(f"  - {next_key}: {question.get(next_key)}")
        shown = 0
        for cid, payload in list(mapping.items())[:20]:
            display = ""
            if isinstance(payload, dict):
                display = _text_preview(str(payload.get("Display") or ""))
            lines.append(f"  - {cid}: {display}")
            shown += 1
        if len(mapping) > shown:
            lines.append(f"  - … ({len(mapping) - shown} more)")

    _render_block(
        "Choices", question.get("Choices") or {}, "ChoiceOrder", "NextChoiceId"
    )
    _render_block(
        "Answers", question.get("Answers") or {}, "AnswerOrder", "NextAnswerId"
    )

    registry_path = _registry_path()
    lines.append(f"- Registry: {registry_path}")

    return "\n".join(lines)


class ItemsStructuralError(RuntimeError):
    pass


def _normalize_choice_order_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return str(int(value))
    return str(value).strip()


def _existing_numeric_ids(mapping: dict) -> list[int]:
    out: list[int] = []
    for k in mapping.keys():
        try:
            out.append(int(str(k)))
        except (TypeError, ValueError):
            continue
    return sorted(set(out))


def _allocate_choice_id(question: dict) -> str:
    choices = question.get("Choices") or {}
    existing = set(_existing_numeric_ids(choices))
    next_id = question.get("NextChoiceId")
    candidate: int | None = None
    if isinstance(next_id, int):
        candidate = next_id
    elif isinstance(next_id, str) and next_id.strip().isdigit():
        candidate = int(next_id.strip())
    if candidate is None:
        candidate = (max(existing) + 1) if existing else 1
    while candidate in existing:
        candidate += 1
    question["NextChoiceId"] = candidate + 1
    return str(candidate)


def _allocate_answer_id(question: dict) -> str:
    answers = question.get("Answers") or {}
    existing = set(_existing_numeric_ids(answers))
    next_id = question.get("NextAnswerId")
    candidate: int | None = None
    if isinstance(next_id, int):
        candidate = next_id
    elif isinstance(next_id, str) and next_id.strip().isdigit():
        candidate = int(next_id.strip())
    if candidate is None:
        candidate = (max(existing) + 1) if existing else 1
    while candidate in existing:
        candidate += 1
    question["NextAnswerId"] = candidate + 1
    return str(candidate)


def _append_choice_order(question: dict, choice_id: str) -> None:
    order = question.get("ChoiceOrder")
    if order is None:
        question["ChoiceOrder"] = [choice_id]
        return
    if not isinstance(order, list):
        # Qualtrics sometimes uses unexpected types; normalize.
        question["ChoiceOrder"] = [choice_id]
        return
    normalized = [_normalize_choice_order_value(v) for v in order]
    if choice_id in normalized:
        return
    # Preserve existing element types where possible: if all are ints, append int.
    if all(isinstance(v, int) for v in order):
        order.append(int(choice_id))
    else:
        order.append(choice_id)


def _append_answer_order(question: dict, answer_id: str) -> None:
    order = question.get("AnswerOrder")
    if order is None:
        question["AnswerOrder"] = [answer_id]
        return
    if not isinstance(order, list):
        question["AnswerOrder"] = [answer_id]
        return
    normalized = [_normalize_choice_order_value(v) for v in order]
    if answer_id in normalized:
        return
    if all(isinstance(v, int) for v in order):
        order.append(int(answer_id))
    else:
        order.append(answer_id)


def _remove_choice_from_order(question: dict, choice_id: str) -> None:
    order = question.get("ChoiceOrder")
    if not isinstance(order, list):
        return
    new_order = []
    for v in order:
        if _normalize_choice_order_value(v) == choice_id:
            continue
        new_order.append(v)
    question["ChoiceOrder"] = new_order


def _remove_answer_from_order(question: dict, answer_id: str) -> None:
    order = question.get("AnswerOrder")
    if not isinstance(order, list):
        return
    new_order = []
    for v in order:
        if _normalize_choice_order_value(v) == answer_id:
            continue
        new_order.append(v)
    question["AnswerOrder"] = new_order


def _cleanup_choice_translations(
    question: dict, choice_id: str, enabled_langs: list[str]
) -> None:
    lang_block = question.get("Language")
    if not isinstance(lang_block, dict):
        return
    for lang in enabled_langs:
        section = lang_block.get(lang)
        if not isinstance(section, dict):
            continue
        choices = section.get("Choices")
        if isinstance(choices, dict) and choice_id in choices:
            del choices[choice_id]
            if not choices:
                section.pop("Choices", None)
        # If the section becomes empty, prune it.
        if not section:
            lang_block.pop(lang, None)
    if not lang_block:
        question.pop("Language", None)


def _cleanup_answer_translations(
    question: dict, answer_id: str, enabled_langs: list[str]
) -> None:
    lang_block = question.get("Language")
    if not isinstance(lang_block, dict):
        return
    for lang in enabled_langs:
        section = lang_block.get(lang)
        if not isinstance(section, dict):
            continue
        answers = section.get("Answers")
        if isinstance(answers, dict) and answer_id in answers:
            del answers[answer_id]
            if not answers:
                section.pop("Answers", None)
        if not section:
            lang_block.pop(lang, None)
    if not lang_block:
        question.pop("Language", None)


def _require_supported_question(
    question: dict, *, qid: str, target: str, experimental_unsupported: bool
) -> None:
    if target == "choices":
        if not isinstance(question.get("Choices"), dict):
            raise ItemsStructuralError(
                f"[qsync:items:edit] QID {qid} has no Choices section; unsupported for target=choices."
            )
        return

    if target == "answers":
        if not isinstance(question.get("Answers"), dict):
            raise ItemsStructuralError(
                f"[qsync:items:edit] QID {qid} has no Answers section; unsupported for target=answers."
            )
        return

    raise ItemsStructuralError(f"[qsync:items:edit] Unknown target: {target}")


def _validate_html_or_raise(html: str) -> None:
    errors = validate_html_fragment(html)
    if errors:
        msg = "; ".join(errors[:3])
        raise ItemsStructuralError(f"[qsync:items:edit] Invalid HTML fragment: {msg}")


def preflight_items_edit(
    *,
    survey_id: str,
    ignore_workbook_drift: bool,
    interactive: bool,
) -> None:
    """Refresh cache and (optionally) block on workbook drift before structural edits."""

    info(
        "[qsync:items:edit]",
        f"Refreshing cache from API for {format_survey_ref(survey_id)}…",
    )
    survey, _changed = refresh_survey_cache(survey_id)

    resolver = WorkbookResolver()
    xlsx_path = resolver.resolve(survey_id)
    if not xlsx_path.exists():
        return

    from .items_core import preview_changes

    diffs = preview_changes(
        survey_id,
        xlsx_path,
        check_drift=False,
        self_heal_system_columns=False,
        annotate_dirty=False,
    )
    if not diffs:
        return

    if not interactive and not ignore_workbook_drift:
        raise ItemsStructuralError(
            f"[qsync:items:edit] Workbook differs from refreshed cache at {xlsx_path}. "
            "Resolve workbook changes first, or re-run with --ignore-workbook-drift (dangerous)."
        )

    if not interactive and ignore_workbook_drift:
        warn(
            "[qsync:items:edit]",
            f"Proceeding despite workbook drift ({len(diffs)} diff(s)) because --ignore-workbook-drift was set.",
        )
        return

    choices = [
        "Exit and let me handle my Excel changes first",
        "Show diffs",
        "Re-pull workbook from Qualtrics (overwrite local workbook)",
    ]
    if ignore_workbook_drift:
        choices.append("Proceed anyway (dangerous)")

    while True:
        warn(
            "[qsync:items:edit]",
            f"Workbook differs from refreshed cache ({len(diffs)} diff(s)) at {xlsx_path}.",
        )
        selected = select_from_list(
            "Resolve workbook drift before structural edits", choices
        )
        if selected is None or selected.startswith("Exit"):
            raise ItemsStructuralError(
                "[qsync:items:edit] Aborted due to workbook drift."
            )
        if selected.startswith("Show diffs"):
            from ..terminal_colors import colorize_unified_diff_lines

            print("[qsync:items:edit] Diffs (cached vs workbook):")
            for change in diffs:
                print("-" * 80)
                header = f"{change.kind.upper()} qid={change.qid}"
                if change.choice_id is not None:
                    header += f", choice_id={change.choice_id}"
                if change.answer_id is not None:
                    header += f", answer_id={change.answer_id}"
                print(header)
                for line in colorize_unified_diff_lines(change.diff_lines or []):
                    print("  " + line)
            continue
        if selected.startswith("Re-pull workbook"):
            if not confirm(
                f"Overwrite workbook at {xlsx_path}? (will move current file aside)",
                default=False,
            ):
                continue
            backup = xlsx_path.with_suffix(".bak.xlsx")
            counter = 1
            while backup.exists():
                backup = xlsx_path.with_suffix(f".bak{counter}.xlsx")
                counter += 1
            xlsx_path.rename(backup)
            excel_io.init_workbook_from_survey(survey_id, survey.payload, xlsx_path)
            info(
                "[qsync:items:edit]",
                f"Re-pulled workbook to {xlsx_path} (backup: {backup})",
            )
            diffs = preview_changes(
                survey_id,
                xlsx_path,
                check_drift=False,
                self_heal_system_columns=False,
                annotate_dirty=False,
            )
            if not diffs:
                return
            continue
        if selected.startswith("Proceed anyway") and ignore_workbook_drift:
            warn(
                "[qsync:items:edit]",
                "Proceeding despite workbook drift (dangerous; workbook will be stale after structural edits).",
            )
            return


def stage_choice_op(
    *,
    survey_id: str,
    qid: str,
    action: str,
    html: str | None,
    choice_id: str | None,
    allow_delete: bool,
    interactive: bool,
    experimental_unsupported: bool,
) -> dict[str, Any]:
    """Apply one structural choice operation to the local cache and return an op record."""

    survey = load_cached_survey(survey_id)
    question = survey.questions.get(qid)
    if not isinstance(question, dict):
        raise ItemsStructuralError(
            f"[qsync:items:edit] QID {qid} not found in cached survey."
        )

    tag = (question.get("DataExportTag") or "").strip() or None
    owner = external_owner_for(qid=qid, data_export_tag=tag)
    if owner:
        raise ItemsStructuralError(
            f"[qsync:items:edit] QID {qid} is externally managed (owner={owner})."
        )

    if experimental_unsupported:
        log_push_event(
            action="qsync.items.edit.experimental",
            method="LOCAL",
            path="dimensions.items_structural.stage_choice_op",
            survey_id=survey_id,
            meta={"qid": qid, "tag": tag},
        )

    _require_supported_question(
        question,
        qid=qid,
        target="choices",
        experimental_unsupported=experimental_unsupported,
    )

    op: dict[str, Any] = {"qid": qid, "target": "choices"}
    if action == "add":
        if not html:
            raise ItemsStructuralError("[qsync:items:edit] add requires --text (HTML).")
        _validate_html_or_raise(html)
        new_id = _allocate_choice_id(question)
        choices = question.setdefault("Choices", {})
        choices[str(new_id)] = {"Display": html, "Display_Unsafe": html}
        _append_choice_order(question, str(new_id))
        # SBSMatrix: keep duplicated statements under AdditionalQuestions[*].Choices consistent.
        if _is_sbs_matrix_question(question):
            additional = question.get("AdditionalQuestions") or {}
            if isinstance(additional, dict):
                for aq in additional.values():
                    if not isinstance(aq, dict):
                        continue
                    aq_choices = aq.get("Choices")
                    if not isinstance(aq_choices, dict):
                        continue
                    aq_choices[str(new_id)] = {"Display": html, "Display_Unsafe": html}
                    _append_choice_order(aq, str(new_id))
        op.update(
            {"op": "choice_add", "choice_id": str(new_id), "html": html, "tag": tag}
        )
    elif action == "edit":
        if not choice_id:
            raise ItemsStructuralError(
                "[qsync:items:edit] edit requires --id (choice_id)."
            )
        if not html:
            raise ItemsStructuralError(
                "[qsync:items:edit] edit requires --text (HTML)."
            )
        _validate_html_or_raise(html)
        choices = question.get("Choices") or {}
        entry = choices.get(str(choice_id))
        if not isinstance(entry, dict):
            raise ItemsStructuralError(
                f"[qsync:items:edit] ChoiceId {choice_id} not found."
            )
        prev = str(entry.get("Display") or "")
        entry["Display"] = html
        if "Display_Unsafe" in entry:
            entry["Display_Unsafe"] = html
        if _is_sbs_matrix_question(question):
            additional = question.get("AdditionalQuestions") or {}
            if isinstance(additional, dict):
                for aq in additional.values():
                    if not isinstance(aq, dict):
                        continue
                    aq_choices = aq.get("Choices")
                    if not isinstance(aq_choices, dict):
                        continue
                    aq_entry = aq_choices.get(str(choice_id))
                    if isinstance(aq_entry, dict):
                        aq_entry["Display"] = html
                        if "Display_Unsafe" in aq_entry:
                            aq_entry["Display_Unsafe"] = html
        op.update(
            {
                "op": "choice_edit",
                "choice_id": str(choice_id),
                "html": html,
                "prev_html": prev,
                "tag": tag,
            }
        )
    elif action == "remove":
        if not choice_id:
            raise ItemsStructuralError(
                "[qsync:items:edit] remove requires --id (choice_id)."
            )
        if not allow_delete:
            if not interactive:
                raise ItemsStructuralError(
                    "[qsync:items:edit] remove requires --allow-delete in non-interactive mode."
                )
            if not confirm(
                f"Remove ChoiceId {choice_id} from {qid}? This is destructive.",
                default=False,
            ):
                raise ItemsStructuralError("[qsync:items:edit] Cancelled.")
            if not confirm(
                "Confirm delete again (last chance).",
                default=False,
            ):
                raise ItemsStructuralError("[qsync:items:edit] Cancelled.")
        enabled_langs = list_enabled_languages(survey_id)
        choices = question.get("Choices") or {}
        prev_entry = choices.get(str(choice_id))
        prev = None
        if isinstance(prev_entry, dict):
            prev = str(prev_entry.get("Display") or "")
        if str(choice_id) in choices:
            del choices[str(choice_id)]
        _remove_choice_from_order(question, str(choice_id))
        _cleanup_choice_translations(question, str(choice_id), enabled_langs)
        if _is_sbs_matrix_question(question):
            additional = question.get("AdditionalQuestions") or {}
            if isinstance(additional, dict):
                for aq in additional.values():
                    if not isinstance(aq, dict):
                        continue
                    aq_choices = aq.get("Choices")
                    if not isinstance(aq_choices, dict):
                        continue
                    if str(choice_id) in aq_choices:
                        del aq_choices[str(choice_id)]
                    _remove_choice_from_order(aq, str(choice_id))
        op.update(
            {
                "op": "choice_remove",
                "choice_id": str(choice_id),
                "prev_html": prev,
                "tag": tag,
            }
        )
    else:
        raise ItemsStructuralError(f"[qsync:items:edit] Unknown action: {action}")

    survey.save()
    log_push_event(
        action="qsync.items.edit.stage",
        method="LOCAL",
        path="dimensions.items_structural.stage_choice_op",
        survey_id=survey_id,
        meta=op,
    )
    return op


def stage_answer_op(
    *,
    survey_id: str,
    qid: str,
    action: str,
    html: str | None,
    answer_id: str | None,
    allow_delete: bool,
    interactive: bool,
    experimental_unsupported: bool,
) -> dict[str, Any]:
    """Apply one structural answer operation to the local cache and return an op record."""

    survey = load_cached_survey(survey_id)
    question = survey.questions.get(qid)
    if not isinstance(question, dict):
        raise ItemsStructuralError(
            f"[qsync:items:edit] QID {qid} not found in cached survey."
        )

    tag = (question.get("DataExportTag") or "").strip() or None
    owner = external_owner_for(qid=qid, data_export_tag=tag)
    if owner:
        raise ItemsStructuralError(
            f"[qsync:items:edit] QID {qid} is externally managed (owner={owner})."
        )

    if experimental_unsupported:
        log_push_event(
            action="qsync.items.edit.experimental",
            method="LOCAL",
            path="dimensions.items_structural.stage_answer_op",
            survey_id=survey_id,
            meta={"qid": qid, "tag": tag},
        )

    _require_supported_question(
        question,
        qid=qid,
        target="answers",
        experimental_unsupported=experimental_unsupported,
    )

    op: dict[str, Any] = {"qid": qid, "target": "answers"}
    if action == "add":
        if not html:
            raise ItemsStructuralError("[qsync:items:edit] add requires --text (HTML).")
        _validate_html_or_raise(html)
        new_id = _allocate_answer_id(question)
        answers = question.setdefault("Answers", {})
        answers[str(new_id)] = {"Display": html, "Display_Unsafe": html}
        _append_answer_order(question, str(new_id))
        op.update(
            {"op": "answer_add", "answer_id": str(new_id), "html": html, "tag": tag}
        )
    elif action == "edit":
        if not answer_id:
            raise ItemsStructuralError(
                "[qsync:items:edit] edit requires --id (answer_id)."
            )
        if not html:
            raise ItemsStructuralError(
                "[qsync:items:edit] edit requires --text (HTML)."
            )
        _validate_html_or_raise(html)
        answers = question.get("Answers") or {}
        entry = answers.get(str(answer_id))
        if not isinstance(entry, dict):
            raise ItemsStructuralError(
                f"[qsync:items:edit] AnswerId {answer_id} not found."
            )
        prev = str(entry.get("Display") or "")
        entry["Display"] = html
        if "Display_Unsafe" in entry:
            entry["Display_Unsafe"] = html
        op.update(
            {
                "op": "answer_edit",
                "answer_id": str(answer_id),
                "html": html,
                "prev_html": prev,
                "tag": tag,
            }
        )
    elif action == "remove":
        if not answer_id:
            raise ItemsStructuralError(
                "[qsync:items:edit] remove requires --id (answer_id)."
            )
        if not allow_delete:
            if not interactive:
                raise ItemsStructuralError(
                    "[qsync:items:edit] remove requires --allow-delete in non-interactive mode."
                )
            if not confirm(
                f"Remove AnswerId {answer_id} from {qid}? This is destructive.",
                default=False,
            ):
                raise ItemsStructuralError("[qsync:items:edit] Cancelled.")
            if not confirm(
                "Confirm delete again (last chance).",
                default=False,
            ):
                raise ItemsStructuralError("[qsync:items:edit] Cancelled.")
        enabled_langs = list_enabled_languages(survey_id)
        answers = question.get("Answers") or {}
        prev_entry = answers.get(str(answer_id))
        prev = None
        if isinstance(prev_entry, dict):
            prev = str(prev_entry.get("Display") or "")
        if str(answer_id) in answers:
            del answers[str(answer_id)]
        _remove_answer_from_order(question, str(answer_id))
        _cleanup_answer_translations(question, str(answer_id), enabled_langs)
        op.update(
            {
                "op": "answer_remove",
                "answer_id": str(answer_id),
                "prev_html": prev,
                "tag": tag,
            }
        )
    else:
        raise ItemsStructuralError(f"[qsync:items:edit] Unknown action: {action}")

    survey.save()
    log_push_event(
        action="qsync.items.edit.stage",
        method="LOCAL",
        path="dimensions.items_structural.stage_answer_op",
        survey_id=survey_id,
        meta=op,
    )
    return op


def stage_question_text_op(
    *,
    survey_id: str,
    qid: str,
    text: str,
    text_format: str,
    interactive: bool,
    experimental_unsupported: bool,
) -> dict[str, Any]:
    """Apply a question text edit to the local cache and return an op record."""

    survey = load_cached_survey(survey_id)
    question = survey.questions.get(qid)
    if not isinstance(question, dict):
        raise ItemsStructuralError(
            f"[qsync:items:edit] QID {qid} not found in cached survey."
        )

    tag = (question.get("DataExportTag") or "").strip() or None
    owner = external_owner_for(qid=qid, data_export_tag=tag)
    if owner:
        raise ItemsStructuralError(
            f"[qsync:items:edit] QID {qid} is externally managed (owner={owner})."
        )

    if experimental_unsupported:
        log_push_event(
            action="qsync.items.edit.experimental",
            method="LOCAL",
            path="dimensions.items_structural.stage_question_text_op",
            survey_id=survey_id,
            meta={"qid": qid, "tag": tag},
        )

    fmt = (text_format or "").strip().lower()
    if fmt in {"markdown", "md"}:
        new_html = normalize_text(md_to_html(text or ""))
        fmt = "md"
    elif fmt in {"html"}:
        new_html = normalize_text(text or "")
        errors = validate_html_fragment(new_html)
        if errors:
            msg = "; ".join(errors[:3])
            raise ItemsStructuralError(
                f"[qsync:items:edit] Invalid HTML fragment: {msg}"
            )
        fmt = "html"
    else:
        raise ItemsStructuralError(
            "[qsync:items:edit] Unsupported --text-format. Use 'md' or 'html'."
        )

    prev_html = normalize_text(str(question.get("QuestionText") or ""))
    question["QuestionText"] = new_html
    if "QuestionText_Unsafe" in question:
        question["QuestionText_Unsafe"] = new_html

    op: dict[str, Any] = {
        "qid": qid,
        "target": "question_text",
        "op": "question_text_edit",
        "format": fmt,
        "html": new_html,
        "prev_html": prev_html,
        "tag": tag,
    }

    survey.save()
    log_push_event(
        action="qsync.items.edit.stage",
        method="LOCAL",
        path="dimensions.items_structural.stage_question_text_op",
        survey_id=survey_id,
        meta=op,
    )
    return op


def stage_sbs_column_op(
    *,
    survey_id: str,
    qid: str,
    action: str,
    html: str | None,
    column_id: str | None,
    allow_delete: bool,
    interactive: bool,
    experimental_unsupported: bool,
) -> dict[str, Any]:
    """Stage an SBS column header edit (AdditionalQuestions[*].QuestionText)."""

    if not column_id:
        raise ItemsStructuralError(
            "[qsync:items:edit] SBS column op requires --id (ColumnId)."
        )
    action_norm = (action or "").strip().lower()
    if action_norm not in {"add", "edit", "remove"}:
        raise ItemsStructuralError(
            "[qsync:items:edit] SBS columns support action=add|edit|remove."
        )
    if action_norm in {"add", "edit"}:
        if not html:
            raise ItemsStructuralError(
                "[qsync:items:edit] SBS column add/edit requires --text (HTML)."
            )
        _validate_html_or_raise(html)
    if action_norm == "remove" and not allow_delete:
        if not interactive:
            raise ItemsStructuralError(
                "[qsync:items:edit] remove requires --allow-delete in non-interactive mode."
            )
        if not confirm(
            f"Remove SBS ColumnId {column_id} from {qid}? This is destructive.",
            default=False,
        ):
            raise ItemsStructuralError("[qsync:items:edit] Cancelled.")
        if not confirm("Confirm delete again (last chance).", default=False):
            raise ItemsStructuralError("[qsync:items:edit] Cancelled.")

    survey = load_cached_survey(survey_id)
    question = survey.questions.get(qid)
    if not isinstance(question, dict):
        raise ItemsStructuralError(
            f"[qsync:items:edit] QID {qid} not found in cached survey."
        )
    if not _is_sbs_matrix_question(question):
        raise ItemsStructuralError(
            f"[qsync:items:edit] QID {qid} is not an SBSMatrix question."
        )

    tag = (question.get("DataExportTag") or "").strip() or None
    owner = external_owner_for(qid=qid, data_export_tag=tag)
    if owner:
        raise ItemsStructuralError(
            f"[qsync:items:edit] QID {qid} is externally managed (owner={owner})."
        )

    additional = question.get("AdditionalQuestions") or {}
    if not isinstance(additional, dict):
        raise ItemsStructuralError(
            f"[qsync:items:edit] QID {qid} has no AdditionalQuestions; cannot edit SBS columns."
        )

    col_key = str(column_id)
    aq = additional.get(col_key)
    if action_norm == "add":
        if col_key in additional:
            raise ItemsStructuralError(
                f"[qsync:items:edit] SBS ColumnId {column_id} already exists for {qid}."
            )
        if not additional:
            raise ItemsStructuralError(
                f"[qsync:items:edit] QID {qid} has no existing SBS columns to clone from."
            )
        import copy

        template = next((v for v in additional.values() if isinstance(v, dict)), None)
        if not isinstance(template, dict):
            raise ItemsStructuralError(
                f"[qsync:items:edit] QID {qid} AdditionalQuestions template missing."
            )
        new_col = copy.deepcopy(template)
        prev_html = ""
        new_html = normalize_text(html or "")
        new_col["QuestionText"] = new_html
        if "QuestionText_Unsafe" in new_col:
            new_col["QuestionText_Unsafe"] = new_html
        additional[col_key] = new_col
        op_type = "sbs_column_add"
    elif action_norm == "remove":
        if not isinstance(aq, dict):
            raise ItemsStructuralError(
                f"[qsync:items:edit] SBS ColumnId {column_id} not found for {qid}."
            )
        prev_html = normalize_text(str(aq.get("QuestionText") or ""))
        if col_key in additional:
            del additional[col_key]
        new_html = ""
        op_type = "sbs_column_remove"
    else:
        if not isinstance(aq, dict):
            raise ItemsStructuralError(
                f"[qsync:items:edit] SBS ColumnId {column_id} not found for {qid}."
            )
        prev_html = normalize_text(str(aq.get("QuestionText") or ""))
        new_html = normalize_text(html or "")
        aq["QuestionText"] = new_html
        if "QuestionText_Unsafe" in aq:
            aq["QuestionText_Unsafe"] = new_html
        op_type = "sbs_column_edit"

    op: dict[str, Any] = {
        "qid": qid,
        "target": "sbs_columns",
        "op": op_type,
        "choice_id": col_key,
        "html": new_html,
        "prev_html": prev_html,
        "tag": tag,
    }

    survey.save()
    log_push_event(
        action="qsync.items.edit.stage",
        method="LOCAL",
        path="dimensions.items_structural.stage_sbs_column_op",
        survey_id=survey_id,
        meta=op,
    )
    return op


def stage_sbs_column_answer_op(
    *,
    survey_id: str,
    qid: str,
    action: str,
    html: str | None,
    column_id: str | None,
    answer_id: str | None,
    allow_delete: bool,
    interactive: bool,
    experimental_unsupported: bool,
) -> dict[str, Any]:
    """Stage an SBS per-column answer label edit (AdditionalQuestions[*].Answers[*].Display)."""

    action_norm = (action or "").strip().lower()
    if action_norm not in {"add", "edit", "remove"}:
        raise ItemsStructuralError(
            "[qsync:items:edit] SBS column answers support action=add|edit|remove."
        )
    if not column_id:
        raise ItemsStructuralError(
            "[qsync:items:edit] SBS column answer op requires --column-id."
        )
    if not answer_id:
        raise ItemsStructuralError(
            "[qsync:items:edit] SBS column answer op requires --id (AnswerId)."
        )
    if action_norm in {"add", "edit"}:
        if not html:
            raise ItemsStructuralError(
                "[qsync:items:edit] SBS column answer add/edit requires --text (HTML)."
            )
        _validate_html_or_raise(html)
    if action_norm == "remove" and not allow_delete:
        if not interactive:
            raise ItemsStructuralError(
                "[qsync:items:edit] remove requires --allow-delete in non-interactive mode."
            )
        if not confirm(
            f"Remove SBS AnswerId {answer_id} from column {column_id}? This is destructive.",
            default=False,
        ):
            raise ItemsStructuralError("[qsync:items:edit] Cancelled.")
        if not confirm("Confirm delete again (last chance).", default=False):
            raise ItemsStructuralError("[qsync:items:edit] Cancelled.")

    survey = load_cached_survey(survey_id)
    question = survey.questions.get(qid)
    if not isinstance(question, dict):
        raise ItemsStructuralError(
            f"[qsync:items:edit] QID {qid} not found in cached survey."
        )
    if not _is_sbs_matrix_question(question):
        raise ItemsStructuralError(
            f"[qsync:items:edit] QID {qid} is not an SBSMatrix question."
        )

    tag = (question.get("DataExportTag") or "").strip() or None
    owner = external_owner_for(qid=qid, data_export_tag=tag)
    if owner:
        raise ItemsStructuralError(
            f"[qsync:items:edit] QID {qid} is externally managed (owner={owner})."
        )

    additional = question.get("AdditionalQuestions") or {}
    if not isinstance(additional, dict):
        raise ItemsStructuralError(
            f"[qsync:items:edit] QID {qid} has no AdditionalQuestions; cannot edit SBS columns."
        )
    aq = additional.get(str(column_id))
    if not isinstance(aq, dict):
        raise ItemsStructuralError(
            f"[qsync:items:edit] SBS ColumnId {column_id} not found for {qid}."
        )
    answers = aq.get("Answers") or {}
    if not isinstance(answers, dict):
        raise ItemsStructuralError(
            f"[qsync:items:edit] SBS ColumnId {column_id} has no Answers map."
        )
 
    def _append_order(obj: dict, key: str, item: str) -> None:
        order = obj.get(key)
        if isinstance(order, list):
            if item not in order:
                order.append(item)
        else:
            obj[key] = [item]

    def _remove_order(obj: dict, key: str, item: str) -> None:
        order = obj.get(key)
        if isinstance(order, list) and item in order:
            obj[key] = [x for x in order if str(x) != item]

    aid_key = str(answer_id)
    if action_norm == "add":
        if aid_key in answers:
            raise ItemsStructuralError(
                f"[qsync:items:edit] SBS AnswerId {answer_id} already exists for {qid}#{column_id}."
            )
        import copy

        template = next((v for v in answers.values() if isinstance(v, dict)), None)
        entry = copy.deepcopy(template) if isinstance(template, dict) else {}
        prev_html = ""
        new_html = normalize_text(html or "")
        entry["Display"] = new_html
        entry["Display_Unsafe"] = new_html
        answers[aid_key] = entry
        _append_order(aq, "AnswerOrder", aid_key)
        try:
            next_candidate = int(aid_key) + 1
            current_next = aq.get("NextAnswerId")
            if isinstance(current_next, int):
                aq["NextAnswerId"] = max(current_next, next_candidate)
            else:
                aq["NextAnswerId"] = next_candidate
        except (TypeError, ValueError):
            pass
        op_type = "sbs_column_answer_add"
    elif action_norm == "remove":
        entry = answers.get(aid_key)
        if not isinstance(entry, dict):
            raise ItemsStructuralError(
                f"[qsync:items:edit] SBS AnswerId {answer_id} not found for {qid}#{column_id}."
            )
        prev_html = normalize_text(str(entry.get("Display") or ""))
        if aid_key in answers:
            del answers[aid_key]
        _remove_order(aq, "AnswerOrder", aid_key)
        new_html = ""
        op_type = "sbs_column_answer_remove"
    else:
        entry = answers.get(aid_key)
        if not isinstance(entry, dict):
            raise ItemsStructuralError(
                f"[qsync:items:edit] SBS AnswerId {answer_id} not found for {qid}#{column_id}."
            )
        prev_html = normalize_text(str(entry.get("Display") or ""))
        new_html = normalize_text(html or "")
        entry["Display"] = new_html
        if "Display_Unsafe" in entry:
            entry["Display_Unsafe"] = new_html
        op_type = "sbs_column_answer_edit"

    op: dict[str, Any] = {
        "qid": qid,
        "target": "sbs_column_answers",
        "op": op_type,
        "choice_id": str(column_id),
        "answer_id": aid_key,
        "html": new_html,
        "prev_html": prev_html,
        "tag": tag,
    }

    survey.save()
    log_push_event(
        action="qsync.items.edit.stage",
        method="LOCAL",
        path="dimensions.items_structural.stage_sbs_column_answer_op",
        survey_id=survey_id,
        meta=op,
    )
    return op


def stage_structural_op(
    *,
    survey_id: str,
    qid: str,
    target: str,
    action: str,
    html: str | None,
    text_format: str | None,
    item_id: str | None,
    allow_delete: bool,
    interactive: bool,
    experimental_unsupported: bool,
) -> dict[str, Any]:
    target_norm = (target or "").strip().lower()
    if target_norm in {"options", "option"}:
        target_norm = "options"
    if target_norm in {"subitems", "subitem"}:
        target_norm = "subitems"
    if target_norm == "choices":
        return stage_choice_op(
            survey_id=survey_id,
            qid=qid,
            action=action,
            html=html,
            choice_id=item_id,
            allow_delete=allow_delete,
            interactive=interactive,
            experimental_unsupported=experimental_unsupported,
        )
    if target_norm == "options":
        # Workbook terminology: "Options" may map to Answers for Matrix questions.
        survey = load_cached_survey(survey_id)
        question = survey.questions.get(qid)
        if not isinstance(question, dict):
            raise ItemsStructuralError(f"[qsync:items:edit] QID {qid} not found.")
        if _is_matrix_question(question):
            op = stage_answer_op(
                survey_id=survey_id,
                qid=qid,
                action=action,
                html=html,
                answer_id=item_id,
                allow_delete=allow_delete,
                interactive=interactive,
                experimental_unsupported=experimental_unsupported,
            )
            op.setdefault("surface", "options")
            return op
        op = stage_choice_op(
            survey_id=survey_id,
            qid=qid,
            action=action,
            html=html,
            choice_id=item_id,
            allow_delete=allow_delete,
            interactive=interactive,
            experimental_unsupported=experimental_unsupported,
        )
        op.setdefault("surface", "options")
        return op
    if target_norm == "answers":
        return stage_answer_op(
            survey_id=survey_id,
            qid=qid,
            action=action,
            html=html,
            answer_id=item_id,
            allow_delete=allow_delete,
            interactive=interactive,
            experimental_unsupported=experimental_unsupported,
        )
    if target_norm == "subitems":
        # Workbook terminology: "Subitems" map to:
        # - Matrix/SBSMatrix: Choices (rows/statements)
        # - Else: Answers
        survey = load_cached_survey(survey_id)
        question = survey.questions.get(qid)
        if not isinstance(question, dict):
            raise ItemsStructuralError(f"[qsync:items:edit] QID {qid} not found.")
        if _is_matrix_question(question) or _is_sbs_matrix_question(question):
            op = stage_choice_op(
                survey_id=survey_id,
                qid=qid,
                action=action,
                html=html,
                choice_id=item_id,
                allow_delete=allow_delete,
                interactive=interactive,
                experimental_unsupported=experimental_unsupported,
            )
            op.setdefault("surface", "subitems")
            return op
        op = stage_answer_op(
            survey_id=survey_id,
            qid=qid,
            action=action,
            html=html,
            answer_id=item_id,
            allow_delete=allow_delete,
            interactive=interactive,
            experimental_unsupported=experimental_unsupported,
        )
        op.setdefault("surface", "subitems")
        return op
    if target_norm in {"question", "question_text", "question-text"}:
        if action != "edit":
            raise ItemsStructuralError(
                "[qsync:items:edit] Question text supports action=edit only."
            )
        if html is None:
            raise ItemsStructuralError("[qsync:items:edit] edit requires --text.")
        return stage_question_text_op(
            survey_id=survey_id,
            qid=qid,
            text=html or "",
            text_format=text_format or "md",
            interactive=interactive,
            experimental_unsupported=experimental_unsupported,
        )
    if target_norm in {"sbs_columns", "sbs_column"}:
        return stage_sbs_column_op(
            survey_id=survey_id,
            qid=qid,
            action=action,
            html=html,
            column_id=item_id,
            allow_delete=allow_delete,
            interactive=interactive,
            experimental_unsupported=experimental_unsupported,
        )
    if target_norm in {"sbs_column_answers", "sbs_column_answer"}:
        # For non-interactive calls, pass ColumnId via `--column-id` and AnswerId via `--id`.
        # This wrapper expects `item_id` to be the AnswerId.
        raise ItemsStructuralError(
            "[qsync:items:edit] Use stage_sbs_column_answer_op directly (requires column_id + answer_id)."
        )
    raise ItemsStructuralError(f"[qsync:items:edit] Unknown target: {target}")


def _qid_workbook_diffs(
    *, survey_id: str, qid: str
) -> list[Any]:
    """Return workbook diffs affecting a single QID (best-effort, no writes)."""
    resolver = WorkbookResolver()
    xlsx_path = resolver.resolve(survey_id)
    if not xlsx_path.exists():
        return []
    from .items_core import preview_changes

    return list(
        preview_changes(
            survey_id,
            xlsx_path,
            include_qids={qid},
            check_drift=False,
            self_heal_system_columns=False,
            annotate_dirty=False,
        )
        or []
    )


def _wipe_workbook_qid_cells(
    *,
    survey_id: str,
    qid: str,
    surfaces: set[str] | None = None,
    dry_run: bool = True,
) -> list[str]:
    """Clear base-language workbook cells for a QID so init_workbook_from_survey can refill them.

    Strategy:
    - Clear `Text_<BASE>_MD` / `Text_<BASE>_IsHTML` on Questions for the QID.
    - Clear `Label_<BASE>_MD` / `Label_<BASE>_IsHTML` on Options/Subitems/SBS sheets for the QID.
    - Then re-run `init_workbook_from_survey` to repopulate those now-empty cells from cache.
    """

    resolver = WorkbookResolver()
    xlsx_path = resolver.resolve(survey_id)
    if not xlsx_path.exists():
        return [f"No workbook found at {xlsx_path}"]

    survey = load_cached_survey(survey_id)
    base_lang = _get_base_language(survey.payload)

    try:
        from openpyxl import load_workbook
    except Exception as exc:
        return [f"openpyxl unavailable: {exc}"]

    wb = load_workbook(xlsx_path)
    notes: list[str] = []

    def _clear_sheet_rows(sheet: str, qid_col: str, extra_match: dict[str, str] | None, cols_to_clear: list[str]) -> None:
        if sheet not in wb.sheetnames:
            return
        ws = wb[sheet]
        rows = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        if not rows:
            return
        headers = [str(h or "").strip() for h in rows[0]]
        idx = {h: i for i, h in enumerate(headers) if h}
        if qid_col not in idx:
            return
        clear_idxs = [idx[c] for c in cols_to_clear if c in idx]
        if not clear_idxs:
            return
        match_idxs = {}
        if extra_match:
            for k, v in extra_match.items():
                if k in idx:
                    match_idxs[k] = (idx[k], v)
        cleared = 0
        for r in ws.iter_rows(min_row=2):
            cell_qid = r[idx[qid_col]].value
            if str(cell_qid or "").strip() != qid:
                continue
            ok = True
            for _k, (mi, expected) in match_idxs.items():
                if str(r[mi].value or "").strip() != expected:
                    ok = False
                    break
            if not ok:
                continue
            for ci in clear_idxs:
                if ci < len(r):
                    r[ci].value = ""
            cleared += 1
        if cleared:
            notes.append(f"{sheet}: cleared {cleared} row(s) for QID={qid}")

    # Decide which surfaces to wipe.
    surfaces = set(surfaces or {"question_text", "options", "subitems", "sbs_columns", "sbs_column_answers"})

    if "question_text" in surfaces:
        _clear_sheet_rows(
            "Questions",
            "QID",
            None,
            [f"Text_{base_lang}_MD", f"Text_{base_lang}_IsHTML"],
        )
    if "options" in surfaces:
        _clear_sheet_rows(
            "Options",
            "QID",
            None,
            [f"Label_{base_lang}_MD", f"Label_{base_lang}_IsHTML"],
        )
    if "subitems" in surfaces:
        _clear_sheet_rows(
            "Subitems",
            "QID",
            None,
            [f"Label_{base_lang}_MD", f"Label_{base_lang}_IsHTML"],
        )
    if "sbs_columns" in surfaces:
        _clear_sheet_rows(
            "SBS_Columns",
            "QID",
            None,
            [f"Label_{base_lang}_MD", f"Label_{base_lang}_IsHTML"],
        )
    if "sbs_column_answers" in surfaces:
        _clear_sheet_rows(
            "SBS_ColumnAnswers",
            "QID",
            None,
            [f"Label_{base_lang}_MD", f"Label_{base_lang}_IsHTML"],
        )

    if dry_run:
        return notes or [f"No matching rows/columns found to clear for QID={qid}"]

    wb.save(xlsx_path)
    excel_io.init_workbook_from_survey(survey_id, survey.payload, xlsx_path)
    return notes or [f"No matching rows/columns found to clear for QID={qid}"]


def interactive_choice_wizard(
    *,
    survey_id: str,
    qid: str | None,
    allow_delete: bool,
    experimental_unsupported: bool,
) -> dict[str, Any]:
    survey = load_cached_survey(survey_id)
    active = list(iter_active_qids_in_flow(survey))
    all_qids = iter_all_qids(survey)
    if not qid:
        if not active and not all_qids:
            raise ItemsStructuralError(
                "[qsync:items:edit] No QIDs found in cached survey."
            )
        active_set = set(active)
        labels = []
        if active:
            for aqid in active:
                labels.append(
                    _format_qid_label(
                        survey=survey,
                        qid=aqid,
                        active_set=active_set,
                        include_flow_status=False,
                    )
                )
        if active and len(active) > 30:
            from ..interactive_menu import autocomplete_from_list

            choices = [
                "Browse active-in-flow (arrow list)",
                "Search active by tag/text (autocomplete)",
                "Filter active by ExportTag (autocomplete)",
            ]
            if all_qids and len(active) < len(all_qids):
                choices.append("Show all questions (includes not-in-flow)")
            choices.append("↩ Cancel")

            mode = select_from_list(
                "How do you want to select a QID?",
                choices,
            )
            if not mode or "Cancel" in mode:
                raise ItemsStructuralError("[qsync:items:edit] Cancelled.")
            if mode.startswith("Show all questions"):
                labels = []
                for aqid in all_qids:
                    labels.append(
                        _format_qid_label(
                            survey=survey,
                            qid=aqid,
                            active_set=active_set,
                            include_flow_status=True,
                        )
                    )
                selected = autocomplete_from_list(
                    message="Search QID (all)",
                    choices=labels,
                    instruction="type to filter, enter to select",
                )
                if not selected:
                    raise ItemsStructuralError("[qsync:items:edit] Cancelled.")
                qid = selected.split()[0]
                question = survey.questions.get(qid)
                if not isinstance(question, dict):
                    raise ItemsStructuralError(
                        f"[qsync:items:edit] QID {qid} not found."
                    )
                tag = (question.get("DataExportTag") or "").strip() or None
                owner = external_owner_for(qid=qid, data_export_tag=tag)
                if owner:
                    raise ItemsStructuralError(
                        f"[qsync:items:edit] QID {qid} is externally managed (owner={owner})."
                    )
                # Continue with target/action selection below.
            else:
                if mode.startswith("Search active"):
                    selected = autocomplete_from_list(
                        message="Search QID (active)",
                        choices=labels,
                        instruction="type to filter, enter to select",
                    )
                    if not selected:
                        raise ItemsStructuralError("[qsync:items:edit] Cancelled.")
                elif mode.startswith("Filter active"):
                    tags: list[str] = []
                    tag_to_qids: dict[str, list[str]] = {}
                    for aqid in active:
                        q = survey.questions.get(aqid) or {}
                        tag = (q.get("DataExportTag") or "").strip()
                        if not tag:
                            continue
                        tag_to_qids.setdefault(tag, []).append(aqid)
                    tags = sorted(tag_to_qids.keys())
                    if not tags:
                        raise ItemsStructuralError(
                            "[qsync:items:edit] No DataExportTag values found in active questions."
                        )
                    chosen_tag = autocomplete_from_list(
                        message="Select ExportTag",
                        choices=tags,
                        instruction="type to filter, enter to select",
                    )
                    if not chosen_tag:
                        raise ItemsStructuralError("[qsync:items:edit] Cancelled.")
                    filtered_labels = [
                        _format_qid_label(
                            survey=survey,
                            qid=aqid,
                            active_set=active_set,
                            include_flow_status=False,
                        )
                        for aqid in tag_to_qids.get(chosen_tag, [])
                    ]
                    selected = select_from_list("Select a QID to edit", filtered_labels)
                    if not selected:
                        raise ItemsStructuralError("[qsync:items:edit] Cancelled.")
                else:
                    selected = select_from_list("Select a QID to edit", labels)
                    if not selected:
                        raise ItemsStructuralError("[qsync:items:edit] Cancelled.")
                qid = selected.split()[0]
        else:
            if all_qids and len(active) < len(all_qids):
                labels.append("─" * 40)
                labels.append("Show all questions (includes not-in-flow)")
            selected = select_from_list("Select a QID to edit", labels)
        if not selected:
            raise ItemsStructuralError("[qsync:items:edit] Cancelled.")
        if selected.startswith("Show all questions"):
            labels = []
            for aqid in all_qids:
                labels.append(
                    _format_qid_label(
                        survey=survey,
                        qid=aqid,
                        active_set=active_set,
                        include_flow_status=True,
                    )
                )
            selected = select_from_list("Select a QID to edit", labels)
            if not selected:
                raise ItemsStructuralError("[qsync:items:edit] Cancelled.")
        qid = selected.split()[0]

    question = survey.questions.get(qid)
    if not isinstance(question, dict):
        raise ItemsStructuralError(f"[qsync:items:edit] QID {qid} not found.")

    # Workbook drift check (scoped to this QID) before editing.
    diffs = _qid_workbook_diffs(survey_id=survey_id, qid=qid)
    if diffs:
        warn(
            "[qsync:items:edit]",
            f"Workbook has {len(diffs)} unstaged diff(s) for {qid}.",
        )
        sel = select_from_list(
            "Workbook edits detected for this QID. What next?",
            [
                "Enter items sync cycle for this QID (recommended)",
                "Overwrite workbook edits for this QID (dangerous)",
                "↩ Cancel",
            ],
        )
        if not sel or "Cancel" in sel:
            raise ItemsStructuralError("[qsync:items:edit] Cancelled.")
        if sel.startswith("Enter items sync cycle"):
            from ..scope_filter import ScopeFilter
            from ..sync_orchestrator import sync_survey

            scope = ScopeFilter.parse(f"qid:{qid}")
            sync_survey(
                survey_id=survey_id,
                dimensions=["items"],
                interactive=True,
                scope=scope,
                per_dimension=False,
                auto_yes=False,
                allow_drift=False,
            )
            diffs = _qid_workbook_diffs(survey_id=survey_id, qid=qid)
            if diffs:
                warn(
                    "[qsync:items:edit]",
                    f"Workbook still differs for {qid} after sync ({len(diffs)} diff(s)). Continuing anyway.",
                )
        else:
            if not confirm(
                "This will overwrite workbook cells for this QID to match cache. Proceed?",
                default=False,
            ):
                raise ItemsStructuralError("[qsync:items:edit] Cancelled.")
            notes = _wipe_workbook_qid_cells(
                survey_id=survey_id,
                qid=qid,
                dry_run=False,
            )
            for line in notes[:6]:
                info("[qsync:items:edit]", line)

    tag = (question.get("DataExportTag") or "").strip() or None
    owner = external_owner_for(qid=qid, data_export_tag=tag)
    if owner:
        raise ItemsStructuralError(
            f"[qsync:items:edit] QID {qid} is externally managed (owner={owner})."
        )

    is_matrix = _is_matrix_question(question)
    is_sbs_matrix = _is_sbs_matrix_question(question)

    supported_targets: list[str] = ["Question text"]
    if is_sbs_matrix:
        # SBSMatrix does not use the Options sheet; statements live under Subitems.
        supported_targets.append("Subitems (rows/statements)")
        supported_targets.append("SBS Columns (headers)")
        supported_targets.append("SBS Column Answers (per-column)")
    elif is_matrix:
        supported_targets.append("Options (columns)")
        supported_targets.append("Subitems (rows/statements)")
    else:
        if isinstance(question.get("Choices"), dict):
            supported_targets.append("Options")
        if isinstance(question.get("Answers"), dict):
            supported_targets.append("Subitems")
    if len(supported_targets) == 1:
        target = supported_targets[0]
    else:
        target = select_from_list("Select edit target", supported_targets)
        if not target:
            raise ItemsStructuralError("[qsync:items:edit] Cancelled.")
    if target.lower().startswith("question"):
        target_norm = "question-text"
    elif target.lower().startswith("sbs columns"):
        target_norm = "sbs_columns"
    elif target.lower().startswith("sbs column answers"):
        target_norm = "sbs_column_answers"
    elif target.lower().startswith("options"):
        target_norm = "options"
    else:
        target_norm = "subitems"

    if target_norm == "question-text":
        current_html = normalize_text(str(question.get("QuestionText") or ""))
        default_fmt = "html" if should_treat_as_html(current_html) else "md"
        fmt_choices = (
            ["Markdown (recommended)", "HTML (advanced)"]
            if default_fmt == "md"
            else ["HTML (advanced)", "Markdown (recommended)"]
        )
        chosen = select_from_list("Choose input format", fmt_choices)
        if not chosen:
            raise ItemsStructuralError("[qsync:items:edit] Cancelled.")
        fmt = "md" if chosen.startswith("Markdown") else "html"

        if fmt == "md" and should_treat_as_html(current_html):
            warn(
                "[qsync:items:edit]",
                "Current QuestionText looks like complex HTML; editing in Markdown may lose structure. Prefer HTML.",
            )

        default_text = html_to_md(current_html) if fmt == "md" else current_html
        input_mode = select_from_list(
            "How do you want to edit the text?",
            [
                "Edit in $EDITOR (multiline)",
                "Enter inline (single line)",
                "↩ Cancel",
            ],
        )
        if not input_mode or "Cancel" in input_mode:
            raise ItemsStructuralError("[qsync:items:edit] Cancelled.")

        if input_mode.startswith("Edit in"):
            suffix = ".md" if fmt == "md" else ".html"
            new_text = edit_text_in_editor(
                "Edit QuestionText",
                initial_text=default_text,
                suffix=suffix,
            )
        else:
            new_text = text_input("Enter new QuestionText", default=default_text)

        if new_text is None:
            raise ItemsStructuralError("[qsync:items:edit] Cancelled.")

        return stage_structural_op(
            survey_id=survey_id,
            qid=qid,
            target="question-text",
            action="edit",
            html=new_text,
            text_format=fmt,
            item_id=None,
            allow_delete=allow_delete,
            interactive=True,
            experimental_unsupported=experimental_unsupported,
        )

    if target_norm == "sbs_columns":
        if not is_sbs_matrix:
            raise ItemsStructuralError(
                "[qsync:items:edit] SBS Columns are only supported for SBSMatrix questions."
            )
        additional = question.get("AdditionalQuestions") or {}
        if not isinstance(additional, dict) or not additional:
            raise ItemsStructuralError(
                "[qsync:items:edit] No AdditionalQuestions found for SBSMatrix question."
            )
        columns = sorted(str(k) for k in additional.keys())
        labels = []
        for cid in columns:
            aq = additional.get(cid) or {}
            txt = _text_preview(str((aq.get("QuestionText") or "")), max_len=60) if isinstance(aq, dict) else ""
            labels.append(f"{cid}: {txt}".strip())
        action = select_from_list("Select an action", ["add", "edit", "remove"])
        if not action:
            raise ItemsStructuralError("[qsync:items:edit] Cancelled.")
        action_norm = action.strip().lower()

        if action_norm == "add":
            nums: list[int] = []
            for c in columns:
                try:
                    nums.append(int(str(c)))
                except Exception:
                    continue
            suggested = str(max(nums) + 1) if nums else ""
            col_id = text_input("Enter new ColumnId", default=suggested)
            if col_id is None:
                raise ItemsStructuralError("[qsync:items:edit] Cancelled.")
            col_id = str(col_id).strip()
            if not col_id:
                raise ItemsStructuralError("[qsync:items:edit] ColumnId is required.")
            new_html = text_input(f"Enter HTML for new SBS ColumnId {col_id}")
            if new_html is None:
                raise ItemsStructuralError("[qsync:items:edit] Cancelled.")
            _validate_html_or_raise(new_html)
            return stage_sbs_column_op(
                survey_id=survey_id,
                qid=qid,
                action="add",
                html=new_html,
                column_id=col_id,
                allow_delete=allow_delete,
                interactive=True,
                experimental_unsupported=experimental_unsupported,
            )

        if action_norm == "edit":
            sel = select_from_list("Select a ColumnId to edit", labels)
            if not sel:
                raise ItemsStructuralError("[qsync:items:edit] Cancelled.")
            col_id = sel.split(":", 1)[0].strip()
            new_html = text_input(f"Enter new HTML for SBS ColumnId {col_id}")
            if new_html is None:
                raise ItemsStructuralError("[qsync:items:edit] Cancelled.")
            _validate_html_or_raise(new_html)
            return stage_sbs_column_op(
                survey_id=survey_id,
                qid=qid,
                action="edit",
                html=new_html,
                column_id=col_id,
                allow_delete=allow_delete,
                interactive=True,
                experimental_unsupported=experimental_unsupported,
            )

        if not allow_delete:
            raise ItemsStructuralError(
                "[qsync:items:edit] Remove requires --allow-delete."
            )
        sel = select_from_list("Select a ColumnId to remove", labels)
        if not sel:
            raise ItemsStructuralError("[qsync:items:edit] Cancelled.")
        col_id = sel.split(":", 1)[0].strip()
        if not confirm(
            f"Delete SBS ColumnId {col_id} from {qid}? This is destructive.",
            default=False,
        ):
            raise ItemsStructuralError("[qsync:items:edit] Cancelled.")
        return stage_sbs_column_op(
            survey_id=survey_id,
            qid=qid,
            action="remove",
            html=None,
            column_id=col_id,
            allow_delete=allow_delete,
            interactive=True,
            experimental_unsupported=experimental_unsupported,
        )

    if target_norm == "sbs_column_answers":
        if not is_sbs_matrix:
            raise ItemsStructuralError(
                "[qsync:items:edit] SBS Column Answers are only supported for SBSMatrix questions."
            )
        additional = question.get("AdditionalQuestions") or {}
        if not isinstance(additional, dict) or not additional:
            raise ItemsStructuralError(
                "[qsync:items:edit] No AdditionalQuestions found for SBSMatrix question."
            )
        columns = sorted(str(k) for k in additional.keys())
        chosen_col = autocomplete_from_list(
            message="Select ColumnId",
            choices=columns,
            instruction="type to filter, enter to select",
        )
        if not chosen_col:
            raise ItemsStructuralError("[qsync:items:edit] Cancelled.")
        aq = additional.get(str(chosen_col)) or {}
        if not isinstance(aq, dict):
            raise ItemsStructuralError("[qsync:items:edit] Invalid AdditionalQuestion.")
        answers = aq.get("Answers") or {}
        if not isinstance(answers, dict) or not answers:
            raise ItemsStructuralError(
                "[qsync:items:edit] No SBS answers found for that column."
            )
        labels = []
        for aid, entry in answers.items():
            txt = ""
            if isinstance(entry, dict):
                txt = _text_preview(str(entry.get("Display") or ""), max_len=60)
            labels.append(f"{aid}: {txt}".strip())
        action = select_from_list("Select an action", ["add", "edit", "remove"])
        if not action:
            raise ItemsStructuralError("[qsync:items:edit] Cancelled.")
        action_norm = action.strip().lower()

        if action_norm == "add":
            existing = list(answers.keys())
            nums: list[int] = []
            for a in existing:
                try:
                    nums.append(int(str(a)))
                except Exception:
                    continue
            suggested = str(max(nums) + 1) if nums else ""
            aid = text_input("Enter new AnswerId", default=suggested)
            if aid is None:
                raise ItemsStructuralError("[qsync:items:edit] Cancelled.")
            aid = str(aid).strip()
            if not aid:
                raise ItemsStructuralError("[qsync:items:edit] AnswerId is required.")
            new_html = text_input(f"Enter HTML for new SBS AnswerId {aid}")
            if new_html is None:
                raise ItemsStructuralError("[qsync:items:edit] Cancelled.")
            _validate_html_or_raise(new_html)
            return stage_sbs_column_answer_op(
                survey_id=survey_id,
                qid=qid,
                action="add",
                html=new_html,
                column_id=str(chosen_col),
                answer_id=aid,
                allow_delete=allow_delete,
                interactive=True,
                experimental_unsupported=experimental_unsupported,
            )

        if action_norm == "edit":
            sel = select_from_list("Select an AnswerId to edit", labels)
            if not sel:
                raise ItemsStructuralError("[qsync:items:edit] Cancelled.")
            aid = sel.split(":", 1)[0].strip()
            new_html = text_input(f"Enter new HTML for SBS AnswerId {aid}")
            if new_html is None:
                raise ItemsStructuralError("[qsync:items:edit] Cancelled.")
            _validate_html_or_raise(new_html)
            return stage_sbs_column_answer_op(
                survey_id=survey_id,
                qid=qid,
                action="edit",
                html=new_html,
                column_id=str(chosen_col),
                answer_id=aid,
                allow_delete=allow_delete,
                interactive=True,
                experimental_unsupported=experimental_unsupported,
            )

        if not allow_delete:
            raise ItemsStructuralError(
                "[qsync:items:edit] Remove requires --allow-delete."
            )
        sel = select_from_list("Select an AnswerId to remove", labels)
        if not sel:
            raise ItemsStructuralError("[qsync:items:edit] Cancelled.")
        aid = sel.split(":", 1)[0].strip()
        if not confirm(
            f"Delete SBS AnswerId {aid} from column {chosen_col}? This is destructive.",
            default=False,
        ):
            raise ItemsStructuralError("[qsync:items:edit] Cancelled.")
        return stage_sbs_column_answer_op(
            survey_id=survey_id,
            qid=qid,
            action="remove",
            html=None,
            column_id=str(chosen_col),
            answer_id=aid,
            allow_delete=allow_delete,
            interactive=True,
            experimental_unsupported=experimental_unsupported,
        )

    # Options/Subitems: add/edit/remove
    action = select_from_list("Select an action", ["add", "edit", "remove"])
    if not action:
        raise ItemsStructuralError("[qsync:items:edit] Cancelled.")

    # Resolve container for list rendering. (Operation staging uses stage_structural_op
    # so it can map by question type.)
    if target_norm == "options":
        mapping_key = "Answers" if is_matrix else "Choices"
    elif target_norm == "subitems":
        mapping_key = "Choices" if (is_matrix or is_sbs_matrix) else "Answers"
    else:
        mapping_key = "Choices"
    existing_map = question.get(mapping_key) or {}
    existing_ids = list(existing_map.keys()) if isinstance(existing_map, dict) else []

    if action == "add":
        html = text_input("Enter HTML for the new Display")
        if html is None:
            raise ItemsStructuralError("[qsync:items:edit] Cancelled.")
        _validate_html_or_raise(html)
        return stage_structural_op(
            survey_id=survey_id,
            qid=qid,
            target=target_norm,
            action="add",
            html=html,
            text_format=None,
            item_id=None,
            allow_delete=allow_delete,
            interactive=True,
            experimental_unsupported=experimental_unsupported,
        )

    if action == "edit":
        if not existing_ids:
            raise ItemsStructuralError(
                "[qsync:items:edit] No existing entries to edit."
            )
        labels = []
        for cid in existing_ids:
            entry = existing_map.get(cid) or {}
            display = ""
            if isinstance(entry, dict):
                display = _text_preview(str(entry.get("Display") or ""), max_len=60)
            labels.append(f"{cid}: {display}".strip())
        sel = select_from_list("Select an ID to edit", labels)
        if not sel:
            raise ItemsStructuralError("[qsync:items:edit] Cancelled.")
        item_id = sel.split(":", 1)[0].strip()
        html = text_input(f"Enter new HTML for ID {item_id}")
        if html is None:
            raise ItemsStructuralError("[qsync:items:edit] Cancelled.")
        _validate_html_or_raise(html)
        return stage_structural_op(
            survey_id=survey_id,
            qid=qid,
            target=target_norm,
            action="edit",
            html=html,
            text_format=None,
            item_id=item_id,
            allow_delete=allow_delete,
            interactive=True,
            experimental_unsupported=experimental_unsupported,
        )

    if action == "remove":
        if not existing_ids:
            raise ItemsStructuralError(
                "[qsync:items:edit] No existing entries to remove."
            )
        if len(existing_ids) >= 50:
            if not confirm(
                f"This question has {len(existing_ids)} existing entries. Proceed to delete one?",
                default=False,
            ):
                raise ItemsStructuralError("[qsync:items:edit] Cancelled.")
        labels = []
        for cid in existing_ids:
            entry = existing_map.get(cid) or {}
            display = ""
            if isinstance(entry, dict):
                display = _text_preview(str(entry.get("Display") or ""), max_len=60)
            labels.append(f"{cid}: {display}".strip())
        sel = select_from_list("Select an ID to remove", labels)
        if not sel:
            raise ItemsStructuralError("[qsync:items:edit] Cancelled.")
        item_id = sel.split(":", 1)[0].strip()
        return stage_structural_op(
            survey_id=survey_id,
            qid=qid,
            target=target_norm,
            action="remove",
            html=None,
            text_format=None,
            item_id=item_id,
            allow_delete=allow_delete,
            interactive=True,
            experimental_unsupported=experimental_unsupported,
        )

    raise ItemsStructuralError("[qsync:items:edit] Cancelled.")


def _ops_affecting_qids(structural_ops: list[dict[str, Any]]) -> list[str]:
    qids: list[str] = []
    seen: set[str] = set()
    for op in structural_ops or []:
        qid = str(op.get("qid") or "").strip()
        if not qid or qid in seen:
            continue
        seen.add(qid)
        qids.append(qid)
    return qids


def push_structural_ops(
    *,
    survey_id: str,
    payload: dict,
    structural_ops: list[dict[str, Any]],
    push_journal: dict[str, Any],
    interactive: bool,
    allow_delete: bool,
    force_live: bool,
    force_preview: bool,
    publish: bool,
    dry_run: bool,
    refresh_cache: bool = True,
    save_journal_cb,
) -> None:
    """Push staged structural operations to Qualtrics with resume journal support.

    `save_journal_cb(push_journal)` must persist updated journal state.
    """

    qids = _ops_affecting_qids(structural_ops)
    if not qids:
        return

    # Re-check delete allow at push time.
    has_deletes = any(
        op.get("op")
        in {
            "choice_remove",
            "answer_remove",
            "sbs_column_remove",
            "sbs_column_answer_remove",
        }
        for op in (structural_ops or [])
    )
    if has_deletes and not allow_delete:
        if not interactive:
            raise ItemsStructuralError(
                "[qsync:items:push] Deletes present; re-run with --allow-delete."
            )
        if not confirm(
            "Deletes are staged. Proceed pushing deletes to Qualtrics?",
            default=False,
        ):
            raise ItemsStructuralError("[qsync:items:push] Cancelled.")

    if dry_run:
        info(
            "[qsync:items:push]",
            f"[dry-run] Would push structural ops affecting {len(qids)} QID(s): {', '.join(sorted(qids))}",
        )
        return

    config = SafeguardConfig(
        survey_id=survey_id,
        dimension="items",
        force_live=force_live,
        force_preview=force_preview,
        auto_yes=not interactive,
    )
    safeguard_result = enforce_push_safeguards(config)
    if safeguard_result.blocked:
        raise ItemsStructuralError(
            f"[qsync:items] Push blocked: {safeguard_result.message}"
        )
    if safeguard_result.warnings:
        for w in safeguard_result.warnings:
            warn("[qsync:items]", w)

    ensure_backup(survey_id)

    # Refresh cache once at push-time to reduce collision risk.
    if refresh_cache:
        info(
            "[qsync:items:push]",
            f"Refreshing cache from API for {format_survey_ref(survey_id)}…",
        )
        refresh_survey_cache(survey_id)
    survey = load_cached_survey(survey_id)

    enabled_langs = list_enabled_languages(survey_id)

    journal_pushed = set((push_journal.get("pushed_qids") or []))
    remaining = [qid for qid in qids if qid not in journal_pushed]
    if not remaining:
        info("[qsync:items:push]", "All structural QIDs already pushed (journal).")
        return

    # Apply + push per QID so we can record progress.
    for qid in remaining:
        question = survey.questions.get(qid)
        if not isinstance(question, dict):
            raise ItemsStructuralError(
                f"[qsync:items:push] QID {qid} not found in refreshed payload."
            )

        tag = (question.get("DataExportTag") or "").strip() or None
        owner = external_owner_for(qid=qid, data_export_tag=tag)
        if owner:
            raise ItemsStructuralError(
                f"[qsync:items:push] QID {qid} is externally managed (owner={owner})."
            )

        ops_for_qid = [
            op
            for op in (structural_ops or [])
            if str(op.get("qid") or "").strip() == qid
        ]
        for op in ops_for_qid:
            op_type = op.get("op")
            cid = str(op.get("choice_id") or "").strip()
            if op_type == "choice_add":
                html = str(op.get("html") or "")
                if not cid or not html:
                    raise ItemsStructuralError(
                        "[qsync:items:push] Malformed choice_add op."
                    )
                choices = question.setdefault("Choices", {})
                if cid in choices:
                    existing = choices.get(cid) or {}
                    if (
                        isinstance(existing, dict)
                        and str(existing.get("Display") or "") == html
                    ):
                        _append_choice_order(question, cid)
                    else:
                        raise ItemsStructuralError(
                            f"[qsync:items:push] ChoiceId {cid} already exists for {qid} with different content."
                        )
                else:
                    choices[cid] = {"Display": html, "Display_Unsafe": html}
                    _append_choice_order(question, cid)
                    # Keep NextChoiceId monotonic.
                    try:
                        next_candidate = int(cid) + 1
                        current_next = question.get("NextChoiceId")
                        if isinstance(current_next, int):
                            question["NextChoiceId"] = max(current_next, next_candidate)
                        else:
                            question["NextChoiceId"] = next_candidate
                    except (TypeError, ValueError):
                        pass
                if _is_sbs_matrix_question(question):
                    additional = question.get("AdditionalQuestions") or {}
                    if isinstance(additional, dict):
                        for aq in additional.values():
                            if not isinstance(aq, dict):
                                continue
                            aq_choices = aq.get("Choices")
                            if not isinstance(aq_choices, dict):
                                continue
                            aq_choices[cid] = {"Display": html, "Display_Unsafe": html}
                            _append_choice_order(aq, cid)
            elif op_type == "choice_edit":
                html = str(op.get("html") or "")
                if not cid or not html:
                    raise ItemsStructuralError(
                        "[qsync:items:push] Malformed choice_edit op."
                    )
                entry = (question.get("Choices") or {}).get(cid)
                if not isinstance(entry, dict):
                    raise ItemsStructuralError(
                        f"[qsync:items:push] ChoiceId {cid} not found for {qid}."
                    )
                entry["Display"] = html
                if "Display_Unsafe" in entry:
                    entry["Display_Unsafe"] = html
                if _is_sbs_matrix_question(question):
                    additional = question.get("AdditionalQuestions") or {}
                    if isinstance(additional, dict):
                        for aq in additional.values():
                            if not isinstance(aq, dict):
                                continue
                            aq_choices = aq.get("Choices")
                            if not isinstance(aq_choices, dict):
                                continue
                            aq_entry = aq_choices.get(cid)
                            if isinstance(aq_entry, dict):
                                aq_entry["Display"] = html
                                if "Display_Unsafe" in aq_entry:
                                    aq_entry["Display_Unsafe"] = html
            elif op_type == "choice_remove":
                if not cid:
                    raise ItemsStructuralError(
                        "[qsync:items:push] Malformed choice_remove op."
                    )
                choices = question.get("Choices") or {}
                if cid in choices:
                    del choices[cid]
                _remove_choice_from_order(question, cid)
                _cleanup_choice_translations(question, cid, enabled_langs)
                if _is_sbs_matrix_question(question):
                    additional = question.get("AdditionalQuestions") or {}
                    if isinstance(additional, dict):
                        for aq in additional.values():
                            if not isinstance(aq, dict):
                                continue
                            aq_choices = aq.get("Choices")
                            if not isinstance(aq_choices, dict):
                                continue
                            if cid in aq_choices:
                                del aq_choices[cid]
                            _remove_choice_from_order(aq, cid)
            elif op_type == "answer_add":
                aid = str(op.get("answer_id") or "").strip()
                html = str(op.get("html") or "")
                if not aid or not html:
                    raise ItemsStructuralError(
                        "[qsync:items:push] Malformed answer_add op."
                    )
                answers = question.setdefault("Answers", {})
                if aid in answers:
                    existing = answers.get(aid) or {}
                    if (
                        isinstance(existing, dict)
                        and str(existing.get("Display") or "") == html
                    ):
                        _append_answer_order(question, aid)
                    else:
                        raise ItemsStructuralError(
                            f"[qsync:items:push] AnswerId {aid} already exists for {qid} with different content."
                        )
                else:
                    answers[aid] = {"Display": html, "Display_Unsafe": html}
                    _append_answer_order(question, aid)
                    try:
                        next_candidate = int(aid) + 1
                        current_next = question.get("NextAnswerId")
                        if isinstance(current_next, int):
                            question["NextAnswerId"] = max(current_next, next_candidate)
                        else:
                            question["NextAnswerId"] = next_candidate
                    except (TypeError, ValueError):
                        pass
            elif op_type == "answer_edit":
                aid = str(op.get("answer_id") or "").strip()
                html = str(op.get("html") or "")
                if not aid or not html:
                    raise ItemsStructuralError(
                        "[qsync:items:push] Malformed answer_edit op."
                    )
                entry = (question.get("Answers") or {}).get(aid)
                if not isinstance(entry, dict):
                    raise ItemsStructuralError(
                        f"[qsync:items:push] AnswerId {aid} not found for {qid}."
                    )
                entry["Display"] = html
                if "Display_Unsafe" in entry:
                    entry["Display_Unsafe"] = html
            elif op_type == "answer_remove":
                aid = str(op.get("answer_id") or "").strip()
                if not aid:
                    raise ItemsStructuralError(
                        "[qsync:items:push] Malformed answer_remove op."
                    )
                answers = question.get("Answers") or {}
                if aid in answers:
                    del answers[aid]
                _remove_answer_from_order(question, aid)
                _cleanup_answer_translations(question, aid, enabled_langs)
            elif op_type == "question_text_edit":
                html = str(op.get("html") or "")
                if not html:
                    raise ItemsStructuralError(
                        "[qsync:items:push] Malformed question_text_edit op."
                    )
                question["QuestionText"] = html
                if "QuestionText_Unsafe" in question:
                    question["QuestionText_Unsafe"] = html
            elif op_type == "sbs_column_edit":
                html = str(op.get("html") or "")
                if not cid or not html:
                    raise ItemsStructuralError(
                        "[qsync:items:push] Malformed sbs_column_edit op."
                    )
                if not _is_sbs_matrix_question(question):
                    raise ItemsStructuralError(
                        f"[qsync:items:push] QID {qid} is not SBSMatrix; cannot apply sbs_column_edit."
                    )
                additional = question.get("AdditionalQuestions") or {}
                if not isinstance(additional, dict):
                    raise ItemsStructuralError(
                        "[qsync:items:push] Missing AdditionalQuestions map."
                    )
                aq = additional.get(str(cid))
                if not isinstance(aq, dict):
                    raise ItemsStructuralError(
                        f"[qsync:items:push] ColumnId {cid} not found for {qid}."
                    )
                aq["QuestionText"] = html
                if "QuestionText_Unsafe" in aq:
                    aq["QuestionText_Unsafe"] = html
            elif op_type == "sbs_column_add":
                html = str(op.get("html") or "")
                if not cid or not html:
                    raise ItemsStructuralError(
                        "[qsync:items:push] Malformed sbs_column_add op."
                    )
                if not _is_sbs_matrix_question(question):
                    raise ItemsStructuralError(
                        f"[qsync:items:push] QID {qid} is not SBSMatrix; cannot apply sbs_column_add."
                    )
                additional = question.setdefault("AdditionalQuestions", {})
                if not isinstance(additional, dict):
                    raise ItemsStructuralError(
                        "[qsync:items:push] Missing AdditionalQuestions map."
                    )
                if str(cid) in additional:
                    raise ItemsStructuralError(
                        f"[qsync:items:push] ColumnId {cid} already exists for {qid}."
                    )
                import copy

                template = next(
                    (v for v in additional.values() if isinstance(v, dict)), None
                )
                if not isinstance(template, dict):
                    raise ItemsStructuralError(
                        "[qsync:items:push] Cannot add SBS column (no template column found)."
                    )
                new_col = copy.deepcopy(template)
                new_col["QuestionText"] = html
                if "QuestionText_Unsafe" in new_col:
                    new_col["QuestionText_Unsafe"] = html
                additional[str(cid)] = new_col
                order = question.get("AdditionalQuestionsOrder")
                if isinstance(order, list) and str(cid) not in order:
                    order.append(str(cid))
            elif op_type == "sbs_column_remove":
                if not cid:
                    raise ItemsStructuralError(
                        "[qsync:items:push] Malformed sbs_column_remove op."
                    )
                if not _is_sbs_matrix_question(question):
                    raise ItemsStructuralError(
                        f"[qsync:items:push] QID {qid} is not SBSMatrix; cannot apply sbs_column_remove."
                    )
                additional = question.get("AdditionalQuestions") or {}
                if not isinstance(additional, dict):
                    raise ItemsStructuralError(
                        "[qsync:items:push] Missing AdditionalQuestions map."
                    )
                if str(cid) in additional:
                    del additional[str(cid)]
                order = question.get("AdditionalQuestionsOrder")
                if isinstance(order, list) and str(cid) in order:
                    question["AdditionalQuestionsOrder"] = [
                        x for x in order if str(x) != str(cid)
                    ]
            elif op_type == "sbs_column_answer_edit":
                col_id = str(op.get("choice_id") or "").strip()
                aid = str(op.get("answer_id") or "").strip()
                html = str(op.get("html") or "")
                if not col_id or not aid or not html:
                    raise ItemsStructuralError(
                        "[qsync:items:push] Malformed sbs_column_answer_edit op."
                    )
                if not _is_sbs_matrix_question(question):
                    raise ItemsStructuralError(
                        f"[qsync:items:push] QID {qid} is not SBSMatrix; cannot apply sbs_column_answer_edit."
                    )
                additional = question.get("AdditionalQuestions") or {}
                if not isinstance(additional, dict):
                    raise ItemsStructuralError(
                        "[qsync:items:push] Missing AdditionalQuestions map."
                    )
                aq = additional.get(str(col_id))
                if not isinstance(aq, dict):
                    raise ItemsStructuralError(
                        f"[qsync:items:push] ColumnId {col_id} not found for {qid}."
                    )
                answers = aq.get("Answers") or {}
                if not isinstance(answers, dict):
                    raise ItemsStructuralError(
                        "[qsync:items:push] Missing SBS answers map."
                    )
                entry = answers.get(str(aid))
                if not isinstance(entry, dict):
                    raise ItemsStructuralError(
                        f"[qsync:items:push] AnswerId {aid} not found for {qid}#{col_id}."
                    )
                entry["Display"] = html
                if "Display_Unsafe" in entry:
                    entry["Display_Unsafe"] = html
            elif op_type == "sbs_column_answer_add":
                col_id = str(op.get("choice_id") or "").strip()
                aid = str(op.get("answer_id") or "").strip()
                html = str(op.get("html") or "")
                if not col_id or not aid or not html:
                    raise ItemsStructuralError(
                        "[qsync:items:push] Malformed sbs_column_answer_add op."
                    )
                if not _is_sbs_matrix_question(question):
                    raise ItemsStructuralError(
                        f"[qsync:items:push] QID {qid} is not SBSMatrix; cannot apply sbs_column_answer_add."
                    )
                additional = question.get("AdditionalQuestions") or {}
                if not isinstance(additional, dict):
                    raise ItemsStructuralError(
                        "[qsync:items:push] Missing AdditionalQuestions map."
                    )
                aq = additional.get(str(col_id))
                if not isinstance(aq, dict):
                    raise ItemsStructuralError(
                        f"[qsync:items:push] ColumnId {col_id} not found for {qid}."
                    )
                answers = aq.setdefault("Answers", {})
                if not isinstance(answers, dict):
                    raise ItemsStructuralError(
                        "[qsync:items:push] Missing SBS answers map."
                    )
                if str(aid) in answers:
                    raise ItemsStructuralError(
                        f"[qsync:items:push] AnswerId {aid} already exists for {qid}#{col_id}."
                    )
                import copy

                template = next(
                    (v for v in answers.values() if isinstance(v, dict)), None
                )
                entry = copy.deepcopy(template) if isinstance(template, dict) else {}
                entry["Display"] = html
                entry["Display_Unsafe"] = html
                answers[str(aid)] = entry
                order = aq.get("AnswerOrder")
                if isinstance(order, list) and str(aid) not in order:
                    order.append(str(aid))
            elif op_type == "sbs_column_answer_remove":
                col_id = str(op.get("choice_id") or "").strip()
                aid = str(op.get("answer_id") or "").strip()
                if not col_id or not aid:
                    raise ItemsStructuralError(
                        "[qsync:items:push] Malformed sbs_column_answer_remove op."
                    )
                if not _is_sbs_matrix_question(question):
                    raise ItemsStructuralError(
                        f"[qsync:items:push] QID {qid} is not SBSMatrix; cannot apply sbs_column_answer_remove."
                    )
                additional = question.get("AdditionalQuestions") or {}
                if not isinstance(additional, dict):
                    raise ItemsStructuralError(
                        "[qsync:items:push] Missing AdditionalQuestions map."
                    )
                aq = additional.get(str(col_id))
                if not isinstance(aq, dict):
                    raise ItemsStructuralError(
                        f"[qsync:items:push] ColumnId {col_id} not found for {qid}."
                    )
                answers = aq.get("Answers") or {}
                if not isinstance(answers, dict):
                    raise ItemsStructuralError(
                        "[qsync:items:push] Missing SBS answers map."
                    )
                if str(aid) in answers:
                    del answers[str(aid)]
                order = aq.get("AnswerOrder")
                if isinstance(order, list) and str(aid) in order:
                    aq["AnswerOrder"] = [x for x in order if str(x) != str(aid)]
            else:
                raise ItemsStructuralError(
                    f"[qsync:items:push] Unknown op type: {op_type}"
                )

        try:
            push_questions(
                survey,
                [qid],
                context={
                    "origin": "qsync.items.push_structural_ops",
                    "ops": ops_for_qid,
                },
            )
        except Exception as e:
            # Persist progress before failing.
            push_journal.setdefault("pushed_qids", [])
            push_journal["pushed_qids"] = sorted(journal_pushed)
            save_journal_cb(push_journal)
            raise ItemsStructuralError(f"[qsync:items:push] Failed pushing {qid}: {e}")

        journal_pushed.add(qid)
        push_journal.setdefault("pushed_qids", [])
        push_journal["pushed_qids"] = sorted(journal_pushed)
        save_journal_cb(push_journal)
        info("[qsync:items:push]", f"Pushed structural changes for {qid}.")
        log_push_event(
            action="qsync.items.push.structural_qid",
            method="PUT",
            path=f"survey-definitions/{survey_id}/questions/{qid}",
            survey_id=survey_id,
            meta={"qid": qid, "ops": ops_for_qid},
        )

    survey.save()

    if publish:
        auto_publish_after_push(
            survey_id=survey_id,
            dimension="items",
            changed_qids=sorted(journal_pushed),
            count=len(journal_pushed),
            skip_publish=False,
            auto_yes=not interactive,
        )
