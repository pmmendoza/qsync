"""
Survey master CSV management for qsync.

Implements focal-only "survey master" workflow:
- Pull selected Qualtrics fields into CSV with per-survey snapshots
- Edit explicitly allowlisted fields in the CSV
- Preview diffs against saved snapshots
- Apply changes back to Qualtrics with audit logging
"""

from __future__ import annotations

import csv
import hashlib
import json
import os
import io
import re
from functools import lru_cache
from datetime import datetime, timezone
from pathlib import Path
from importlib import resources
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from .api_push import send_api_request
from .config import get_active_account, get_client_config, resolve_root, resolve_scoped_dir
from .rich_support import progress_context, should_use_rich
from .survey_naming import resolve_survey_path
from .survey_inventory import load_focal_snapshot, load_existing_metadata
from .survey_master_validation import validate_all_changes, format_validation_errors
from .workspace_paths import mapping_csv_candidates


def _workspace_root() -> Path:
    return resolve_root(required=False) or Path.cwd()


def _surveys_dir() -> Path:
    root = _workspace_root()
    account = get_active_account()
    return resolve_scoped_dir("surveys", root=root, account=account)


def _snapshots_dir() -> Path:
    return _surveys_dir() / "qualtrics_master_snapshots"


def _rollback_dir() -> Path:
    return _surveys_dir() / "qualtrics_master_rollback"


def _rollback_survey_dir(survey_id: str) -> Path:
    root = _workspace_root()
    return resolve_survey_path(
        _rollback_dir(),
        survey_id,
        is_dir=True,
        root=root,
        prefer_existing=True,
        migrate_existing=True,
    )


def _survey_id_from_named_segment(value: str) -> str | None:
    text = str(value or "").strip()
    if not text:
        return None
    if re.fullmatch(r"SV_[A-Za-z0-9]+", text):
        return text
    marker = "-SV_"
    if marker not in text:
        return None
    suffix = text.rsplit(marker, 1)[-1]
    candidate = f"SV_{suffix}"
    if re.fullmatch(r"SV_[A-Za-z0-9]+", candidate):
        return candidate
    return None


def _master_csv_path() -> Path:
    return _surveys_dir() / "qualtrics_master.csv"


def _master_workbook_path() -> Path:
    return _master_csv_path().with_suffix(".xlsx")


def _mapping_csv_path() -> Path:
    override = os.environ.get("QSYNC_MAPPING_CSV")
    if override:
        return Path(override).expanduser().resolve()

    for candidate in mapping_csv_candidates(root=_workspace_root()):
        if candidate.exists():
            return candidate

    # Preserve legacy fallback path for warning messages.
    return mapping_csv_candidates(root=_workspace_root())[0]


def _load_mapping_csv_text() -> tuple[str, str]:
    """Return (csv_text, source_label) for a workspace CSV override, if any.

    Checks workspace paths and QSYNC_MAPPING_CSV env var.
    Returns (csv_text, source_label) or raises FileNotFoundError.
    """
    mapping_path = _mapping_csv_path()
    if os.environ.get("QSYNC_MAPPING_CSV") and not mapping_path.exists():
        raise FileNotFoundError(
            f"Mapping CSV not found: {mapping_path} (from QSYNC_MAPPING_CSV)"
        )
    if mapping_path.exists():
        return mapping_path.read_text(encoding="utf-8"), str(mapping_path)

    raise FileNotFoundError("No workspace mapping CSV found")


def _parse_csv_text(csv_text: str) -> Dict[str, Dict[str, Any]]:
    """Parse CSV text into a field mapping dict.

    Returns a dict mapping field_name -> field_info (dict with CSV columns).
    Only includes fields with survey_master=read or write.
    """
    fields: Dict[str, Dict[str, Any]] = {}
    fh = io.StringIO(csv_text)
    reader = csv.DictReader(fh)
    for row in reader:
        field_name = (row.get("field_name") or "").strip()
        survey_master = (row.get("survey_master") or "").strip().lower()

        if survey_master not in ("read", "write"):
            continue
        if not field_name:
            continue

        if field_name not in fields:
            fields[field_name] = row
    return fields


def _load_packaged_json() -> Dict[str, Dict[str, Any]]:
    """Load the packaged field mapping JSON shipped with qsync.

    Returns a dict mapping field_name -> field_info, same shape as CSV parsing.
    """
    packaged = resources.files("qsync").joinpath("resources/field_mapping.json")
    data = json.loads(packaged.read_text(encoding="utf-8"))
    fields: Dict[str, Dict[str, Any]] = {}
    for entry in data:
        field_name = (entry.get("field_name") or "").strip()
        survey_master = (entry.get("survey_master") or "").strip().lower()

        if survey_master not in ("read", "write"):
            continue
        if not field_name:
            continue

        if field_name not in fields:
            fields[field_name] = entry
    return fields


def _parse_mapping_csv() -> Dict[str, Dict[str, Any]]:
    """Parse the field mapping and return fields with survey_master=read or write.

    Resolution order:
    1. QSYNC_MAPPING_CSV env var (CSV)
    2. surveys/qualtrics_api_key_mapping.csv (CSV, legacy shared path)
    3. accounts/default/qualtrics_api_key_mapping.csv (CSV, account-root fallback)
    4. appendices/qualtrics_api_key_mapping.csv (CSV, legacy)
    5. Packaged field_mapping.json (JSON, shipped with qsync)

    Returns a dict mapping field_name -> field_info.
    """
    # Try workspace CSV override first
    mapping_path = _mapping_csv_path()
    override = os.environ.get("QSYNC_MAPPING_CSV")

    if override and not mapping_path.exists():
        raise FileNotFoundError(
            f"Mapping CSV not found: {mapping_path} (from QSYNC_MAPPING_CSV)"
        )

    if mapping_path.exists():
        try:
            csv_text = mapping_path.read_text(encoding="utf-8")
            return _parse_csv_text(csv_text)
        except Exception as exc:
            raise FileNotFoundError(
                f"Failed to parse mapping CSV ({mapping_path}): {exc}"
            ) from exc

    # Fall through to packaged JSON
    return _load_packaged_json()


def _derive_endpoint(field_info: Dict[str, str]) -> str:
    """Derive the endpoint (metadata, options, status) from field_info.

    Uses the domain column to determine endpoint:
    - survey_metadata -> metadata
    - survey_options -> options
    - survey_detail -> status
    - Others -> read-only, no apply endpoint
    """
    domain = field_info.get("domain", "").strip().lower()

    if "metadata" in domain:
        return "metadata"
    elif "options" in domain:
        return "options"
    elif "detail" in domain:
        return "status"
    else:
        return None  # Read-only, no apply endpoint


def _compute_schema_version() -> str:
    """Compute schema version hash from the active mapping source.

    Uses the workspace CSV if present, otherwise the packaged JSON.

    NOTE: This must be stable across days. Older snapshots included a date prefix
    (YYYYMMDD-<hash>); we intentionally no longer do that because it caused
    constant daily mismatches even when the mapping was unchanged.
    """
    try:
        mapping_path = _mapping_csv_path()
        if mapping_path.exists():
            source_text = mapping_path.read_text(encoding="utf-8")
        else:
            packaged = resources.files("qsync").joinpath("resources/field_mapping.json")
            source_text = packaged.read_text(encoding="utf-8")
        source_hash = hashlib.md5(source_text.encode("utf-8")).hexdigest()[:8]
    except Exception:
        return "unknown"

    return source_hash


def _schema_version_hash(schema_version: str) -> str:
    """Extract the stable hash portion from a schema version string.

    Supports both:
    - new format: "<hash>"
    - legacy format: "YYYYMMDD-<hash>"
    """
    raw = (schema_version or "").strip()
    if not raw:
        return ""
    if raw == "unknown":
        return raw
    if len(raw) == 8 and all(c in "0123456789abcdef" for c in raw.lower()):
        return raw.lower()
    if len(raw) >= 9 and raw[8] == "-":
        suffix = raw.split("-", 1)[-1].strip().lower()
        if len(suffix) == 8 and all(c in "0123456789abcdef" for c in suffix):
            return suffix
    return raw.lower()


def _fetch_endpoint(
    base_url: str, headers: Dict[str, str], survey_id: str, endpoint_type: str
) -> Tuple[dict, str]:
    """Fetch data from a specific endpoint and return (data, timestamp).

    Args:
        base_url: Qualtrics base URL
        headers: API headers
        survey_id: Survey ID to fetch
        endpoint_type: One of 'status', 'metadata', 'options', 'versions'

    Returns:
        (data_dict, iso_timestamp) where data_dict is the API response "result"
    """
    now_str = datetime.now(timezone.utc).isoformat() + "Z"

    if endpoint_type == "status":
        # GET /surveys/{surveyId}
        resp = send_api_request(
            action="qsync.master.fetch.status",
            method="GET",
            base_url=base_url,
            headers=headers,
            path=f"surveys/{survey_id}",
            log_event=False,
            timeout=30,
        )
        return resp.json().get("result", {}), now_str

    elif endpoint_type == "metadata":
        # GET /survey-definitions/{surveyId}/metadata
        resp = send_api_request(
            action="qsync.master.fetch.metadata",
            method="GET",
            base_url=base_url,
            headers=headers,
            path=f"survey-definitions/{survey_id}/metadata",
            log_event=False,
            timeout=30,
        )
        return resp.json().get("result", {}), now_str

    elif endpoint_type == "options":
        # GET /survey-definitions/{surveyId}/options
        resp = send_api_request(
            action="qsync.master.fetch.options",
            method="GET",
            base_url=base_url,
            headers=headers,
            path=f"survey-definitions/{survey_id}/options",
            log_event=False,
            timeout=30,
        )
        return resp.json().get("result", {}), now_str

    elif endpoint_type == "versions":
        # GET /survey-definitions/{surveyId}/versions
        resp = send_api_request(
            action="qsync.master.fetch.versions",
            method="GET",
            base_url=base_url,
            headers=headers,
            path=f"survey-definitions/{survey_id}/versions",
            log_event=False,
            timeout=30,
        )
        # Reduce to latest published version
        elements = resp.json().get("result", {}).get("elements", [])
        latest_published = _reduce_to_latest_published(elements)
        return latest_published, now_str

    else:
        raise ValueError(f"Unknown endpoint type: {endpoint_type}")


def _reduce_to_latest_published(versions_list: List[dict]) -> dict:
    """Filter to latest published version (highest versionNumber where published=true).

    Returns a dict with versionID, versionNumber, description, creationDate, published.
    Returns empty dict if no published versions found.
    """
    published_versions = [
        v for v in versions_list if v.get("metadata", {}).get("published") is True
    ]

    if not published_versions:
        return {}

    # Sort by versionNumber descending, take first
    latest = max(
        published_versions,
        key=lambda v: int(v.get("metadata", {}).get("versionNumber", 0)),
    )

    meta = latest.get("metadata", {})
    return {
        "versionID": meta.get("versionID"),
        "versionNumber": meta.get("versionNumber"),
        "description": meta.get("description"),
        "creationDate": meta.get("creationDate"),
        "published": meta.get("published"),
    }


def _fetch_survey_name(base_url: str, headers: Dict[str, str], survey_id: str) -> str:
    """Fetch survey name from status endpoint."""
    resp = send_api_request(
        action="qsync.master.fetch.survey.name",
        method="GET",
        base_url=base_url,
        headers=headers,
        path=f"surveys/{survey_id}",
        log_event=False,
        timeout=30,
    )
    return resp.json().get("result", {}).get("name", survey_id)


def create_snapshot(
    survey_id: str,
    survey_name: str,
    status_data: dict,
    metadata_data: dict,
    options_data: dict,
    versions_data: dict,
) -> dict:
    """Build a snapshot JSON with all sections and metadata.

    Returns the snapshot dict (ready to be written as JSON).
    """
    schema_version = _compute_schema_version()
    now_str = datetime.now(timezone.utc).isoformat() + "Z"

    return {
        "survey_id": survey_id,
        "survey_name": survey_name,
        "schema_version": schema_version,
        "pulled_at": now_str,
        "sections": {
            "status": {
                "source_endpoint": "GET /surveys/{surveyId}",
                "pulled_at": now_str,
                "data": status_data,
            },
            "metadata": {
                "source_endpoint": "GET /survey-definitions/{surveyId}/metadata",
                "pulled_at": now_str,
                "data": metadata_data,
            },
            "options": {
                "source_endpoint": "GET /survey-definitions/{surveyId}/options",
                "pulled_at": now_str,
                "data": options_data,
            },
            "versions": {
                "source_endpoint": "GET /survey-definitions/{surveyId}/versions",
                "pulled_at": now_str,
                "data": versions_data,
            },
        },
    }


def save_snapshot(survey_id: str, snapshot: dict) -> Path:
    """Save a snapshot JSON to disk under qualtrics_master_snapshots/.

    Returns the path to the saved file.
    """
    snapshots_dir = _snapshots_dir()
    snapshots_dir.mkdir(parents=True, exist_ok=True)

    snapshot_path = resolve_survey_path(
        snapshots_dir,
        survey_id,
        suffix=".json",
        is_dir=False,
        root=_workspace_root(),
        prefer_existing=False,
        migrate_existing=True,
    )
    snapshot_path.write_text(json.dumps(snapshot, indent=2), encoding="utf-8")

    return snapshot_path


def load_snapshot(survey_id: str) -> Optional[dict]:
    """Load a snapshot JSON from disk, if it exists.

    Returns the snapshot dict, or None if not found.
    """
    snapshot_path = resolve_survey_path(
        _snapshots_dir(),
        survey_id,
        suffix=".json",
        is_dir=False,
        root=_workspace_root(),
        prefer_existing=True,
        migrate_existing=False,
    )
    if not snapshot_path.exists():
        return None

    try:
        return json.loads(snapshot_path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return None


def refresh_snapshot_from_live(
    survey_id: str,
    *,
    base_url: str | None = None,
    headers: Dict[str, str] | None = None,
) -> Path:
    """Refresh the on-disk master snapshot for `survey_id` from live Qualtrics.

    This updates only the read-only baseline under `surveys/qualtrics_master_snapshots/`
    and must not touch any editing surfaces (e.g. the master CSV).
    """
    if base_url is None or headers is None:
        base_url, headers = get_client_config()

    status_data, _ = _fetch_endpoint(base_url, headers, survey_id, "status")
    survey_name = str(status_data.get("name") or survey_id)
    metadata_data, _ = _fetch_endpoint(base_url, headers, survey_id, "metadata")
    options_data, _ = _fetch_endpoint(base_url, headers, survey_id, "options")
    versions_data, _ = _fetch_endpoint(base_url, headers, survey_id, "versions")

    snapshot = create_snapshot(
        survey_id=survey_id,
        survey_name=survey_name,
        status_data=status_data,
        metadata_data=metadata_data,
        options_data=options_data,
        versions_data=versions_data,
    )
    return save_snapshot(survey_id, snapshot)


def _scalar_to_string(value: Any) -> str:
    """Normalize values to the same string form used by master CSV comparisons."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (dict, list)):
        return json.dumps(value, sort_keys=True, ensure_ascii=False, separators=(",", ":"))
    return str(value)


def _rollback_filename_timestamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")


def _load_rollback_snapshot_file(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def _list_rollback_snapshot_paths(survey_id: str) -> List[Path]:
    survey_dir = _rollback_survey_dir(survey_id)
    if not survey_dir.exists():
        return []
    # Filenames are UTC timestamps, so lexical sort is chronological.
    return sorted(survey_dir.glob("*-pre-apply.json"), reverse=True)


def capture_pre_apply_snapshot(
    survey_id: str,
    changes: List[Dict[str, Any]],
) -> Path:
    """Capture a rollback snapshot before writing survey-master apply changes."""
    base_url, headers = get_client_config()
    captured_at = datetime.now(timezone.utc).isoformat() + "Z"

    mapping = _parse_mapping_csv()

    # Only fetch endpoints needed to compute pre-apply values for the fields we're writing.
    endpoints_needed: set[str] = set()
    for change in changes:
        field_name = str(change.get("field") or "").strip()
        if not field_name:
            continue
        field_info = mapping.get(field_name)
        endpoint = (
            str(change.get("endpoint") or "").strip()
            or (_derive_endpoint(field_info) if field_info else None)
        )
        if endpoint in {"metadata", "options", "status"}:
            endpoints_needed.add(endpoint)

    # Prefer local snapshot survey_name to avoid an extra API call.
    snap = load_snapshot(survey_id) or {}
    survey_name = str(snap.get("survey_name") or "").strip() or survey_id

    status_data: dict = {}
    metadata_data: dict = {}
    options_data: dict = {}
    versions_data: dict = {}

    if "status" in endpoints_needed:
        status_data, _ = _fetch_endpoint(base_url, headers, survey_id, "status")
        survey_name = str(status_data.get("name") or survey_name)
    if "metadata" in endpoints_needed:
        metadata_data, _ = _fetch_endpoint(base_url, headers, survey_id, "metadata")
    if "options" in endpoints_needed:
        options_data, _ = _fetch_endpoint(base_url, headers, survey_id, "options")

    snapshot = create_snapshot(
        survey_id=survey_id,
        survey_name=survey_name,
        status_data=status_data,
        metadata_data=metadata_data,
        options_data=options_data,
        versions_data=versions_data,
    )

    rollback_changes: List[Dict[str, Any]] = []
    for change in changes:
        field_name = str(change.get("field") or "").strip()
        if not field_name:
            continue
        field_info = mapping.get(field_name)
        pre_apply_value = (
            _extract_value_from_snapshot(snapshot, field_info) if field_info else None
        )
        rollback_changes.append(
            {
                "field": field_name,
                "endpoint": change.get("endpoint")
                or (_derive_endpoint(field_info) if field_info else "unknown"),
                "is_dangerous": bool(change.get("is_dangerous", False)),
                "pre_apply_value": _scalar_to_string(pre_apply_value),
                "target_value": _scalar_to_string(change.get("new_value")),
            }
        )

    snapshot["rollback"] = {
        "snapshot_type": "pre_apply",
        "captured_at": captured_at,
        "apply_timestamp": captured_at,
        "schema_version": _compute_schema_version(),
        "applied_changes": rollback_changes,
    }

    survey_dir = _rollback_survey_dir(survey_id)
    survey_dir.mkdir(parents=True, exist_ok=True)
    path = survey_dir / f"{_rollback_filename_timestamp()}-pre-apply.json"
    path.write_text(json.dumps(snapshot, indent=2), encoding="utf-8")
    return path


def list_rollback_versions(survey_id: Optional[str] = None) -> List[Dict[str, Any]]:
    """List available rollback snapshots, newest first."""
    root = _rollback_dir()
    if not root.exists():
        return []

    survey_ids: List[str]
    if survey_id:
        survey_ids = [survey_id]
    else:
        survey_ids = sorted(
            {
                sid
                for p in root.iterdir()
                if p.is_dir()
                for sid in [(_survey_id_from_named_segment(p.name))]
                if sid
            }
        )

    entries: List[Dict[str, Any]] = []
    for sid in survey_ids:
        paths = _list_rollback_snapshot_paths(sid)
        for idx, path in enumerate(paths, start=1):
            try:
                snap = _load_rollback_snapshot_file(path)
            except (json.JSONDecodeError, OSError):
                continue
            rollback_meta = snap.get("rollback", {})
            applied_changes = rollback_meta.get("applied_changes", [])
            captured_at = (
                rollback_meta.get("captured_at")
                or snap.get("pulled_at")
                or path.name.split("-pre-apply.json")[0]
            )
            fields = [
                str(c.get("field") or "").strip()
                for c in applied_changes
                if str(c.get("field") or "").strip()
            ]
            entries.append(
                {
                    "survey_id": sid,
                    "version": idx,
                    "captured_at": captured_at,
                    "changes_count": len(fields),
                    "fields": fields,
                    "path": path,
                }
            )

    entries.sort(
        key=lambda row: (
            str(row.get("survey_id") or ""),
            int(row.get("version") or 0),
        )
    )
    return entries


def load_rollback_snapshot(survey_id: str, version: int = 1) -> Tuple[Path, dict]:
    """Load the Nth most recent rollback snapshot for a survey."""
    if version < 1:
        raise ValueError("version must be >= 1")

    paths = _list_rollback_snapshot_paths(survey_id)
    if not paths:
        raise FileNotFoundError(f"No rollback snapshots found for {survey_id}")
    if version > len(paths):
        raise ValueError(
            f"Requested version {version}, but only {len(paths)} snapshot(s) exist for {survey_id}"
        )
    path = paths[version - 1]
    try:
        snap = _load_rollback_snapshot_file(path)
    except (json.JSONDecodeError, OSError) as exc:
        raise ValueError(f"Invalid rollback snapshot: {path}") from exc
    return path, snap


def _get_default_column_order(fields: Dict[str, Dict[str, Any]]) -> List[str]:
    """Determine column order based on mapping 'order' values (config-free default)."""

    ordered_fields: list[tuple[int, str]] = []
    unordered_fields: list[str] = []
    readonly_fields: list[str] = []

    for field_name, field_info in (fields or {}).items():
        order_str = str(field_info.get("order", "") or "").strip()
        domain = str(field_info.get("domain", "") or "").strip().lower()

        if order_str and order_str.isdigit():
            ordered_fields.append((int(order_str), field_name))
        else:
            # external_metadata fields (_focal, _locked, etc.) should not go to readonly section
            if field_name.startswith("_") and "external_metadata" not in domain:
                readonly_fields.append(field_name)
            else:
                unordered_fields.append(field_name)

    # Sort ordered by order value, then append unordered alphabetically
    ordered_fields.sort(key=lambda x: x[0])
    unordered_fields.sort()
    readonly_fields.sort()

    result = [f[1] for f in ordered_fields] + unordered_fields + readonly_fields
    return result


def _get_column_order() -> List[str]:
    """Return the effective Survey Master column order (respects user config).

    Default behavior (no config file): preserve mapping 'order' semantics:
    1. Fields with numeric order values, sorted by order
    2. Remaining fields, sorted alphabetically
    3. Read-only derived fields prefixed with _ (appended at end, except external_metadata)

    If `survey_master_columns.yaml` exists in the workspace root (or the override
    env var is set), order and visibility are driven by that config.
    """

    fields = _parse_mapping_csv()
    default_order = _get_default_column_order(fields)

    # Optional user override (two editing surfaces: YAML + TUI).
    try:
        from .survey_master_columns import (
            master_columns_config_path,
            load_master_columns_yaml,
            resolve_master_columns,
        )

        config_path = master_columns_config_path(root=_workspace_root())
        config_data = load_master_columns_yaml(config_path)
        if config_data is None:
            return default_order

        columns, warnings = resolve_master_columns(
            available_in_default_order=default_order,
            config_data=config_data,
        )
        if warnings and not os.environ.get("QSYNC_JSON_MODE", "").strip():
            for warning in warnings:
                print(f"[qsync:master-columns] WARNING: {warning}", flush=True)

        return [c.name for c in columns if c.enabled]
    except Exception:
        # Config is best-effort: if parsing fails, fall back to the default order.
        return default_order


def _extract_nested_value(obj: dict, path_str: str) -> Any:
    """Extract a value from a nested dict using a dotted path string.

    Example: obj={'a': {'b': 5}}, path='a.b' -> returns 5
    """
    parts = path_str.split(".")
    current = obj
    for part in parts:
        if isinstance(current, dict):
            current = current.get(part)
        else:
            return None
    return current


def _extract_value_from_snapshot(snapshot: dict, field_info: Dict[str, str]) -> Any:
    """Extract a baseline value from a snapshot using mapping field_info."""

    endpoint = _derive_endpoint(field_info)
    object_path = (field_info.get("object_path") or "").strip()
    if not object_path:
        return None

    if endpoint == "metadata":
        section_data = snapshot.get("sections", {}).get("metadata", {}).get("data", {})
        field_key = object_path.split(".")[-1]
        return section_data.get(field_key)

    if endpoint == "options":
        section_data = snapshot.get("sections", {}).get("options", {}).get("data", {})
        parts = object_path.split(".")
        if parts and parts[0] == "result":
            parts = parts[1:]
        current: Any = section_data
        for part in parts:
            if isinstance(current, dict):
                current = current.get(part)
            else:
                return None
        return current

    if endpoint == "status":
        section_data = snapshot.get("sections", {}).get("status", {}).get("data", {})
        parts = object_path.split(".")
        if parts and parts[0] == "result":
            parts = parts[1:]
        current = section_data
        for part in parts:
            if isinstance(current, dict):
                current = current.get(part)
            else:
                return None
        return current

    return None


def _extract_value_from_live(
    live_result: dict, field_info: Dict[str, str], *, endpoint: str
) -> Any:
    """Extract a value from a live endpoint `result` payload using mapping field_info."""

    object_path = (field_info.get("object_path") or "").strip()
    if not object_path:
        return None

    parts = object_path.split(".")
    if parts and parts[0] == "result":
        parts = parts[1:]

    current: Any = live_result
    for part in parts:
        if isinstance(current, dict):
            current = current.get(part)
        else:
            return None
    return current


def generate_master_csv_from_snapshots(survey_ids: List[str]) -> List[List[str]]:
    """Generate master CSV rows from loaded snapshots.

    Args:
        survey_ids: List of focal survey IDs to include

    Returns:
        List of rows (first row is headers, remaining are data)
    """
    mapping = _parse_mapping_csv()
    column_order = _get_column_order()

    # Load external project metadata from qualtrics_surveys.csv
    external_metadata = load_existing_metadata()

    # Filter columns to only those in the mapping and marked for master
    columns = [c for c in column_order if c in mapping]

    # Build header row
    rows: List[List[str]] = [columns]

    # Build data rows
    for survey_id in survey_ids:
        snapshot = load_snapshot(survey_id)
        if not snapshot:
            continue

        row_data = {}
        for field_name in columns:
            field_info = mapping.get(field_name)
            if not field_info:
                row_data[field_name] = ""
                continue

            # Find which section this field comes from
            endpoint = _derive_endpoint(field_info)
            domain = field_info.get("domain", "").strip().lower()

            # Get the source path from mapping
            object_path = field_info.get("object_path", "").strip()
            if not object_path:
                row_data[field_name] = ""
                continue

            # Extract from snapshot based on domain and endpoint
            # Check external_metadata FIRST (before endpoint checks, since domain contains "metadata")
            if "external_metadata" in domain:
                # External project metadata from qualtrics_surveys.csv (_focal, _locked, etc.)
                survey_metadata = external_metadata.get(survey_id, {})
                metadata_key = field_name[1:]  # Remove _ prefix
                value = survey_metadata.get(metadata_key)
                if value is None:
                    value = ""
            elif endpoint == "metadata":
                section_data = (
                    snapshot.get("sections", {}).get("metadata", {}).get("data", {})
                )
                # object_path is like "result.SurveyName", we need "SurveyName"
                field_key = object_path.split(".")[-1]
                value = section_data.get(field_key)
            elif endpoint == "options":
                section_data = (
                    snapshot.get("sections", {}).get("options", {}).get("data", {})
                )
                # Handle nested like "result.CustomStyles.customCSS" -> ["CustomStyles", "customCSS"]
                parts = object_path.split(".")[1:]  # Skip "result"
                value = section_data
                for part in parts:
                    if isinstance(value, dict):
                        value = value.get(part)
                    else:
                        value = None
                        break
            elif endpoint == "status":
                section_data = (
                    snapshot.get("sections", {}).get("status", {}).get("data", {})
                )
                # object_path like "result.responseCounts.generated"
                parts = object_path.split(".")[1:]  # Skip "result"
                value = section_data
                for part in parts:
                    if isinstance(value, dict):
                        value = value.get(part)
                    else:
                        value = None
                        break
            elif "survey_def" in domain:
                # survey_def fields like SurveyID come from metadata endpoint in master
                section_data = (
                    snapshot.get("sections", {}).get("metadata", {}).get("data", {})
                )
                field_key = object_path.split(".")[-1]
                value = section_data.get(field_key)
            elif "versions_list" in domain:
                # Version fields are read-only
                versions_data = (
                    snapshot.get("sections", {}).get("versions", {}).get("data", {})
                )
                field_key = object_path.split(".")[-1]
                value = versions_data.get(field_key)
            elif field_name.startswith("_"):
                # Derived read-only fields from versions (_versionID, _versionNumber, etc.)
                versions_data = (
                    snapshot.get("sections", {}).get("versions", {}).get("data", {})
                )
                if field_name == "_versionID":
                    value = versions_data.get("versionID")
                elif field_name == "_versionNumber":
                    value = versions_data.get("versionNumber")
                elif field_name == "_lastPublishedDate":
                    value = versions_data.get("creationDate")
                elif field_name == "_publishedDescription":
                    value = versions_data.get("description")
                else:
                    value = None
            else:
                value = None

            row_data[field_name] = _scalar_to_string(value)

        row_values = [row_data.get(col, "") for col in columns]
        rows.append(row_values)

    return rows


_MASTER_MAIN_SHEET = "Survey_Master"
_MASTER_GUIDE_SHEET = "Survey_Master_Guide"
_MASTER_HEADER_FILL = PatternFill(fill_type="solid", fgColor="FFE8EEF7")
_MASTER_READONLY_FILL = PatternFill(fill_type="solid", fgColor="FFECECEC")
_MASTER_COMMENT_AUTHOR = "qsync"
_MASTER_HEADER_DOCS_RESOURCE = "resources/survey_master_header_docs.json"
_MASTER_COMMENT_MAX_LENGTH = 32000
_DEFAULT_OPTIONS_DOC_LINKS: list[dict[str, str]] = [
    {
        "label": "Get options (response body)",
        "url": "https://api.qualtrics.com/021740be5b5b6-get-options#response-body",
    },
    {
        "label": "Update options (request body)",
        "url": "https://api.qualtrics.com/5d9e865296ce5-update-options",
    },
]
_DEFAULT_OPTIONS_ENDPOINTS: list[str] = [
    "GET /survey-definitions/{surveyId}/options",
    "PUT /survey-definitions/{surveyId}/options",
]


def _is_master_editable(field_info: Dict[str, Any] | None) -> bool:
    if not field_info:
        return False
    mode = str(field_info.get("survey_master") or "").strip().lower()
    return mode == "write"


def _master_allowed_values(field_info: Dict[str, Any] | None) -> List[str]:
    if not field_info:
        return []
    raw = str(field_info.get("allowed_values") or "").strip()
    if raw:
        vals = [v.strip() for v in raw.split(";") if v.strip()]
        if vals:
            if {v.lower() for v in vals} == {"true", "false"}:
                return ["true", "false"]
            return vals
    data_type = str(field_info.get("data_type") or "").strip().lower()
    if data_type == "bool":
        return ["true", "false"]
    return []


@lru_cache(maxsize=1)
def _load_master_header_docs() -> Dict[str, Any]:
    """Load optional packaged docs links used for header comments."""

    try:
        packaged = resources.files("qsync").joinpath(_MASTER_HEADER_DOCS_RESOURCE)
        data = json.loads(packaged.read_text(encoding="utf-8"))
        if isinstance(data, dict):
            return data
    except Exception:
        pass
    return {}


def _normalize_comment_links(raw: Any) -> List[tuple[str, str]]:
    """Normalize docs link objects from config into (label, url) tuples."""

    if not isinstance(raw, list):
        return []

    links: List[tuple[str, str]] = []
    seen_urls: set[str] = set()
    for item in raw:
        label = ""
        url = ""
        if isinstance(item, str):
            url = item.strip()
            label = "Documentation"
        elif isinstance(item, dict):
            label = str(item.get("label") or "").strip() or "Documentation"
            url = str(item.get("url") or "").strip()
        if not url or url in seen_urls:
            continue
        seen_urls.add(url)
        links.append((label, url))
    return links


def _normalize_comment_endpoints(raw: Any) -> List[str]:
    """Normalize endpoint lists from config into endpoint strings."""

    if not isinstance(raw, list):
        return []

    endpoints: List[str] = []
    seen: set[str] = set()
    for item in raw:
        endpoint = ""
        if isinstance(item, str):
            endpoint = item.strip()
        elif isinstance(item, dict):
            method = str(item.get("method") or "").strip().upper()
            path = str(item.get("path") or "").strip()
            if method and path:
                endpoint = f"{method} {path}"
            else:
                endpoint = str(item.get("value") or "").strip()
        if not endpoint or endpoint in seen:
            continue
        seen.add(endpoint)
        endpoints.append(endpoint)
    return endpoints


def _master_comment_doc_links(
    field_name: str, field_info: Dict[str, Any] | None
) -> List[tuple[str, str]]:
    docs_config = _load_master_header_docs()
    links: List[tuple[str, str]] = []

    if isinstance(docs_config, dict):
        # Field-level links override and extend domain defaults.
        field_cfg = (docs_config.get("fields") or {}).get(field_name)
        if isinstance(field_cfg, dict):
            links.extend(_normalize_comment_links(field_cfg.get("links")))

        domain = str((field_info or {}).get("domain") or "").strip().lower()
        domain_cfg = (docs_config.get("domains") or {}).get(domain)
        if isinstance(domain_cfg, dict):
            links.extend(_normalize_comment_links(domain_cfg.get("links")))

    # Always provide Options docs for survey_options fields if absent from config.
    domain = str((field_info or {}).get("domain") or "").strip().lower()
    if domain == "survey_options" and not links:
        links.extend(_normalize_comment_links(_DEFAULT_OPTIONS_DOC_LINKS))

    deduped: List[tuple[str, str]] = []
    seen_urls: set[str] = set()
    for label, url in links:
        if url in seen_urls:
            continue
        seen_urls.add(url)
        deduped.append((label, url))
    return deduped


def _master_comment_endpoints(
    field_name: str, field_info: Dict[str, Any] | None
) -> List[str]:
    docs_config = _load_master_header_docs()
    endpoints: List[str] = []

    if isinstance(docs_config, dict):
        field_cfg = (docs_config.get("fields") or {}).get(field_name)
        if isinstance(field_cfg, dict):
            endpoints.extend(_normalize_comment_endpoints(field_cfg.get("endpoints")))

        domain = str((field_info or {}).get("domain") or "").strip().lower()
        domain_cfg = (docs_config.get("domains") or {}).get(domain)
        if isinstance(domain_cfg, dict):
            endpoints.extend(_normalize_comment_endpoints(domain_cfg.get("endpoints")))

    # Fallback for options when no config is available.
    domain = str((field_info or {}).get("domain") or "").strip().lower()
    if domain == "survey_options" and not endpoints:
        endpoints.extend(_DEFAULT_OPTIONS_ENDPOINTS)

    # Include direct mapping endpoint hints when present (workspace CSV mappings).
    read_endpoint = str((field_info or {}).get("read_endpoint") or "").strip()
    write_endpoint = str((field_info or {}).get("write_endpoint") or "").strip()
    if read_endpoint:
        endpoints.append(read_endpoint)
    if write_endpoint and _is_master_editable(field_info):
        endpoints.append(write_endpoint)

    deduped: List[str] = []
    seen: set[str] = set()
    for endpoint in endpoints:
        if endpoint in seen:
            continue
        seen.add(endpoint)
        deduped.append(endpoint)
    return deduped


def _build_master_header_comment(
    field_name: str, field_info: Dict[str, Any] | None
) -> str | None:
    """Build a compact per-column help comment shown on the header cell."""

    if not field_info:
        return None

    lines: List[str] = [field_name]
    lines.append(
        "Editable in Survey Master: "
        + ("Yes" if _is_master_editable(field_info) else "No (read-only)")
    )

    description = str(field_info.get("description") or "").strip()
    if description:
        lines.append(f"What it controls: {description}")

    allowed_values = _master_allowed_values(field_info)
    if allowed_values:
        lines.append("Allowed values:")
        for value in allowed_values:
            lines.append(f"- {value}")
    else:
        data_type = str(field_info.get("data_type") or "").strip()
        if data_type:
            lines.append(f"Expected type: {data_type}")

    format_notes = str(field_info.get("format_notes") or "").strip()
    if format_notes:
        lines.append(f"Format notes: {format_notes}")

    endpoints = _master_comment_endpoints(field_name, field_info)
    if endpoints:
        lines.append("API endpoints:")
        for endpoint in endpoints:
            lines.append(f"- {endpoint}")

    doc_links = _master_comment_doc_links(field_name, field_info)
    if doc_links:
        lines.append("Documentation:")
        for label, url in doc_links:
            lines.append(f"- {label}: {url}")

    text = "\n".join(line for line in lines if line.strip()).strip()
    if not text:
        return None
    if len(text) > _MASTER_COMMENT_MAX_LENGTH:
        return text[: _MASTER_COMMENT_MAX_LENGTH - 3].rstrip() + "..."
    return text


def _master_stringify_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (dict, list)):
        return json.dumps(value, sort_keys=True, ensure_ascii=False, separators=(",", ":"))
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _master_sheet_headers(ws) -> List[str]:
    if ws.max_row < 1 or ws.max_column < 1:
        return []
    raw_headers = [
        _master_stringify_cell(v) for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    ]
    last = 0
    for idx, header in enumerate(raw_headers, start=1):
        if header.strip():
            last = idx
    if last <= 0:
        return []
    return [raw_headers[idx - 1].strip() for idx in range(1, last + 1)]


def _apply_master_list_validation(
    ws,
    column_idx: int,
    allowed_values: List[str],
) -> None:
    if ws.max_row <= 1 or not allowed_values:
        return
    values = [str(v).strip() for v in allowed_values if str(v).strip()]
    if not values:
        return
    if any("," in v or '"' in v for v in values):
        return
    formula = '"' + ",".join(values) + '"'
    if len(formula) > 255:
        return
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    col_letter = get_column_letter(column_idx)
    dv.add(f"{col_letter}2:{col_letter}{ws.max_row}")


def _estimate_master_col_width(values: List[str]) -> float:
    max_len = 0
    for value in values:
        max_len = max(max_len, len(str(value or "")))
    if max_len <= 0:
        return 12.0
    return float(min(80, max(12, int(max_len * 1.08) + 2)))


def write_master_workbook(rows: List[List[str]]) -> Path:
    """Write survey-master workbook surface (`qualtrics_master.xlsx`)."""
    workbook_path = _master_workbook_path()
    workbook_path.parent.mkdir(parents=True, exist_ok=True)

    mapping = _parse_mapping_csv()
    if rows and rows[0]:
        headers = [str(h or "").strip() for h in rows[0]]
        data_rows = rows[1:]
    else:
        headers = _get_column_order()
        data_rows = []

    wb = Workbook()
    ws = wb.active
    ws.title = _MASTER_MAIN_SHEET
    ws.append(headers)

    for row in data_rows:
        values = list(row)
        if len(values) < len(headers):
            values.extend([""] * (len(headers) - len(values)))
        elif len(values) > len(headers):
            values = values[: len(headers)]
        ws.append([_master_stringify_cell(v) for v in values])

    ws.freeze_panes = "A2"

    editable_headers: set[str] = set()
    for col_idx, header in enumerate(headers, start=1):
        field_info = mapping.get(header)
        if _is_master_editable(field_info):
            editable_headers.add(header)
        header_cell = ws.cell(row=1, column=col_idx)
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(
            vertical="center",
            horizontal="left",
            wrap_text=False,
        )
        header_cell.fill = _MASTER_HEADER_FILL
        header_comment = _build_master_header_comment(header, field_info)
        if header_comment:
            header_cell.comment = Comment(header_comment, _MASTER_COMMENT_AUTHOR)

    for row_idx in range(2, ws.max_row + 1):
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(
                vertical="center",
                horizontal="left",
                wrap_text=False,
            )
            if header not in editable_headers:
                cell.fill = _MASTER_READONLY_FILL

    for col_idx, header in enumerate(headers, start=1):
        field_info = mapping.get(header)
        if not _is_master_editable(field_info):
            continue
        _apply_master_list_validation(ws, col_idx, _master_allowed_values(field_info))

    sample_rows = min(ws.max_row, 500)
    for col_idx, header in enumerate(headers, start=1):
        column_values = [header]
        for row_idx in range(2, sample_rows + 1):
            column_values.append(_master_stringify_cell(ws.cell(row=row_idx, column=col_idx).value))
        ws.column_dimensions[get_column_letter(col_idx)].width = _estimate_master_col_width(column_values)

    if headers:
        table_last_col = get_column_letter(len(headers))
        table_ref = f"A1:{table_last_col}{max(1, ws.max_row)}"
        table = Table(displayName="SurveyMasterTable", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    guide = wb.create_sheet(title=_MASTER_GUIDE_SHEET)
    guide_headers = [
        "Column",
        "Editability",
        "Description",
        "AllowedValues",
        "FormatNotes",
    ]
    guide.append(guide_headers)
    for idx, header in enumerate(guide_headers, start=1):
        cell = guide.cell(row=1, column=idx)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
        cell.fill = _MASTER_HEADER_FILL

    for header in headers:
        field_info = mapping.get(header, {})
        editability = "Editable" if _is_master_editable(field_info) else "Read-only"
        allowed_values = "; ".join(_master_allowed_values(field_info))
        description = str(field_info.get("description") or "").strip()
        format_notes = str(field_info.get("format_notes") or "").strip()
        guide.append([header, editability, description, allowed_values, format_notes])

    for row_idx in range(2, guide.max_row + 1):
        for col_idx in range(1, guide.max_column + 1):
            cell = guide.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)
            cell.fill = _MASTER_READONLY_FILL

    guide.column_dimensions["A"].width = 30.0
    guide.column_dimensions["B"].width = 14.0
    guide.column_dimensions["C"].width = 52.0
    guide.column_dimensions["D"].width = 28.0
    guide.column_dimensions["E"].width = 40.0
    guide.freeze_panes = "A2"

    wb.save(workbook_path)
    return workbook_path


def _load_master_csv_file() -> Tuple[List[str], List[Dict[str, str]]]:
    csv_path = _master_csv_path()
    if not csv_path.exists():
        return [], []

    rows: List[Dict[str, str]] = []
    headers: List[str] = []
    # Use utf-8-sig so files generated with a BOM are read correctly.
    with csv_path.open(newline="", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh)
        if reader.fieldnames:
            headers = [str(h or "") for h in reader.fieldnames]
        for row in reader:
            rows.append({str(k or ""): _master_stringify_cell(v) for k, v in row.items()})

    return headers, rows


def load_master_workbook() -> Tuple[List[str], List[Dict[str, str]]]:
    """Load survey-master rows from workbook surface (`qualtrics_master.xlsx`)."""
    workbook_path = _master_workbook_path()
    if not workbook_path.exists():
        return [], []

    wb = load_workbook(workbook_path, data_only=False)
    sheet_name = _MASTER_MAIN_SHEET if _MASTER_MAIN_SHEET in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    headers = _master_sheet_headers(ws)
    if not headers:
        return [], []

    rows: List[Dict[str, str]] = []
    for values in ws.iter_rows(
        min_row=2,
        max_row=ws.max_row,
        min_col=1,
        max_col=len(headers),
        values_only=True,
    ):
        values_list = list(values)
        if not any(_master_stringify_cell(v).strip() for v in values_list):
            continue
        row = {
            headers[idx]: _master_stringify_cell(values_list[idx] if idx < len(values_list) else "")
            for idx in range(len(headers))
        }
        rows.append(row)

    return headers, rows


def _latest_master_surface() -> str | None:
    csv_path = _master_csv_path()
    workbook_path = _master_workbook_path()
    has_csv = csv_path.exists()
    has_xlsx = workbook_path.exists()
    if not has_csv and not has_xlsx:
        return None
    if has_csv and not has_xlsx:
        return "csv"
    if has_xlsx and not has_csv:
        return "workbook"
    csv_mtime = csv_path.stat().st_mtime
    xlsx_mtime = workbook_path.stat().st_mtime
    return "workbook" if xlsx_mtime >= csv_mtime else "csv"


def write_master_csv(rows: List[List[str]]) -> Path:
    """Write survey-master rows to CSV + workbook surfaces.

    Returns:
        Path to the CSV surface for backwards compatibility.
    """
    csv_path = _master_csv_path()
    csv_path.parent.mkdir(parents=True, exist_ok=True)

    # Write with a UTF-8 BOM so spreadsheet apps (notably Excel) auto-detect UTF-8.
    # Our reader uses utf-8-sig, so the BOM will never leak into header names.
    with csv_path.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.writer(fh)
        for row in rows:
            writer.writerow(row)

    write_master_workbook(rows)
    return csv_path


def load_master_csv() -> Tuple[List[str], List[Dict[str, str]]]:
    """Load survey-master rows from the latest local editing surface.

    Source resolution:
    - If only one of CSV/XLSX exists, load that file.
    - If both exist, load the newer file by modification timestamp.
    """
    source = _latest_master_surface()
    if source == "workbook":
        return load_master_workbook()
    if source == "csv":
        return _load_master_csv_file()
    return [], []


def validate_master_csv(
    csv_headers: List[str], csv_rows: List[Dict[str, str]]
) -> List[str]:
    """Validate master CSV against schema.

    Returns list of error messages (empty if valid).
    """
    def _try_parse_boolish(raw: object) -> Optional[bool]:
        """Best-effort parse for boolean-ish values commonly produced by spreadsheets.

        We only use this for fields that are explicitly boolean in the mapping
        (data_type=bool) or have allowed_values of true/false.
        """
        if raw is None:
            return None
        if isinstance(raw, bool):
            return raw
        if not isinstance(raw, str):
            raw = str(raw)

        s = raw.strip().lower()
        if s == "":
            return None

        if s in ("true", "t", "yes", "y", "1", "on"):
            return True
        if s in ("false", "f", "no", "n", "0", "off"):
            return False
        return None

    def _is_booleanish_field(field_name: str, field_info: dict) -> bool:
        if field_name.startswith("_"):
            return False
        if field_name == "SurveyID":
            return False

        data_type = str(field_info.get("data_type") or "").strip().lower()
        if data_type == "bool":
            return True

        allowed_values = str(field_info.get("allowed_values") or "").strip()
        if not allowed_values:
            return False

        allowed = [v.strip().lower() for v in allowed_values.split(";") if v.strip()]
        return set(allowed) == {"true", "false"}

    errors: List[str] = []
    mapping = _parse_mapping_csv()
    valid_field_names = set(mapping.keys())

    # Normalize boolean-ish cells in-place so case differences like TRUE/FALSE
    # don't create spurious diffs and so validation isn't tripped by spreadsheets
    # that rewrite booleans in uppercase.
    for row in csv_rows:
        for field_name, field_value in list(row.items()):
            field_info = mapping.get(field_name)
            if not field_info:
                continue
            if not _is_booleanish_field(field_name, field_info):
                continue
            parsed = _try_parse_boolish(field_value)
            if parsed is None:
                continue
            row[field_name] = "true" if parsed else "false"

    # Check for unknown columns
    for header in csv_headers:
        if header not in valid_field_names:
            errors.append(f"Unknown column: '{header}' (not in mapping)")

    # Check for attempts to edit read-only fields
    for header in csv_headers:
        if header in mapping:
            field_info = mapping[header]
            survey_master = field_info.get("survey_master", "").strip().lower()
            if survey_master == "read":
                # Read-only field - warn but don't error (user might be editing, but apply will reject)
                pass

    # Validate field values against schema (data types, allowed values, etc.)
    validation_errors = validate_all_changes(csv_rows, mapping)
    if validation_errors:
        errors.append(format_validation_errors(validation_errors))

    return errors


def compute_diff(survey_id: str, csv_row: Dict[str, str]) -> Dict[str, Any]:
    """Compute diff between CSV row (desired) and snapshot (baseline).

    Returns dict with:
    - survey_id
    - survey_name
    - changes: list of {field, old_value, new_value, endpoint, is_dangerous}
    - publish_required: bool (True if metadata/options fields changed)
    - has_dangerous_changes: bool (True if any dangerous field changed)
    """
    snapshot = load_snapshot(survey_id)
    if not snapshot:
        return {
            "survey_id": survey_id,
            "survey_name": "UNKNOWN",
            "changes": [],
            "publish_required": False,
            "has_dangerous_changes": False,
            "error": f"No snapshot found for {survey_id}",
        }

    mapping = _parse_mapping_csv()
    survey_name = snapshot.get("survey_name", survey_id)

    # Dangerous fields that require explicit override
    dangerous_fields = {
        "isActive",  # Activation status
        "EOSRedirectURL",  # End-of-survey redirect
        "BallotBoxStuffingPreventionURL",  # Ballot box stuffing prevention URL
        "RefererURL",  # Referrer URL
        "PasswordProtection",  # Password protection toggle
        "SurveyStatus",  # Survey status (Active/Inactive)
    }

    changes = []
    publish_required = False
    has_dangerous_changes = False

    for field_name in csv_row.keys():
        # Preserve exact CSV cell contents (leading/trailing spaces can be meaningful,
        # e.g., button labels); do not strip.
        csv_value = csv_row.get(field_name, "")

        # Get field info
        if field_name not in mapping:
            continue

        field_info = mapping[field_name]
        survey_master = field_info.get("survey_master", "").strip().lower()

        # Skip read-only fields
        if survey_master != "write":
            continue

        # Get snapshot baseline value
        endpoint = _derive_endpoint(field_info)
        baseline_value = _extract_value_from_snapshot(snapshot, field_info)

        # Normalize baseline for comparison
        baseline_str = _scalar_to_string(baseline_value)

        # Compare
        if csv_value != baseline_str:
            is_dangerous = field_name in dangerous_fields
            changes.append(
                {
                    "field": field_name,
                    "old_value": baseline_str,
                    "new_value": csv_value,
                    "endpoint": endpoint or "unknown",
                    "is_dangerous": is_dangerous,
                }
            )

            if endpoint in ("metadata", "options"):
                publish_required = True
            if is_dangerous:
                has_dangerous_changes = True

    return {
        "survey_id": survey_id,
        "survey_name": survey_name,
        "changes": changes,
        "publish_required": publish_required,
        "has_dangerous_changes": has_dangerous_changes,
        "error": None,
    }


def _format_preview_output(diff: Dict[str, Any], use_color: bool = True) -> str:
    """Format diff for display with grouping by endpoint.

    Returns formatted string showing changes grouped by endpoint type.

    Args:
        diff: Diff dictionary with changes
        use_color: If True, apply ANSI color codes

    Returns:
        Formatted string for display
    """
    from .terminal_colors import colored, Colors, diff_colored

    if diff.get("error"):
        error_msg = diff["error"]
        if use_color:
            error_msg = colored(error_msg, Colors.RED)
        return f"  ❌ ERROR: {error_msg}"

    if not diff.get("changes"):
        return "  ✓ No changes"

    lines = []

    # Group changes by endpoint
    changes_by_endpoint = {}
    for change in diff["changes"]:
        endpoint = change.get("endpoint", "unknown")
        if endpoint not in changes_by_endpoint:
            changes_by_endpoint[endpoint] = []
        changes_by_endpoint[endpoint].append(change)

    # Show metadata changes
    if "metadata" in changes_by_endpoint:
        header = "    📝 Metadata changes (will trigger publish):"
        if use_color:
            header = colored(header, Colors.CYAN, bold=True)
        lines.append(header)
        for change in changes_by_endpoint["metadata"]:
            field = change.get("field_name", "unknown")
            old_val = change.get("old_value", "")
            new_val = change.get("new_value", "")

            # Format values with color
            if use_color:
                old_val_fmt, new_val_fmt = diff_colored(old_val, new_val, max_width=60)
            else:
                old_val_fmt = old_val[:60] + ("..." if len(old_val) > 60 else "")
                new_val_fmt = new_val[:60] + ("..." if len(new_val) > 60 else "")

            is_dangerous = change.get("dangerous")
            prefix = "⚠️ " if is_dangerous else "✏️ "
            if use_color and is_dangerous:
                prefix = colored(prefix, Colors.RED, bold=True)

            lines.append(f"      {prefix}{field}: '{old_val_fmt}' → '{new_val_fmt}'")

    # Show options changes
    if "options" in changes_by_endpoint:
        header = "    ⚙️  Options changes (will trigger publish):"
        if use_color:
            header = colored(header, Colors.BLUE, bold=True)
        lines.append(header)
        for change in changes_by_endpoint["options"]:
            field = change.get("field_name", "unknown")
            old_val = change.get("old_value", "")
            new_val = change.get("new_value", "")

            # Format values with color
            if use_color:
                old_val_fmt, new_val_fmt = diff_colored(old_val, new_val, max_width=50)
            else:
                old_val_fmt = old_val[:50] + ("..." if len(old_val) > 50 else "")
                new_val_fmt = new_val[:50] + ("..." if len(new_val) > 50 else "")

            is_dangerous = change.get("dangerous")
            prefix = "⚠️ " if is_dangerous else "✏️ "
            if use_color and is_dangerous:
                prefix = colored(prefix, Colors.RED, bold=True)

            lines.append(f"      {prefix}{field}: '{old_val_fmt}' → '{new_val_fmt}'")

    # Show status changes
    if "status" in changes_by_endpoint:
        header = "    🔄 Status changes (no publish):"
        if use_color:
            header = colored(header, Colors.YELLOW, bold=True)
        lines.append(header)
        for change in changes_by_endpoint["status"]:
            field = change.get("field_name", "unknown")
            old_val = change.get("old_value", "")
            new_val = change.get("new_value", "")

            is_dangerous = change.get("dangerous")
            prefix = "⚠️ " if is_dangerous else "✏️ "
            if use_color and is_dangerous:
                prefix = colored(prefix, Colors.RED, bold=True)

            lines.append(f"      {prefix}{field}: {old_val} → {new_val}")

    return "\n".join(lines)


def preview_master(
    csv_headers: Optional[List[str]] = None,
    csv_rows: Optional[List[Dict[str, str]]] = None,
    verbose: bool = False,
    survey_id: Optional[str] = None,
) -> Dict[str, Any]:
    """Preview changes that would be applied to Qualtrics.

    If csv_headers/rows are None, loads from disk.

    Args:
        csv_headers: CSV header row (optional)
        csv_rows: CSV data rows (optional)
        verbose: If True, print status messages
        survey_id: If provided, only preview this survey (None = all surveys)

    Returns dict with:
    - validation_errors: list of error messages
    - survey_diffs: list of diff dicts (one per survey)
    - summary: {total_surveys, surveys_with_changes, total_changes, requires_publish, has_dangerous}
    """
    # Load CSV if not provided
    if csv_headers is None or csv_rows is None:
        csv_headers, csv_rows = load_master_csv()

    # Filter to specific survey if requested
    if survey_id:
        csv_rows = [r for r in csv_rows if r.get("SurveyID", "").strip() == survey_id]

    if verbose:
        print(f"[qsync:master-preview] Validating {len(csv_rows)} surveys...")

    # Validate CSV
    validation_errors = validate_master_csv(csv_headers, csv_rows)
    if validation_errors:
        if verbose:
            for err in validation_errors:
                print(f"[qsync:master-preview] VALIDATION ERROR: {err}")
        return {
            "validation_errors": validation_errors,
            "survey_diffs": [],
            "summary": None,
        }

    if verbose:
        print(f"[qsync:master-preview] Computing diffs for {len(csv_rows)} surveys...")

    # Compute diffs for each survey
    survey_diffs = []
    total_changes = 0
    surveys_with_changes = 0
    requires_publish = False
    has_dangerous = False

    show_progress = not verbose and len(csv_rows) > 1 and should_use_rich()
    if show_progress:
        total = len(csv_rows)
        with progress_context("Previewing Survey Master changes", total=total) as prog:
            for idx, row in enumerate(csv_rows, start=1):
                if prog:
                    progress, task_id = prog
                    progress.update(
                        task_id,
                        description=f"Previewing Survey Master changes ({idx}/{total})",
                    )

                survey_id_from_row = row.get("SurveyID", "").strip()
                if not survey_id_from_row:
                    if prog:
                        progress, task_id = prog
                        progress.advance(task_id)
                    continue

                diff = compute_diff(survey_id_from_row, row)
                survey_diffs.append(diff)

                if diff.get("error"):
                    pass
                elif diff["changes"]:
                    surveys_with_changes += 1
                    total_changes += len(diff["changes"])
                    if diff["publish_required"]:
                        requires_publish = True
                    if diff["has_dangerous_changes"]:
                        has_dangerous = True

                if prog:
                    progress, task_id = prog
                    progress.advance(task_id)
    else:
        for row in csv_rows:
            survey_id_from_row = row.get("SurveyID", "").strip()
            if not survey_id_from_row:
                continue

            diff = compute_diff(survey_id_from_row, row)
            survey_diffs.append(diff)

            if diff.get("error"):
                if verbose:
                    print(
                        f"[qsync:master-preview]   {survey_id_from_row}: ERROR - {diff['error']}"
                    )
            elif diff["changes"]:
                surveys_with_changes += 1
                total_changes += len(diff["changes"])
                if diff["publish_required"]:
                    requires_publish = True
                if diff["has_dangerous_changes"]:
                    has_dangerous = True

                if verbose:
                    print(
                        f"[qsync:master-preview]   {survey_id_from_row}: {len(diff['changes'])} change(s)"
                    )

    summary = {
        "total_surveys": len([d for d in survey_diffs if not d.get("error")]),
        "surveys_with_changes": surveys_with_changes,
        "total_changes": total_changes,
        "requires_publish": requires_publish,
        "has_dangerous": has_dangerous,
    }

    if verbose:
        print(
            f"[qsync:master-preview] Summary: {surveys_with_changes}/{summary['total_surveys']} surveys have changes"
        )
        print(f"[qsync:master-preview] Total changes: {total_changes}")
        if requires_publish:
            print("[qsync:master-preview] ⚠️  Publishing required after push")
        if has_dangerous:
            print(
                "[qsync:master-preview] ⚠️  Some changes involve dangerous fields (requires --allow-dangerous)"
            )

    return {
        "validation_errors": [],
        "survey_diffs": survey_diffs,
        "summary": summary,
    }


def _compute_snapshot_hash(survey_ids: List[str]) -> str:
    """Compute SHA256 hash of snapshots for drift detection.

    Args:
        survey_ids: List of survey IDs to include in hash

    Returns:
        SHA256 hex digest of snapshot content
    """
    import hashlib
    import json

    snapshot_data = []
    for sid in sorted(survey_ids):
        snapshot = load_snapshot(sid)
        if snapshot:
            sections = snapshot.get("sections", {})
            snapshot_data.append({
                "survey_id": sid,
                "metadata": sections.get("metadata", {}),
                "options": sections.get("options", {}),
                "status": sections.get("status", {}),
            })

    json_str = json.dumps(snapshot_data, sort_keys=True)
    return hashlib.sha256(json_str.encode()).hexdigest()


def stage_master(
    csv_headers: Optional[List[str]] = None,
    csv_rows: Optional[List[Dict[str, str]]] = None,
    verbose: bool = False,
    survey_id: Optional[str] = None,
) -> Dict[str, Any]:
    """Stage changes from CSV to pending (no API writes).

    Args:
        csv_headers: CSV header row (optional)
        csv_rows: CSV data rows (optional)
        verbose: If True, print status messages
        survey_id: If provided, only stage this survey

    Returns dict with:
        - staged_surveys: int
        - total_changes: int
        - validation_errors: list
        - survey_diffs: list
    """
    from .pending_stage import (
        PendingStagedChanges,
        MasterPendingPayload,
        save_pending,
        clear_pending,
    )

    # Load CSV if not provided
    if csv_headers is None or csv_rows is None:
        csv_headers, csv_rows = load_master_csv()

    # Filter to specific survey if requested
    if survey_id:
        csv_rows = [r for r in csv_rows if r.get("SurveyID", "").strip() == survey_id]

    if verbose:
        print(f"[qsync:master-stage] Validating {len(csv_rows)} surveys...")

    # Validate CSV
    validation_errors = validate_master_csv(csv_headers, csv_rows)
    if validation_errors:
        if verbose:
            for err in validation_errors:
                print(f"[qsync:master-stage] VALIDATION ERROR: {err}")
        return {
            "staged_surveys": 0,
            "total_changes": 0,
            "validation_errors": validation_errors,
            "survey_diffs": [],
        }

    if verbose:
        print(f"[qsync:master-stage] Computing diffs for {len(csv_rows)} surveys...")

    # Compute diffs for each survey
    survey_diffs = []
    surveys_with_changes = []
    total_changes = 0

    for row in csv_rows:
        sid = row.get("SurveyID", "").strip()
        if not sid:
            continue

        diff = compute_diff(sid, row)

        # Skip surveys with errors or no changes
        if diff.get("error") or not diff.get("changes"):
            if verbose and diff.get("error"):
                print(f"[qsync:master-stage]   {sid}: ERROR - {diff['error']}")
            continue

        survey_diffs.append(diff)
        surveys_with_changes.append(sid)
        total_changes += len(diff["changes"])

        if verbose:
            print(f"[qsync:master-stage]   {sid}: {len(diff['changes'])} change(s)")

    # No changes: clear stale pending
    if not survey_diffs:
        for row in csv_rows:
            sid = row.get("SurveyID", "").strip()
            if sid:
                clear_pending(sid, "master")

        if verbose:
            print("[qsync:master-stage] No changes to stage")

        return {
            "staged_surveys": 0,
            "total_changes": 0,
            "validation_errors": [],
            "survey_diffs": [],
        }

    # Compute per-survey snapshot hashes for drift detection.
    # Each pending record is survey-scoped, so hash must also be survey-scoped.
    snapshot_hashes = {
        sid: _compute_snapshot_hash([sid]) for sid in sorted(set(surveys_with_changes))
    }

    if verbose:
        print(f"[qsync:master-stage] Saving pending records for {len(surveys_with_changes)} survey(s)...")

    # Save pending (one per survey)
    for diff in survey_diffs:
        sid = diff["survey_id"]
        payload = MasterPendingPayload(
            survey_ids=[sid],
            snapshot_hash=snapshot_hashes.get(sid, ""),
            changes=[diff],
        )
        record = PendingStagedChanges(
            survey_id=sid,
            dimension="master",
            payload=payload,
            schema_version=1,
        )
        save_pending(record)

        if verbose:
            print(f"[qsync:master-stage]   Staged: surveys/pending/master/{sid}.json")

    if verbose:
        print(f"[qsync:master-stage] ✅ Staged {len(surveys_with_changes)} survey(s) with {total_changes} change(s)")

    return {
        "staged_surveys": len(surveys_with_changes),
        "total_changes": total_changes,
        "validation_errors": [],
        "survey_diffs": survey_diffs,
    }


def pull_master(
    survey_ids: Optional[List[str]] = None,
    verbose: bool = False,
    force_overwrite: bool = False,
) -> Tuple[int, Path]:
    """Pull survey master snapshots and generate/merge master CSV.

    By default, preserves user edits by merging overrides from existing CSV.
    Use --force-overwrite to discard local edits and generate fresh CSV.

    Args:
        survey_ids: Specific survey IDs to pull. If None, pulls all focal surveys.
        verbose: If True, print status messages.
        force_overwrite: If True, skip merge logic and overwrite existing CSV.

    Returns:
        (snapshots_created, csv_path) tuple
    """
    base_url, headers = get_client_config()

    # Determine which surveys to pull
    if survey_ids is None:
        focal_snapshot = load_focal_snapshot()
        survey_ids = [sid for sid, is_focal in focal_snapshot.items() if is_focal]

    if verbose:
        print(f"[qsync:master-pull] Pulling {len(survey_ids)} focal surveys...")

    snapshots_created = 0
    fetch_errors: list[str] = []
    show_progress = not verbose and len(survey_ids) > 1 and should_use_rich()

    def _pull_single(survey_id: str) -> None:
        try:
            if verbose:
                from .survey_ref import format_survey_ref

                print(
                    f"[qsync:master-pull]   Fetching {format_survey_ref(survey_id)}..."
                )

            # Fetch all required endpoints
            survey_name = _fetch_survey_name(base_url, headers, survey_id)
            status_data, _ = _fetch_endpoint(base_url, headers, survey_id, "status")
            metadata_data, _ = _fetch_endpoint(base_url, headers, survey_id, "metadata")
            options_data, _ = _fetch_endpoint(base_url, headers, survey_id, "options")
            versions_data, _ = _fetch_endpoint(base_url, headers, survey_id, "versions")

            # Create and save snapshot
            snapshot = create_snapshot(
                survey_id,
                survey_name,
                status_data,
                metadata_data,
                options_data,
                versions_data,
            )
            save_snapshot(survey_id, snapshot)
            nonlocal snapshots_created
            snapshots_created += 1

            if verbose:
                print("[qsync:master-pull]     ✓ Snapshot saved")

        except Exception as e:
            from .survey_ref import format_survey_ref

            fetch_errors.append(
                f"[qsync:master-pull] ERROR fetching {format_survey_ref(survey_id)}: {e}",
            )
            # Continue with next survey

    # Fetch snapshots for each survey
    if show_progress:
        total = len(survey_ids)
        with progress_context("Pulling Survey Master snapshots", total=total) as prog:
            for idx, survey_id in enumerate(survey_ids, start=1):
                if prog:
                    progress, task_id = prog
                    progress.update(
                        task_id,
                        description=f"Pulling Survey Master snapshots ({idx}/{total})",
                    )
                _pull_single(survey_id)
                if prog:
                    progress, task_id = prog
                    progress.advance(task_id)
    else:
        for survey_id in survey_ids:
            _pull_single(survey_id)

    for msg in fetch_errors:
        print(msg, flush=True)

    # Generate master CSV from snapshots
    if verbose:
        print(
            f"[qsync:master-pull] Generating master CSV from {snapshots_created} snapshots..."
        )

    fresh_rows = generate_master_csv_from_snapshots(survey_ids)
    csv_path = _master_csv_path()
    existing_csv_exists = csv_path.exists()

    # MERGE LOGIC: Preserve user overrides by default
    if existing_csv_exists and not force_overwrite:
        try:
            if verbose:
                print("[qsync:master-pull] Merging with existing CSV to preserve user edits...")

            # Load existing CSV
            existing_headers, existing_rows = load_master_csv()
            existing_by_id = {row.get("SurveyID"): row for row in existing_rows if row.get("SurveyID")}

            # Index fresh CSV by SurveyID
            fresh_by_id = {}
            if fresh_rows and len(fresh_rows) > 1:
                fresh_headers = fresh_rows[0]
                for fresh_row_list in fresh_rows[1:]:
                    fresh_row = dict(zip(fresh_headers, fresh_row_list))
                    sid = fresh_row.get("SurveyID")
                    if sid:
                        fresh_by_id[sid] = fresh_row

            # Compute user overrides (fields that differ from baseline)
            user_overrides = {}
            for sid, existing_row in existing_by_id.items():
                if sid not in fresh_by_id:
                    continue

                fresh_row = fresh_by_id[sid]
                overrides = {}

                for field, existing_value in existing_row.items():
                    if field.startswith("_"):
                        # Skip read-only fields
                        continue

                    fresh_value = fresh_row.get(field, "")
                    if existing_value != fresh_value:
                        overrides[field] = existing_value

                if overrides:
                    user_overrides[sid] = overrides

            if user_overrides:
                if verbose:
                    print(f"[qsync:master-pull] Found overrides for {len(user_overrides)} survey(s)")

                # Apply overrides to fresh CSV
                merged_rows = [fresh_rows[0]]  # Headers
                for fresh_row_list in fresh_rows[1:]:
                    fresh_row = dict(zip(fresh_rows[0], fresh_row_list))
                    sid = fresh_row.get("SurveyID")

                    if sid in user_overrides:
                        for field, override_value in user_overrides[sid].items():
                            fresh_row[field] = override_value

                    merged_row_list = [fresh_row.get(col, "") for col in fresh_rows[0]]
                    merged_rows.append(merged_row_list)

                fresh_rows = merged_rows

            # Backup existing CSV before writing
            import shutil
            backup_path = csv_path.with_suffix(".csv.bak")
            shutil.copy2(csv_path, backup_path)
            if verbose:
                print(f"[qsync:master-pull] Backed up existing CSV to {backup_path}")

        except Exception as exc:
            if verbose:
                print(f"[qsync:master-pull] ⚠️  Merge failed: {exc}. Using fresh CSV.")
            # On merge failure, fall through to write fresh CSV
    elif existing_csv_exists and force_overwrite:
        if verbose:
            print("[qsync:master-pull] --force-overwrite: Discarding existing CSV")

    # Write final CSV + workbook surfaces
    csv_path = write_master_csv(fresh_rows)
    workbook_path = _master_workbook_path()

    if verbose:
        print(f"[qsync:master-pull] ✓ Master CSV written to {csv_path}")
        print(f"[qsync:master-pull] ✓ Master workbook written to {workbook_path}")
        print(
            f"[qsync:master-pull] Complete: {snapshots_created} snapshots, 1 CSV + 1 workbook ({len(fresh_rows)-1} rows)"
        )

    return snapshots_created, csv_path


# ===== MVP Stage 4: Apply (Write Back Safely) =====


def _get_dangerous_fields() -> set:
    """Return set of field names that require --allow-dangerous to change."""
    return {
        "isActive",  # Activation status
        "EOSRedirectURL",  # End-of-survey redirect
        "BallotBoxStuffingPreventionURL",  # Ballot box stuffing prevention URL
        "RefererURL",  # Referrer URL
        "PasswordProtection",  # Password protection toggle
        "SurveyStatus",  # Survey status (Active/Inactive)
    }


def detect_drift(survey_id: str, csv_row: Dict[str, str]) -> Dict[str, Any]:
    """Detect drift between snapshot baseline and current live values.

    For all fields that differ between CSV and snapshot, fetch current live
    values from Qualtrics and compare to snapshot baseline.

    Returns dict with:
    - survey_id
    - drifted_fields: list of {field, baseline, current} where baseline != current
    - schema_version_matches: bool
    - schema_mismatch_warning: str or None
    """
    snapshot = load_snapshot(survey_id)
    if not snapshot:
        return {
            "survey_id": survey_id,
            "drifted_fields": [],
            "schema_version_matches": False,
            "schema_mismatch_warning": f"No snapshot found for {survey_id}; cannot detect drift",
        }

    # Check schema version (compare stable hash portion; legacy snapshots had a date prefix)
    current_schema = _compute_schema_version()
    snapshot_schema = snapshot.get("schema_version", "") or ""
    schema_matches = _schema_version_hash(current_schema) == _schema_version_hash(
        snapshot_schema
    )

    mapping = _parse_mapping_csv()
    base_url, headers = get_client_config()
    drifted_fields: list[dict] = []

    # Determine which fields we intend to change (csv != snapshot baseline),
    # grouped by endpoint, so we fetch each endpoint at most once.
    fields_by_endpoint: Dict[str, List[Tuple[str, str]]] = {}

    for field_name in csv_row.keys():
        csv_value = csv_row.get(field_name, "")
        field_info = mapping.get(field_name)
        if not field_info:
            continue

        survey_master = field_info.get("survey_master", "").strip().lower()
        if survey_master != "write":
            continue

        endpoint = _derive_endpoint(field_info) or "unknown"
        baseline = _extract_value_from_snapshot(snapshot, field_info)
        baseline_str = _scalar_to_string(baseline)

        if csv_value != baseline_str:
            fields_by_endpoint.setdefault(endpoint, []).append((field_name, baseline_str))

    # Fetch each endpoint once and evaluate drift for all relevant fields.
    live_payload_by_endpoint: Dict[str, Optional[dict]] = {}
    for endpoint in sorted(fields_by_endpoint.keys()):
        if endpoint not in {"metadata", "options", "status"}:
            live_payload_by_endpoint[endpoint] = None
            continue
        try:
            payload, _ = _fetch_endpoint(base_url, headers, survey_id, endpoint)
            live_payload_by_endpoint[endpoint] = payload
        except Exception:
            live_payload_by_endpoint[endpoint] = None

    for endpoint, items in fields_by_endpoint.items():
        payload = live_payload_by_endpoint.get(endpoint)
        for field_name, baseline_str in items:
            field_info = mapping.get(field_name)
            if not field_info:
                continue

            if payload is None:
                drifted_fields.append(
                    {
                        "field": field_name,
                        "baseline": baseline_str,
                        "current": "UNKNOWN",
                        "endpoint": endpoint,
                    }
                )
                continue

            current = _extract_value_from_live(payload, field_info, endpoint=endpoint)
            current_str = _scalar_to_string(current)
            if baseline_str != current_str:
                drifted_fields.append(
                    {
                        "field": field_name,
                        "baseline": baseline_str,
                        "current": current_str,
                        "endpoint": endpoint,
                    }
                )

    schema_mismatch_warning = None
    if not schema_matches:
        schema_mismatch_warning = (
            f"Schema version mismatch: snapshot={snapshot_schema}, current={current_schema}. "
            "Consider running 'qsync survey master pull' to refresh snapshots."
        )

    return {
        "survey_id": survey_id,
        "drifted_fields": drifted_fields,
        "schema_version_matches": schema_matches,
        "schema_mismatch_warning": schema_mismatch_warning,
    }


def _write_metadata(
    base_url: str, headers: Dict[str, str], survey_id: str, changes: Dict[str, str]
) -> bool:
    """Write metadata fields using PATCH-like semantics (send only changed keys).

    Never sends null for date fields; omits them if not provided.

    Returns True if successful, False otherwise.
    """
    if not changes:
        return True

    # Build payload with only changed keys
    payload = {}
    mapping = _parse_mapping_csv()

    for field_name, csv_value in changes.items():
        if field_name not in mapping:
            continue

        field_info = mapping[field_name]
        endpoint = _derive_endpoint(field_info)

        # Only process metadata fields
        if endpoint != "metadata":
            continue

        object_path = (field_info.get("object_path") or "").strip()
        key = object_path.split(".")[-1] if object_path else field_name

        # Skip empty values for date fields (don't send null)
        if not csv_value or csv_value.strip() == "":
            # Check if this is a date field by convention (name contains Date)
            if "Date" in key or "date" in key:
                continue  # Omit null dates
            # For non-date fields, include empty string or skip based on field type
            csv_value = ""

        # Convert string representation back to proper type
        if csv_value.lower() in ("true", "false"):
            payload[key] = csv_value.lower() == "true"
        else:
            payload[key] = csv_value

    if not payload:
        return True

    try:
        resp = send_api_request(
            action="qsync.master.write.metadata",
            method="PUT",
            base_url=base_url,
            headers=headers,
            path=f"survey-definitions/{survey_id}/metadata",
            survey_id=survey_id,
            log_meta={
                "operation": "master_apply",
                "endpoints_written": ["/survey-definitions/{id}/metadata"],
                "changed_fields": sorted(payload.keys()),
            },
            json=payload,
            timeout=30,
        )
        return resp.status_code in (200, 204)
    except Exception:
        return False


def _write_options(
    base_url: str, headers: Dict[str, str], survey_id: str, changes: Dict[str, str]
) -> bool:
    """Write options using GET→merge→PUT semantics.

    Fetches current options, merges changes, and PUTs the full object back.
    This prevents accidentally dropping unrelated option keys.

    Returns True if successful, False otherwise.
    """
    if not changes:
        return True

    mapping = _parse_mapping_csv()

    # Fetch current options
    try:
        current_data, _ = _fetch_endpoint(base_url, headers, survey_id, "options")
    except Exception:
        return False

    # Merge changes into current
    merged = dict(current_data)
    for field_name, csv_value in changes.items():
        if field_name not in mapping:
            continue

        field_info = mapping[field_name]
        endpoint = _derive_endpoint(field_info)

        # Only process options fields
        if endpoint != "options":
            continue

        object_path = (field_info.get("object_path") or "").strip()
        parts = object_path.split(".")
        if parts and parts[0] == "result":
            parts = parts[1:]
        if not parts:
            continue

        current = merged
        for part in parts[:-1]:
            if part not in current or not isinstance(current.get(part), dict):
                current[part] = {}
            current = current[part]
        final_key = parts[-1]
        data_type = str(field_info.get("data_type") or "").strip().lower()
        if data_type == "object":
            raw_json = csv_value.strip() if isinstance(csv_value, str) else str(csv_value).strip()
            if not raw_json:
                current[final_key] = {}
                continue
            try:
                decoded = json.loads(raw_json)
            except Exception as exc:
                raise ValueError(
                    f"{field_name}: invalid JSON object text ({exc})"
                ) from exc
            if not isinstance(decoded, dict):
                raise ValueError(
                    f"{field_name}: expected JSON object (decoded type={type(decoded).__name__})"
                )
            current[final_key] = decoded
        elif isinstance(csv_value, str) and csv_value.lower() in ("true", "false"):
            current[final_key] = csv_value.lower() == "true"
        else:
            current[final_key] = csv_value

    # PUT merged object back
    try:
        resp = send_api_request(
            action="qsync.master.write.options",
            method="PUT",
            base_url=base_url,
            headers=headers,
            path=f"survey-definitions/{survey_id}/options",
            survey_id=survey_id,
            log_meta={
                "operation": "master_apply",
                "endpoints_written": ["/survey-definitions/{id}/options"],
                "changed_fields": sorted(changes.keys()),
            },
            json=merged,
            timeout=30,
        )
        return resp.status_code in (200, 204)
    except Exception as exc:
        print(f"[qsync:master-write] options payload error: {exc}")
        return False


def _write_status(
    base_url: str, headers: Dict[str, str], survey_id: str, changes: Dict[str, str]
) -> bool:
    """Write status fields (e.g., isActive) using PATCH-like semantics.

    Sends only the changed status keys.

    Returns True if successful, False otherwise.
    """
    if not changes:
        return True

    mapping = _parse_mapping_csv()
    payload = {}

    for field_name, csv_value in changes.items():
        if field_name not in mapping:
            continue

        field_info = mapping[field_name]
        endpoint = _derive_endpoint(field_info)

        # Only process status fields
        if endpoint != "status":
            continue

        object_path = (field_info.get("object_path") or "").strip()
        parts = object_path.split(".")
        if parts and parts[0] == "result":
            parts = parts[1:]
        if not parts:
            continue

        current = payload
        for part in parts[:-1]:
            if part not in current or not isinstance(current.get(part), dict):
                current[part] = {}
            current = current[part]
        final_key = parts[-1]
        if isinstance(csv_value, str) and csv_value.lower() in ("true", "false"):
            current[final_key] = csv_value.lower() == "true"
        else:
            current[final_key] = csv_value

    if not payload:
        return True

    try:
        resp = send_api_request(
            action="qsync.master.write.status",
            method="PUT",
            base_url=base_url,
            headers=headers,
            path=f"surveys/{survey_id}",
            survey_id=survey_id,
            log_meta={
                "operation": "master_apply",
                "endpoints_written": ["/surveys/{id}"],
                "changed_fields": sorted(changes.keys()),
            },
            json=payload,
            timeout=30,
        )
        return resp.status_code in (200, 204)
    except Exception:
        return False


def _write_audit_log(survey_id: str, applied_changes: Dict[str, Any]) -> None:
    """Write audit log entry to JSONL file for applied changes.

    Log format matches the shared write log schema from qsync_api_fixes.md.
    """
    try:
        from .api_push import get_write_log_path

        log_entry = {
            "timestamp": datetime.now(timezone.utc).isoformat() + "Z",
            "action": "qsync.master.apply",
            "survey_id": survey_id,
            "changes": applied_changes,
        }

        log_path = get_write_log_path()
        log_path.parent.mkdir(parents=True, exist_ok=True)

        with log_path.open("a", encoding="utf-8") as fh:
            fh.write(json.dumps(log_entry) + "\n")
    except Exception:
        pass  # Log write failures don't block the main operation


def _write_rollback_audit_log(
    survey_id: str,
    snapshot_path: Path,
    restored_changes: List[Dict[str, Any]],
    *,
    dry_run: bool,
    published: bool,
) -> None:
    """Write an audit entry for survey-master rollback operations."""
    try:
        from .api_push import get_write_log_path

        log_entry = {
            "timestamp": datetime.now(timezone.utc).isoformat() + "Z",
            "action": "qsync.master.rollback",
            "survey_id": survey_id,
            "snapshot_path": str(snapshot_path),
            "dry_run": bool(dry_run),
            "published": bool(published),
            "restored_changes": restored_changes,
        }
        log_path = get_write_log_path()
        log_path.parent.mkdir(parents=True, exist_ok=True)
        with log_path.open("a", encoding="utf-8") as fh:
            fh.write(json.dumps(log_entry) + "\n")
    except Exception:
        pass


def rollback_master(
    survey_id: str,
    *,
    version: int = 1,
    dry_run: bool = False,
    force: bool = False,
    allow_dangerous: bool = False,
    publish: bool = True,
    publish_description: Optional[str] = None,
    verbose: bool = False,
) -> Dict[str, Any]:
    """Rollback one survey to a pre-apply snapshot captured by survey master."""
    from .qualtrics_client import publish_survey_definition

    result: Dict[str, Any] = {
        "survey_id": survey_id,
        "version": version,
        "snapshot_path": None,
        "dry_run": dry_run,
        "applied": False,
        "published": False,
        "changes": [],
        "drifted_fields": [],
        "error": None,
    }

    try:
        snapshot_path, snapshot = load_rollback_snapshot(survey_id, version=version)
    except Exception as exc:
        result["error"] = str(exc)
        return result

    result["snapshot_path"] = str(snapshot_path)
    rollback_meta = snapshot.get("rollback", {})
    applied_changes = rollback_meta.get("applied_changes", [])
    if not applied_changes:
        result["error"] = (
            f"Rollback snapshot {snapshot_path.name} has no applied_changes metadata"
        )
        return result

    mapping = _parse_mapping_csv()
    dangerous_fields = _get_dangerous_fields()
    base_url, headers = get_client_config()

    # Ensure survey is still accessible before attempting rollback.
    try:
        _fetch_survey_name(base_url, headers, survey_id)
    except Exception as exc:
        result["error"] = f"Unable to access survey {survey_id}: {exc}"
        return result

    target_fields: List[str] = []
    endpoint_by_field: Dict[str, str] = {}
    target_by_field: Dict[str, str] = {}
    expected_post_apply_by_field: Dict[str, str] = {}
    for change in applied_changes:
        field_name = str(change.get("field") or "").strip()
        if not field_name:
            continue
        field_info = mapping.get(field_name)
        if not field_info:
            continue
        if field_info.get("survey_master", "").strip().lower() != "write":
            continue
        endpoint = _derive_endpoint(field_info)
        if endpoint not in ("metadata", "options", "status"):
            continue
        target_fields.append(field_name)
        endpoint_by_field[field_name] = endpoint
        target_by_field[field_name] = _scalar_to_string(change.get("pre_apply_value"))
        expected_post_apply_by_field[field_name] = _scalar_to_string(
            change.get("target_value")
        )

    if not target_fields:
        result["error"] = "No writable fields were found in rollback snapshot metadata"
        return result

    # Fetch each endpoint once for current-live comparisons.
    live_endpoint_payloads: Dict[str, dict] = {}
    for endpoint in sorted(set(endpoint_by_field.values())):
        try:
            data, _ = _fetch_endpoint(base_url, headers, survey_id, endpoint)
            live_endpoint_payloads[endpoint] = data
        except Exception as exc:
            result["error"] = f"Failed to fetch live {endpoint} payload: {exc}"
            return result

    current_by_field: Dict[str, str] = {}
    for field_name in target_fields:
        endpoint = endpoint_by_field[field_name]
        live_payload = live_endpoint_payloads.get(endpoint, {})
        current_value = _extract_value_from_live(
            live_payload, mapping[field_name], endpoint=endpoint
        )
        current_by_field[field_name] = _scalar_to_string(current_value)

    drifted_fields: List[Dict[str, Any]] = []
    for field_name in target_fields:
        expected = expected_post_apply_by_field.get(field_name, "")
        current = current_by_field.get(field_name, "")
        if current != expected:
            drifted_fields.append(
                {
                    "field": field_name,
                    "expected_post_apply": expected,
                    "current": current,
                    "endpoint": endpoint_by_field[field_name],
                }
            )
    result["drifted_fields"] = drifted_fields
    if drifted_fields and not force:
        fields = ", ".join(item["field"] for item in drifted_fields)
        result["error"] = (
            f"Drift detected before rollback ({fields}). Re-run with --force after review."
        )
        return result

    changes_by_endpoint: Dict[str, Dict[str, str]] = {
        "metadata": {},
        "options": {},
        "status": {},
    }
    rollback_changes: List[Dict[str, Any]] = []
    has_dangerous = False
    for field_name in target_fields:
        current = current_by_field.get(field_name, "")
        target = target_by_field.get(field_name, "")
        if current == target:
            continue
        endpoint = endpoint_by_field[field_name]
        changes_by_endpoint[endpoint][field_name] = target
        is_dangerous = field_name in dangerous_fields
        has_dangerous = has_dangerous or is_dangerous
        rollback_changes.append(
            {
                "field": field_name,
                "endpoint": endpoint,
                "old_value": current,
                "new_value": target,
                "is_dangerous": is_dangerous,
            }
        )

    result["changes"] = rollback_changes
    if has_dangerous and not allow_dangerous:
        result["error"] = "Dangerous changes require --allow-dangerous"
        return result

    if not rollback_changes:
        result["applied"] = True
        return result

    if dry_run:
        _write_rollback_audit_log(
            survey_id,
            snapshot_path,
            rollback_changes,
            dry_run=True,
            published=False,
        )
        result["applied"] = True
        return result

    metadata_ok = _write_metadata(
        base_url, headers, survey_id, changes_by_endpoint.get("metadata", {})
    )
    options_ok = _write_options(
        base_url, headers, survey_id, changes_by_endpoint.get("options", {})
    )
    status_ok = _write_status(
        base_url, headers, survey_id, changes_by_endpoint.get("status", {})
    )

    if not (metadata_ok and options_ok and status_ok):
        result["error"] = "One or more endpoint writes failed during rollback"
        return result

    published = False
    if publish:
        description = (publish_description or "").strip() or "qsync master rollback"
        try:
            publish_survey_definition(survey_id, description=description)
            published = True
        except Exception as exc:
            result["error"] = f"Rollback applied, but publish failed: {exc}"
            return result

    _write_rollback_audit_log(
        survey_id,
        snapshot_path,
        rollback_changes,
        dry_run=False,
        published=published,
    )

    result["applied"] = True
    result["published"] = published
    if verbose:
        print(
            f"[qsync:master-rollback] Restored {len(rollback_changes)} field(s) for {survey_id}"
        )
    return result


def apply_master(
    allow_dangerous: bool = False,
    force: bool = False,
    verbose: bool = False,
    survey_id: Optional[str] = None,
    skip_drift: bool = False,
    dry_run: bool = False,
    csv_rows: Optional[List[Dict[str, str]]] = None,
) -> Dict[str, Any]:
    """Apply changes from master CSV to Qualtrics.

    Enforces:
    - Drift detection (refuses unless --force or --skip-drift)
    - Schema version mismatch warning (recommends re-pull)
    - Dangerous field policy (refuses unless --allow-dangerous)
    - Endpoint-specific write semantics
    - Publishing after definition changes

    Args:
        allow_dangerous: Allow changes to dangerous fields (isActive, URLs, etc.)
        force: Override drift detection
        verbose: Print status messages
        survey_id: If provided, only apply to this survey (None = all surveys)
        skip_drift: Skip drift detection (faster but riskier)
        dry_run: Preview changes without actually writing them
        csv_rows: If provided, use these CSV rows instead of loading from disk (for tag filtering)

    Returns dict with:
    - total_surveys: number of surveys processed
    - surveys_applied: number of surveys with successful applies
    - surveys_failed: number of surveys with failed applies
    - details: list of {survey_id, applied, reason}
    - errors: list of error messages
    - dry_run: bool (True if this was a dry run)
    """
    base_url: Optional[str] = None
    headers: Optional[Dict[str, str]] = None

    # Load and validate CSV
    if csv_rows is None:
        csv_headers, csv_rows = load_master_csv()
    else:
        # If csv_rows provided, load headers from disk
        _, all_rows = load_master_csv()
        csv_headers = all_rows[0].keys() if all_rows else []

    if not csv_headers or not csv_rows:
        return {
            "total_surveys": 0,
            "surveys_applied": 0,
            "surveys_failed": 0,
            "details": [],
            "errors": ["No master CSV found or it is empty"],
            "dry_run": dry_run,
        }

    # Filter to specific survey if requested
    if survey_id:
        csv_rows = [r for r in csv_rows if r.get("SurveyID", "").strip() == survey_id]

    validation_errors = validate_master_csv(csv_headers, csv_rows)
    if validation_errors:
        return {
            "total_surveys": 0,
            "surveys_applied": 0,
            "surveys_failed": 0,
            "details": [],
            "errors": validation_errors,
            "dry_run": dry_run,
        }

    if verbose:
        mode = "DRY RUN: " if dry_run else ""
        print(f"[qsync:master-apply] {mode}Processing {len(csv_rows)} surveys...")
        if skip_drift:
            print("[qsync:master-apply] ⚠️  Drift detection is DISABLED (--skip-drift)")

    details = []
    surveys_applied = 0
    surveys_failed = 0
    show_progress = not verbose and len(csv_rows) > 1 and should_use_rich()

    def _process_row(row: Dict[str, str]) -> None:
        nonlocal surveys_applied, surveys_failed, base_url, headers
        survey_id = row.get("SurveyID", "").strip()
        if not survey_id:
            return
        from .survey_ref import format_survey_ref

        survey_ref = format_survey_ref(
            survey_id,
            str(row.get("SurveyName") or row.get("name") or "").strip() or None,
        )

        # Compute diffs
        diff = compute_diff(survey_id, row)
        if diff.get("error") or not diff.get("changes"):
            details.append(
                {
                    "survey_id": survey_id,
                    "applied": False,
                    "reason": diff.get("error") or "No changes",
                }
            )
            return

        # Check dangerous fields
        changes = diff.get("changes", [])
        has_dangerous = any(c.get("is_dangerous") for c in changes)

        if has_dangerous and not allow_dangerous:
            details.append(
                {
                    "survey_id": survey_id,
                    "applied": False,
                    "reason": "Dangerous changes require --allow-dangerous flag",
                }
            )
            return

        # Check for drift (unless skipped)
        drift_info = None
        if not skip_drift:
            drift_info = detect_drift(survey_id, row)
            if drift_info.get("drifted_fields") and not force:
                drifted = ", ".join(
                    f["field"] for f in drift_info.get("drifted_fields", [])
                )
                details.append(
                    {
                        "survey_id": survey_id,
                        "applied": False,
                        "reason": f"Drift detected in fields: {drifted}",
                    }
                )
                return

            # Warn about schema mismatch
            if not drift_info.get("schema_version_matches") and verbose:
                warning = drift_info.get("schema_mismatch_warning")
                if warning:
                    print(f"[qsync:master-apply]   ⚠️  {warning}")

        rollback_snapshot_path = None
        if not dry_run:
            try:
                rollback_snapshot_path = capture_pre_apply_snapshot(survey_id, changes)
                if verbose and rollback_snapshot_path is not None:
                    print(
                        f"[qsync:master-apply]   ✓ Rollback snapshot saved: {rollback_snapshot_path.name}"
                    )
            except Exception as exc:
                surveys_failed += 1
                details.append(
                    {
                        "survey_id": survey_id,
                        "applied": False,
                        "reason": f"Failed to capture pre-apply rollback snapshot: {exc}",
                    }
                )
                if verbose:
                    print(
                        f"[qsync:master-apply]   ✗ {survey_ref} rollback snapshot failed: {exc}"
                    )
                return

        # Group changes by endpoint
        changes_by_endpoint = {}
        for change in changes:
            endpoint = change.get("endpoint")
            field = change.get("field")
            new_value = change.get("new_value")

            if endpoint not in changes_by_endpoint:
                changes_by_endpoint[endpoint] = {}
            changes_by_endpoint[endpoint][field] = new_value

        # Apply changes per endpoint
        try:
            if base_url is None or headers is None:
                base_url, headers = get_client_config()

            metadata_ok = True
            options_ok = True
            status_ok = True
            applied_changes = []

            if "metadata" in changes_by_endpoint:
                if verbose:
                    print("[qsync:master-apply]   Writing metadata...")
                if not dry_run:
                    metadata_ok = _write_metadata(
                        base_url, headers, survey_id, changes_by_endpoint["metadata"]
                    )
                else:
                    metadata_ok = True  # In dry-run, assume success
                if metadata_ok:
                    applied_changes.extend(
                        [
                            {"field": f, "new_value": v}
                            for f, v in changes_by_endpoint["metadata"].items()
                        ]
                    )
                    if verbose:
                        marker = "[DRY RUN]" if dry_run else "✓"
                        print(f"[qsync:master-apply]     {marker} Metadata written")

            if "options" in changes_by_endpoint:
                if verbose:
                    print("[qsync:master-apply]   Writing options...")
                if not dry_run:
                    options_ok = _write_options(
                        base_url, headers, survey_id, changes_by_endpoint["options"]
                    )
                else:
                    options_ok = True  # In dry-run, assume success
                if options_ok:
                    applied_changes.extend(
                        [
                            {"field": f, "new_value": v}
                            for f, v in changes_by_endpoint["options"].items()
                        ]
                    )
                    if verbose:
                        marker = "[DRY RUN]" if dry_run else "✓"
                        print(f"[qsync:master-apply]     {marker} Options written")

            if "status" in changes_by_endpoint:
                if verbose:
                    print("[qsync:master-apply]   Writing status...")
                if not dry_run:
                    status_ok = _write_status(
                        base_url, headers, survey_id, changes_by_endpoint["status"]
                    )
                else:
                    status_ok = True  # In dry-run, assume success
                if status_ok:
                    applied_changes.extend(
                        [
                            {"field": f, "new_value": v}
                            for f, v in changes_by_endpoint["status"].items()
                        ]
                    )
                    if verbose:
                        marker = "[DRY RUN]" if dry_run else "✓"
                        print(f"[qsync:master-apply]     {marker} Status written")

            # Check if all writes succeeded
            all_ok = metadata_ok and options_ok and status_ok
            if all_ok:
                # Write audit log (only in actual apply, not dry-run)
                if not dry_run:
                    _write_audit_log(survey_id, applied_changes)

                # Note: Publishing is now a separate step (qsync survey master push)
                # Changes are written to Qualtrics but not published until push is called

                surveys_applied += 1
                details.append(
                    {
                        "survey_id": survey_id,
                        "applied": True,
                        "reason": f"{len(applied_changes)} changes applied",
                    }
                )
            else:
                surveys_failed += 1
                details.append(
                    {
                        "survey_id": survey_id,
                        "applied": False,
                        "reason": "One or more endpoint writes failed",
                    }
                )
                if verbose:
                    print(f"[qsync:master-apply]   ✗ {survey_ref} failed")

        except Exception as e:
            surveys_failed += 1
            details.append(
                {
                    "survey_id": survey_id,
                    "applied": False,
                    "reason": str(e),
                }
            )
            if verbose:
                print(f"[qsync:master-apply]   ✗ {survey_ref} error: {e}")

    if show_progress:
        total = len(csv_rows)
        with progress_context("Applying Survey Master changes", total=total) as prog:
            for idx, row in enumerate(csv_rows, start=1):
                if prog:
                    progress, task_id = prog
                    progress.update(
                        task_id,
                        description=f"Applying Survey Master changes ({idx}/{total})",
                    )
                _process_row(row)
                if prog:
                    progress, task_id = prog
                    progress.advance(task_id)
    else:
        for row in csv_rows:
            _process_row(row)

    return {
        "total_surveys": len(csv_rows),
        "surveys_applied": surveys_applied,
        "surveys_failed": surveys_failed,
        "details": details,
        "errors": [],
        "dry_run": dry_run,
    }


def push_master(
    description: Optional[str] = None,
    verbose: bool = False,
    survey_id: Optional[str] = None,
    all_surveys: bool = False,
    no_publish: bool = False,
    force_live: bool = False,
    force_preview: bool = False,
    auto_yes: bool = False,
    allow_dangerous: bool = False,
    allow_locked: bool = False,
) -> Dict[str, Any]:
    """Push staged master changes to API (NEW BEHAVIOR).

    Flow: load pending → enforce safeguards → write API → publish → clear pending

    Args:
        description: Description for the published version (default: "qsync master push")
        verbose: Print status messages
        survey_id: If provided, only push this survey (None = all with pending)
        all_surveys: If True and survey_id is not provided, scan all pending master records
                     (including non-focal). Default is focal-only.
        no_publish: If True, skip publish step (API write only)
        force_live: Allow push even with live responses
        force_preview: Skip preview response warnings
        auto_yes: Skip confirmation prompts
        allow_dangerous: Allow changes to dangerous fields
        allow_locked: Allow push to locked surveys

    Returns dict with:
    - total_surveys: number of surveys processed
    - surveys_pushed: number of surveys with API writes applied
    - surveys_published: number of surveys published
    - surveys_failed: number of surveys that failed
    - details: list of {survey_id, pushed, published, reason}
    - errors: list of error messages
    """
    from .pending_stage import load_pending, clear_pending, MasterPendingPayload
    from .push_safeguards import enforce_push_safeguards, SafeguardConfig
    from .qualtrics_client import publish_survey_definition, ensure_backup
    from .survey_inventory import load_focal_snapshot

    if description is None:
        description = "qsync master push"

    # Find surveys with staged changes
    if survey_id:
        survey_ids_to_check = [survey_id]
    elif all_surveys:
        from .config import resolve_root, resolve_scoped_dir

        root = resolve_root(required=False) or Path.cwd()
        pending_dir = resolve_scoped_dir("surveys", root=root) / "pending" / "master"
        if pending_dir.exists():
            survey_ids_to_check = sorted(
                {p.stem for p in pending_dir.glob("*.json") if p.is_file()}
            )
        else:
            survey_ids_to_check = []
    else:
        focal_snapshot = load_focal_snapshot()
        survey_ids_to_check = [sid for sid, is_focal in focal_snapshot.items() if is_focal]

    # Filter to only surveys with pending records
    survey_ids = []
    for sid in survey_ids_to_check:
        pending = load_pending(sid, "master")
        if pending and isinstance(pending.payload, MasterPendingPayload):
            survey_ids.append(sid)

    if not survey_ids:
        return {
            "total_surveys": 0,
            "surveys_pushed": 0,
            "surveys_published": 0,
            "surveys_failed": 0,
            "details": [],
            "errors": ["No staged master changes found. Run 'qsync survey master stage' first."],
        }

    if verbose:
        print(f"[qsync:master-push] Pushing {len(survey_ids)} survey(s) with staged changes...")

    details = []
    surveys_pushed = 0
    surveys_published = 0
    surveys_failed = 0
    base_url = None
    headers = None
    show_progress = not verbose and len(survey_ids) > 1 and should_use_rich()

    def _push_single(sid: str) -> None:
        nonlocal surveys_pushed, surveys_published, surveys_failed, base_url, headers
        from .survey_ref import format_survey_ref

        pending = load_pending(sid, "master")
        if not pending or not isinstance(pending.payload, MasterPendingPayload):
            details.append(
                {
                    "survey_id": sid,
                    "pushed": False,
                    "published": False,
                    "reason": "No staged changes",
                }
            )
            return

        survey_diffs = pending.payload.changes
        if not survey_diffs or len(survey_diffs) != 1:
            details.append(
                {
                    "survey_id": sid,
                    "pushed": False,
                    "published": False,
                    "reason": "Invalid pending payload",
                }
            )
            surveys_failed += 1
            return

        diff = survey_diffs[0]
        changes = list(diff.get("changes") or [])
        if not changes:
            clear_pending(sid, "master")
            return

        has_dangerous = any(c.get("is_dangerous") for c in changes)
        if has_dangerous and not allow_dangerous:
            if verbose:
                print(
                    f"[qsync:master-push]   {format_survey_ref(sid)}: ✗ Dangerous changes require --allow-dangerous"
                )
            details.append(
                {
                    "survey_id": sid,
                    "pushed": False,
                    "published": False,
                    "reason": "Dangerous changes require --allow-dangerous",
                }
            )
            surveys_failed += 1
            return

        current_hash = _compute_snapshot_hash([sid])
        if current_hash != pending.payload.snapshot_hash:
            if verbose:
                print(
                    f"[qsync:master-push]   {format_survey_ref(sid)}: ✗ Drift detected (snapshots changed since staging)"
                )
            details.append(
                {
                    "survey_id": sid,
                    "pushed": False,
                    "published": False,
                    "reason": "Drift detected: snapshots changed since staging. Re-run 'qsync survey master pull' and 'qsync survey master stage'.",
                }
            )
            surveys_failed += 1
            return

        try:
            config = SafeguardConfig(
                survey_id=sid,
                dimension="master",
                force_live=force_live,
                force_preview=force_preview,
                allow_locked=allow_locked,
                auto_yes=auto_yes,
            )
            enforce_push_safeguards(config)
        except SystemExit as e:
            if verbose:
                print(
                    f"[qsync:master-push]   {format_survey_ref(sid)}: ✗ Blocked by safeguards"
                )
            details.append(
                {
                    "survey_id": sid,
                    "pushed": False,
                    "published": False,
                    "reason": f"Blocked by safeguards: {e}",
                }
            )
            surveys_failed += 1
            return

        try:
            ensure_backup(sid)
        except Exception as e:
            if verbose:
                print(f"[qsync:master-push]   {format_survey_ref(sid)}: ✗ Backup failed: {e}")
            details.append(
                {
                    "survey_id": sid,
                    "pushed": False,
                    "published": False,
                    "reason": f"Backup failed: {e}",
                }
            )
            surveys_failed += 1
            return

        try:
            capture_pre_apply_snapshot(sid, changes)
            if verbose:
                print(
                    f"[qsync:master-push]   {format_survey_ref(sid)}: Rollback snapshot saved"
                )
        except Exception as e:
            if verbose:
                print(
                    f"[qsync:master-push]   {format_survey_ref(sid)}: ✗ Rollback snapshot failed: {e}"
                )
            details.append(
                {
                    "survey_id": sid,
                    "pushed": False,
                    "published": False,
                    "reason": f"Rollback snapshot failed: {e}",
                }
            )
            surveys_failed += 1
            return

        changes_by_endpoint: Dict[str, Dict[str, str]] = {}
        for change in changes:
            endpoint = change.get("endpoint")
            field = change.get("field")
            new_value = change.get("new_value")
            if endpoint not in changes_by_endpoint:
                changes_by_endpoint[endpoint] = {}
            changes_by_endpoint[endpoint][field] = new_value

        try:
            if base_url is None or headers is None:
                base_url, headers = get_client_config()

            write_success = True
            if "metadata" in changes_by_endpoint:
                if verbose:
                    print(
                        f"[qsync:master-push]   {format_survey_ref(sid)}: Writing metadata..."
                    )
                if not _write_metadata(
                    base_url, headers, sid, changes_by_endpoint["metadata"]
                ):
                    write_success = False

            if "options" in changes_by_endpoint and write_success:
                if verbose:
                    print(
                        f"[qsync:master-push]   {format_survey_ref(sid)}: Writing options..."
                    )
                if not _write_options(
                    base_url, headers, sid, changes_by_endpoint["options"]
                ):
                    write_success = False

            if "status" in changes_by_endpoint and write_success:
                if verbose:
                    print(
                        f"[qsync:master-push]   {format_survey_ref(sid)}: Writing status..."
                    )
                if not _write_status(
                    base_url, headers, sid, changes_by_endpoint["status"]
                ):
                    write_success = False

            if not write_success:
                if verbose:
                    print(
                        f"[qsync:master-push]   {format_survey_ref(sid)}: ✗ API write failed"
                    )
                details.append(
                    {
                        "survey_id": sid,
                        "pushed": False,
                        "published": False,
                        "reason": "API write failed",
                    }
                )
                surveys_failed += 1
                return

            surveys_pushed += 1
            if verbose:
                print(
                    f"[qsync:master-push]   {format_survey_ref(sid)}: ✓ API write successful ({len(changes)} change(s))"
                )
        except Exception as e:
            if verbose:
                print(
                    f"[qsync:master-push]   {format_survey_ref(sid)}: ✗ API write error: {e}"
                )
            details.append(
                {
                    "survey_id": sid,
                    "pushed": False,
                    "published": False,
                    "reason": f"API write error: {e}",
                }
            )
            surveys_failed += 1
            return

        published = False
        if not no_publish and diff.get("publish_required"):
            try:
                if verbose:
                    print(f"[qsync:master-push]   {format_survey_ref(sid)}: Publishing...")
                publish_survey_definition(sid, description=description)
                surveys_published += 1
                published = True
                if verbose:
                    print(f"[qsync:master-push]   {format_survey_ref(sid)}: ✓ Published")
            except Exception as e:
                if verbose:
                    print(
                        f"[qsync:master-push]   {format_survey_ref(sid)}: ✗ Publish failed: {e}"
                    )
                details.append(
                    {
                        "survey_id": sid,
                        "pushed": True,
                        "published": False,
                        "reason": f"Pushed successfully but publish failed: {e}",
                    }
                )
                surveys_failed += 1
                return
        # Postcondition for qsync sync: refresh read-only caches from live Qualtrics.
        # This prevents cache drift (and repeated diffs) after a successful push.
        try:
            from .qualtrics_client import refresh_survey_cache

            # Refresh the generic survey cache used by other dimensions.
            refresh_survey_cache(sid)
            # Refresh the master snapshot baseline used for master diffs.
            refresh_snapshot_from_live(sid, base_url=base_url, headers=headers)
        except Exception as e:
            # Keep pending so the user can retry and we don't silently accept drift.
            details.append(
                {
                    "survey_id": sid,
                    "pushed": True,
                    "published": published,
                    "reason": f"Pushed successfully but cache refresh failed: {e}",
                }
            )
            surveys_failed += 1
            return

        clear_pending(sid, "master")
        details.append(
            {
                "survey_id": sid,
                "pushed": True,
                "published": published,
                "reason": f"Pushed {len(changes)} change(s)"
                + (" and published" if published else ""),
            }
        )
        if verbose:
            print(f"[qsync:master-push]   {format_survey_ref(sid)}: ✓ Complete")

    if show_progress:
        total = len(survey_ids)
        with progress_context("Pushing Survey Master changes", total=total) as prog:
            for idx, sid in enumerate(survey_ids, start=1):
                if prog:
                    progress, task_id = prog
                    progress.update(
                        task_id,
                        description=f"Pushing Survey Master changes ({idx}/{total})",
                    )
                _push_single(sid)
                if prog:
                    progress, task_id = prog
                    progress.advance(task_id)
    else:
        for sid in survey_ids:
            _push_single(sid)

    if verbose:
        print(f"[qsync:master-push] Summary: {surveys_pushed}/{len(survey_ids)} pushed, {surveys_published} published, {surveys_failed} failed")

    return {
        "total_surveys": len(survey_ids),
        "surveys_pushed": surveys_pushed,
        "surveys_published": surveys_published,
        "surveys_failed": surveys_failed,
        "details": details,
        "errors": [],
    }
