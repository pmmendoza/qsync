"""Excel workbook IO for qsync workbooks.

This module owns the schema and transformations for the per-survey Excel
workbooks used by `qsync items ...` to preview/stage/push edits back to
Qualtrics.

Key sheets:
- `Questions`: question text
- `Options`: MC choices and Matrix answer scales
- `Subitems`: Matrix rows/statements (and SBSMatrix statements)
- `SBS_Columns` / `SBS_ColumnAnswers`: SBSMatrix column headers + per-column answer
  labels (Qualtrics `AdditionalQuestions`)
- `Embedded_Data`: SurveyFlow embedded defaults

SBSMatrix note: Qualtrics encodes side-by-side matrices as
`QuestionType="SBS"` with `Selector="SBSMatrix"`. In JSON, statements live
under `Choices` but are edited via the `Subitems` sheet; SBS column headers and
per-column answer labels live under `AdditionalQuestions` and are edited via
the SBS sheets.
"""

from __future__ import annotations

from copy import copy
from dataclasses import dataclass
import json
import re
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

from .markdown_codec import (
    html_to_md,
    normalize_text,
    should_treat_as_html,
    md_to_html,
    is_markdown_safe_html,
)

QUESTION_SHEET = "Questions"
OPTIONS_SHEET = "Options"
SUBITEMS_SHEET = "Subitems"
SBS_COLUMNS_SHEET = "SBS_Columns"
SBS_COLUMN_ANSWERS_SHEET = "SBS_ColumnAnswers"
EMBEDDED_DATA_SHEET = "Embedded_Data"
SYSTEM_SHEET = "System"
INSTRUCTIONS_SHEET = "Instructions"
TRANSLATION_KEY_SHEET = "TranslationKeyMap"
SURVEY_METADATA_SHEET = "Survey_Metadata"

EMBEDDED_EMPTY_VALUE = "---"
SURVEY_METADATA_KEYS = [
    "SurveyTitle",
    "SurveyDescription",
    "SurveyMetaDescription",
]
FLOW_METADATA_COLUMNS = (
    "BlockName",
    "BlockID",
    "BlockOrder",
    "QuestionOrder",
    "QuestionOrderInBlock",
)
QUESTION_CONFIG_JSON_COLUMN = "QuestionConfigJSON"
LEGACY_QUESTION_CONFIG_COLUMNS = (
    "ForceResponseMode",
    "ValidationType",
    "ValidationSettingsJSON",
    "RandomizationType",
    "RandomizationSettingsJSON",
)
CENTER_ALIGN_HEADER_NAMES = {
    "QID",
    "BlockName",
    "BlockID",
    "BlockOrder",
    "QuestionOrder",
    "QuestionOrderInBlock",
    "QuestionType",
    "DataExportTag",
    "ExportTag",
    "RequiredResponse",
    "ForceResponseMode",
    "ValidationType",
    "ValidationSettingsJSON",
    "RandomizationType",
    "RandomizationSettingsJSON",
}
_QUESTION_VALIDATION_CORE_KEYS = {"ForceResponse", "Type"}
_QUESTION_VALIDATION_TYPE_DEFAULT = "None"
_QUESTION_RANDOMIZATION_CORE_KEYS = {"Type"}
_QUESTION_RANDOMIZATION_TYPE_DEFAULT = "None"


def _normalize_force_response_mode(value: object) -> str:
    raw = str(value or "").strip()
    if not raw:
        return "OFF"
    folded = raw.replace("_", "").replace("-", "").replace(" ", "").lower()
    if folded in {"on", "true", "1", "yes", "y"}:
        return "ON"
    if folded in {"off", "false", "0", "no", "n", "none"}:
        return "OFF"
    if folded == "requestresponse":
        return "RequestResponse"
    if raw.upper() in {"ON", "OFF"}:
        return raw.upper()
    if raw.lower() == "requestresponse":
        return "RequestResponse"
    return raw


def _normalize_validation_type(value: object) -> str:
    raw = str(value or "").strip()
    if not raw:
        return _QUESTION_VALIDATION_TYPE_DEFAULT
    if raw.lower() == "none":
        return _QUESTION_VALIDATION_TYPE_DEFAULT
    return raw


def _normalize_randomization_type(value: object) -> str:
    raw = str(value or "").strip()
    if not raw:
        return _QUESTION_RANDOMIZATION_TYPE_DEFAULT
    if raw.lower() == "none":
        return _QUESTION_RANDOMIZATION_TYPE_DEFAULT
    return raw


def _is_required_response(force_response_mode: object) -> bool:
    mode = _normalize_force_response_mode(force_response_mode)
    return mode in {"ON", "RequestResponse"}


def _validation_settings_dict(question: dict) -> dict[str, object]:
    validation = question.get("Validation") or {}
    if not isinstance(validation, dict):
        return {}
    settings = validation.get("Settings") or {}
    if not isinstance(settings, dict):
        return {}
    return dict(settings)


def _validation_settings_extra_dict(settings: dict[str, object]) -> dict[str, object]:
    out: dict[str, object] = {}
    for key, value in (settings or {}).items():
        key_str = str(key or "").strip()
        if not key_str or key_str in _QUESTION_VALIDATION_CORE_KEYS:
            continue
        if value is None:
            continue
        out[key_str] = value
    return out


def _dump_validation_settings_json(extras: dict[str, object]) -> str:
    if not extras:
        return ""
    return json.dumps(extras, ensure_ascii=True, sort_keys=True, separators=(",", ":"))


def _randomization_settings_dict(question: dict) -> dict[str, object]:
    randomization = question.get("Randomization") or {}
    if not isinstance(randomization, dict):
        return {}
    return dict(randomization)


def _randomization_settings_extra_dict(settings: dict[str, object]) -> dict[str, object]:
    out: dict[str, object] = {}
    for key, value in (settings or {}).items():
        key_str = str(key or "").strip()
        if not key_str or key_str in _QUESTION_RANDOMIZATION_CORE_KEYS:
            continue
        if value is None:
            continue
        out[key_str] = value
    return out


def _dump_randomization_settings_json(extras: dict[str, object]) -> str:
    if not extras:
        return ""
    return json.dumps(extras, ensure_ascii=True, sort_keys=True, separators=(",", ":"))


def _normalize_validation_settings_for_config(
    settings: dict[str, object] | None,
) -> dict[str, object]:
    raw = settings if isinstance(settings, dict) else {}
    force_mode = _normalize_force_response_mode(raw.get("ForceResponse"))
    validation_type = _normalize_validation_type(raw.get("Type"))
    normalized: dict[str, object] = {
        "ForceResponse": force_mode,
        "Type": validation_type,
    }
    extras = _validation_settings_extra_dict(raw)
    if force_mode == "OFF":
        extras.pop("ForceResponseType", None)
    for key in sorted(extras.keys()):
        normalized[key] = extras[key]
    return normalized


def _normalize_randomization_settings_for_config(
    settings: dict[str, object] | None,
) -> dict[str, object]:
    raw = settings if isinstance(settings, dict) else {}
    randomization_type = _normalize_randomization_type(raw.get("Type"))
    normalized: dict[str, object] = {"Type": randomization_type}
    extras = _randomization_settings_extra_dict(raw)
    for key in sorted(extras.keys()):
        normalized[key] = extras[key]
    return normalized


def _question_config_dict(
    *,
    validation_settings: dict[str, object] | None,
    randomization_settings: dict[str, object] | None,
) -> dict[str, dict[str, object]]:
    return {
        "Validation": _normalize_validation_settings_for_config(validation_settings),
        "Randomization": _normalize_randomization_settings_for_config(
            randomization_settings
        ),
    }


def _dump_question_config_json(
    *,
    validation_settings: dict[str, object] | None,
    randomization_settings: dict[str, object] | None,
) -> str:
    config = _question_config_dict(
        validation_settings=validation_settings,
        randomization_settings=randomization_settings,
    )
    return json.dumps(config, ensure_ascii=True, sort_keys=True, separators=(",", ":"))


def _load_json_object_lenient(raw: object) -> dict[str, object]:
    text = str(raw or "").strip()
    if not text:
        return {}
    try:
        parsed = json.loads(text)
    except Exception:
        return {}
    return parsed if isinstance(parsed, dict) else {}


def _build_question_config_json_from_legacy_values(
    *,
    force_response_mode: object,
    validation_type: object,
    validation_settings_json: object,
    randomization_type: object,
    randomization_settings_json: object,
) -> str:
    validation_settings: dict[str, object] = {
        "ForceResponse": _normalize_force_response_mode(force_response_mode),
        "Type": _normalize_validation_type(validation_type),
    }
    for key, value in _load_json_object_lenient(validation_settings_json).items():
        key_str = str(key or "").strip()
        if not key_str or key_str in {"ForceResponse", "Type"}:
            continue
        if value is None:
            continue
        validation_settings[key_str] = value

    randomization_settings: dict[str, object] = {
        "Type": _normalize_randomization_type(randomization_type),
    }
    for key, value in _load_json_object_lenient(randomization_settings_json).items():
        key_str = str(key or "").strip()
        if not key_str or key_str == "Type":
            continue
        if value is None:
            continue
        randomization_settings[key_str] = value

    return _dump_question_config_json(
        validation_settings=validation_settings,
        randomization_settings=randomization_settings,
    )


def _extract_force_response_from_question_config_json(
    config_json: object, *, fallback_mode: object = "OFF"
) -> str:
    fallback = _normalize_force_response_mode(fallback_mode)
    text = str(config_json or "").strip()
    if not text:
        return fallback
    try:
        parsed = json.loads(text)
    except Exception:
        return fallback
    if not isinstance(parsed, dict):
        return fallback
    validation = parsed.get("Validation")
    if isinstance(validation, dict):
        return _normalize_force_response_mode(validation.get("ForceResponse"))
    return _normalize_force_response_mode(parsed.get("ForceResponse"))


def _coerce_bool_cell(value: object, *, default: bool = False) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return default
    text = str(value).strip().lower()
    if text in {"true", "1", "yes", "y", "on"}:
        return True
    if text in {"false", "0", "no", "n", "off", ""}:
        return False
    return default


def _normalize_language_code(lang: str) -> str:
    raw = str(lang or "").strip().replace("_", "-")
    if not raw:
        return ""
    parts = [part.strip() for part in raw.split("-") if part.strip()]
    return "-".join(part.upper() for part in parts)


def _language_suffix(lang: str) -> str:
    code = _normalize_language_code(lang)
    if not code:
        return ""
    return code.lower().replace("-", "_")


def _language_from_suffix(suffix: str) -> str:
    raw = str(suffix or "").strip()
    if not raw:
        return ""
    return raw.replace("_", "-").upper()


def _lookup_language_block(language_blocks: object, lang_code: str) -> dict:
    if not isinstance(language_blocks, dict):
        return {}
    code = _normalize_language_code(lang_code)
    if not code:
        return {}

    candidates: list[str] = []
    for variant in (code, code.replace("-", "_")):
        candidates.extend([variant, variant.upper(), variant.lower()])

    seen: set[str] = set()
    for key in candidates:
        if key in seen:
            continue
        seen.add(key)
        block = language_blocks.get(key)
        if isinstance(block, dict):
            return block
    return {}


def _normalize_subitem_field(value: str | None) -> str:
    raw = str(value or "").strip()
    if not raw:
        return "Answer"
    lowered = raw.lower()
    if lowered == "label":
        return "Label"
    if lowered == "answer":
        return "Answer"
    return raw


def _normalize_language_list(languages: Sequence[str] | None) -> List[str]:
    cleaned: List[str] = []
    seen: set[str] = set()
    for lang in languages or []:
        code = _normalize_language_code(lang)
        if not code or code in seen:
            continue
        seen.add(code)
        cleaned.append(code)
    return cleaned


def _ordered_languages(
    languages: Sequence[str] | None,
    *,
    base_language: str | None = None,
) -> List[str]:
    ordered = _normalize_language_list(languages)
    if base_language:
        base = _normalize_language_code(base_language)
        if base in ordered:
            ordered = [base] + [lang for lang in ordered if lang != base]
        else:
            ordered = [base] + ordered
    else:
        # Legacy fallback: inject EN at front
        if "EN" in ordered:
            ordered = ["EN"] + [lang for lang in ordered if lang != "EN"]
        else:
            ordered = ["EN"] + ordered
    return ordered


def _translation_columns(
    prefix: str,
    languages: Sequence[str] | None,
    *,
    base_language: str | None = None,
) -> List[str]:
    columns: List[str] = []
    for lang in _ordered_languages(languages, base_language=base_language):
        suffix = _language_suffix(lang)
        if not suffix:
            continue
        columns.append(f"{prefix}_{suffix}_MD")
        columns.append(f"{prefix}_{suffix}_IsHTML")
    return columns


def _question_text_md_column(lang: str) -> str:
    suffix = _language_suffix(lang) or "en"
    return f"text_{suffix}"


def _question_text_ishtml_column(lang: str) -> str:
    suffix = _language_suffix(lang) or "en"
    return f"ishtml_{suffix}"


def _question_text_columns(
    languages: Sequence[str] | None,
    *,
    base_language: str | None = None,
) -> List[str]:
    columns: List[str] = []
    for lang in _ordered_languages(languages, base_language=base_language):
        columns.append(_question_text_md_column(lang))
        columns.append(_question_text_ishtml_column(lang))
    return columns


def _question_text_legacy_md_column(lang: str) -> str:
    suffix = _language_suffix(lang) or "en"
    return f"Text_{suffix}_MD"


def _question_text_legacy_ishtml_column(lang: str) -> str:
    suffix = _language_suffix(lang) or "en"
    return f"Text_{suffix}_IsHTML"


def _question_text_lang_columns_from_headers(
    headers: Sequence[object],
    *,
    include_legacy: bool = True,
) -> dict[str, tuple[str, str | None]]:
    """Map language code -> (text col, ishtml col) for Questions sheet headers."""

    header_names = [str(name or "") for name in headers]
    header_set = set(header_names)
    mapping: dict[str, tuple[str, str | None]] = {}

    # Canonical headers.
    for header in header_names:
        if not header.startswith("text_"):
            continue
        suffix = header[len("text_") :].strip()
        if not suffix:
            continue
        lang_code = _language_from_suffix(suffix)
        html_name = f"ishtml_{suffix}"
        mapping[lang_code] = (header, html_name if html_name in header_set else None)

    # Backward compatibility with legacy Text_*_MD / Text_*_IsHTML headers.
    if include_legacy:
        for header in header_names:
            if not header.startswith("Text_") or not header.endswith("_MD"):
                continue
            suffix = header[len("Text_") : -len("_MD")]
            lang_code = _language_from_suffix(suffix)
            if lang_code in mapping:
                continue
            html_name = f"Text_{suffix}_IsHTML"
            mapping[lang_code] = (
                header,
                html_name if html_name in header_set else None,
            )
    return mapping


def _is_question_text_md_header(header: str) -> bool:
    name = str(header or "")
    return name.startswith("text_") or (
        name.startswith("Text_") and name.endswith("_MD")
    )


def _is_question_text_html_header(header: str) -> bool:
    name = str(header or "")
    return name.startswith("ishtml_") or (
        name.startswith("Text_") and name.endswith("_IsHTML")
    )


def _question_text_md_header_from_html_header(header: str) -> str:
    name = str(header or "")
    if name.startswith("ishtml_"):
        return f"text_{name[len('ishtml_'):]}"
    if name.startswith("Text_") and name.endswith("_IsHTML"):
        suffix = name[len("Text_") : -len("_IsHTML")]
        return f"Text_{suffix}_MD"
    return ""


def _metadata_columns(keys: Sequence[str] | None = None) -> List[str]:
    columns: List[str] = []
    for key in keys or SURVEY_METADATA_KEYS:
        columns.append(f"{key}_MD")
        columns.append(f"{key}_IsHTML")
    return columns


def _column_guide(base_language: str = "EN") -> dict:
    """Build the column guide dict, using *base_language* for the base text columns."""
    base_suffix = _language_suffix(base_language) or "en"
    base_upper = _normalize_language_code(base_language) or "EN"
    text_md = _question_text_md_column(base_language)
    text_html = _question_text_ishtml_column(base_language)
    label_md = f"Label_{base_suffix}_MD"
    label_html = f"Label_{base_suffix}_IsHTML"
    return {
        QUESTION_SHEET: [
            ("SurveyID", "System", "Qualtrics Survey ID. Read-only."),
            ("QID", "System", "Qualtrics Question ID. Read-only."),
            ("BlockName", "System", "Qualtrics block name. Read-only."),
            ("BlockID", "System", "Qualtrics block ID from SurveyFlow. Read-only."),
            ("BlockOrder", "System", "1-based block position in SurveyFlow. Read-only."),
            (
                "QuestionOrder",
                "System",
                "1-based question position in SurveyFlow. Read-only.",
            ),
            (
                "QuestionOrderInBlock",
                "System",
                "1-based question position inside the block. Read-only.",
            ),
            ("QuestionType", "System", "Qualtrics question type (MC, TE, etc.)."),
            ("DataExportTag", "System", "Qualtrics DataExportTag / variable name."),
            (
                "RequiredResponse",
                "System",
                "Read-only marker derived from ForceResponseMode "
                "(TRUE when ON/RequestResponse).",
            ),
            (
                "ForceResponseMode",
                "Editable",
                "Question-level response requirement mode (OFF, ON, RequestResponse).",
            ),
            (
                "ValidationType",
                "Editable",
                "Question-level validation type (for example None, MinChoices, CustomValidation).",
            ),
            (
                "ValidationSettingsJSON",
                "Editable",
                "JSON object for additional Validation settings (excluding ForceResponse and Type).",
            ),
            (
                "RandomizationType",
                "Editable",
                "Question-level randomization mode (for example None, All, Subset, Advanced).",
            ),
            (
                "RandomizationSettingsJSON",
                "Editable",
                "JSON object for additional Randomization settings (excluding Type).",
            ),
            (
                QUESTION_CONFIG_JSON_COLUMN,
                "System",
                "Read-only canonical mirror of response settings. Canonical shape: "
                '{"Validation":{"ForceResponse":"OFF|ON|RequestResponse","Type":"None|..."},'
                '"Randomization":{"Type":"None|..."}}.',
            ),
            (text_md, "Editable", f"{base_upper} wording in restricted Markdown."),
            (
                text_html,
                "Flag",
                f"TRUE when {text_md} should be treated as raw HTML.",
            ),
            ("OptionsPreview", "System", "Read-only preview of answer options."),
            (
                "SubitemsPreview",
                "System",
                "Read-only preview of subitems / statements.",
            ),
            (
                "Dirty",
                "System",
                "Auto-flag set by qsync items preview/stage when a row has pending pushes.",
            ),
        ],
        OPTIONS_SHEET: [
            ("SurveyID", "System", "Qualtrics Survey ID. Read-only."),
            ("QID", "System", "Qualtrics Question ID. Read-only."),
            ("BlockName", "System", "Qualtrics block name from SurveyFlow. Read-only."),
            ("BlockID", "System", "Qualtrics block ID from SurveyFlow. Read-only."),
            ("BlockOrder", "System", "1-based block position in SurveyFlow. Read-only."),
            (
                "QuestionOrder",
                "System",
                "1-based question position in SurveyFlow. Read-only.",
            ),
            (
                "QuestionOrderInBlock",
                "System",
                "1-based question position inside the block. Read-only.",
            ),
            ("ChoiceId", "System", "Qualtrics choice ID for this option."),
            ("QuestionType", "System", "Qualtrics question type (MC, Matrix, etc.)."),
            ("ExportTag", "System", "Qualtrics DataExportTag / variable name."),
            ("Code", "System", "Choice code or recode value."),
            (
                label_md,
                "Editable",
                f"{base_upper} option label in restricted Markdown.",
            ),
            (
                label_html,
                "Flag",
                f"TRUE when {label_md} should be treated as raw HTML.",
            ),
            (
                "MetaComment",
                "Note",
                "Auto-generated or manual notes (e.g., externally managed scripts).",
            ),
            (
                "Dirty",
                "System",
                "Auto-flag set by qsync items preview/stage when a row has pending pushes.",
            ),
        ],
        SUBITEMS_SHEET: [
            ("SurveyID", "System", "Qualtrics Survey ID. Read-only."),
            ("QID", "System", "Qualtrics Question ID. Read-only."),
            ("BlockName", "System", "Qualtrics block name from SurveyFlow. Read-only."),
            ("BlockID", "System", "Qualtrics block ID from SurveyFlow. Read-only."),
            ("BlockOrder", "System", "1-based block position in SurveyFlow. Read-only."),
            (
                "QuestionOrder",
                "System",
                "1-based question position in SurveyFlow. Read-only.",
            ),
            (
                "QuestionOrderInBlock",
                "System",
                "1-based question position inside the block. Read-only.",
            ),
            ("AnswerId", "System", "Qualtrics sub-item / statement ID."),
            (
                "Field",
                "System",
                "Disambiguator for subitem meaning (Answer | Label).",
            ),
            ("QuestionType", "System", "Qualtrics question type."),
            ("ExportTag", "System", "Qualtrics DataExportTag / variable name."),
            (
                label_md,
                "Editable",
                f"{base_upper} sub-item text in restricted Markdown.",
            ),
            (
                label_html,
                "Flag",
                f"TRUE when {label_md} should be treated as raw HTML.",
            ),
            (
                "Dirty",
                "System",
                "Auto-flag set by qsync items preview/stage when a row has pending pushes.",
            ),
        ],
        SBS_COLUMNS_SHEET: [
            ("SurveyID", "System", "Qualtrics Survey ID. Read-only."),
            ("QID", "System", "Qualtrics Question ID. Read-only."),
            ("BlockName", "System", "Qualtrics block name from SurveyFlow. Read-only."),
            ("BlockID", "System", "Qualtrics block ID from SurveyFlow. Read-only."),
            ("BlockOrder", "System", "1-based block position in SurveyFlow. Read-only."),
            (
                "QuestionOrder",
                "System",
                "1-based question position in SurveyFlow. Read-only.",
            ),
            (
                "QuestionOrderInBlock",
                "System",
                "1-based question position inside the block. Read-only.",
            ),
            (
                "ColumnId",
                "System",
                "SBS column ID (AdditionalQuestions key). Read-only.",
            ),
            ("QuestionType", "System", "Qualtrics question type (SBS)."),
            ("ExportTag", "System", "Qualtrics DataExportTag / variable name."),
            (
                label_md,
                "Editable",
                f"{base_upper} SBS column header in restricted Markdown.",
            ),
            (
                label_html,
                "Flag",
                f"TRUE when {label_md} should be treated as raw HTML.",
            ),
            (
                "MetaComment",
                "Note",
                "Auto-generated or manual notes (e.g., externally managed scripts).",
            ),
            (
                "Dirty",
                "System",
                "Auto-flag set by qsync items preview/stage when a row has pending pushes.",
            ),
        ],
        SBS_COLUMN_ANSWERS_SHEET: [
            ("SurveyID", "System", "Qualtrics Survey ID. Read-only."),
            ("QID", "System", "Qualtrics Question ID. Read-only."),
            ("BlockName", "System", "Qualtrics block name from SurveyFlow. Read-only."),
            ("BlockID", "System", "Qualtrics block ID from SurveyFlow. Read-only."),
            ("BlockOrder", "System", "1-based block position in SurveyFlow. Read-only."),
            (
                "QuestionOrder",
                "System",
                "1-based question position in SurveyFlow. Read-only.",
            ),
            (
                "QuestionOrderInBlock",
                "System",
                "1-based question position inside the block. Read-only.",
            ),
            ("ColumnId", "System", "SBS column ID (AdditionalQuestions key)."),
            ("AnswerId", "System", "Qualtrics answer ID inside the SBS column."),
            ("QuestionType", "System", "Qualtrics question type (SBS)."),
            ("ExportTag", "System", "Qualtrics DataExportTag / variable name."),
            (
                label_md,
                "Editable",
                f"{base_upper} SBS column answer label in restricted Markdown.",
            ),
            (
                label_html,
                "Flag",
                f"TRUE when {label_md} should be treated as raw HTML.",
            ),
            (
                "MetaComment",
                "Note",
                "Auto-generated or manual notes (e.g., externally managed scripts).",
            ),
            (
                "Dirty",
                "System",
                "Auto-flag set by qsync items preview/stage when a row has pending pushes.",
            ),
        ],
        EMBEDDED_DATA_SHEET: [
            ("SurveyID", "System", "Qualtrics Survey ID. Read-only."),
            (
                "FlowID",
                "System",
                "SurveyFlow node ID (disambiguates duplicate fields).",
            ),
            ("FlowOrder", "System", "Survey flow order (0 for JS-only fields)."),
            ("Field", "System", "Embedded data field name. Read-only."),
            ("Value", "Editable", "Default value (--- for fields without defaults)."),
            ("Type", "System", "Embedded data type (Custom, Recipient, JS-only)."),
            ("WrittenByQIDs", "System", "Comma-separated QIDs that set this field."),
            (
                "Dirty",
                "System",
                "Auto-flag set by qsync items preview/stage when a row has pending pushes.",
            ),
        ],
        SYSTEM_SHEET: [
            ("SurveyID", "System", "Qualtrics Survey ID for Timing/meta items."),
            ("QID", "System", "Timing question ID."),
            ("QuestionType", "System", "Qualtrics question type."),
            ("DataExportTag", "System", "DataExportTag for the Timing item."),
            ("ChoiceId", "System", "Choice ID inside the Timing question."),
            (
                "Display",
                "System",
                "HTML returned by Qualtrics for the Timing display.",
            ),
        ],
        SURVEY_METADATA_SHEET: [
            ("SurveyID", "System", "Qualtrics Survey ID. Read-only."),
            ("Language", "System", "Language code for this metadata row."),
            ("SurveyTitle_MD", "Editable", "Survey title in restricted Markdown."),
            (
                "SurveyTitle_IsHTML",
                "Flag",
                "TRUE when SurveyTitle_MD should be treated as raw HTML.",
            ),
            (
                "SurveyDescription_MD",
                "Editable",
                "Survey description in restricted Markdown.",
            ),
            (
                "SurveyDescription_IsHTML",
                "Flag",
                "TRUE when SurveyDescription_MD should be treated as raw HTML.",
            ),
            (
                "SurveyMetaDescription_MD",
                "Editable",
                "Survey meta description in restricted Markdown.",
            ),
            (
                "SurveyMetaDescription_IsHTML",
                "Flag",
                "TRUE when SurveyMetaDescription_MD should be treated as raw HTML.",
            ),
        ],
    }


# Backward-compatible alias — default EN base.
COLUMN_GUIDE = _column_guide()


def _make_options_preview_formula(
    cell_ref: str,
    *,
    base_language: str = "EN",
    question_type: str | None = None,
) -> ArrayFormula:
    """Generate array formula for dynamic option preview.

    For regular questions this previews answer options from `OptionsTable`.
    For SBSMatrix questions (QuestionType="SBS") it previews SBS columns from
    `SBSColumnsTable` since SBS options live under AdditionalQuestions.
    """
    label_col = f"Label_{_language_suffix(base_language) or 'en'}_MD"
    if (question_type or "").strip().upper() == "SBS":
        table = "SBSColumnsTable"
        id_col = "ColumnId"
    else:
        table = "OptionsTable"
        id_col = "ChoiceId"
    formula = (
        "=_xlfn.LET(_xlpm.q,QuestionsTable[[#This Row],[QID]],\n"
        "     IFERROR(_xlfn.TEXTJOIN(CHAR(10), TRUE,\n"
        f'         "[" & _xlfn._xlws.FILTER({table}[{id_col}], {table}[QID]=_xlpm.q) & "] " &\n'
        f"         _xlfn._xlws.FILTER({table}[{label_col}], {table}[QID]=_xlpm.q)\n"
        '     ), "")\n'
        ")"
    )
    return ArrayFormula(ref=cell_ref, text=formula)


def _make_subitems_preview_formula(
    cell_ref: str, *, base_language: str = "EN"
) -> ArrayFormula:
    """Generate array formula for dynamic subitem preview from SubitemsTable."""
    label_col = f"Label_{_language_suffix(base_language) or 'en'}_MD"
    formula = (
        "=_xlfn.LET(_xlpm.q,QuestionsTable[[#This Row],[QID]],\n"
        "     IFERROR(_xlfn.TEXTJOIN(CHAR(10), TRUE,\n"
        '         "[" & _xlfn._xlws.FILTER(SubitemsTable[AnswerId], SubitemsTable[QID]=_xlpm.q) & "] " &\n'
        f"         _xlfn._xlws.FILTER(SubitemsTable[{label_col}], SubitemsTable[QID]=_xlpm.q)\n"
        '     ), "")\n'
        ")"
    )
    return ArrayFormula(ref=cell_ref, text=formula)


def _update_instructions_sheet(
    wb: Workbook,
    languages: Sequence[str] | None = None,
    *,
    base_language: str | None = None,
) -> None:
    """Build the Instructions sheet with column ownership + descriptions."""

    if INSTRUCTIONS_SHEET in wb.sheetnames:
        ws = wb[INSTRUCTIONS_SHEET]
        wb.remove(ws)
    ws = wb.create_sheet(title=INSTRUCTIONS_SHEET)
    ws.append(["Sheet", "Column", "Ownership", "Description"])

    base = _normalize_language_code(base_language or "") or "EN"
    guide = _column_guide(base)
    for sheet_name, columns in guide.items():
        for column_name, ownership, description in columns:
            ws.append([sheet_name, column_name, ownership, description])

    extra_langs = [
        lang
        for lang in _ordered_languages(languages, base_language=base_language)
        if lang != base
    ]
    for lang in extra_langs:
        suffix = _language_suffix(lang)
        if not suffix:
            continue
        text_md_col = _question_text_md_column(lang)
        text_html_col = _question_text_ishtml_column(lang)
        ws.append(
            [
                QUESTION_SHEET,
                text_md_col,
                "Editable",
                f"{lang} wording in restricted Markdown.",
            ]
        )
        ws.append(
            [
                QUESTION_SHEET,
                text_html_col,
                "Flag",
                f"TRUE when {text_md_col} should be treated as raw HTML.",
            ]
        )
        ws.append(
            [
                OPTIONS_SHEET,
                f"Label_{suffix}_MD",
                "Editable",
                f"{lang} option label in restricted Markdown.",
            ]
        )
        ws.append(
            [
                OPTIONS_SHEET,
                f"Label_{suffix}_IsHTML",
                "Flag",
                f"TRUE when Label_{suffix}_MD should be treated as raw HTML.",
            ]
        )
        ws.append(
            [
                SUBITEMS_SHEET,
                f"Label_{suffix}_MD",
                "Editable",
                f"{lang} sub-item text in restricted Markdown.",
            ]
        )
        ws.append(
            [
                SUBITEMS_SHEET,
                f"Label_{suffix}_IsHTML",
                "Flag",
                f"TRUE when Label_{suffix}_MD should be treated as raw HTML.",
            ]
        )
        ws.append(
            [
                SBS_COLUMNS_SHEET,
                f"Label_{suffix}_MD",
                "Editable",
                f"{lang} SBS column header in restricted Markdown.",
            ]
        )
        ws.append(
            [
                SBS_COLUMNS_SHEET,
                f"Label_{suffix}_IsHTML",
                "Flag",
                f"TRUE when Label_{suffix}_MD should be treated as raw HTML.",
            ]
        )
        ws.append(
            [
                SBS_COLUMN_ANSWERS_SHEET,
                f"Label_{suffix}_MD",
                "Editable",
                f"{lang} SBS column answer label in restricted Markdown.",
            ]
        )
        ws.append(
            [
                SBS_COLUMN_ANSWERS_SHEET,
                f"Label_{suffix}_IsHTML",
                "Flag",
                f"TRUE when Label_{suffix}_MD should be treated as raw HTML.",
            ]
        )

    # Make header row bold
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    for cell in header_row:
        _make_bold(cell)

    # Apply alignment to all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(
                vertical="center", horizontal="left", wrap_text=True
            )
    _apply_readonly_fill(ws, ["Sheet", "Column", "Ownership", "Description"], set())

    # Set widths
    widths = {
        "Sheet": 18.0,
        "Column": 26.0,
        "Ownership": 14.0,
        "Description": 90.0,
    }
    headers_list = [
        h[0] for h in [("Sheet",), ("Column",), ("Ownership",), ("Description",)]
    ]
    for idx, header in enumerate(headers_list, start=1):
        w = widths.get(header, 20.0)
        ws.column_dimensions[get_column_letter(idx)].width = w

    # Create table
    max_row = ws.max_row
    max_col = ws.max_column
    if max_row >= 2 and max_col >= 1:
        table = Table(
            displayName="InstructionsTable",
            ref=f"A1:{get_column_letter(max_col)}{max_row}",
        )
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        # Always refresh the table so its ref covers all rows/columns.
        if "InstructionsTable" in ws._tables:
            del ws._tables["InstructionsTable"]
        ws.add_table(table)


def _translation_key_for_question(qid: str) -> str:
    return f"{qid}_QuestionText"


def _translation_key_for_option(row: "OptionRow") -> tuple[str, str]:
    if row.question_type == "Matrix":
        return "Answer", f"{row.qid}_Answer{row.choice_id}"
    return "Choice", f"{row.qid}_Choice{row.choice_id}"


def _translation_key_for_subitem(row: "SubitemRow") -> tuple[str, str]:
    if row.field == "Label":
        return "Label", f"{row.qid}_Label{row.answer_id}"
    if row.question_type == "Matrix":
        return "Choice", f"{row.qid}_Choice{row.answer_id}"
    return "Answer", f"{row.qid}_Answer{row.answer_id}"


def _update_translation_key_map(
    wb: Workbook,
    questions_map: Dict[str, "QuestionRow"],
    options_map: Dict[Tuple[str, str], "OptionRow"],
    subitems_map: Dict[Tuple[str, str, str], "SubitemRow"],
) -> None:
    if TRANSLATION_KEY_SHEET in wb.sheetnames:
        ws = wb[TRANSLATION_KEY_SHEET]
        wb.remove(ws)
    ws = wb.create_sheet(title=TRANSLATION_KEY_SHEET)

    headers = [
        "Sheet",
        "QuestionType",
        "QID",
        "ChoiceId",
        "AnswerId",
        "Field",
        "TranslationKey",
        "BaseText",
    ]
    ws.append(headers)

    for qid, row in questions_map.items():
        ws.append(
            [
                QUESTION_SHEET,
                row.question_type,
                qid,
                "",
                "",
                "QuestionText",
                _translation_key_for_question(qid),
                row.text_en_md or "",
            ]
        )

    for (qid, choice_id), row in options_map.items():
        field, key = _translation_key_for_option(row)
        ws.append(
            [
                OPTIONS_SHEET,
                row.question_type,
                qid,
                choice_id,
                "",
                field,
                key,
                row.label_en_md or "",
            ]
        )

    for row in subitems_map.values():
        field, key = _translation_key_for_subitem(row)
        ws.append(
            [
                SUBITEMS_SHEET,
                row.question_type,
                row.qid,
                "",
                row.answer_id,
                field,
                key,
                row.label_en_md or "",
            ]
        )

    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    for cell in header_row:
        _make_bold(cell)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(
                vertical="center", horizontal="left", wrap_text=True
            )

    widths = {
        "Sheet": 14.0,
        "QuestionType": 12.0,
        "QID": 10.0,
        "ChoiceId": 10.0,
        "AnswerId": 10.0,
        "Field": 12.0,
        "TranslationKey": 36.0,
        "BaseText": 80.0,
    }
    for idx, name in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(name, 20.0)

    ws.sheet_state = "hidden"


@dataclass
class QuestionRow:
    """Typed representation of a row in the `Questions` sheet."""

    survey_id: str
    qid: str
    block_name: str
    question_type: str
    data_export_tag: str
    required_response: bool
    question_config_json: str | None
    force_response_mode: str
    validation_type: str
    validation_settings_json: str | None
    randomization_type: str
    randomization_settings_json: str | None
    text_en_md: str | None
    text_en_is_html: bool
    externally_managed_by: str | None = None


@dataclass
class OptionRow:
    """Typed representation of a row in the `Options` sheet."""

    survey_id: str
    qid: str
    choice_id: str
    question_type: str
    export_tag: str
    code: str | None
    label_en_md: str | None
    label_en_is_html: bool
    externally_managed_by: str | None = None


@dataclass
class SubitemRow:
    """Typed representation of a row in the `Subitems` sheet."""

    survey_id: str
    qid: str
    answer_id: str
    field: str
    question_type: str
    export_tag: str
    label_en_md: str | None
    label_en_is_html: bool


@dataclass
class SbsColumnRow:
    """Typed representation of a row in the `SBS_Columns` sheet."""

    survey_id: str
    qid: str
    column_id: str
    question_type: str
    export_tag: str
    label_en_md: str | None
    label_en_is_html: bool
    externally_managed_by: str | None = None


@dataclass
class SbsColumnAnswerRow:
    """Typed representation of a row in the `SBS_ColumnAnswers` sheet."""

    survey_id: str
    qid: str
    column_id: str
    answer_id: str
    question_type: str
    export_tag: str
    label_en_md: str | None
    label_en_is_html: bool
    externally_managed_by: str | None = None


@dataclass
class EmbeddedDataRow:
    """Typed representation of a row in the `Embedded_Data` sheet."""

    survey_id: str
    flow_id: str | None
    flow_order: int
    field: str
    value: str | None
    ed_type: str
    written_by_qids: str | None


_SET_EMBEDDED_RE = re.compile(
    r"\bsetEmbeddedData\s*\(\s*(['\"])(?P<field>(?:\\.|(?!\1).)*?)\1\s*,",
    re.IGNORECASE,
)
_SET_EMBEDDED_EXPR_RE = re.compile(
    r"\bsetEmbeddedData\s*\(\s*(?P<expr>.+?)\s*,",
    re.IGNORECASE | re.S,
)


def _strip_js_comments(js_text: str) -> str:
    """Remove JS comments for simple regex scanning."""

    without_block = re.sub(r"/\*.*?\*/", "", js_text, flags=re.S)
    return re.sub(r"//.*", "", without_block)


def _unescape_js_string(value: str) -> str:
    """Unescape simple JS string literals (best-effort)."""

    return value.replace("\\\\", "\\").replace("\\'", "'").replace('\\"', '"').strip()


def _detect_dynamic_field(expr: str) -> str | None:
    expr = (expr or "").strip()
    if not expr:
        return None
    if expr[0] in {"'", '"'}:
        match = re.match(
            r"^(['\"])(?P<prefix>(?:\\.|(?!\1).)*?)\1\s*\+",
            expr,
        )
        if match:
            prefix = _unescape_js_string(match.group("prefix") or "")
            return f"{prefix}*" if prefix else None
        return None
    if expr[0] == "`":
        match = re.match(r"^`(?P<prefix>[^`$]*)\$\{", expr)
        if match:
            prefix = match.group("prefix") or ""
            return f"{prefix}*" if prefix else None
    return None


def _collect_js_embedded_data_fields(survey_payload: dict) -> Dict[str, List[str]]:
    """Map embedded data fields to QIDs that set them via QuestionJS."""

    result = survey_payload.get("result", {})
    questions = result.get("Questions") or {}
    valid_qids = _get_valid_qids(survey_payload)
    field_to_qids: Dict[str, set[str]] = {}

    for qid, details in questions.items():
        if valid_qids and qid not in valid_qids:
            continue
        js_text = (
            details.get("QuestionJS") or details.get("QuestionJSContent") or ""
        ).strip()
        if not js_text:
            continue
        cleaned = _strip_js_comments(js_text)
        for match in _SET_EMBEDDED_RE.finditer(cleaned):
            raw = match.group("field") or ""
            field = _unescape_js_string(raw)
            if not field:
                continue
            field_to_qids.setdefault(field, set()).add(qid)
        for match in _SET_EMBEDDED_EXPR_RE.finditer(cleaned):
            expr = match.group("expr") or ""
            dynamic_field = _detect_dynamic_field(expr)
            if not dynamic_field:
                continue
            field_to_qids.setdefault(dynamic_field, set()).add(qid)

    return {field: sorted(qids) for field, qids in field_to_qids.items()}


def _iter_embedded_data_nodes(survey_payload: dict) -> List[Tuple[int, dict]]:
    """Return ordered (flow_order, node) pairs for EmbeddedData SurveyFlow nodes."""

    flow = survey_payload.get("result", {}).get("SurveyFlow", {}).get("Flow", [])
    nodes: List[Tuple[int, dict]] = []
    order = 0

    def walk(flow_list):
        nonlocal order
        if not isinstance(flow_list, list):
            return
        for node in flow_list:
            if not isinstance(node, dict):
                continue
            if node.get("Type") == "EmbeddedData":
                order += 1
                nodes.append((order, node))
            subflow = node.get("Flow")
            if isinstance(subflow, list):
                walk(subflow)

    walk(flow)
    return nodes


def build_embedded_data_rows(
    survey_id: str, survey_payload: dict
) -> List[EmbeddedDataRow]:
    """Build EmbeddedDataRow objects from SurveyFlow + QuestionJS."""

    js_map = _collect_js_embedded_data_fields(survey_payload)
    rows: List[EmbeddedDataRow] = []
    fields_in_flow: set[str] = set()

    for flow_order, node in _iter_embedded_data_nodes(survey_payload):
        flow_id = str(node.get("FlowID") or "").strip() or None
        for entry in node.get("EmbeddedData", []) or []:
            field = str(entry.get("Field") or "").strip()
            if not field:
                continue
            fields_in_flow.add(field)
            value = entry.get("Value")
            value_str = str(value) if value is not None else None
            ed_type = str(entry.get("Type") or "").strip() or "Custom"
            written_by = js_map.get(field) or []
            written_by_qids = ",".join(written_by) if written_by else ""
            rows.append(
                EmbeddedDataRow(
                    survey_id=survey_id,
                    flow_id=flow_id,
                    flow_order=flow_order,
                    field=field,
                    value=value_str,
                    ed_type=ed_type,
                    written_by_qids=written_by_qids,
                )
            )

    for field in sorted(js_map.keys()):
        if field in fields_in_flow:
            continue
        written_by = js_map.get(field) or []
        written_by_qids = ",".join(written_by) if written_by else ""
        rows.append(
            EmbeddedDataRow(
                survey_id=survey_id,
                flow_id=None,
                flow_order=0,
                field=field,
                value=None,
                ed_type="JS-only",
                written_by_qids=written_by_qids,
            )
        )

    rows.sort(
        key=lambda r: (
            r.flow_order == 0,
            r.flow_order,
            (r.flow_id or ""),
            r.field,
        )
    )
    return rows


def _get_or_create_sheet(wb: Workbook, name: str) -> Worksheet:
    if name in wb.sheetnames:
        return wb[name]
    return wb.create_sheet(title=name)


def _strip_all_comments(wb: Workbook) -> None:
    """Remove legacy Excel comments so the workbook saves cleanly."""

    for ws in wb.worksheets:
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        if max_row == 0 or max_col == 0:
            continue
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                if cell.comment is not None:
                    cell.comment = None


def _iter_sheet_rows(ws: Worksheet):
    rows = list(ws.iter_rows(values_only=False))
    if not rows:
        return [], []
    header_row = rows[0]
    headers = [cell.value or "" for cell in header_row]
    data_rows = rows[1:]
    return headers, data_rows


def _reorder_columns(ws: Worksheet, ordered_headers: List[str]) -> None:
    """Reorder worksheet columns to match ordered_headers, preserving values."""

    headers, data_rows = _iter_sheet_rows(ws)
    if not headers:
        ws.append(ordered_headers)
        return
    if headers == ordered_headers:
        return

    index_map = {name: idx for idx, name in enumerate(headers)}
    new_rows: List[List[object]] = [ordered_headers]
    for row in data_rows:
        new_rows.append(
            [
                row[index_map[name]].value if name in index_map else None
                for name in ordered_headers
            ]
        )

    if ws.max_row:
        ws.delete_rows(1, ws.max_row)
    for row_values in new_rows:
        ws.append(row_values)


def _drop_stale_translation_columns(
    ws: Worksheet,
    required_cols: List[str],
    prefixes: Sequence[str] = ("Text", "Label"),
) -> None:
    """Remove translation columns (``{prefix}_*_MD``, ``{prefix}_*_IsHTML``)
    that are NOT in *required_cols*.

    This cleans up columns left behind by a previous workbook init that used a
    different base language (e.g. ``Text_en_MD`` when the survey base is now FR).
    """
    headers, _ = _iter_sheet_rows(ws)
    if not headers:
        return
    required_set = set(required_cols)
    stale_indices: list[int] = []
    for idx, name in enumerate(headers):
        header = str(name or "")
        for prefix in prefixes:
            if header.startswith(f"{prefix}_") and (
                header.endswith("_MD") or header.endswith("_IsHTML")
            ):
                if header not in required_set:
                    stale_indices.append(idx)
                break
    # Delete rightmost first so indices stay valid.
    for col_idx in sorted(stale_indices, reverse=True):
        ws.delete_cols(col_idx + 1)


def _drop_stale_question_text_columns(
    ws: Worksheet,
    required_cols: List[str],
) -> None:
    """Remove stale question text columns (canonical + legacy naming)."""

    headers, _ = _iter_sheet_rows(ws)
    if not headers:
        return
    required_set = set(required_cols)
    stale_indices: list[int] = []
    for idx, name in enumerate(headers):
        header = str(name or "")
        is_question_text_col = (
            header.startswith("text_")
            or header.startswith("ishtml_")
            or (
                header.startswith("Text_")
                and (header.endswith("_MD") or header.endswith("_IsHTML"))
            )
        )
        if is_question_text_col and header not in required_set:
            stale_indices.append(idx)
    for col_idx in sorted(stale_indices, reverse=True):
        ws.delete_cols(col_idx + 1)


def _migrate_legacy_question_text_columns(
    ws: Worksheet,
    *,
    languages: Sequence[str] | None,
    base_language: str | None,
) -> None:
    """Copy legacy `Text_*_MD` values into canonical `text_*` columns."""

    headers, _ = _iter_sheet_rows(ws)
    if not headers:
        return
    idx = {str(name or ""): i + 1 for i, name in enumerate(headers)}
    langs = _ordered_languages(languages, base_language=base_language)
    for lang in langs:
        legacy_md = _question_text_legacy_md_column(lang)
        legacy_html = _question_text_legacy_ishtml_column(lang)
        new_md = _question_text_md_column(lang)
        new_html = _question_text_ishtml_column(lang)
        if legacy_md not in idx or new_md not in idx:
            continue
        legacy_md_col = idx[legacy_md]
        legacy_html_col = idx.get(legacy_html)
        new_md_col = idx[new_md]
        new_html_col = idx.get(new_html)
        for row_idx in range(2, ws.max_row + 1):
            old_text = ws.cell(row=row_idx, column=legacy_md_col).value
            new_text_cell = ws.cell(row=row_idx, column=new_md_col)
            if (
                (new_text_cell.value is None or str(new_text_cell.value).strip() == "")
                and old_text is not None
                and str(old_text).strip() != ""
            ):
                new_text_cell.value = old_text
            if legacy_html_col and new_html_col:
                old_html = ws.cell(row=row_idx, column=legacy_html_col).value
                new_html_cell = ws.cell(row=row_idx, column=new_html_col)
                if (
                    new_html_cell.value is None
                    and old_html is not None
                    and str(old_html).strip() != ""
                ):
                    new_html_cell.value = old_html


def _drop_columns_by_name(ws: Worksheet, names: Sequence[str]) -> None:
    headers, _ = _iter_sheet_rows(ws)
    if not headers:
        return
    targets = set(str(name or "") for name in names)
    indices = [
        idx + 1
        for idx, header in enumerate(headers)
        if str(header or "") in targets
    ]
    for col_idx in sorted(indices, reverse=True):
        ws.delete_cols(col_idx)


def _migrate_legacy_question_config_columns(ws: Worksheet) -> None:
    headers, _ = _iter_sheet_rows(ws)
    if not headers:
        return
    idx = {str(name or ""): i + 1 for i, name in enumerate(headers)}
    if QUESTION_CONFIG_JSON_COLUMN not in idx:
        return
    if not any(name in idx for name in LEGACY_QUESTION_CONFIG_COLUMNS):
        return

    config_col = idx[QUESTION_CONFIG_JSON_COLUMN]
    for row_idx in range(2, ws.max_row + 1):
        config_cell = ws.cell(row=row_idx, column=config_col)
        if config_cell.value is not None and str(config_cell.value).strip():
            continue
        config_cell.value = _build_question_config_json_from_legacy_values(
            force_response_mode=(
                ws.cell(row=row_idx, column=idx["ForceResponseMode"]).value
                if "ForceResponseMode" in idx
                else "OFF"
            ),
            validation_type=(
                ws.cell(row=row_idx, column=idx["ValidationType"]).value
                if "ValidationType" in idx
                else "None"
            ),
            validation_settings_json=(
                ws.cell(row=row_idx, column=idx["ValidationSettingsJSON"]).value
                if "ValidationSettingsJSON" in idx
                else ""
            ),
            randomization_type=(
                ws.cell(row=row_idx, column=idx["RandomizationType"]).value
                if "RandomizationType" in idx
                else "None"
            ),
            randomization_settings_json=(
                ws.cell(row=row_idx, column=idx["RandomizationSettingsJSON"]).value
                if "RandomizationSettingsJSON" in idx
                else ""
            ),
        )


def _ensure_columns(ws: Worksheet, required: List[str]) -> Dict[str, int]:
    """Ensure required columns exist in order; return mapping name -> 0-based index."""

    header_cells = list(ws.iter_rows(min_row=1, max_row=1, values_only=False))
    if header_cells:
        header_cells = header_cells[0]
        headers = [c.value or "" for c in header_cells]
        # If the header row is effectively empty (all blanks), treat as no header.
        if not any(headers):
            headers = []
    else:
        headers = []

    if not headers:
        for col in required:
            ws.cell(row=1, column=len(headers) + 1, value=col)
            headers.append(col)
        return {name: idx for idx, name in enumerate(headers)}

    extras = [h for h in headers if h not in required]
    ordered_headers = list(dict.fromkeys(required + extras))
    _reorder_columns(ws, ordered_headers)
    return {name: idx for idx, name in enumerate(ordered_headers)}


EXTERNALLY_MANAGED_TAGS: Dict[str, str] = {
    # DataExportTag -> controlling script
    "newsmem_recognition": "scripts/update_newsmem_recognition.py",
    "newsmem_salience": "scripts/update_salience_items.py",
    "newsmem_recall_cued": "scripts/update_salience_items.py",
}
_EXTERNALLY_MANAGED_NOTE_PREFIX = "Externally managed – edit via "


def _is_externally_managed_question(data_export_tag: str | None) -> Optional[str]:
    if not data_export_tag:
        return None
    return EXTERNALLY_MANAGED_TAGS.get(data_export_tag)


def _externally_managed_note(script: str) -> str:
    return f"{_EXTERNALLY_MANAGED_NOTE_PREFIX}{script}, not directly in Excel."


@dataclass(frozen=True)
class _QuestionFlowMeta:
    block_id: str
    block_name: str
    block_order: int
    question_order: int
    question_order_in_block: int


def _question_flow_meta_by_qid(survey_payload: dict) -> Dict[str, _QuestionFlowMeta]:
    """Return SurveyFlow position metadata for each in-flow/non-trash QID."""

    result = survey_payload.get("result", {})
    questions = result.get("Questions", {})
    blocks = result.get("Blocks", {})
    if not isinstance(questions, dict) or not isinstance(blocks, dict):
        return {}

    flow_meta: Dict[str, _QuestionFlowMeta] = {}
    seen_qids: set[str] = set()
    question_order = 0
    block_order = 0

    for block_id in _iter_block_ids_in_flow(survey_payload):
        block = blocks.get(block_id)
        if not isinstance(block, dict):
            continue
        if str(block.get("Type") or "").strip() == "Trash":
            continue
        block_order += 1
        block_name = str(block.get("Description") or "")
        elements = block.get("BlockElements") or block.get("Elements") or []
        if not isinstance(elements, list):
            continue

        question_order_in_block = 0
        for elem in elements:
            if not isinstance(elem, dict):
                continue
            elem_type = str(elem.get("Type") or "").strip()
            if elem_type not in {"", "Question"}:
                continue
            qid = str(elem.get("QuestionID") or "").strip()
            if not qid or qid in seen_qids or qid not in questions:
                continue
            seen_qids.add(qid)
            question_order += 1
            question_order_in_block += 1
            flow_meta[qid] = _QuestionFlowMeta(
                block_id=str(block_id),
                block_name=block_name,
                block_order=block_order,
                question_order=question_order,
                question_order_in_block=question_order_in_block,
            )

    return flow_meta


def _flow_meta_values(
    qid: str,
    flow_meta_by_qid: Dict[str, _QuestionFlowMeta],
) -> Dict[str, str | int | None]:
    """Return workbook column values for flow metadata columns."""

    meta = flow_meta_by_qid.get(str(qid))
    if not meta:
        return {
            "BlockName": "",
            "BlockID": "",
            "BlockOrder": None,
            "QuestionOrder": None,
            "QuestionOrderInBlock": None,
        }
    return {
        "BlockName": meta.block_name,
        "BlockID": meta.block_id,
        "BlockOrder": meta.block_order,
        "QuestionOrder": meta.question_order,
        "QuestionOrderInBlock": meta.question_order_in_block,
    }


def _write_flow_meta_cells(
    ws: Worksheet,
    row_idx: int,
    col_index: Dict[str, int],
    qid: str,
    flow_meta_by_qid: Dict[str, _QuestionFlowMeta],
) -> None:
    values = _flow_meta_values(qid, flow_meta_by_qid)
    for col_name, value in values.items():
        if col_name in col_index:
            ws.cell(row=row_idx, column=col_index[col_name] + 1, value=value)


def _iter_block_ids_in_flow(survey_payload: dict) -> List[str]:
    """Return block IDs in survey-flow order."""

    result = survey_payload.get("result", {})
    flow = result.get("SurveyFlow", {})
    ordered_block_ids: List[str] = []

    def walk(node):
        if isinstance(node, dict):
            # Qualtrics uses both `Standard` and nested `Block` nodes in SurveyFlow.
            # Treat both as block references for ordering.
            if node.get("Type") in {"Standard", "Block"} and "ID" in node:
                bid = node["ID"]
                if bid not in ordered_block_ids:
                    ordered_block_ids.append(bid)
            for v in node.values():
                if isinstance(v, (dict, list)):
                    walk(v)
        elif isinstance(node, list):
            for v in node:
                if isinstance(v, (dict, list)):
                    walk(v)

    walk(flow.get("Flow", []))
    return ordered_block_ids


def _iter_question_entries_in_flow(survey_payload: dict) -> List[Tuple[str, str]]:
    """Return `(qid, block_name)` pairs in SurveyFlow order (non-Trash blocks only)."""

    flow_meta = _question_flow_meta_by_qid(survey_payload)
    ordered = sorted(flow_meta.items(), key=lambda item: item[1].question_order)
    return [(qid, meta.block_name) for qid, meta in ordered]


def _ordered_qids_in_flow(survey_payload: dict) -> List[str]:
    """Return active/non-trash QIDs in SurveyFlow order."""

    return [qid for qid, _ in _iter_question_entries_in_flow(survey_payload)]


def _build_option_previews(survey_payload: dict) -> Dict[str, str]:
    """Map QID -> newline-joined option labels (Markdown).

    For Matrix questions, uses Answers (response scale).
    For MC/SC questions, uses Choices.
    For SBSMatrix questions, previews the SBS column headers from
    AdditionalQuestions[*].QuestionText (since SBS options are per-column).
    """

    result: Dict[str, List[str]] = {}
    questions = survey_payload.get("result", {}).get("Questions", {})
    for qid, q in questions.items():
        qtype = q.get("QuestionType") or ""

        if _is_sbs_matrix_question(q):
            additional = q.get("AdditionalQuestions") or {}
            if not isinstance(additional, dict) or not additional:
                continue
            lines: List[str] = []
            for column_id in _ordered_numeric_keys(additional):
                aq = additional.get(column_id)
                if not isinstance(aq, dict):
                    continue
                display = aq.get("QuestionText") or ""
                text = html_to_md(display)
                if text:
                    lines.append(text)
            if lines:
                result[qid] = lines
            continue

        if qtype == "Matrix":
            # For Matrix, options are the Answers (response scale)
            answers = q.get("Answers") or {}
            if not answers:
                continue
            order: List[str] = []
            answer_order = q.get("AnswerOrder")
            if isinstance(answer_order, list):
                order.extend([str(aid) for aid in answer_order if str(aid) in answers])
            for aid in answers.keys():
                said = str(aid)
                if said not in order:
                    order.append(said)
            lines: List[str] = []
            for aid in order:
                answer = answers.get(aid)
                if not answer:
                    continue
                display = answer.get("Display") or ""
                text = html_to_md(display)
                if text:
                    lines.append(text)
            if lines:
                result[qid] = lines
        else:
            # For MC/SC, options are Choices
            choices = q.get("Choices") or {}
            if not choices:
                continue
            order: List[str] = []
            choice_order = q.get("ChoiceOrder")
            if isinstance(choice_order, list):
                order.extend([str(cid) for cid in choice_order if str(cid) in choices])
            for cid in choices.keys():
                scid = str(cid)
                if scid not in order:
                    order.append(scid)
            lines: List[str] = []
            for cid in order:
                choice = choices.get(cid)
                if not choice:
                    continue
                display = choice.get("Display") or ""
                text = html_to_md(display)
                if text:
                    lines.append(text)
            if lines:
                result[qid] = lines

    return {qid: "\n".join(lines) for qid, lines in result.items()}


def _build_subitem_previews(survey_payload: dict) -> Dict[str, str]:
    """Map QID -> newline-joined subitem labels (Markdown).

    For Matrix questions, uses Choices (matrix rows/statements).
    For other questions, uses Answers.
    """

    result: Dict[str, List[str]] = {}
    questions = survey_payload.get("result", {}).get("Questions", {})
    for qid, q in questions.items():
        qtype = q.get("QuestionType") or ""

        if qtype == "Matrix" or _is_sbs_matrix_question(q):
            # For Matrix, subitems are Choices (the statements/headlines)
            choices = q.get("Choices") or {}
            if not choices:
                continue
            order: List[str] = []
            choice_order = q.get("ChoiceOrder")
            if isinstance(choice_order, list):
                order.extend([str(cid) for cid in choice_order if str(cid) in choices])
            for cid in choices.keys():
                scid = str(cid)
                if scid not in order:
                    order.append(scid)
            lines: List[str] = []
            for cid in order:
                choice = choices.get(cid)
                if not choice:
                    continue
                display = choice.get("Display") or ""
                text = html_to_md(display)
                if text:
                    lines.append(text)
            if lines:
                result[qid] = lines
        else:
            # For other questions, subitems are Answers
            answers = q.get("Answers") or {}
            if not answers:
                continue
            lines: List[str] = []
            for answer_id, answer in answers.items():
                display = answer.get("Display")
                text = html_to_md(str(display) if display is not None else "")
                if text:
                    lines.append(text)
            if lines:
                result[qid] = lines

    return {qid: "\n".join(lines) for qid, lines in result.items()}


def _get_valid_qids(survey_payload: dict) -> set[str]:
    """
    Extract active QIDs from SurveyFlow (excluding Trash blocks).

    Returns the QIDs that should appear in workbook item sheets. A question is
    considered valid only when it is referenced by a non-Trash block that is
    reachable via SurveyFlow.
    """
    return set(_ordered_qids_in_flow(survey_payload))


def build_question_rows(
    survey_id: str,
    survey_payload: dict,
) -> Dict[str, QuestionRow]:
    """Build QuestionRow objects from a survey JSON payload.

    Questions are ordered by SurveyFlow block order and BlockElements order and
    include only QIDs reachable in SurveyFlow (excluding Trash blocks).
    """

    result = survey_payload.get("result", {})
    questions = result.get("Questions", {})

    rows: Dict[str, QuestionRow] = {}
    question_entries = _iter_question_entries_in_flow(survey_payload)
    for qid, block_name in question_entries:
        q = questions[qid]
        qtype = q.get("QuestionType") or ""
        tag = q.get("DataExportTag") or ""
        text_html = q.get("QuestionText") or ""
        settings = _validation_settings_dict(q)
        force_mode = _normalize_force_response_mode(settings.get("ForceResponse"))
        validation_extras = _validation_settings_extra_dict(settings)
        if force_mode == "OFF":
            validation_extras.pop("ForceResponseType", None)
        validation_type = _normalize_validation_type(settings.get("Type"))
        validation_settings_json = _dump_validation_settings_json(validation_extras)
        randomization = _randomization_settings_dict(q)
        randomization_type = _normalize_randomization_type(randomization.get("Type"))
        randomization_settings_json = _dump_randomization_settings_json(
            _randomization_settings_extra_dict(randomization)
        )
        question_config_json = _dump_question_config_json(
            validation_settings=settings,
            randomization_settings=randomization,
        )

        if is_markdown_safe_html(text_html):
            text_md = html_to_md(text_html)
            is_html = False
        elif should_treat_as_html(text_html):
            text_md = normalize_text(text_html)
            is_html = True
        else:
            # Plaintext or odd edge cases – treat as Markdown.
            text_md = html_to_md(text_html)
            is_html = False

        rows[qid] = QuestionRow(
            survey_id=survey_id,
            qid=qid,
            block_name=block_name,
            question_type=qtype,
            data_export_tag=tag,
            required_response=_is_required_response(force_mode),
            question_config_json=question_config_json,
            force_response_mode=force_mode,
            validation_type=validation_type,
            validation_settings_json=validation_settings_json or None,
            randomization_type=randomization_type,
            randomization_settings_json=randomization_settings_json or None,
            text_en_md=text_md,
            text_en_is_html=is_html,
            externally_managed_by=None,
        )

    return rows


def _is_sbs_matrix_question(question: dict) -> bool:
    """Return True for Qualtrics SBS side-by-side matrix questions.

    Qualtrics encodes these as QuestionType="SBS" with Selector="SBSMatrix".
    """

    if not isinstance(question, dict):
        return False
    return (
        str(question.get("QuestionType") or "").strip() == "SBS"
        and str(question.get("Selector") or "").strip() == "SBSMatrix"
    )


def _ordered_numeric_keys(items: dict) -> List[str]:
    """Return dict keys ordered numerically when possible (stable fallback to string)."""

    keys = [str(k) for k in (items or {}).keys()]

    def key_fn(value: str) -> tuple[int, int | None, str]:
        value = str(value)
        try:
            return (0, int(value), value)
        except ValueError:
            return (1, None, value)

    return sorted(keys, key=key_fn)


def _sbs_matrix_qids(survey_payload: dict) -> set[str]:
    """Return QIDs for non-Trash SBSMatrix questions."""

    questions = (survey_payload.get("result") or {}).get("Questions") or {}
    if not isinstance(questions, dict):
        return set()
    valid_qids = _get_valid_qids(survey_payload)
    return {
        str(qid)
        for qid, q in questions.items()
        if qid in valid_qids and isinstance(q, dict) and _is_sbs_matrix_question(q)
    }


def build_option_rows(
    survey_id: str,
    survey_payload: dict,
) -> Dict[Tuple[str, str], OptionRow]:
    """Build OptionRow objects from a survey JSON payload.

    Mapping rules (must stay in sync with sync_core.py):
    - For MC/SC questions: options come from `Choices` (one row per choice).
    - For Matrix questions: options come from `Answers` (the response scale).
    - For SBSMatrix questions: options are per-column under `AdditionalQuestions` and
      are edited via the `SBS_ColumnAnswers` sheet (so this function skips SBSMatrix).
    Options for each question are ordered by `ChoiceOrder` or `AnswerOrder`
    where available so they match the respondent-facing order.

    Only includes options for questions that are active in SurveyFlow to
    maintain referential integrity with the Questions sheet.
    """

    questions = survey_payload.get("result", {}).get("Questions", {})
    ordered_qids = _ordered_qids_in_flow(survey_payload)

    rows: Dict[Tuple[str, str], OptionRow] = {}
    for qid in ordered_qids:
        q = questions.get(qid)
        if not isinstance(q, dict):
            continue

        qtype = q.get("QuestionType") or ""
        if _is_sbs_matrix_question(q):
            # SBSMatrix statements live in Choices, but are edited via Subitems.
            # SBS columns/answer scales live in AdditionalQuestions.
            continue
        tag = q.get("DataExportTag") or ""
        externally_by = _is_externally_managed_question(tag)

        # For Matrix questions, options come from Answers (the response scale)
        if qtype == "Matrix":
            answers = q.get("Answers")
            if not answers:
                continue

            ordered_ids: List[str] = []
            answer_order = q.get("AnswerOrder")
            if isinstance(answer_order, list) and answer_order:
                ordered_ids.extend(
                    [str(aid) for aid in answer_order if str(aid) in answers]
                )

            for aid in answers.keys():
                if str(aid) not in ordered_ids:
                    ordered_ids.append(str(aid))

            for answer_id in ordered_ids:
                answer = answers.get(answer_id)
                if not answer:
                    continue
                display = answer.get("Display") or ""

                if is_markdown_safe_html(display):
                    label_md = html_to_md(display)
                    is_html = False
                elif should_treat_as_html(display):
                    label_md = normalize_text(display)
                    is_html = True
                else:
                    label_md = html_to_md(display)
                    is_html = False

                key = (qid, answer_id)
                rows[key] = OptionRow(
                    survey_id=survey_id,
                    qid=qid,
                    choice_id=str(answer_id),
                    question_type=qtype,
                    export_tag=tag,
                    code=None,
                    label_en_md=label_md,
                    label_en_is_html=is_html,
                    externally_managed_by=externally_by,
                )
        else:
            # For MC/SC questions, options come from Choices
            choices = q.get("Choices")
            if not choices:
                continue

            # Determine iteration order for choices.
            ordered_ids: List[str] = []
            choice_order = q.get("ChoiceOrder")
            if isinstance(choice_order, list) and choice_order:
                # ChoiceOrder may be numeric; normalise to strings for lookup.
                ordered_ids.extend(
                    [str(cid) for cid in choice_order if str(cid) in choices]
                )

            # Append any remaining choices not in ChoiceOrder, preserving dict order.
            for cid in choices.keys():
                if str(cid) not in ordered_ids:
                    ordered_ids.append(str(cid))

            for choice_id in ordered_ids:
                choice = choices.get(choice_id)
                if not choice:
                    continue
                display = choice.get("Display") or ""
                code = choice.get("Recode") or None

                if is_markdown_safe_html(display):
                    label_md = html_to_md(display)
                    is_html = False
                elif should_treat_as_html(display):
                    label_md = normalize_text(display)
                    is_html = True
                else:
                    label_md = html_to_md(display)
                    is_html = False

                key = (qid, choice_id)
                rows[key] = OptionRow(
                    survey_id=survey_id,
                    qid=qid,
                    choice_id=str(choice_id),
                    question_type=qtype,
                    export_tag=tag,
                    code=code,
                    label_en_md=label_md,
                    label_en_is_html=is_html,
                    externally_managed_by=externally_by,
                )
    return rows


def build_subitem_rows(
    survey_id: str,
    survey_payload: dict,
) -> Dict[Tuple[str, str, str], SubitemRow]:
    """Build SubitemRow objects from a survey JSON payload.

    Mapping rules (must stay in sync with sync_core.py):
    - For Matrix questions: subitems come from `Choices` (matrix rows/statements).
    - For SBSMatrix questions (QuestionType="SBS", Selector="SBSMatrix"): subitems
      come from `Choices` (the SBS statements/rows) and are edited via the Subitems
      sheet (not Options).
    - For other question types: subitems come from `Answers` (if present).
    - Label rows (Field=Label) come from `Labels` when present.

    Only includes subitems for questions that are active in SurveyFlow to
    maintain referential integrity with the Questions sheet.
    """

    questions = survey_payload.get("result", {}).get("Questions", {})
    ordered_qids = _ordered_qids_in_flow(survey_payload)

    rows: Dict[Tuple[str, str, str], SubitemRow] = {}
    for qid in ordered_qids:
        q = questions.get(qid)
        if not isinstance(q, dict):
            continue

        qtype = q.get("QuestionType") or ""
        tag = q.get("DataExportTag") or ""

        # For Matrix questions (and SBSMatrix), subitems come from Choices (rows/statements).
        if qtype == "Matrix" or _is_sbs_matrix_question(q):
            choices = q.get("Choices")
            if not choices:
                continue

            ordered_ids: List[str] = []
            choice_order = q.get("ChoiceOrder")
            if isinstance(choice_order, list) and choice_order:
                ordered_ids.extend(
                    [str(cid) for cid in choice_order if str(cid) in choices]
                )

            for cid in choices.keys():
                if str(cid) not in ordered_ids:
                    ordered_ids.append(str(cid))

            for choice_id in ordered_ids:
                choice = choices.get(choice_id)
                if not choice:
                    continue
                display = choice.get("Display") or ""

                if is_markdown_safe_html(display):
                    label_md = html_to_md(display)
                    is_html = False
                elif should_treat_as_html(display):
                    label_md = normalize_text(display)
                    is_html = True
                else:
                    label_md = html_to_md(display)
                    is_html = False

                rows[(qid, "Answer", str(choice_id))] = SubitemRow(
                    survey_id=survey_id,
                    qid=qid,
                    answer_id=str(choice_id),
                    field="Answer",
                    question_type=qtype,
                    export_tag=tag,
                    label_en_md=label_md,
                    label_en_is_html=is_html,
                )
        else:
            # For other question types, subitems come from Answers
            answers = q.get("Answers")
            if answers:
                for answer_id, answer in answers.items():
                    raw_display = answer.get("Display")
                    display = str(raw_display) if raw_display is not None else ""

                    if is_markdown_safe_html(display):
                        label_md = html_to_md(display)
                        is_html = False
                    elif should_treat_as_html(display):
                        label_md = normalize_text(display)
                        is_html = True
                    else:
                        label_md = html_to_md(display)
                        is_html = False

                    rows[(qid, "Answer", str(answer_id))] = SubitemRow(
                        survey_id=survey_id,
                        qid=qid,
                        answer_id=str(answer_id),
                        field="Answer",
                        question_type=qtype,
                        export_tag=tag,
                        label_en_md=label_md,
                        label_en_is_html=is_html,
                    )

        labels = q.get("Labels") or {}
        if labels:
            for label_id, label in labels.items():
                raw_display = label.get("Display")
                display = str(raw_display) if raw_display is not None else ""

                if is_markdown_safe_html(display):
                    label_md = html_to_md(display)
                    is_html = False
                elif should_treat_as_html(display):
                    label_md = normalize_text(display)
                    is_html = True
                else:
                    label_md = html_to_md(display)
                    is_html = False

                rows[(qid, "Label", str(label_id))] = SubitemRow(
                    survey_id=survey_id,
                    qid=qid,
                    answer_id=str(label_id),
                    field="Label",
                    question_type=qtype,
                    export_tag=tag,
                    label_en_md=label_md,
                    label_en_is_html=is_html,
                )

    return rows


def build_sbs_column_rows(
    survey_id: str,
    survey_payload: dict,
) -> Dict[Tuple[str, str], SbsColumnRow]:
    """Build SbsColumnRow objects for Qualtrics SBSMatrix questions."""

    questions = survey_payload.get("result", {}).get("Questions", {})
    ordered_qids = _ordered_qids_in_flow(survey_payload)
    rows: Dict[Tuple[str, str], SbsColumnRow] = {}

    for qid in ordered_qids:
        q = questions.get(qid)
        if not isinstance(q, dict):
            continue
        if not _is_sbs_matrix_question(q):
            continue

        qtype = str(q.get("QuestionType") or "").strip()
        tag = str(q.get("DataExportTag") or "").strip()
        externally_by = _is_externally_managed_question(tag)
        additional = q.get("AdditionalQuestions") or {}
        if not isinstance(additional, dict) or not additional:
            continue

        for column_id in _ordered_numeric_keys(additional):
            aq = additional.get(column_id)
            if not isinstance(aq, dict):
                continue
            display = aq.get("QuestionText") or ""

            if is_markdown_safe_html(display):
                label_md = html_to_md(display)
                is_html = False
            elif should_treat_as_html(display):
                label_md = normalize_text(display)
                is_html = True
            else:
                label_md = html_to_md(display)
                is_html = False

            rows[(qid, str(column_id))] = SbsColumnRow(
                survey_id=survey_id,
                qid=qid,
                column_id=str(column_id),
                question_type=qtype,
                export_tag=tag,
                label_en_md=label_md,
                label_en_is_html=is_html,
                externally_managed_by=externally_by,
            )

    return rows


def build_sbs_column_answer_rows(
    survey_id: str,
    survey_payload: dict,
) -> Dict[Tuple[str, str, str], SbsColumnAnswerRow]:
    """Build SbsColumnAnswerRow objects for SBSMatrix AdditionalQuestions answer scales."""

    questions = survey_payload.get("result", {}).get("Questions", {})
    ordered_qids = _ordered_qids_in_flow(survey_payload)
    rows: Dict[Tuple[str, str, str], SbsColumnAnswerRow] = {}

    for qid in ordered_qids:
        q = questions.get(qid)
        if not isinstance(q, dict):
            continue
        if not _is_sbs_matrix_question(q):
            continue

        qtype = str(q.get("QuestionType") or "").strip()
        tag = str(q.get("DataExportTag") or "").strip()
        externally_by = _is_externally_managed_question(tag)
        additional = q.get("AdditionalQuestions") or {}
        if not isinstance(additional, dict) or not additional:
            continue

        for column_id in _ordered_numeric_keys(additional):
            aq = additional.get(column_id)
            if not isinstance(aq, dict):
                continue
            answers = aq.get("Answers") or {}
            if not isinstance(answers, dict) or not answers:
                continue

            ordered_ids: list[str] = []
            answer_order = aq.get("AnswerOrder")
            if isinstance(answer_order, list) and answer_order:
                ordered_ids.extend(
                    [str(aid) for aid in answer_order if str(aid) in answers]
                )
            for aid in answers.keys():
                said = str(aid)
                if said not in ordered_ids:
                    ordered_ids.append(said)

            for answer_id in ordered_ids:
                answer = answers.get(answer_id)
                if not isinstance(answer, dict):
                    continue
                display = answer.get("Display") or ""

                if is_markdown_safe_html(display):
                    label_md = html_to_md(display)
                    is_html = False
                elif should_treat_as_html(display):
                    label_md = normalize_text(display)
                    is_html = True
                else:
                    label_md = html_to_md(display)
                    is_html = False

                rows[(qid, str(column_id), str(answer_id))] = SbsColumnAnswerRow(
                    survey_id=survey_id,
                    qid=qid,
                    column_id=str(column_id),
                    answer_id=str(answer_id),
                    question_type=qtype,
                    export_tag=tag,
                    label_en_md=label_md,
                    label_en_is_html=is_html,
                    externally_managed_by=externally_by,
                )

    return rows


def _extract_base_language(survey_payload: dict) -> str:
    """Extract the base survey language from the payload, defaulting to EN."""
    result = survey_payload.get("result", {})
    if not isinstance(result, dict):
        result = survey_payload
    options = result.get("SurveyOptions") or {}
    lang = _normalize_language_code(options.get("SurveyLanguage") or "")
    return lang or "EN"


@dataclass
class WorkbookOrphanRowsReport:
    """Summary of workbook rows that do not map to the current survey JSON."""

    survey_id: str
    workbook_path: Path
    row_indices_by_sheet: dict[str, list[int]]
    unknown_qids: set[str]
    sample_labels: list[str]

    @property
    def total_rows(self) -> int:
        return sum(len(rows) for rows in self.row_indices_by_sheet.values())

    @property
    def has_orphans(self) -> bool:
        return self.total_rows > 0

    def counts_by_sheet(self) -> dict[str, int]:
        return {
            sheet: len(rows)
            for sheet, rows in self.row_indices_by_sheet.items()
            if rows
        }

    def counts_text(self) -> str:
        counts = self.counts_by_sheet()
        if not counts:
            return "none"
        ordered = sorted(counts.items(), key=lambda item: item[0])
        return ", ".join(f"{sheet}: {count}" for sheet, count in ordered)

    def unknown_qids_text(self, *, limit: int = 8) -> str:
        if not self.unknown_qids:
            return "none"
        ordered = sorted(
            self.unknown_qids,
            key=lambda qid: _qid_order_key(qid),
        )
        if len(ordered) <= limit:
            return ", ".join(ordered)
        remaining = len(ordered) - limit
        return f"{', '.join(ordered[:limit])}, +{remaining} more"

    def sample_text(self, *, limit: int = 3) -> str:
        samples = self.sample_labels[:limit]
        if not samples:
            return ""
        return ", ".join(samples)


def _collect_workbook_orphan_rows(
    wb: Workbook,
    *,
    survey_id: str,
    survey_payload: dict,
    workbook_path: Path,
) -> WorkbookOrphanRowsReport:
    valid_questions = set(build_question_rows(survey_id, survey_payload).keys())
    valid_options = set(build_option_rows(survey_id, survey_payload).keys())
    valid_subitems = set(build_subitem_rows(survey_id, survey_payload).keys())
    valid_sbs_columns = set(build_sbs_column_rows(survey_id, survey_payload).keys())
    valid_sbs_answers = set(build_sbs_column_answer_rows(survey_id, survey_payload).keys())

    row_indices_by_sheet: dict[str, list[int]] = {
        QUESTION_SHEET: [],
        OPTIONS_SHEET: [],
        SUBITEMS_SHEET: [],
        SBS_COLUMNS_SHEET: [],
        SBS_COLUMN_ANSWERS_SHEET: [],
    }
    unknown_qids: set[str] = set()
    sample_labels: list[str] = []

    def _record(sheet: str, row_idx: int, label: str, qid: str | None = None) -> None:
        row_indices_by_sheet[sheet].append(int(row_idx))
        if qid:
            unknown_qids.add(str(qid))
        if len(sample_labels) < 12:
            sample_labels.append(label)

    if QUESTION_SHEET in wb.sheetnames:
        ws = wb[QUESTION_SHEET]
        headers, data_rows = _iter_sheet_rows(ws)
        if headers and "QID" in headers:
            qid_idx = headers.index("QID")
            for row in data_rows:
                qid = str(row[qid_idx].value or "").strip()
                if not qid:
                    continue
                if qid not in valid_questions:
                    _record(
                        QUESTION_SHEET,
                        int(row[0].row),
                        f"{QUESTION_SHEET} row {row[0].row} (QID={qid})",
                        qid=qid,
                    )

    if OPTIONS_SHEET in wb.sheetnames:
        ws = wb[OPTIONS_SHEET]
        headers, data_rows = _iter_sheet_rows(ws)
        if headers and "QID" in headers and "ChoiceId" in headers:
            qid_idx = headers.index("QID")
            choice_idx = headers.index("ChoiceId")
            for row in data_rows:
                qid = str(row[qid_idx].value or "").strip()
                choice_id = str(row[choice_idx].value or "").strip()
                if not qid or not choice_id:
                    continue
                key = (qid, choice_id)
                if key not in valid_options:
                    _record(
                        OPTIONS_SHEET,
                        int(row[0].row),
                        f"{OPTIONS_SHEET} row {row[0].row} (QID={qid}, ChoiceId={choice_id})",
                        qid=qid,
                    )

    if SUBITEMS_SHEET in wb.sheetnames:
        ws = wb[SUBITEMS_SHEET]
        headers, data_rows = _iter_sheet_rows(ws)
        if headers and "QID" in headers and "AnswerId" in headers:
            qid_idx = headers.index("QID")
            answer_idx = headers.index("AnswerId")
            field_idx = headers.index("Field") if "Field" in headers else None
            for row in data_rows:
                qid = str(row[qid_idx].value or "").strip()
                answer_id = str(row[answer_idx].value or "").strip()
                if not qid or not answer_id:
                    continue
                field_val = row[field_idx].value if field_idx is not None else "Answer"
                field = _normalize_subitem_field(field_val)
                key = (qid, field, answer_id)
                if key not in valid_subitems:
                    _record(
                        SUBITEMS_SHEET,
                        int(row[0].row),
                        f"{SUBITEMS_SHEET} row {row[0].row} (QID={qid}, Field={field}, AnswerId={answer_id})",
                        qid=qid,
                    )

    if SBS_COLUMNS_SHEET in wb.sheetnames:
        ws = wb[SBS_COLUMNS_SHEET]
        headers, data_rows = _iter_sheet_rows(ws)
        if headers and "QID" in headers and "ColumnId" in headers:
            qid_idx = headers.index("QID")
            col_idx = headers.index("ColumnId")
            for row in data_rows:
                qid = str(row[qid_idx].value or "").strip()
                column_id = str(row[col_idx].value or "").strip()
                if not qid or not column_id:
                    continue
                key = (qid, column_id)
                if key not in valid_sbs_columns:
                    _record(
                        SBS_COLUMNS_SHEET,
                        int(row[0].row),
                        f"{SBS_COLUMNS_SHEET} row {row[0].row} (QID={qid}, ColumnId={column_id})",
                        qid=qid,
                    )

    if SBS_COLUMN_ANSWERS_SHEET in wb.sheetnames:
        ws = wb[SBS_COLUMN_ANSWERS_SHEET]
        headers, data_rows = _iter_sheet_rows(ws)
        if (
            headers
            and "QID" in headers
            and "ColumnId" in headers
            and "AnswerId" in headers
        ):
            qid_idx = headers.index("QID")
            col_idx = headers.index("ColumnId")
            answer_idx = headers.index("AnswerId")
            for row in data_rows:
                qid = str(row[qid_idx].value or "").strip()
                column_id = str(row[col_idx].value or "").strip()
                answer_id = str(row[answer_idx].value or "").strip()
                if not qid or not column_id or not answer_id:
                    continue
                key = (qid, column_id, answer_id)
                if key not in valid_sbs_answers:
                    _record(
                        SBS_COLUMN_ANSWERS_SHEET,
                        int(row[0].row),
                        (
                            f"{SBS_COLUMN_ANSWERS_SHEET} row {row[0].row} "
                            f"(QID={qid}, ColumnId={column_id}, AnswerId={answer_id})"
                        ),
                        qid=qid,
                    )

    return WorkbookOrphanRowsReport(
        survey_id=survey_id,
        workbook_path=Path(workbook_path),
        row_indices_by_sheet=row_indices_by_sheet,
        unknown_qids=unknown_qids,
        sample_labels=sample_labels,
    )


def inspect_workbook_orphan_rows(
    survey_id: str,
    survey_payload: dict,
    xlsx_path: Path,
) -> WorkbookOrphanRowsReport:
    """Inspect workbook rows that no longer exist in the current survey payload."""

    path = Path(xlsx_path)
    if not path.exists():
        return WorkbookOrphanRowsReport(
            survey_id=survey_id,
            workbook_path=path,
            row_indices_by_sheet={},
            unknown_qids=set(),
            sample_labels=[],
        )

    wb = load_workbook(path, read_only=True, data_only=True)
    return _collect_workbook_orphan_rows(
        wb,
        survey_id=survey_id,
        survey_payload=survey_payload,
        workbook_path=path,
    )


def prune_workbook_orphan_rows(
    survey_id: str,
    survey_payload: dict,
    xlsx_path: Path,
) -> WorkbookOrphanRowsReport:
    """Delete workbook rows that no longer exist in the current survey payload."""

    path = Path(xlsx_path)
    if not path.exists():
        return WorkbookOrphanRowsReport(
            survey_id=survey_id,
            workbook_path=path,
            row_indices_by_sheet={},
            unknown_qids=set(),
            sample_labels=[],
        )

    wb = load_workbook(path)
    report = _collect_workbook_orphan_rows(
        wb,
        survey_id=survey_id,
        survey_payload=survey_payload,
        workbook_path=path,
    )

    if report.has_orphans:
        for sheet_name, row_indices in report.row_indices_by_sheet.items():
            if not row_indices or sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            for row_idx in sorted(set(int(r) for r in row_indices), reverse=True):
                ws.delete_rows(row_idx, 1)
        wb.save(path)

    return report


def init_workbook_from_survey(
    survey_id: str,
    survey_payload: dict,
    xlsx_path: Path,
    *,
    languages: Sequence[str] | None = None,
    prune_orphans: bool = False,
) -> WorkbookOrphanRowsReport | None:
    """Create or update an Excel workbook from a survey JSON payload.

    This is the workbook initializer used by `qsync items pull` (and legacy
    `qsync init`).

    - Ensures the core sheets exist: Questions, Options, Subitems, SBS_Columns,
      SBS_ColumnAnswers, Embedded_Data, System, and Instructions.
    - Adds rows for new QIDs/choices/subitems/SBS columns/SBS answers/embedded fields.
    - Preserves non-empty user-entered Markdown/value cells (`*_MD` and
      `Embedded_Data.Value`).
    - Refreshes system-owned metadata columns and workflow flags (for example
      `*_IsHTML`) from the cached survey JSON.
    - Writes externally managed notes via `MetaComment` (Options and SBS sheets)
      when options/columns/answers are owned by scripts (see `EXTERNALLY_MANAGED_TAGS`).
    - Applies an SBSMatrix migration for legacy/broken workbooks that placed
      SBS statements in the Options sheet (moves them into Subitems).
    - Rebuilds the Instructions sheet with up-to-date column guidance.
    - Optionally prunes orphan rows from item sheets when `prune_orphans=True`.

    Args:
        survey_id: Qualtrics survey ID (e.g., `SV_xxx`).
        survey_payload: Survey JSON payload (as returned by the Qualtrics API).
        xlsx_path: Where to write the workbook.
        languages: Optional list of language codes to add as translation columns.
        prune_orphans: Remove stale workbook rows whose item keys are no longer
            present in the current survey payload.

    Returns:
        `WorkbookOrphanRowsReport` when `prune_orphans=True`, otherwise `None`.

    Example:
        >>> from pathlib import Path
        >>> from qsync.excel_io import init_workbook_from_survey
        >>> from qsync.qualtrics_client import load_cached_survey
        >>> survey = load_cached_survey("SV_xxx")  # requires surveys/*__SV_xxx.json
        >>> init_workbook_from_survey("SV_xxx", survey.payload, Path("excel/workbook.xlsx"))
    """

    xlsx_path = Path(xlsx_path)
    if xlsx_path.exists():
        wb = load_workbook(xlsx_path)
    else:
        wb = Workbook()
        # Remove default sheet if present; we will create our own.
        default_sheet = wb.active
        wb.remove(default_sheet)

    # Ensure no lingering Excel comments survive from older exports.
    _strip_all_comments(wb)

    base_language = _extract_base_language(survey_payload)

    # SBSMatrix migration: older workbooks incorrectly placed SBS statements in Options.
    # Snapshot those values before we re-init sheets so we can move them into Subitems.
    sbs_qids = _sbs_matrix_qids(survey_payload)
    sbs_broken_qids: set[str] = set()
    sbs_option_label_snapshot: dict[tuple[str, str], dict[str, object]] = {}

    if sbs_qids and OPTIONS_SHEET in wb.sheetnames:
        ws_opt = wb[OPTIONS_SHEET]
        opt_headers, opt_rows = _iter_sheet_rows(ws_opt)
        if opt_headers and "QID" in opt_headers and "ChoiceId" in opt_headers:
            opt_qid_idx = opt_headers.index("QID")
            opt_choice_idx = opt_headers.index("ChoiceId")
            option_qids_with_rows: set[str] = set()
            for row in opt_rows:
                qid = str(row[opt_qid_idx].value or "").strip()
                cid = str(row[opt_choice_idx].value or "").strip()
                if qid in sbs_qids and cid:
                    option_qids_with_rows.add(qid)

            subitems_qids_with_rows: set[str] = set()
            if SUBITEMS_SHEET in wb.sheetnames:
                ws_sub = wb[SUBITEMS_SHEET]
                sub_headers, sub_rows = _iter_sheet_rows(ws_sub)
                if sub_headers and "QID" in sub_headers and "AnswerId" in sub_headers:
                    sub_qid_idx = sub_headers.index("QID")
                    sub_ans_idx = sub_headers.index("AnswerId")
                    sub_field_idx = (
                        sub_headers.index("Field") if "Field" in sub_headers else None
                    )
                    for row in sub_rows:
                        qid = str(row[sub_qid_idx].value or "").strip()
                        aid = str(row[sub_ans_idx].value or "").strip()
                        if qid not in sbs_qids or not aid:
                            continue
                        field_val = (
                            row[sub_field_idx].value
                            if sub_field_idx is not None
                            else ""
                        )
                        field = _normalize_subitem_field(field_val)
                        if field != "Label":
                            subitems_qids_with_rows.add(qid)

            sbs_broken_qids = option_qids_with_rows - subitems_qids_with_rows

            label_cols = [
                str(h or "")
                for h in opt_headers
                if str(h or "").startswith("Label_")
                and (str(h or "").endswith("_MD") or str(h or "").endswith("_IsHTML"))
            ]
            idx = {str(name or ""): i for i, name in enumerate(opt_headers)}
            if sbs_broken_qids and label_cols:
                for row in opt_rows:
                    qid = str(row[opt_qid_idx].value or "").strip()
                    cid = str(row[opt_choice_idx].value or "").strip()
                    if qid not in sbs_broken_qids or not cid:
                        continue
                    snapshot: dict[str, object] = {}
                    for col in label_cols:
                        j = idx.get(col)
                        if j is None or j >= len(row):
                            continue
                        snapshot[col] = row[j].value
                    sbs_option_label_snapshot[(qid, cid)] = snapshot

    questions_map = build_question_rows(survey_id, survey_payload)
    options_map = build_option_rows(survey_id, survey_payload)
    subitems_map = build_subitem_rows(survey_id, survey_payload)
    sbs_columns_map = build_sbs_column_rows(survey_id, survey_payload)
    sbs_column_answers_map = build_sbs_column_answer_rows(survey_id, survey_payload)
    embedded_rows = build_embedded_data_rows(survey_id, survey_payload)
    option_previews = _build_option_previews(survey_payload)
    subitem_previews = _build_subitem_previews(survey_payload)
    flow_meta_by_qid = _question_flow_meta_by_qid(survey_payload)

    _init_questions_sheet(
        wb,
        questions_map,
        option_previews,
        subitem_previews,
        flow_meta_by_qid,
        survey_payload,
        languages=languages,
        base_language=base_language,
    )
    _init_options_sheet(
        wb,
        options_map,
        flow_meta_by_qid,
        survey_payload,
        languages=languages,
        base_language=base_language,
    )
    _init_subitems_sheet(
        wb,
        subitems_map,
        flow_meta_by_qid,
        survey_payload,
        languages=languages,
        base_language=base_language,
    )
    _init_sbs_columns_sheet(
        wb,
        sbs_columns_map,
        flow_meta_by_qid,
        survey_payload,
        languages=languages,
        base_language=base_language,
    )
    _init_sbs_column_answers_sheet(
        wb,
        sbs_column_answers_map,
        flow_meta_by_qid,
        survey_payload,
        languages=languages,
        base_language=base_language,
    )
    _init_survey_metadata_sheet(wb, survey_payload, languages=languages)
    _init_embedded_data_sheet(wb, embedded_rows)

    # SBSMatrix migration: move SBS statements out of Options and into Subitems.
    # We only merge values when the workbook was in the "broken" state (Options had rows,
    # Subitems did not) to avoid clobbering real Subitems edits.
    if sbs_qids and OPTIONS_SHEET in wb.sheetnames and SUBITEMS_SHEET in wb.sheetnames:
        if sbs_option_label_snapshot:
            ws_sub = wb[SUBITEMS_SHEET]
            sub_headers, _ = _iter_sheet_rows(ws_sub)
            if sub_headers and "QID" in sub_headers and "AnswerId" in sub_headers:
                sub_idx = {str(name or ""): i for i, name in enumerate(sub_headers)}
                snapshot_cols: set[str] = set()
                for snap in sbs_option_label_snapshot.values():
                    snapshot_cols.update([str(k) for k in snap.keys()])
                label_cols = [str(h or "") for h in sub_headers if str(h or "") in snapshot_cols]

                # Build a row lookup for Subitems (Answer field only).
                sub_row_for_key: dict[tuple[str, str], int] = {}
                sub_qid_idx = sub_idx["QID"]
                sub_ans_idx = sub_idx["AnswerId"]
                sub_field_idx = sub_idx.get("Field")
                for row in ws_sub.iter_rows(min_row=2, values_only=False):
                    qid = str(row[sub_qid_idx].value or "").strip()
                    aid = str(row[sub_ans_idx].value or "").strip()
                    if not qid or not aid:
                        continue
                    field_val = (
                        row[sub_field_idx].value if sub_field_idx is not None else ""
                    )
                    field = _normalize_subitem_field(field_val)
                    if field == "Label":
                        continue
                    sub_row_for_key[(qid, aid)] = int(row[0].row)

                for (qid, cid), snap in sbs_option_label_snapshot.items():
                    target_row = sub_row_for_key.get((qid, cid))
                    if not target_row:
                        continue
                    for col in label_cols:
                        val = snap.get(col)
                        if val is None or str(val).strip() == "":
                            continue
                        ws_sub.cell(row=target_row, column=sub_idx[col] + 1, value=val)

        # Always delete SBS option rows: SBS does not use the Options sheet.
        ws_opt = wb[OPTIONS_SHEET]
        opt_headers, _ = _iter_sheet_rows(ws_opt)
        if opt_headers and "QID" in opt_headers:
            opt_qid_idx = opt_headers.index("QID")
            delete_rows: list[int] = []
            for row in ws_opt.iter_rows(min_row=2, values_only=False):
                qid = str(row[opt_qid_idx].value or "").strip()
                if qid in sbs_qids:
                    delete_rows.append(int(row[0].row))
            for row_idx in sorted(set(delete_rows), reverse=True):
                ws_opt.delete_rows(row_idx, 1)

    # Normalise ordering so item sheets follow SurveyFlow question order first
    # (fallback: numeric QID), then per-row IDs. This keeps previews and manual
    # inspection predictable, even after multiple init runs.
    _sort_questions_sheet(wb[QUESTION_SHEET])
    _sort_sheet_by_qid_and_id(wb[OPTIONS_SHEET], "ChoiceId")
    _sort_sheet_by_qid_and_id(wb[SUBITEMS_SHEET], "AnswerId")
    _sort_sheet_by_qid_and_id(wb[SBS_COLUMNS_SHEET], "ColumnId")
    _sort_sheet_by_qid_and_two_ids(wb[SBS_COLUMN_ANSWERS_SHEET], "ColumnId", "AnswerId")
    _sort_sheet_by_flow_order(wb[EMBEDDED_DATA_SHEET])

    # Apply table styles, wrapping, colours, and validations.
    _format_questions_sheet(wb[QUESTION_SHEET])
    _format_options_sheet(wb[OPTIONS_SHEET])
    _format_subitems_sheet(wb[SUBITEMS_SHEET])
    _format_sbs_columns_sheet(wb[SBS_COLUMNS_SHEET])
    _format_sbs_column_answers_sheet(wb[SBS_COLUMN_ANSWERS_SHEET])
    _format_survey_metadata_sheet(wb[SURVEY_METADATA_SHEET])
    _format_embedded_data_sheet(wb[EMBEDDED_DATA_SHEET])
    _set_optional_sheet_visibility(wb[SBS_COLUMNS_SHEET], hide_when_empty=True)
    _set_optional_sheet_visibility(wb[SBS_COLUMN_ANSWERS_SHEET], hide_when_empty=True)
    _set_optional_sheet_visibility(wb[EMBEDDED_DATA_SHEET], hide_when_empty=True)

    # Optional: add a System sheet for inspection of Timing/meta options.
    _populate_system_sheet(wb, survey_id, survey_payload)

    # Document the workbook layout so we no longer rely on Excel comments.
    _update_translation_key_map(wb, questions_map, options_map, subitems_map)
    _update_instructions_sheet(wb, languages=languages, base_language=base_language)

    orphan_report: WorkbookOrphanRowsReport | None = None
    if prune_orphans:
        orphan_report = _collect_workbook_orphan_rows(
            wb,
            survey_id=survey_id,
            survey_payload=survey_payload,
            workbook_path=xlsx_path,
        )
        if orphan_report.has_orphans:
            for sheet_name, row_indices in orphan_report.row_indices_by_sheet.items():
                if not row_indices or sheet_name not in wb.sheetnames:
                    continue
                ws = wb[sheet_name]
                for row_idx in sorted(set(int(r) for r in row_indices), reverse=True):
                    ws.delete_rows(row_idx, 1)

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
    return orphan_report


def _init_questions_sheet(
    wb: Workbook,
    questions_map: Dict[str, QuestionRow],
    option_previews: Dict[str, str],
    subitem_previews: Dict[str, str],
    flow_meta_by_qid: Dict[str, _QuestionFlowMeta],
    survey_payload: dict,
    *,
    languages: Sequence[str] | None = None,
    base_language: str = "EN",
) -> None:
    ws = _get_or_create_sheet(wb, QUESTION_SHEET)

    base_text_col = _question_text_md_column(base_language)
    base_html_col = _question_text_ishtml_column(base_language)

    text_columns = _question_text_columns(languages, base_language=base_language)
    required_cols = [
        "SurveyID",
        "QID",
        "BlockName",
        "BlockID",
        "BlockOrder",
        "QuestionOrder",
        "QuestionOrderInBlock",
        "QuestionType",
        "DataExportTag",
        "RequiredResponse",
        "ForceResponseMode",
        "ValidationType",
        "ValidationSettingsJSON",
        "RandomizationType",
        "RandomizationSettingsJSON",
        QUESTION_CONFIG_JSON_COLUMN,
        *text_columns,
        "OptionsPreview",
        "SubitemsPreview",
    ]
    # Legacy clean-up: remove deprecated QuestionKey column.
    headers, _ = _iter_sheet_rows(ws)
    if headers and "QuestionKey" in headers:
        ws.delete_cols(headers.index("QuestionKey") + 1)
    col_index = _ensure_columns(ws, required_cols)
    _migrate_legacy_question_config_columns(ws)
    _migrate_legacy_question_text_columns(
        ws,
        languages=languages,
        base_language=base_language,
    )
    _drop_stale_question_text_columns(ws, required_cols)
    col_index = _ensure_columns(ws, required_cols)

    # Build index of existing rows by QID
    headers, data_rows = _iter_sheet_rows(ws)
    qid_col = col_index["QID"]
    existing_rows: Dict[str, int] = {}
    for idx, row in enumerate(data_rows, start=2):
        cell = row[qid_col]
        qid = str(cell.value).strip() if cell.value is not None else ""
        if qid:
            existing_rows[qid] = idx

    questions = survey_payload.get("result", {}).get("Questions", {})
    text_lang_columns = _question_text_lang_columns_from_headers(
        headers,
        include_legacy=True,
    )

    for qid, row_data in questions_map.items():
        q_json = questions.get(qid, {}) if isinstance(questions, dict) else {}
        language_blocks = q_json.get("Language") or {}

        if qid in existing_rows:
            row_idx = existing_rows[qid]
            # Update read-only metadata; do not touch user-managed cells unless blank.
            ws.cell(
                row=row_idx, column=col_index["SurveyID"] + 1, value=row_data.survey_id
            )
            ws.cell(row=row_idx, column=col_index["QID"] + 1, value=row_data.qid)
            _write_flow_meta_cells(
                ws,
                row_idx,
                col_index,
                row_data.qid,
                flow_meta_by_qid,
            )
            ws.cell(
                row=row_idx,
                column=col_index["QuestionType"] + 1,
                value=row_data.question_type,
            )
            ws.cell(
                row=row_idx,
                column=col_index["DataExportTag"] + 1,
                value=row_data.data_export_tag,
            )
            force_mode_cell = ws.cell(
                row=row_idx, column=col_index["ForceResponseMode"] + 1
            )
            if force_mode_cell.value is None or str(force_mode_cell.value).strip() == "":
                force_mode_cell.value = row_data.force_response_mode
            validation_type_cell = ws.cell(
                row=row_idx, column=col_index["ValidationType"] + 1
            )
            if (
                validation_type_cell.value is None
                or str(validation_type_cell.value).strip() == ""
            ):
                validation_type_cell.value = row_data.validation_type
            validation_settings_cell = ws.cell(
                row=row_idx, column=col_index["ValidationSettingsJSON"] + 1
            )
            if (
                validation_settings_cell.value is None
                or str(validation_settings_cell.value).strip() == ""
            ) and row_data.validation_settings_json:
                validation_settings_cell.value = row_data.validation_settings_json
            randomization_type_cell = ws.cell(
                row=row_idx, column=col_index["RandomizationType"] + 1
            )
            if (
                randomization_type_cell.value is None
                or str(randomization_type_cell.value).strip() == ""
            ):
                randomization_type_cell.value = row_data.randomization_type
            randomization_settings_cell = ws.cell(
                row=row_idx, column=col_index["RandomizationSettingsJSON"] + 1
            )
            if (
                randomization_settings_cell.value is None
                or str(randomization_settings_cell.value).strip() == ""
            ) and row_data.randomization_settings_json:
                randomization_settings_cell.value = (
                    row_data.randomization_settings_json
                )
            config_cell = ws.cell(
                row=row_idx, column=col_index[QUESTION_CONFIG_JSON_COLUMN] + 1
            )
            config_cell.value = _build_question_config_json_from_legacy_values(
                force_response_mode=force_mode_cell.value,
                validation_type=validation_type_cell.value,
                validation_settings_json=validation_settings_cell.value,
                randomization_type=randomization_type_cell.value,
                randomization_settings_json=randomization_settings_cell.value,
            )
            effective_force_mode = (
                force_mode_cell.value
                if force_mode_cell.value is not None
                else row_data.force_response_mode
            )
            ws.cell(
                row=row_idx,
                column=col_index["RequiredResponse"] + 1,
                value=_is_required_response(effective_force_mode),
            )

            text_cell = ws.cell(row=row_idx, column=col_index[base_text_col] + 1)
            is_html_cell = ws.cell(row=row_idx, column=col_index[base_html_col] + 1)
            if (
                text_cell.value is None or str(text_cell.value).strip() == ""
            ) and row_data.text_en_md:
                text_cell.value = row_data.text_en_md
            is_html_cell.value = bool(row_data.text_en_is_html)

            for lang_code, (md_col, html_col) in text_lang_columns.items():
                if lang_code == _normalize_language_code(base_language):
                    continue
                lang_block = _lookup_language_block(language_blocks, lang_code)
                if not lang_block:
                    continue
                lang_text = lang_block.get("QuestionText")
                text_md, is_html = _metadata_cell_value(lang_text)
                if not text_md:
                    continue
                lang_cell = ws.cell(row=row_idx, column=col_index[md_col] + 1)
                if lang_cell.value is not None and str(lang_cell.value).strip() != "":
                    continue
                lang_cell.value = text_md
                if html_col:
                    ws.cell(
                        row=row_idx,
                        column=col_index[html_col] + 1,
                        value=bool(is_html),
                    )
            if "OptionsPreview" in col_index:
                col_letter = get_column_letter(col_index["OptionsPreview"] + 1)
                cell_ref = f"{col_letter}{row_idx}"
                ws.cell(
                    row=row_idx,
                    column=col_index["OptionsPreview"] + 1,
                    value=_make_options_preview_formula(
                        cell_ref,
                        base_language=base_language,
                        question_type=row_data.question_type,
                    ),
                )
            if "SubitemsPreview" in col_index:
                col_letter = get_column_letter(col_index["SubitemsPreview"] + 1)
                cell_ref = f"{col_letter}{row_idx}"
                ws.cell(
                    row=row_idx,
                    column=col_index["SubitemsPreview"] + 1,
                    value=_make_subitems_preview_formula(
                        cell_ref, base_language=base_language
                    ),
                )
        else:
            # Append new row
            new_row_idx = ws.max_row + 1
            ws.cell(
                row=new_row_idx,
                column=col_index["SurveyID"] + 1,
                value=row_data.survey_id,
            )
            ws.cell(row=new_row_idx, column=col_index["QID"] + 1, value=row_data.qid)
            _write_flow_meta_cells(
                ws,
                new_row_idx,
                col_index,
                row_data.qid,
                flow_meta_by_qid,
            )
            ws.cell(
                row=new_row_idx,
                column=col_index["QuestionType"] + 1,
                value=row_data.question_type,
            )
            ws.cell(
                row=new_row_idx,
                column=col_index["DataExportTag"] + 1,
                value=row_data.data_export_tag,
            )
            ws.cell(
                row=new_row_idx,
                column=col_index["RequiredResponse"] + 1,
                value=bool(row_data.required_response),
            )
            ws.cell(
                row=new_row_idx,
                column=col_index["ForceResponseMode"] + 1,
                value=row_data.force_response_mode,
            )
            ws.cell(
                row=new_row_idx,
                column=col_index["ValidationType"] + 1,
                value=row_data.validation_type,
            )
            if row_data.validation_settings_json:
                ws.cell(
                    row=new_row_idx,
                    column=col_index["ValidationSettingsJSON"] + 1,
                    value=row_data.validation_settings_json,
                )
            ws.cell(
                row=new_row_idx,
                column=col_index["RandomizationType"] + 1,
                value=row_data.randomization_type,
            )
            if row_data.randomization_settings_json:
                ws.cell(
                    row=new_row_idx,
                    column=col_index["RandomizationSettingsJSON"] + 1,
                    value=row_data.randomization_settings_json,
                )
            ws.cell(
                row=new_row_idx,
                column=col_index[QUESTION_CONFIG_JSON_COLUMN] + 1,
                value=row_data.question_config_json,
            )

            text_cell = ws.cell(
                row=new_row_idx,
                column=col_index[base_text_col] + 1,
                value=row_data.text_en_md,
            )
            is_html_cell = ws.cell(
                row=new_row_idx,
                column=col_index[base_html_col] + 1,
                value=bool(row_data.text_en_is_html),
            )

            for lang_code, (md_col, html_col) in text_lang_columns.items():
                if lang_code == _normalize_language_code(base_language):
                    continue
                lang_block = _lookup_language_block(language_blocks, lang_code)
                if not lang_block:
                    continue
                lang_text = lang_block.get("QuestionText")
                text_md, is_html = _metadata_cell_value(lang_text)
                if not text_md:
                    continue
                lang_cell = ws.cell(row=new_row_idx, column=col_index[md_col] + 1)
                if lang_cell.value is not None and str(lang_cell.value).strip() != "":
                    continue
                lang_cell.value = text_md
                if html_col:
                    ws.cell(
                        row=new_row_idx,
                        column=col_index[html_col] + 1,
                        value=bool(is_html),
                    )
            # Routing flags start empty/False

            if "OptionsPreview" in col_index:
                col_letter = get_column_letter(col_index["OptionsPreview"] + 1)
                cell_ref = f"{col_letter}{new_row_idx}"
                ws.cell(
                    row=new_row_idx,
                    column=col_index["OptionsPreview"] + 1,
                    value=_make_options_preview_formula(
                        cell_ref,
                        base_language=base_language,
                        question_type=row_data.question_type,
                    ),
                )
            if "SubitemsPreview" in col_index:
                col_letter = get_column_letter(col_index["SubitemsPreview"] + 1)
                cell_ref = f"{col_letter}{new_row_idx}"
                ws.cell(
                    row=new_row_idx,
                    column=col_index["SubitemsPreview"] + 1,
                    value=_make_subitems_preview_formula(
                        cell_ref, base_language=base_language
                    ),
                )


def _init_options_sheet(
    wb: Workbook,
    options_map: Dict[Tuple[str, str], OptionRow],
    flow_meta_by_qid: Dict[str, _QuestionFlowMeta],
    survey_payload: dict,
    *,
    languages: Sequence[str] | None = None,
    base_language: str = "EN",
) -> None:
    ws = _get_or_create_sheet(wb, OPTIONS_SHEET)
    base_suffix = _language_suffix(base_language) or "en"
    base_label_col = f"Label_{base_suffix}_MD"
    base_label_html_col = f"Label_{base_suffix}_IsHTML"
    label_columns = _translation_columns(
        "Label", languages, base_language=base_language
    )
    required_cols = [
        "SurveyID",
        "QID",
        "BlockName",
        "BlockID",
        "BlockOrder",
        "QuestionOrder",
        "QuestionOrderInBlock",
        "ChoiceId",
        "QuestionType",
        "ExportTag",
        "Code",
        *label_columns,
        "MetaComment",
    ]
    # Legacy clean-up: drop deprecated DisplayPreview column if it exists.
    headers, _ = _iter_sheet_rows(ws)
    if headers and "DisplayPreview" in headers:
        disp_idx = headers.index("DisplayPreview") + 1
        ws.delete_cols(disp_idx)

    _drop_stale_translation_columns(ws, required_cols, prefixes=["Label"])
    col_index = _ensure_columns(ws, required_cols)

    headers, data_rows = _iter_sheet_rows(ws)
    label_lang_columns: dict[str, tuple[str, str | None]] = {}
    for name in headers:
        header = str(name or "")
        if not header.startswith("Label_") or not header.endswith("_MD"):
            continue
        suffix = header[len("Label_") : -len("_MD")]
        lang_code = _language_from_suffix(suffix)
        is_html_name = f"Label_{suffix}_IsHTML"
        label_lang_columns[lang_code] = (
            header,
            is_html_name if is_html_name in col_index else None,
        )

    qid_col = col_index["QID"]
    choice_col = col_index["ChoiceId"]
    existing_rows: Dict[Tuple[str, str], int] = {}
    for idx, row in enumerate(data_rows, start=2):
        qid_val = row[qid_col].value
        choice_val = row[choice_col].value
        if qid_val is None or choice_val is None:
            continue
        key = (str(qid_val).strip(), str(choice_val).strip())
        existing_rows[key] = idx

    # Build a lookup of question types for system-only routing later
    questions = survey_payload.get("result", {}).get("Questions", {})

    for key, row_data in options_map.items():
        qid, choice_id = key
        # Skip options for pure Timing/meta questions; they go to System sheet.
        q_json = questions.get(qid, {})
        qtype = q_json.get("QuestionType")
        if qtype == "Timing":
            continue
        if key in existing_rows:
            row_idx = existing_rows[key]
            ws.cell(
                row=row_idx, column=col_index["SurveyID"] + 1, value=row_data.survey_id
            )
            ws.cell(row=row_idx, column=col_index["QID"] + 1, value=row_data.qid)
            _write_flow_meta_cells(
                ws,
                row_idx,
                col_index,
                row_data.qid,
                flow_meta_by_qid,
            )
            ws.cell(
                row=row_idx, column=col_index["ChoiceId"] + 1, value=row_data.choice_id
            )
            ws.cell(
                row=row_idx,
                column=col_index["QuestionType"] + 1,
                value=row_data.question_type,
            )
            ws.cell(
                row=row_idx,
                column=col_index["ExportTag"] + 1,
                value=row_data.export_tag,
            )
            if row_data.code is not None:
                ws.cell(row=row_idx, column=col_index["Code"] + 1, value=row_data.code)

            label_cell = ws.cell(row=row_idx, column=col_index[base_label_col] + 1)
            is_html_cell = ws.cell(
                row=row_idx, column=col_index[base_label_html_col] + 1
            )
            if (
                label_cell.value is None or str(label_cell.value).strip() == ""
            ) and row_data.label_en_md:
                label_cell.value = row_data.label_en_md
            is_html_cell.value = bool(row_data.label_en_is_html)

            meta_cell = ws.cell(row=row_idx, column=col_index["MetaComment"] + 1)
            if row_data.externally_managed_by:
                meta_cell.value = _externally_managed_note(
                    row_data.externally_managed_by
                )
            else:
                current = str(meta_cell.value or "").strip()
                if current.startswith(_EXTERNALLY_MANAGED_NOTE_PREFIX):
                    meta_cell.value = ""
        else:
            new_row_idx = ws.max_row + 1
            ws.cell(
                row=new_row_idx,
                column=col_index["SurveyID"] + 1,
                value=row_data.survey_id,
            )
            ws.cell(row=new_row_idx, column=col_index["QID"] + 1, value=row_data.qid)
            _write_flow_meta_cells(
                ws,
                new_row_idx,
                col_index,
                row_data.qid,
                flow_meta_by_qid,
            )
            ws.cell(
                row=new_row_idx,
                column=col_index["ChoiceId"] + 1,
                value=row_data.choice_id,
            )
            ws.cell(
                row=new_row_idx,
                column=col_index["QuestionType"] + 1,
                value=row_data.question_type,
            )
            ws.cell(
                row=new_row_idx,
                column=col_index["ExportTag"] + 1,
                value=row_data.export_tag,
            )
            if row_data.code is not None:
                ws.cell(
                    row=new_row_idx, column=col_index["Code"] + 1, value=row_data.code
                )

            label_cell = ws.cell(
                row=new_row_idx,
                column=col_index[base_label_col] + 1,
                value=row_data.label_en_md,
            )
            is_html_cell = ws.cell(
                row=new_row_idx,
                column=col_index[base_label_html_col] + 1,
                value=bool(row_data.label_en_is_html),
            )
            meta_cell = ws.cell(row=new_row_idx, column=col_index["MetaComment"] + 1)
            if row_data.externally_managed_by:
                meta_cell.value = _externally_managed_note(
                    row_data.externally_managed_by
                )

            row_idx = new_row_idx

        q_json = questions.get(qid, {}) if isinstance(questions, dict) else {}
        language_blocks = q_json.get("Language") or {}
        section = "Choices"
        if str(row_data.question_type or "").strip().lower() == "matrix":
            section = "Answers"

        for lang_code, (md_col, html_col) in label_lang_columns.items():
            if lang_code == _normalize_language_code(base_language):
                continue
            lang_block = _lookup_language_block(language_blocks, lang_code)
            if not lang_block:
                continue
            items = lang_block.get(section) if isinstance(lang_block, dict) else None
            if not isinstance(items, dict):
                continue
            entry = items.get(str(choice_id))
            if not isinstance(entry, dict):
                continue
            lang_display = entry.get("Display")
            text_md, is_html = _metadata_cell_value(lang_display)
            if not text_md:
                continue
            lang_cell = ws.cell(row=row_idx, column=col_index[md_col] + 1)
            if lang_cell.value is not None and str(lang_cell.value).strip() != "":
                continue
            lang_cell.value = text_md
            if html_col:
                ws.cell(
                    row=row_idx,
                    column=col_index[html_col] + 1,
                    value=bool(is_html),
                )


def _init_subitems_sheet(
    wb: Workbook,
    subitems_map: Dict[Tuple[str, str, str], SubitemRow],
    flow_meta_by_qid: Dict[str, _QuestionFlowMeta],
    survey_payload: dict,
    *,
    languages: Sequence[str] | None = None,
    base_language: str = "EN",
) -> None:
    ws = _get_or_create_sheet(wb, SUBITEMS_SHEET)
    survey_id = str((survey_payload.get("result") or {}).get("SurveyID") or "").strip()
    base_suffix = _language_suffix(base_language) or "en"
    base_label_col = f"Label_{base_suffix}_MD"
    base_label_html_col = f"Label_{base_suffix}_IsHTML"
    label_columns = _translation_columns(
        "Label", languages, base_language=base_language
    )
    required_cols = [
        "SurveyID",
        "QID",
        "BlockName",
        "BlockID",
        "BlockOrder",
        "QuestionOrder",
        "QuestionOrderInBlock",
        "AnswerId",
        "Field",
        "QuestionType",
        "ExportTag",
        *label_columns,
    ]
    _drop_stale_translation_columns(ws, required_cols, prefixes=["Label"])
    col_index = _ensure_columns(ws, required_cols)

    headers, data_rows = _iter_sheet_rows(ws)
    label_lang_columns: dict[str, tuple[str, str | None]] = {}
    for name in headers:
        header = str(name or "")
        if not header.startswith("Label_") or not header.endswith("_MD"):
            continue
        suffix = header[len("Label_") : -len("_MD")]
        lang_code = _language_from_suffix(suffix)
        is_html_name = f"Label_{suffix}_IsHTML"
        label_lang_columns[lang_code] = (
            header,
            is_html_name if is_html_name in col_index else None,
        )

    qid_col = col_index["QID"]
    answer_col = col_index["AnswerId"]
    field_col = col_index.get("Field")
    existing_rows: Dict[Tuple[str, str, str], int] = {}
    for idx, row in enumerate(data_rows, start=2):
        qid_val = row[qid_col].value
        answer_val = row[answer_col].value
        if qid_val is None or answer_val is None:
            continue
        field_val = row[field_col].value if field_col is not None else "Answer"
        field = _normalize_subitem_field(field_val)
        key = (str(qid_val).strip(), field, str(answer_val).strip())
        existing_rows[key] = idx

    questions = survey_payload.get("result", {}).get("Questions", {})

    for key, row_data in subitems_map.items():
        qid, field_key, answer_id = key
        field_key = _normalize_subitem_field(field_key)
        row_key = (qid, field_key, answer_id)
        if row_key in existing_rows:
            row_idx = existing_rows[row_key]
            ws.cell(
                row=row_idx, column=col_index["SurveyID"] + 1, value=row_data.survey_id
            )
            ws.cell(row=row_idx, column=col_index["QID"] + 1, value=row_data.qid)
            _write_flow_meta_cells(
                ws,
                row_idx,
                col_index,
                row_data.qid,
                flow_meta_by_qid,
            )
            ws.cell(
                row=row_idx, column=col_index["AnswerId"] + 1, value=row_data.answer_id
            )
            if field_col is not None:
                field_cell = ws.cell(row=row_idx, column=col_index["Field"] + 1)
                if field_cell.value is None or str(field_cell.value).strip() == "":
                    field_cell.value = field_key
            ws.cell(
                row=row_idx,
                column=col_index["QuestionType"] + 1,
                value=row_data.question_type,
            )
            ws.cell(
                row=row_idx,
                column=col_index["ExportTag"] + 1,
                value=row_data.export_tag,
            )
            label_cell = ws.cell(row=row_idx, column=col_index[base_label_col] + 1)
            is_html_cell = ws.cell(
                row=row_idx, column=col_index[base_label_html_col] + 1
            )
            if (
                label_cell.value is None or str(label_cell.value).strip() == ""
            ) and row_data.label_en_md:
                label_cell.value = row_data.label_en_md
            is_html_cell.value = bool(row_data.label_en_is_html)
        else:
            new_row_idx = ws.max_row + 1
            existing_rows[row_key] = new_row_idx
            ws.cell(
                row=new_row_idx,
                column=col_index["SurveyID"] + 1,
                value=row_data.survey_id,
            )
            ws.cell(row=new_row_idx, column=col_index["QID"] + 1, value=row_data.qid)
            _write_flow_meta_cells(
                ws,
                new_row_idx,
                col_index,
                row_data.qid,
                flow_meta_by_qid,
            )
            ws.cell(
                row=new_row_idx,
                column=col_index["AnswerId"] + 1,
                value=row_data.answer_id,
            )
            ws.cell(
                row=new_row_idx,
                column=col_index["Field"] + 1,
                value=field_key,
            )
            ws.cell(
                row=new_row_idx,
                column=col_index["QuestionType"] + 1,
                value=row_data.question_type,
            )
            ws.cell(
                row=new_row_idx,
                column=col_index["ExportTag"] + 1,
                value=row_data.export_tag,
            )
            ws.cell(
                row=new_row_idx,
                column=col_index[base_label_col] + 1,
                value=row_data.label_en_md,
            )
            ws.cell(
                row=new_row_idx,
                column=col_index[base_label_html_col] + 1,
                value=bool(row_data.label_en_is_html),
            )

            row_idx = new_row_idx

        q_json = questions.get(qid, {}) if isinstance(questions, dict) else {}
        language_blocks = q_json.get("Language") or {}
        section = "Answers"
        if field_key == "Label":
            section = "Labels"
        elif str(row_data.question_type or "").strip().lower() == "matrix" or _is_sbs_matrix_question(
            q_json
        ):
            section = "Choices"

        for lang_code, (md_col, html_col) in label_lang_columns.items():
            if lang_code == _normalize_language_code(base_language):
                continue
            lang_block = _lookup_language_block(language_blocks, lang_code)
            if not lang_block:
                continue
            items = lang_block.get(section) if isinstance(lang_block, dict) else None
            if not isinstance(items, dict):
                continue
            entry = items.get(str(answer_id))
            if not isinstance(entry, dict):
                continue
            lang_display = entry.get("Display")
            text_md, is_html = _metadata_cell_value(lang_display)
            if not text_md:
                continue
            lang_cell = ws.cell(row=row_idx, column=col_index[md_col] + 1)
            if lang_cell.value is not None and str(lang_cell.value).strip() != "":
                continue
            lang_cell.value = text_md
            if html_col:
                ws.cell(
                    row=row_idx,
                    column=col_index[html_col] + 1,
                    value=bool(is_html),
                )

    # Backfill label endpoints for slider/scale questions (Field=Label).
    # Keep this aligned with the rest of the workbook surfaces by restricting
    # to QIDs present in SurveyFlow (ordered_qids_in_flow).
    for qid in _ordered_qids_in_flow(survey_payload):
        q = questions.get(qid)
        if not isinstance(q, dict):
            continue
        labels = q.get("Labels") or {}
        if not labels:
            continue
        tag = q.get("DataExportTag") or ""
        qtype = q.get("QuestionType") or ""
        language_blocks = q.get("Language") or {}

        for label_id, label in labels.items():
            label_key = (qid, "Label", str(label_id))
            if label_key in existing_rows:
                row_idx = existing_rows[label_key]
            else:
                row_idx = ws.max_row + 1
                existing_rows[label_key] = row_idx
                ws.cell(row=row_idx, column=col_index["SurveyID"] + 1, value=survey_id)
                ws.cell(row=row_idx, column=col_index["QID"] + 1, value=qid)
                _write_flow_meta_cells(
                    ws,
                    row_idx,
                    col_index,
                    qid,
                    flow_meta_by_qid,
                )
                ws.cell(
                    row=row_idx,
                    column=col_index["AnswerId"] + 1,
                    value=str(label_id),
                )
                ws.cell(
                    row=row_idx,
                    column=col_index["Field"] + 1,
                    value="Label",
                )
                ws.cell(
                    row=row_idx,
                    column=col_index["QuestionType"] + 1,
                    value=qtype,
                )
                ws.cell(
                    row=row_idx,
                    column=col_index["ExportTag"] + 1,
                    value=tag,
                )
            _write_flow_meta_cells(
                ws,
                row_idx,
                col_index,
                qid,
                flow_meta_by_qid,
            )

            # Always enforce Field=Label for label rows
            if "Field" in col_index:
                ws.cell(row=row_idx, column=col_index["Field"] + 1, value="Label")

            base_display = label.get("Display")
            base_text = str(base_display) if base_display is not None else ""
            if base_text:
                if is_markdown_safe_html(base_text):
                    base_md, base_is_html = html_to_md(base_text), False
                elif should_treat_as_html(base_text):
                    base_md, base_is_html = normalize_text(base_text), True
                else:
                    base_md, base_is_html = html_to_md(base_text), False

                base_cell = ws.cell(row=row_idx, column=col_index[base_label_col] + 1)
                base_html_cell = ws.cell(
                    row=row_idx, column=col_index[base_label_html_col] + 1
                )
                if base_cell.value is None or str(base_cell.value).strip() == "":
                    base_cell.value = base_md
                    base_html_cell.value = bool(base_is_html)

            # Non-base languages: fill only when columns exist + cell empty
            for lang_code, (md_col, html_col) in label_lang_columns.items():
                if lang_code == _normalize_language_code(base_language):
                    continue
                lang_block = _lookup_language_block(language_blocks, lang_code)
                if not lang_block:
                    continue
                lang_labels = lang_block.get("Labels") or {}
                lang_entry = lang_labels.get(str(label_id)) or {}
                lang_display = lang_entry.get("Display")
                if lang_display is None:
                    continue
                lang_text = str(lang_display)
                lang_cell = ws.cell(row=row_idx, column=col_index[md_col] + 1)
                if lang_cell.value is not None and str(lang_cell.value).strip() != "":
                    continue
                if is_markdown_safe_html(lang_text):
                    lang_md, lang_is_html = html_to_md(lang_text), False
                elif should_treat_as_html(lang_text):
                    lang_md, lang_is_html = normalize_text(lang_text), True
                else:
                    lang_md, lang_is_html = html_to_md(lang_text), False
                lang_cell.value = lang_md
                if html_col:
                    ws.cell(
                        row=row_idx,
                        column=col_index[html_col] + 1,
                        value=bool(lang_is_html),
                    )


def _init_sbs_columns_sheet(
    wb: Workbook,
    columns_map: Dict[Tuple[str, str], SbsColumnRow],
    flow_meta_by_qid: Dict[str, _QuestionFlowMeta],
    survey_payload: dict,
    *,
    languages: Sequence[str] | None = None,
    base_language: str = "EN",
) -> None:
    ws = _get_or_create_sheet(wb, SBS_COLUMNS_SHEET)
    base_suffix = _language_suffix(base_language) or "en"
    base_label_col = f"Label_{base_suffix}_MD"
    base_label_html_col = f"Label_{base_suffix}_IsHTML"
    label_columns = _translation_columns(
        "Label", languages, base_language=base_language
    )
    required_cols = [
        "SurveyID",
        "QID",
        "BlockName",
        "BlockID",
        "BlockOrder",
        "QuestionOrder",
        "QuestionOrderInBlock",
        "ColumnId",
        "QuestionType",
        "ExportTag",
        *label_columns,
        "MetaComment",
    ]
    _drop_stale_translation_columns(ws, required_cols, prefixes=["Label"])
    col_index = _ensure_columns(ws, required_cols)

    headers, data_rows = _iter_sheet_rows(ws)
    label_lang_columns: dict[str, tuple[str, str | None]] = {}
    for name in headers:
        header = str(name or "")
        if not header.startswith("Label_") or not header.endswith("_MD"):
            continue
        suffix = header[len("Label_") : -len("_MD")]
        lang_code = _language_from_suffix(suffix)
        is_html_name = f"Label_{suffix}_IsHTML"
        label_lang_columns[lang_code] = (
            header,
            is_html_name if is_html_name in col_index else None,
        )

    qid_col = col_index["QID"]
    column_col = col_index["ColumnId"]
    existing_rows: Dict[Tuple[str, str], int] = {}
    for idx, row in enumerate(data_rows, start=2):
        qid_val = row[qid_col].value
        col_val = row[column_col].value
        if qid_val is None or col_val is None:
            continue
        key = (str(qid_val).strip(), str(col_val).strip())
        existing_rows[key] = idx

    questions = survey_payload.get("result", {}).get("Questions", {})

    for key, row_data in columns_map.items():
        qid, column_id = key
        if key in existing_rows:
            row_idx = existing_rows[key]
            ws.cell(
                row=row_idx, column=col_index["SurveyID"] + 1, value=row_data.survey_id
            )
            ws.cell(row=row_idx, column=col_index["QID"] + 1, value=row_data.qid)
            _write_flow_meta_cells(
                ws,
                row_idx,
                col_index,
                row_data.qid,
                flow_meta_by_qid,
            )
            ws.cell(
                row=row_idx, column=col_index["ColumnId"] + 1, value=row_data.column_id
            )
            ws.cell(
                row=row_idx,
                column=col_index["QuestionType"] + 1,
                value=row_data.question_type,
            )
            ws.cell(
                row=row_idx,
                column=col_index["ExportTag"] + 1,
                value=row_data.export_tag,
            )

            label_cell = ws.cell(row=row_idx, column=col_index[base_label_col] + 1)
            is_html_cell = ws.cell(
                row=row_idx, column=col_index[base_label_html_col] + 1
            )
            if (
                label_cell.value is None or str(label_cell.value).strip() == ""
            ) and row_data.label_en_md:
                label_cell.value = row_data.label_en_md
            is_html_cell.value = bool(row_data.label_en_is_html)

            meta_cell = ws.cell(row=row_idx, column=col_index["MetaComment"] + 1)
            if row_data.externally_managed_by:
                meta_cell.value = _externally_managed_note(
                    row_data.externally_managed_by
                )
            else:
                current = str(meta_cell.value or "").strip()
                if current.startswith(_EXTERNALLY_MANAGED_NOTE_PREFIX):
                    meta_cell.value = ""
        else:
            new_row_idx = ws.max_row + 1
            ws.cell(
                row=new_row_idx,
                column=col_index["SurveyID"] + 1,
                value=row_data.survey_id,
            )
            ws.cell(row=new_row_idx, column=col_index["QID"] + 1, value=row_data.qid)
            _write_flow_meta_cells(
                ws,
                new_row_idx,
                col_index,
                row_data.qid,
                flow_meta_by_qid,
            )
            ws.cell(
                row=new_row_idx,
                column=col_index["ColumnId"] + 1,
                value=row_data.column_id,
            )
            ws.cell(
                row=new_row_idx,
                column=col_index["QuestionType"] + 1,
                value=row_data.question_type,
            )
            ws.cell(
                row=new_row_idx,
                column=col_index["ExportTag"] + 1,
                value=row_data.export_tag,
            )
            ws.cell(
                row=new_row_idx,
                column=col_index[base_label_col] + 1,
                value=row_data.label_en_md,
            )
            ws.cell(
                row=new_row_idx,
                column=col_index[base_label_html_col] + 1,
                value=bool(row_data.label_en_is_html),
            )
            meta_cell = ws.cell(row=new_row_idx, column=col_index["MetaComment"] + 1)
            if row_data.externally_managed_by:
                meta_cell.value = _externally_managed_note(
                    row_data.externally_managed_by
                )
            row_idx = new_row_idx

        q_json = questions.get(qid, {}) if isinstance(questions, dict) else {}
        language_blocks = q_json.get("Language") or {}

        for lang_code, (md_col, html_col) in label_lang_columns.items():
            if lang_code == _normalize_language_code(base_language):
                continue
            lang_block = _lookup_language_block(language_blocks, lang_code)
            if not lang_block:
                continue
            lang_display = None
            answers = lang_block.get("Answers") if isinstance(lang_block, dict) else None
            if isinstance(answers, dict):
                entry = answers.get(str(column_id))
                if isinstance(entry, dict):
                    lang_display = entry.get("Display")
            if lang_display is None:
                aq_block = (
                    lang_block.get("AdditionalQuestions")
                    if isinstance(lang_block, dict)
                    else None
                )
                if isinstance(aq_block, dict):
                    entry = aq_block.get(str(column_id))
                    if isinstance(entry, dict):
                        lang_display = entry.get("QuestionText")

            text_md, is_html = _metadata_cell_value(
                str(lang_display) if lang_display is not None else None
            )
            if not text_md:
                continue
            lang_cell = ws.cell(row=row_idx, column=col_index[md_col] + 1)
            if lang_cell.value is not None and str(lang_cell.value).strip() != "":
                continue
            lang_cell.value = text_md
            if html_col:
                ws.cell(
                    row=row_idx,
                    column=col_index[html_col] + 1,
                    value=bool(is_html),
                )


def _init_sbs_column_answers_sheet(
    wb: Workbook,
    answers_map: Dict[Tuple[str, str, str], SbsColumnAnswerRow],
    flow_meta_by_qid: Dict[str, _QuestionFlowMeta],
    survey_payload: dict,
    *,
    languages: Sequence[str] | None = None,
    base_language: str = "EN",
) -> None:
    ws = _get_or_create_sheet(wb, SBS_COLUMN_ANSWERS_SHEET)
    base_suffix = _language_suffix(base_language) or "en"
    base_label_col = f"Label_{base_suffix}_MD"
    base_label_html_col = f"Label_{base_suffix}_IsHTML"
    label_columns = _translation_columns(
        "Label", languages, base_language=base_language
    )
    required_cols = [
        "SurveyID",
        "QID",
        "BlockName",
        "BlockID",
        "BlockOrder",
        "QuestionOrder",
        "QuestionOrderInBlock",
        "ColumnId",
        "AnswerId",
        "QuestionType",
        "ExportTag",
        *label_columns,
        "MetaComment",
    ]
    _drop_stale_translation_columns(ws, required_cols, prefixes=["Label"])
    col_index = _ensure_columns(ws, required_cols)

    headers, data_rows = _iter_sheet_rows(ws)
    label_lang_columns: dict[str, tuple[str, str | None]] = {}
    for name in headers:
        header = str(name or "")
        if not header.startswith("Label_") or not header.endswith("_MD"):
            continue
        suffix = header[len("Label_") : -len("_MD")]
        lang_code = _language_from_suffix(suffix)
        is_html_name = f"Label_{suffix}_IsHTML"
        label_lang_columns[lang_code] = (
            header,
            is_html_name if is_html_name in col_index else None,
        )

    qid_col = col_index["QID"]
    column_col = col_index["ColumnId"]
    answer_col = col_index["AnswerId"]
    existing_rows: Dict[Tuple[str, str, str], int] = {}
    for idx, row in enumerate(data_rows, start=2):
        qid_val = row[qid_col].value
        col_val = row[column_col].value
        ans_val = row[answer_col].value
        if qid_val is None or col_val is None or ans_val is None:
            continue
        key = (str(qid_val).strip(), str(col_val).strip(), str(ans_val).strip())
        existing_rows[key] = idx

    questions = survey_payload.get("result", {}).get("Questions", {})

    for key, row_data in answers_map.items():
        qid, column_id, answer_id = key
        if key in existing_rows:
            row_idx = existing_rows[key]
            ws.cell(
                row=row_idx, column=col_index["SurveyID"] + 1, value=row_data.survey_id
            )
            ws.cell(row=row_idx, column=col_index["QID"] + 1, value=row_data.qid)
            _write_flow_meta_cells(
                ws,
                row_idx,
                col_index,
                row_data.qid,
                flow_meta_by_qid,
            )
            ws.cell(
                row=row_idx, column=col_index["ColumnId"] + 1, value=row_data.column_id
            )
            ws.cell(
                row=row_idx, column=col_index["AnswerId"] + 1, value=row_data.answer_id
            )
            ws.cell(
                row=row_idx,
                column=col_index["QuestionType"] + 1,
                value=row_data.question_type,
            )
            ws.cell(
                row=row_idx,
                column=col_index["ExportTag"] + 1,
                value=row_data.export_tag,
            )

            label_cell = ws.cell(row=row_idx, column=col_index[base_label_col] + 1)
            is_html_cell = ws.cell(
                row=row_idx, column=col_index[base_label_html_col] + 1
            )
            if (
                label_cell.value is None or str(label_cell.value).strip() == ""
            ) and row_data.label_en_md:
                label_cell.value = row_data.label_en_md
            is_html_cell.value = bool(row_data.label_en_is_html)

            meta_cell = ws.cell(row=row_idx, column=col_index["MetaComment"] + 1)
            if row_data.externally_managed_by:
                meta_cell.value = _externally_managed_note(
                    row_data.externally_managed_by
                )
            else:
                current = str(meta_cell.value or "").strip()
                if current.startswith(_EXTERNALLY_MANAGED_NOTE_PREFIX):
                    meta_cell.value = ""
        else:
            new_row_idx = ws.max_row + 1
            ws.cell(
                row=new_row_idx,
                column=col_index["SurveyID"] + 1,
                value=row_data.survey_id,
            )
            ws.cell(row=new_row_idx, column=col_index["QID"] + 1, value=row_data.qid)
            _write_flow_meta_cells(
                ws,
                new_row_idx,
                col_index,
                row_data.qid,
                flow_meta_by_qid,
            )
            ws.cell(
                row=new_row_idx,
                column=col_index["ColumnId"] + 1,
                value=row_data.column_id,
            )
            ws.cell(
                row=new_row_idx,
                column=col_index["AnswerId"] + 1,
                value=row_data.answer_id,
            )
            ws.cell(
                row=new_row_idx,
                column=col_index["QuestionType"] + 1,
                value=row_data.question_type,
            )
            ws.cell(
                row=new_row_idx,
                column=col_index["ExportTag"] + 1,
                value=row_data.export_tag,
            )
            ws.cell(
                row=new_row_idx,
                column=col_index[base_label_col] + 1,
                value=row_data.label_en_md,
            )
            ws.cell(
                row=new_row_idx,
                column=col_index[base_label_html_col] + 1,
                value=bool(row_data.label_en_is_html),
            )
            meta_cell = ws.cell(row=new_row_idx, column=col_index["MetaComment"] + 1)
            if row_data.externally_managed_by:
                meta_cell.value = _externally_managed_note(
                    row_data.externally_managed_by
                )

            row_idx = new_row_idx

        q_json = questions.get(qid, {}) if isinstance(questions, dict) else {}
        language_blocks = q_json.get("Language") or {}

        # Best-effort: Qualtrics exports may not include per-language SBS column answer labels.
        for lang_code, (md_col, html_col) in label_lang_columns.items():
            if lang_code == _normalize_language_code(base_language):
                continue
            lang_block = _lookup_language_block(language_blocks, lang_code)
            if not lang_block:
                continue
            aq_block = (
                lang_block.get("AdditionalQuestions")
                if isinstance(lang_block, dict)
                else None
            )
            if not isinstance(aq_block, dict):
                continue
            aq_entry = aq_block.get(str(column_id))
            if not isinstance(aq_entry, dict):
                continue
            lang_answers = aq_entry.get("Answers")
            if not isinstance(lang_answers, dict):
                continue
            ans_entry = lang_answers.get(str(answer_id))
            if not isinstance(ans_entry, dict):
                continue
            lang_display = ans_entry.get("Display")
            text_md, is_html = _metadata_cell_value(
                str(lang_display) if lang_display is not None else None
            )
            if not text_md:
                continue
            lang_cell = ws.cell(row=row_idx, column=col_index[md_col] + 1)
            if lang_cell.value is not None and str(lang_cell.value).strip() != "":
                continue
            lang_cell.value = text_md
            if html_col:
                ws.cell(
                    row=row_idx,
                    column=col_index[html_col] + 1,
                    value=bool(is_html),
                )


def _metadata_language_list(
    survey_payload: dict,
    languages: Sequence[str] | None,
) -> List[str]:
    result = survey_payload.get("result", {})
    options = result.get("SurveyOptions", {}) or {}
    base = _normalize_language_code(options.get("SurveyLanguage") or "")
    available = options.get("AvailableLanguages") or []
    meta = options.get("MetaDataTranslations") or {}

    lang_list = _normalize_language_list(languages or [])
    if not lang_list:
        lang_list = _normalize_language_list(available)
    if isinstance(meta, dict):
        meta_langs = [str(k) for k in meta.keys()]
        lang_list = _normalize_language_list(list(lang_list) + meta_langs)
    if base and base not in lang_list:
        lang_list = [base] + lang_list
    return lang_list


def _metadata_cell_value(text_html: str | None) -> tuple[str | None, bool]:
    if text_html is None or str(text_html).strip() == "":
        return None, False
    raw = str(text_html)
    if is_markdown_safe_html(raw):
        return html_to_md(raw), False
    if should_treat_as_html(raw):
        return normalize_text(raw), True
    return html_to_md(raw), False


def _init_survey_metadata_sheet(
    wb: Workbook,
    survey_payload: dict,
    *,
    languages: Sequence[str] | None = None,
) -> None:
    ws = _get_or_create_sheet(wb, SURVEY_METADATA_SHEET)
    result = survey_payload.get("result", {}) or {}
    options = result.get("SurveyOptions", {}) or {}
    base_lang = _normalize_language_code(options.get("SurveyLanguage") or "")
    meta_translations = options.get("MetaDataTranslations") or {}

    metadata_cols = _metadata_columns()
    required_cols = ["SurveyID", "Language", *metadata_cols]
    col_index = _ensure_columns(ws, required_cols)

    headers, data_rows = _iter_sheet_rows(ws)
    lang_idx = col_index["Language"]
    existing_rows: Dict[str, int] = {}
    for idx, row in enumerate(data_rows, start=2):
        lang_val = row[lang_idx].value
        lang = _normalize_language_code(str(lang_val or ""))
        if lang:
            existing_rows[lang] = idx

    languages = _metadata_language_list(survey_payload, languages)
    survey_id = str(result.get("SurveyID") or "")

    for lang in languages:
        row_idx = existing_rows.get(lang)
        if row_idx is None:
            row_idx = ws.max_row + 1
            ws.cell(row=row_idx, column=col_index["SurveyID"] + 1, value=survey_id)
            ws.cell(row=row_idx, column=col_index["Language"] + 1, value=lang)
        else:
            ws.cell(row=row_idx, column=col_index["SurveyID"] + 1, value=survey_id)
            ws.cell(row=row_idx, column=col_index["Language"] + 1, value=lang)

        entry = {}
        if isinstance(meta_translations, dict):
            entry = (
                meta_translations.get(lang)
                or meta_translations.get(lang.lower())
                or meta_translations.get(lang.upper())
                or {}
            )

        for key in SURVEY_METADATA_KEYS:
            if lang == base_lang:
                raw_value = result.get(key)
            else:
                raw_value = entry.get(key) if isinstance(entry, dict) else None
                if (
                    raw_value is None
                    and key == "SurveyDescription"
                    and isinstance(entry, dict)
                ):
                    raw_value = entry.get("SurveyMetaDescription")
            text_md, is_html = _metadata_cell_value(raw_value)

            md_col = f"{key}_MD"
            html_col = f"{key}_IsHTML"
            md_idx = col_index.get(md_col)
            html_idx = col_index.get(html_col)
            if md_idx is not None:
                cell = ws.cell(row=row_idx, column=md_idx + 1)
                if (cell.value is None or str(cell.value).strip() == "") and text_md:
                    cell.value = text_md
            if html_idx is not None:
                cell = ws.cell(row=row_idx, column=html_idx + 1)
                if cell.value is None or str(cell.value).strip() == "":
                    cell.value = bool(is_html)


def _init_embedded_data_sheet(wb: Workbook, rows: List[EmbeddedDataRow]) -> None:
    ws = _get_or_create_sheet(wb, EMBEDDED_DATA_SHEET)
    required_cols = [
        "SurveyID",
        "FlowID",
        "FlowOrder",
        "Field",
        "Value",
        "Type",
        "WrittenByQIDs",
    ]
    col_index = _ensure_columns(ws, required_cols)

    headers, data_rows = _iter_sheet_rows(ws)
    field_col = col_index["Field"]
    flow_col = col_index["FlowID"]

    seen_keys: set[Tuple[str, str]] = set()
    duplicate_rows: List[int] = []
    for idx, row in enumerate(data_rows, start=2):
        field_val = row[field_col].value
        if field_val is None:
            continue
        field = str(field_val).strip()
        if not field:
            continue
        flow_val = row[flow_col].value
        flow_id = str(flow_val).strip() if flow_val is not None else ""
        key = (flow_id, field)
        if key in seen_keys:
            duplicate_rows.append(idx)
            continue
        seen_keys.add(key)

    for row_idx in sorted(duplicate_rows, reverse=True):
        ws.delete_rows(row_idx, 1)

    headers, data_rows = _iter_sheet_rows(ws)
    existing_rows: Dict[Tuple[str, str], int] = {}
    fallback_rows: Dict[str, int] = {}
    for idx, row in enumerate(data_rows, start=2):
        field_val = row[field_col].value
        if field_val is None:
            continue
        field = str(field_val).strip()
        if not field:
            continue
        flow_val = row[flow_col].value
        flow_id = str(flow_val).strip() if flow_val is not None else ""
        key = (flow_id, field)
        existing_rows[key] = idx
        if not flow_id and field not in fallback_rows:
            fallback_rows[field] = idx

    for row_data in rows:
        flow_id = row_data.flow_id or ""
        key = (flow_id, row_data.field)
        row_idx = existing_rows.get(key)
        if row_idx is None and flow_id and row_data.field in fallback_rows:
            row_idx = fallback_rows[row_data.field]
        if row_idx is not None:
            ws.cell(
                row=row_idx, column=col_index["SurveyID"] + 1, value=row_data.survey_id
            )
            ws.cell(row=row_idx, column=col_index["FlowID"] + 1, value=flow_id or None)
            ws.cell(
                row=row_idx,
                column=col_index["FlowOrder"] + 1,
                value=row_data.flow_order,
            )
            ws.cell(row=row_idx, column=col_index["Field"] + 1, value=row_data.field)
            ws.cell(row=row_idx, column=col_index["Type"] + 1, value=row_data.ed_type)
            ws.cell(
                row=row_idx,
                column=col_index["WrittenByQIDs"] + 1,
                value=row_data.written_by_qids or "",
            )
            value_cell = ws.cell(row=row_idx, column=col_index["Value"] + 1)
            if value_cell.value is None or str(value_cell.value).strip() == "":
                value_cell.value = (
                    row_data.value
                    if row_data.value is not None
                    else EMBEDDED_EMPTY_VALUE
                )
        else:
            new_row_idx = ws.max_row + 1
            ws.cell(
                row=new_row_idx,
                column=col_index["SurveyID"] + 1,
                value=row_data.survey_id,
            )
            ws.cell(
                row=new_row_idx, column=col_index["FlowID"] + 1, value=flow_id or None
            )
            ws.cell(
                row=new_row_idx,
                column=col_index["FlowOrder"] + 1,
                value=row_data.flow_order,
            )
            ws.cell(
                row=new_row_idx, column=col_index["Field"] + 1, value=row_data.field
            )
            ws.cell(
                row=new_row_idx, column=col_index["Value"] + 1, value=row_data.value
            )
            if row_data.value is None:
                ws.cell(
                    row=new_row_idx,
                    column=col_index["Value"] + 1,
                    value=EMBEDDED_EMPTY_VALUE,
                )
            ws.cell(
                row=new_row_idx, column=col_index["Type"] + 1, value=row_data.ed_type
            )
            ws.cell(
                row=new_row_idx,
                column=col_index["WrittenByQIDs"] + 1,
                value=row_data.written_by_qids or "",
            )


_HTML_FILL = PatternFill(fill_type="solid", fgColor="FFFFF4B2")
_DIRTY_FILL = PatternFill(fill_type="solid", fgColor="FFF4B084")
_REQUIRED_FILL = PatternFill(fill_type="solid", fgColor="FFFFE2E2")
_READONLY_FILL = PatternFill(fill_type="solid", fgColor="FFECECEC")


def _make_bold(cell) -> None:
    """Make a cell bold."""
    if cell.font:
        cell.font = Font(
            bold=True, name=cell.font.name, size=cell.font.size, color=cell.font.color
        )
    else:
        cell.font = Font(bold=True)


def _set_horizontal_alignment(cell, horizontal: str) -> None:
    """Set horizontal alignment while preserving existing vertical/wrap settings."""

    current = cell.alignment if cell.alignment is not None else Alignment()
    cell.alignment = Alignment(
        horizontal=horizontal,
        vertical=current.vertical or "center",
        wrap_text=current.wrap_text,
    )


def _center_align_short_columns(ws: Worksheet, headers: Sequence[object]) -> None:
    """Center-align short/system columns and all `*_IsHTML` boolean flag columns."""

    center_headers: set[str] = set()
    for name in headers:
        header = str(name or "").strip()
        if not header:
            continue
        if (
            header in CENTER_ALIGN_HEADER_NAMES
            or header.endswith("_IsHTML")
            or header.startswith("ishtml_")
        ):
            center_headers.add(header)
    if not center_headers:
        return
    for header in center_headers:
        col_idx = next(
            (idx for idx, name in enumerate(headers, start=1) if str(name or "") == header),
            None,
        )
        if not col_idx:
            continue
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            _set_horizontal_alignment(row[0], "center")


def _set_cell_italic(cell, italic: bool) -> None:
    """Toggle italic while preserving the existing font family/size/weight."""

    current = cell.font if cell.font is not None else Font()
    updated = copy(current)
    updated.italic = bool(italic)
    cell.font = updated


def _apply_html_md_italics(ws: Worksheet, headers: Sequence[object], *, prefix: str) -> None:
    """Apply direct italics to `{prefix}_*_MD` cells when `{prefix}_*_IsHTML` is TRUE."""

    if ws.max_row <= 1:
        return
    header_index = {str(name or ""): idx + 1 for idx, name in enumerate(headers)}
    for name in headers:
        header = str(name or "")
        if prefix == "Text":
            if not _is_question_text_html_header(header):
                continue
            md_name = _question_text_md_header_from_html_header(header)
        else:
            if not header.startswith(f"{prefix}_") or not header.endswith("_IsHTML"):
                continue
            md_name = f"{header[:-len('_IsHTML')]}_MD"
        html_col = header_index.get(header)
        md_col = header_index.get(md_name)
        if not html_col or not md_col:
            continue
        for row_idx in range(2, ws.max_row + 1):
            html_val = _coerce_bool_cell(
                ws.cell(row=row_idx, column=html_col).value,
                default=False,
            )
            _set_cell_italic(ws.cell(row=row_idx, column=md_col), html_val)


def _sheet_has_data_rows(ws: Worksheet) -> bool:
    if ws.max_row <= 1:
        return False
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if any(value not in (None, "") for value in row):
            return True
    return False


def _set_optional_sheet_visibility(ws: Worksheet, *, hide_when_empty: bool = True) -> None:
    if hide_when_empty and not _sheet_has_data_rows(ws):
        ws.sheet_state = "hidden"
    else:
        ws.sheet_state = "visible"


def _wrap_column(ws: Worksheet, header_name: str) -> None:
    headers, _ = _iter_sheet_rows(ws)
    if not headers:
        return
    if header_name not in headers:
        return
    col_idx = headers.index(header_name) + 1
    for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
        cell = row[0]
        current = cell.alignment if cell.alignment is not None else Alignment()
        cell.alignment = Alignment(
            wrap_text=True,
            vertical=current.vertical or "top",
            horizontal=current.horizontal or "left",
        )


def _autofit_rows(ws: Worksheet) -> None:
    """Enable auto-fit for all rows by clearing height settings."""
    # Clear all row heights to enable auto-fit
    for row_idx in range(1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = None


def _apply_boolean_validation(ws: Worksheet, header_name: str) -> None:
    _apply_list_validation(ws, header_name, ["TRUE", "FALSE"])


def _apply_list_validation(
    ws: Worksheet, header_name: str, allowed_values: Sequence[str]
) -> None:
    headers, _ = _iter_sheet_rows(ws)
    if not headers or header_name not in headers:
        return
    cleaned_values = [str(value).strip() for value in allowed_values if str(value).strip()]
    if not cleaned_values:
        return
    col_idx = headers.index(header_name) + 1
    col_letter = get_column_letter(col_idx)
    max_row = ws.max_row
    if max_row <= 1:
        return
    formula = '"' + ",".join(cleaned_values) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}2:{col_letter}{max_row}")


def _apply_readonly_fill(
    ws: Worksheet, headers: Sequence[object], editable_headers: set[str]
) -> None:
    if ws.max_row <= 1:
        return
    read_only_indices = []
    editable_indices = []
    for idx, name in enumerate(headers, start=1):
        header = str(name or "")
        if header not in editable_headers:
            read_only_indices.append(idx)
        else:
            editable_indices.append(idx)
    if not read_only_indices:
        return
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for idx in read_only_indices:
            if idx <= len(row):
                row[idx - 1].fill = _READONLY_FILL
        # Clear stale static fills on editable cells so workbook refresh can
        # recover from older schema states that marked them read-only.
        for idx in editable_indices:
            if idx <= len(row):
                row[idx - 1].fill = PatternFill(fill_type=None)


_QID_NUM_RE = re.compile(r"^QID(?P<num>\d+)$", re.IGNORECASE)


def _qid_order_key(qid: str) -> tuple[int, int | None, str]:
    value = str(qid or "").strip()
    match = _QID_NUM_RE.match(value)
    if not match:
        return (1, None, value)
    return (0, int(match.group("num")), value)


def _int_sort_value(value: object) -> int | None:
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return None


def _question_order_sort_key(row, *, headers: Sequence[object], qid_idx: int):
    order_idx = headers.index("QuestionOrder") if "QuestionOrder" in headers else None
    question_order = _int_sort_value(row[order_idx].value) if order_idx is not None else None
    qid = str(row[qid_idx].value or "").strip()
    qid_key = _qid_order_key(qid)
    if question_order is None or question_order <= 0:
        return (1, None, qid_key)
    return (0, question_order, qid_key)


def _sort_questions_sheet(ws: Worksheet) -> None:
    """Sort Questions sheet by QuestionOrder (fallback: numeric QID)."""

    headers, data_rows = _iter_sheet_rows(ws)
    if not headers or "QID" not in headers:
        return
    qid_idx = headers.index("QID")

    rows_with_values = [[cell.value for cell in row] for row in data_rows]
    if not rows_with_values:
        return

    paired = [
        (
            _question_order_sort_key(data_rows[i], headers=headers, qid_idx=qid_idx),
            rows_with_values[i],
        )
        for i in range(len(data_rows))
    ]
    paired.sort(key=lambda x: x[0])

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for _, values in paired:
        ws.append(values)


def _sort_sheet_by_qid_and_id(ws: Worksheet, id_header: str) -> None:
    """Sort a sheet by (QID, id_header) in-place.

    This keeps the header row, reorders all data rows, and preserves cell values.
    """

    headers, data_rows = _iter_sheet_rows(ws)
    if not headers:
        return
    if "QID" not in headers or id_header not in headers:
        return

    qid_idx = headers.index("QID")
    id_idx = headers.index(id_header)
    field_idx = headers.index("Field") if "Field" in headers else None

    def sort_key(row):
        id_val = row[id_idx].value
        q_order_key = _question_order_sort_key(row, headers=headers, qid_idx=qid_idx)
        field_val = row[field_idx].value if field_idx is not None else "Answer"
        field = _normalize_subitem_field(field_val)
        id_str = str(id_val or "")
        try:
            id_num = int(id_str)
        except ValueError:
            id_num = None
        # Numeric IDs first, then non-numeric, all grouped by QID.
        id_key = (0, id_num) if id_num is not None else (1, id_str)
        if field == "Answer":
            field_order = 0
        elif field == "Label":
            field_order = 1
        else:
            field_order = 2
        return (q_order_key, field_order, field, id_key)

    # Extract current values so we can rewrite the rows after sorting.
    rows_with_values = [[cell.value for cell in row] for row in data_rows]
    if not rows_with_values:
        return

    # Pair values with their original sort key.
    paired = [
        (sort_key(data_rows[i]), rows_with_values[i])
        for i in range(len(rows_with_values))
    ]
    paired.sort(key=lambda x: x[0])

    # Clear existing data rows (keep header).
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    # Append rows back in sorted order.
    for _, values in paired:
        ws.append(values)


def _sort_sheet_by_qid_and_two_ids(
    ws: Worksheet, id1_header: str, id2_header: str
) -> None:
    """Sort a sheet by (QID, id1_header, id2_header) in-place."""

    headers, data_rows = _iter_sheet_rows(ws)
    if not headers:
        return
    if "QID" not in headers or id1_header not in headers or id2_header not in headers:
        return

    qid_idx = headers.index("QID")
    id1_idx = headers.index(id1_header)
    id2_idx = headers.index(id2_header)

    def _num_or_str(value: object) -> tuple[int, int | None, str]:
        s = str(value or "")
        try:
            return (0, int(s), s)
        except ValueError:
            return (1, None, s)

    def sort_key(row):
        q_order_key = _question_order_sort_key(row, headers=headers, qid_idx=qid_idx)
        id1 = _num_or_str(row[id1_idx].value)
        id2 = _num_or_str(row[id2_idx].value)
        return (q_order_key, id1, id2)

    rows_with_values = [[cell.value for cell in row] for row in data_rows]
    if not rows_with_values:
        return

    paired = [
        (sort_key(data_rows[i]), rows_with_values[i]) for i in range(len(data_rows))
    ]
    paired.sort(key=lambda x: x[0])

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for _, values in paired:
        ws.append(values)


def _sort_sheet_by_flow_order(ws: Worksheet) -> None:
    """Sort Embedded_Data by FlowOrder and Field (JS-only rows last)."""

    headers, data_rows = _iter_sheet_rows(ws)
    if not headers:
        return
    if "FlowOrder" not in headers or "Field" not in headers:
        return

    order_idx = headers.index("FlowOrder")
    field_idx = headers.index("Field")
    flow_idx = headers.index("FlowID") if "FlowID" in headers else None

    def sort_key(row):
        order_val = row[order_idx].value
        try:
            order = int(order_val)
        except (TypeError, ValueError):
            order = 0
        field = str(row[field_idx].value or "")
        flow_id = str(row[flow_idx].value or "") if flow_idx is not None else ""
        return (order == 0, order, flow_id, field)

    rows_with_values = [[cell.value for cell in row] for row in data_rows]
    if not rows_with_values:
        return

    paired = [
        (sort_key(data_rows[i]), rows_with_values[i])
        for i in range(len(rows_with_values))
    ]
    paired.sort(key=lambda x: x[0])

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for _, values in paired:
        ws.append(values)


def _format_questions_sheet(ws: Worksheet) -> None:
    # Header & system columns
    headers, _ = _iter_sheet_rows(ws)
    if not headers:
        return
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    system_headers = {
        "SurveyID",
        "QID",
        "BlockName",
        "BlockID",
        "BlockOrder",
        "QuestionOrder",
        "QuestionOrderInBlock",
        "QuestionType",
        "DataExportTag",
        "RequiredResponse",
        "OptionsPreview",
        "SubitemsPreview",
    }

    for cell in header_row:
        name = cell.value or ""
        if name in system_headers:
            _make_bold(cell)

    # Make system-owned columns bold in body rows
    system_indices = [headers.index(h) + 1 for h in system_headers if h in headers]
    for row in ws.iter_rows(min_row=2):
        for idx in system_indices:
            if idx <= len(row):
                _make_bold(row[idx - 1])

    # Apply vertical center and horizontal left alignment to all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            # Default alignment: vertically centered, horizontally left
            cell.alignment = Alignment(
                vertical="center", horizontal="left", wrap_text=False
            )

    # Wrap long text (question text and options/subitems preview)
    for name in headers:
        if _is_question_text_md_header(str(name)):
            _wrap_column(ws, str(name))
    _wrap_column(ws, "ValidationSettingsJSON")
    _wrap_column(ws, "RandomizationSettingsJSON")
    _wrap_column(ws, QUESTION_CONFIG_JSON_COLUMN)
    _wrap_column(ws, "OptionsPreview")
    _wrap_column(ws, "SubitemsPreview")

    _center_align_short_columns(ws, headers)

    # Auto-fit row heights
    _autofit_rows(ws)

    # Boolean validations
    for name in headers:
        if _is_question_text_html_header(str(name)):
            _apply_boolean_validation(ws, str(name))
    _apply_list_validation(ws, "ForceResponseMode", ["OFF", "ON", "RequestResponse"])
    _apply_list_validation(
        ws,
        "ValidationType",
        ["None", "MinChoices", "CustomValidation", "ChoicesTotal"],
    )
    _apply_list_validation(
        ws,
        "RandomizationType",
        ["None", "All", "Subset", "Advanced"],
    )

    editable_headers = {
        "ForceResponseMode",
        "ValidationType",
        "ValidationSettingsJSON",
        "RandomizationType",
        "RandomizationSettingsJSON",
    }
    for name in headers:
        header = str(name or "")
        if _is_question_text_md_header(header) or _is_question_text_html_header(header):
            editable_headers.add(header)
    _apply_readonly_fill(ws, headers, editable_headers)

    # Clear any existing conditional formatting (we'll reapply ours)
    try:
        ws.conditional_formatting.clear()
    except AttributeError:
        # Older openpyxl: recreate a fresh container by assigning an empty list
        ws.conditional_formatting._cf_rules = {}

    max_row = ws.max_row
    if "RequiredResponse" in headers and max_row >= 2:
        required_idx = headers.index("RequiredResponse") + 1
        required_col = get_column_letter(required_idx)
        for name in headers:
            if _is_question_text_md_header(str(name)):
                text_idx = headers.index(name) + 1
                text_col = get_column_letter(text_idx)
                formula = f"=${required_col}2=TRUE"
                rule = FormulaRule(formula=[formula], fill=_REQUIRED_FILL)
                ws.conditional_formatting.add(f"{text_col}2:{text_col}{max_row}", rule)

    # Conditional formatting: highlight HTML question text when ishtml_* is TRUE
    for name in headers:
        header = str(name or "")
        if _is_question_text_html_header(header):
            text_name = _question_text_md_header_from_html_header(header)
            if text_name not in headers or max_row < 2:
                continue
            html_idx = headers.index(name) + 1
            text_idx = headers.index(text_name) + 1
            html_col = get_column_letter(html_idx)
            text_col = get_column_letter(text_idx)
            formula = f"=${html_col}2=TRUE"
            rule = FormulaRule(
                formula=[formula],
                fill=_HTML_FILL,
                font=Font(italic=True),
            )
            ws.conditional_formatting.add(f"{text_col}2:{text_col}{max_row}", rule)

    # Conditional formatting: highlight dirty question text when Dirty == 'Y'
    if "Dirty" in headers and max_row >= 2:
        dirty_idx = headers.index("Dirty") + 1
        dirty_col = get_column_letter(dirty_idx)
        for name in headers:
            if _is_question_text_md_header(str(name)):
                text_idx = headers.index(name) + 1
                text_col = get_column_letter(text_idx)
                formula = f'=${dirty_col}2="Y"'
                rule = FormulaRule(formula=[formula], fill=_DIRTY_FILL)
                ws.conditional_formatting.add(f"{text_col}2:{text_col}{max_row}", rule)

    _apply_html_md_italics(ws, headers, prefix="Text")

    # Add a simple Excel table style
    max_row = ws.max_row
    max_col = ws.max_column
    if max_row >= 2 and max_col >= 1:
        table = Table(
            displayName="QuestionsTable",
            ref=f"A1:{get_column_letter(max_col)}{max_row}",
        )
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        # Always refresh the table so its ref/columns cover all rows/columns.
        if "QuestionsTable" in ws._tables:
            del ws._tables["QuestionsTable"]
        ws.add_table(table)

    # Fixed column widths based on header names (tuned for NEWSFLOWS)
    widths = {
        "SurveyID": 18.0,
        "QID": 7.0,
        "BlockName": 19.0,
        "BlockID": 14.0,
        "BlockOrder": 12.0,
        "QuestionOrder": 14.0,
        "QuestionOrderInBlock": 18.0,
        "QuestionType": 14.5,
        "DataExportTag": 19.0,
        "RequiredResponse": 14.0,
        "ForceResponseMode": 20.0,
        "ValidationType": 18.0,
        "ValidationSettingsJSON": 48.0,
        "RandomizationType": 18.0,
        "RandomizationSettingsJSON": 48.0,
        QUESTION_CONFIG_JSON_COLUMN: 70.0,
        "OptionsPreview": 60.0,
        "SubitemsPreview": 60.0,
    }
    for idx, name in enumerate(headers, start=1):
        key = str(name or "")
        w = widths.get(key)
        if w:
            ws.column_dimensions[get_column_letter(idx)].width = w
            continue
        if _is_question_text_md_header(key):
            ws.column_dimensions[get_column_letter(idx)].width = 76.0
        elif _is_question_text_html_header(key):
            ws.column_dimensions[get_column_letter(idx)].width = 16.0


def _format_options_sheet(ws: Worksheet) -> None:
    headers, _ = _iter_sheet_rows(ws)
    if not headers:
        return
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    system_headers = {
        "SurveyID",
        "QID",
        "BlockName",
        "BlockID",
        "BlockOrder",
        "QuestionOrder",
        "QuestionOrderInBlock",
        "ChoiceId",
        "QuestionType",
        "ExportTag",
        "Code",
    }

    for cell in header_row:
        name = cell.value or ""
        if name in system_headers:
            _make_bold(cell)

    system_indices = [headers.index(h) + 1 for h in system_headers if h in headers]
    for row in ws.iter_rows(min_row=2):
        for idx in system_indices:
            if idx <= len(row):
                _make_bold(row[idx - 1])

    # Apply vertical center and horizontal left alignment to all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(
                vertical="center", horizontal="left", wrap_text=False
            )

    # Wrap long text (option labels)
    for name in headers:
        if str(name).startswith("Label_") and str(name).endswith("_MD"):
            _wrap_column(ws, str(name))

    _center_align_short_columns(ws, headers)

    # Auto-fit row heights
    _autofit_rows(ws)

    for name in headers:
        if str(name).startswith("Label_") and str(name).endswith("_IsHTML"):
            _apply_boolean_validation(ws, str(name))

    editable_headers: set[str] = set()
    for name in headers:
        header = str(name or "")
        if header.startswith("Label_") and (
            header.endswith("_MD") or header.endswith("_IsHTML")
        ):
            editable_headers.add(header)
    _apply_readonly_fill(ws, headers, editable_headers)

    # Clear any existing conditional formatting
    try:
        ws.conditional_formatting.clear()
    except AttributeError:
        ws.conditional_formatting._cf_rules = {}

    max_row = ws.max_row
    # Conditional formatting: highlight HTML labels
    for name in headers:
        if str(name).startswith("Label_") and str(name).endswith("_IsHTML"):
            suffix = str(name)[len("Label_") : -len("_IsHTML")]
            text_name = f"Label_{suffix}_MD"
            if text_name not in headers or max_row < 2:
                continue
            html_idx = headers.index(name) + 1
            text_idx = headers.index(text_name) + 1
            html_col = get_column_letter(html_idx)
            text_col = get_column_letter(text_idx)
            formula = f"=${html_col}2=TRUE"
            rule = FormulaRule(
                formula=[formula],
                fill=_HTML_FILL,
                font=Font(italic=True),
            )
            ws.conditional_formatting.add(f"{text_col}2:{text_col}{max_row}", rule)

    # Conditional formatting: highlight dirty option labels
    if "Dirty" in headers and max_row >= 2:
        dirty_idx = headers.index("Dirty") + 1
        dirty_col = get_column_letter(dirty_idx)
        for name in headers:
            if str(name).startswith("Label_") and str(name).endswith("_MD"):
                text_idx = headers.index(name) + 1
                text_col = get_column_letter(text_idx)
                formula = f'=${dirty_col}2="Y"'
                rule = FormulaRule(formula=[formula], fill=_DIRTY_FILL)
                ws.conditional_formatting.add(f"{text_col}2:{text_col}{max_row}", rule)

    _apply_html_md_italics(ws, headers, prefix="Label")

    max_row = ws.max_row
    max_col = ws.max_column
    if max_row >= 2 and max_col >= 1:
        table = Table(
            displayName="OptionsTable",
            ref=f"A1:{get_column_letter(max_col)}{max_row}",
        )
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        # Always refresh the table so its ref covers all rows/columns.
        if "OptionsTable" in ws._tables:
            del ws._tables["OptionsTable"]
        ws.add_table(table)
    # Fixed widths based on header names
    widths = {
        "SurveyID": 20.0,
        "QID": 7.0,
        "BlockName": 19.0,
        "BlockID": 14.0,
        "BlockOrder": 12.0,
        "QuestionOrder": 14.0,
        "QuestionOrderInBlock": 18.0,
        "ChoiceId": 10.0,
        "QuestionType": 14.0,
        "ExportTag": 19.0,
        "Code": 6.0,
        "MetaComment": 42.0,
    }
    for idx, name in enumerate(headers, start=1):
        key = str(name or "")
        w = widths.get(key)
        if w:
            ws.column_dimensions[get_column_letter(idx)].width = w
            continue
        if key.startswith("Label_") and key.endswith("_MD"):
            ws.column_dimensions[get_column_letter(idx)].width = 40.0
        elif key.startswith("Label_") and key.endswith("_IsHTML"):
            ws.column_dimensions[get_column_letter(idx)].width = 17.0


def _format_subitems_sheet(ws: Worksheet) -> None:
    headers, _ = _iter_sheet_rows(ws)
    if not headers:
        return
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    system_headers = {
        "SurveyID",
        "QID",
        "BlockName",
        "BlockID",
        "BlockOrder",
        "QuestionOrder",
        "QuestionOrderInBlock",
        "AnswerId",
        "Field",
        "QuestionType",
        "ExportTag",
    }

    for cell in header_row:
        name = cell.value or ""
        if name in system_headers:
            _make_bold(cell)

    system_indices = [headers.index(h) + 1 for h in system_headers if h in headers]
    for row in ws.iter_rows(min_row=2):
        for idx in system_indices:
            if idx <= len(row):
                _make_bold(row[idx - 1])

    # Apply vertical center and horizontal left alignment to all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(
                vertical="center", horizontal="left", wrap_text=False
            )

    for name in headers:
        if str(name).startswith("Label_") and str(name).endswith("_MD"):
            _wrap_column(ws, str(name))

    _center_align_short_columns(ws, headers)

    # Auto-fit row heights
    _autofit_rows(ws)

    for name in headers:
        if str(name).startswith("Label_") and str(name).endswith("_IsHTML"):
            _apply_boolean_validation(ws, str(name))

    editable_headers: set[str] = set()
    for name in headers:
        header = str(name or "")
        if header.startswith("Label_") and (
            header.endswith("_MD") or header.endswith("_IsHTML")
        ):
            editable_headers.add(header)
    _apply_readonly_fill(ws, headers, editable_headers)
    _apply_html_md_italics(ws, headers, prefix="Label")

    max_row = ws.max_row
    max_col = ws.max_column
    if max_row >= 2 and max_col >= 1:
        table = Table(
            displayName="SubitemsTable",
            ref=f"A1:{get_column_letter(max_col)}{max_row}",
        )
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        # Always refresh the table so its ref covers all rows/columns.
        if "SubitemsTable" in ws._tables:
            del ws._tables["SubitemsTable"]
        ws.add_table(table)

    widths = {
        "SurveyID": 20.0,
        "QID": 7.0,
        "BlockName": 19.0,
        "BlockID": 14.0,
        "BlockOrder": 12.0,
        "QuestionOrder": 14.0,
        "QuestionOrderInBlock": 18.0,
        "AnswerId": 10.0,
        "Field": 9.0,
        "QuestionType": 14.0,
        "ExportTag": 19.0,
    }
    for idx, name in enumerate(headers, start=1):
        key = str(name or "")
        w = widths.get(key)
        if w:
            ws.column_dimensions[get_column_letter(idx)].width = w
            continue
        if key.startswith("Label_") and key.endswith("_MD"):
            ws.column_dimensions[get_column_letter(idx)].width = 40.0
        elif key.startswith("Label_") and key.endswith("_IsHTML"):
            ws.column_dimensions[get_column_letter(idx)].width = 17.0


def _format_sbs_columns_sheet(ws: Worksheet) -> None:
    headers, _ = _iter_sheet_rows(ws)
    if not headers:
        return
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    system_headers = {
        "SurveyID",
        "QID",
        "BlockName",
        "BlockID",
        "BlockOrder",
        "QuestionOrder",
        "QuestionOrderInBlock",
        "ColumnId",
        "QuestionType",
        "ExportTag",
    }

    for cell in header_row:
        name = cell.value or ""
        if name in system_headers:
            _make_bold(cell)

    system_indices = [headers.index(h) + 1 for h in system_headers if h in headers]
    for row in ws.iter_rows(min_row=2):
        for idx in system_indices:
            if idx <= len(row):
                _make_bold(row[idx - 1])

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(
                vertical="center", horizontal="left", wrap_text=False
            )

    for name in headers:
        if str(name).startswith("Label_") and str(name).endswith("_MD"):
            _wrap_column(ws, str(name))

    _center_align_short_columns(ws, headers)

    _autofit_rows(ws)

    for name in headers:
        if str(name).startswith("Label_") and str(name).endswith("_IsHTML"):
            _apply_boolean_validation(ws, str(name))

    editable_headers: set[str] = set()
    for name in headers:
        header = str(name or "")
        if header.startswith("Label_") and (
            header.endswith("_MD") or header.endswith("_IsHTML")
        ):
            editable_headers.add(header)
    _apply_readonly_fill(ws, headers, editable_headers)

    try:
        ws.conditional_formatting.clear()
    except AttributeError:
        ws.conditional_formatting._cf_rules = {}

    max_row = ws.max_row
    for name in headers:
        if str(name).startswith("Label_") and str(name).endswith("_IsHTML"):
            suffix = str(name)[len("Label_") : -len("_IsHTML")]
            text_name = f"Label_{suffix}_MD"
            if text_name not in headers or max_row < 2:
                continue
            html_idx = headers.index(name) + 1
            text_idx = headers.index(text_name) + 1
            html_col = get_column_letter(html_idx)
            text_col = get_column_letter(text_idx)
            formula = f"=${html_col}2=TRUE"
            rule = FormulaRule(
                formula=[formula],
                fill=_HTML_FILL,
                font=Font(italic=True),
            )
            ws.conditional_formatting.add(f"{text_col}2:{text_col}{max_row}", rule)

    if "Dirty" in headers and max_row >= 2:
        dirty_idx = headers.index("Dirty") + 1
        dirty_col = get_column_letter(dirty_idx)
        for name in headers:
            if str(name).startswith("Label_") and str(name).endswith("_MD"):
                text_idx = headers.index(name) + 1
                text_col = get_column_letter(text_idx)
                formula = f'=${dirty_col}2="Y"'
                rule = FormulaRule(formula=[formula], fill=_DIRTY_FILL)
                ws.conditional_formatting.add(f"{text_col}2:{text_col}{max_row}", rule)

    _apply_html_md_italics(ws, headers, prefix="Label")

    max_col = ws.max_column
    if max_row >= 2 and max_col >= 1:
        table = Table(
            displayName="SBSColumnsTable",
            ref=f"A1:{get_column_letter(max_col)}{max_row}",
        )
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        if "SBSColumnsTable" in ws._tables:
            del ws._tables["SBSColumnsTable"]
        ws.add_table(table)

    widths = {
        "SurveyID": 20.0,
        "QID": 7.0,
        "BlockName": 19.0,
        "BlockID": 14.0,
        "BlockOrder": 12.0,
        "QuestionOrder": 14.0,
        "QuestionOrderInBlock": 18.0,
        "ColumnId": 10.0,
        "QuestionType": 14.0,
        "ExportTag": 19.0,
        "MetaComment": 42.0,
    }
    for idx, name in enumerate(headers, start=1):
        key = str(name or "")
        w = widths.get(key)
        if w:
            ws.column_dimensions[get_column_letter(idx)].width = w
            continue
        if key.startswith("Label_") and key.endswith("_MD"):
            ws.column_dimensions[get_column_letter(idx)].width = 40.0
        elif key.startswith("Label_") and key.endswith("_IsHTML"):
            ws.column_dimensions[get_column_letter(idx)].width = 17.0


def _format_sbs_column_answers_sheet(ws: Worksheet) -> None:
    headers, _ = _iter_sheet_rows(ws)
    if not headers:
        return
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    system_headers = {
        "SurveyID",
        "QID",
        "BlockName",
        "BlockID",
        "BlockOrder",
        "QuestionOrder",
        "QuestionOrderInBlock",
        "ColumnId",
        "AnswerId",
        "QuestionType",
        "ExportTag",
    }

    for cell in header_row:
        name = cell.value or ""
        if name in system_headers:
            _make_bold(cell)

    system_indices = [headers.index(h) + 1 for h in system_headers if h in headers]
    for row in ws.iter_rows(min_row=2):
        for idx in system_indices:
            if idx <= len(row):
                _make_bold(row[idx - 1])

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(
                vertical="center", horizontal="left", wrap_text=False
            )

    for name in headers:
        if str(name).startswith("Label_") and str(name).endswith("_MD"):
            _wrap_column(ws, str(name))

    _center_align_short_columns(ws, headers)

    _autofit_rows(ws)

    for name in headers:
        if str(name).startswith("Label_") and str(name).endswith("_IsHTML"):
            _apply_boolean_validation(ws, str(name))

    editable_headers: set[str] = set()
    for name in headers:
        header = str(name or "")
        if header.startswith("Label_") and (
            header.endswith("_MD") or header.endswith("_IsHTML")
        ):
            editable_headers.add(header)
    _apply_readonly_fill(ws, headers, editable_headers)

    try:
        ws.conditional_formatting.clear()
    except AttributeError:
        ws.conditional_formatting._cf_rules = {}

    max_row = ws.max_row
    for name in headers:
        if str(name).startswith("Label_") and str(name).endswith("_IsHTML"):
            suffix = str(name)[len("Label_") : -len("_IsHTML")]
            text_name = f"Label_{suffix}_MD"
            if text_name not in headers or max_row < 2:
                continue
            html_idx = headers.index(name) + 1
            text_idx = headers.index(text_name) + 1
            html_col = get_column_letter(html_idx)
            text_col = get_column_letter(text_idx)
            formula = f"=${html_col}2=TRUE"
            rule = FormulaRule(
                formula=[formula],
                fill=_HTML_FILL,
                font=Font(italic=True),
            )
            ws.conditional_formatting.add(f"{text_col}2:{text_col}{max_row}", rule)

    if "Dirty" in headers and max_row >= 2:
        dirty_idx = headers.index("Dirty") + 1
        dirty_col = get_column_letter(dirty_idx)
        for name in headers:
            if str(name).startswith("Label_") and str(name).endswith("_MD"):
                text_idx = headers.index(name) + 1
                text_col = get_column_letter(text_idx)
                formula = f'=${dirty_col}2="Y"'
                rule = FormulaRule(formula=[formula], fill=_DIRTY_FILL)
                ws.conditional_formatting.add(f"{text_col}2:{text_col}{max_row}", rule)

    _apply_html_md_italics(ws, headers, prefix="Label")

    max_col = ws.max_column
    if max_row >= 2 and max_col >= 1:
        table = Table(
            displayName="SBSColumnAnswersTable",
            ref=f"A1:{get_column_letter(max_col)}{max_row}",
        )
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        if "SBSColumnAnswersTable" in ws._tables:
            del ws._tables["SBSColumnAnswersTable"]
        ws.add_table(table)

    widths = {
        "SurveyID": 20.0,
        "QID": 7.0,
        "BlockName": 19.0,
        "BlockID": 14.0,
        "BlockOrder": 12.0,
        "QuestionOrder": 14.0,
        "QuestionOrderInBlock": 18.0,
        "ColumnId": 10.0,
        "AnswerId": 10.0,
        "QuestionType": 14.0,
        "ExportTag": 19.0,
        "MetaComment": 42.0,
    }
    for idx, name in enumerate(headers, start=1):
        key = str(name or "")
        w = widths.get(key)
        if w:
            ws.column_dimensions[get_column_letter(idx)].width = w
            continue
        if key.startswith("Label_") and key.endswith("_MD"):
            ws.column_dimensions[get_column_letter(idx)].width = 40.0
        elif key.startswith("Label_") and key.endswith("_IsHTML"):
            ws.column_dimensions[get_column_letter(idx)].width = 17.0


def _format_survey_metadata_sheet(ws: Worksheet) -> None:
    headers, _ = _iter_sheet_rows(ws)
    if not headers:
        return
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    system_headers = {"SurveyID", "Language"}

    for cell in header_row:
        name = cell.value or ""
        if name in system_headers:
            _make_bold(cell)

    system_indices = [headers.index(h) + 1 for h in system_headers if h in headers]
    for row in ws.iter_rows(min_row=2):
        for idx in system_indices:
            if idx <= len(row):
                _make_bold(row[idx - 1])

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(
                vertical="center", horizontal="left", wrap_text=False
            )

    for name in headers:
        if str(name).endswith("_MD"):
            _wrap_column(ws, str(name))

    _center_align_short_columns(ws, headers)

    _autofit_rows(ws)

    for name in headers:
        if str(name).endswith("_IsHTML"):
            _apply_boolean_validation(ws, str(name))

    editable_headers: set[str] = set()
    for name in headers:
        header = str(name or "")
        if header.endswith("_MD") or header.endswith("_IsHTML"):
            editable_headers.add(header)
    _apply_readonly_fill(ws, headers, editable_headers)
    for key in SURVEY_METADATA_KEYS:
        _apply_html_md_italics(ws, headers, prefix=key)

    max_row = ws.max_row
    max_col = ws.max_column
    if max_row >= 2 and max_col >= 1:
        table = Table(
            displayName="SurveyMetadataTable",
            ref=f"A1:{get_column_letter(max_col)}{max_row}",
        )
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        if "SurveyMetadataTable" in ws._tables:
            del ws._tables["SurveyMetadataTable"]
        ws.add_table(table)

    widths = {
        "SurveyID": 18.0,
        "Language": 10.0,
    }
    for idx, name in enumerate(headers, start=1):
        key = str(name or "")
        w = widths.get(key)
        if w:
            ws.column_dimensions[get_column_letter(idx)].width = w
        elif key.endswith("_MD"):
            ws.column_dimensions[get_column_letter(idx)].width = 60.0
        elif key.endswith("_IsHTML"):
            ws.column_dimensions[get_column_letter(idx)].width = 18.0


def _format_embedded_data_sheet(ws: Worksheet) -> None:
    headers, _ = _iter_sheet_rows(ws)
    if not headers:
        return
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    system_headers = {
        "SurveyID",
        "FlowID",
        "FlowOrder",
        "Field",
        "Type",
        "WrittenByQIDs",
    }

    for cell in header_row:
        name = cell.value or ""
        if name in system_headers:
            _make_bold(cell)

    system_indices = [headers.index(h) + 1 for h in system_headers if h in headers]
    for row in ws.iter_rows(min_row=2):
        for idx in system_indices:
            if idx <= len(row):
                _make_bold(row[idx - 1])

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(
                vertical="center", horizontal="left", wrap_text=False
            )

    _wrap_column(ws, "Value")
    _wrap_column(ws, "WrittenByQIDs")
    _center_align_short_columns(ws, headers)
    _autofit_rows(ws)
    _apply_readonly_fill(ws, headers, {"Value"})

    # Clear any existing conditional formatting (we'll reapply ours)
    try:
        ws.conditional_formatting.clear()
    except AttributeError:
        ws.conditional_formatting._cf_rules = {}

    max_row = ws.max_row
    if "Dirty" in headers and "Value" in headers and max_row >= 2:
        dirty_idx = headers.index("Dirty") + 1
        value_idx = headers.index("Value") + 1
        dirty_col = get_column_letter(dirty_idx)
        value_col = get_column_letter(value_idx)
        formula = f'=${dirty_col}2="Y"'
        rule = FormulaRule(formula=[formula], fill=_DIRTY_FILL)
        ws.conditional_formatting.add(f"{value_col}2:{value_col}{max_row}", rule)

    max_col = ws.max_column
    if max_row >= 2 and max_col >= 1:
        table = Table(
            displayName="EmbeddedDataTable",
            ref=f"A1:{get_column_letter(max_col)}{max_row}",
        )
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        if "EmbeddedDataTable" in ws._tables:
            del ws._tables["EmbeddedDataTable"]
        ws.add_table(table)

    widths = {
        "SurveyID": 20.0,
        "FlowID": 12.0,
        "FlowOrder": 10.0,
        "Field": 26.0,
        "Value": 30.0,
        "Type": 12.0,
        "WrittenByQIDs": 26.0,
        "Dirty": 8.0,
    }
    for idx, name in enumerate(headers, start=1):
        w = widths.get(name)
        if w:
            ws.column_dimensions[get_column_letter(idx)].width = w


def _populate_system_sheet(wb: Workbook, survey_id: str, survey_payload: dict) -> None:
    """Populate a System sheet with non-editable info such as Timing/meta options."""

    ws = _get_or_create_sheet(wb, SYSTEM_SHEET)
    ws.title = SYSTEM_SHEET

    questions = survey_payload.get("result", {}).get("Questions", {})
    headers = [
        "SurveyID",
        "QID",
        "QuestionType",
        "DataExportTag",
        "ChoiceId",
        "Display",
    ]
    ws.delete_rows(1, ws.max_row)
    ws.append(headers)

    for qid, q in questions.items():
        qtype = q.get("QuestionType")
        if qtype != "Timing":
            continue
        tag = q.get("DataExportTag") or ""
        choices = q.get("Choices") or {}
        for choice_id, choice in choices.items():
            display = choice.get("Display") or ""
            ws.append(
                [
                    survey_id,
                    qid,
                    qtype,
                    tag,
                    choice_id,
                    display,
                ]
            )

    # Make header row bold
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    for cell in header_row:
        _make_bold(cell)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(
                vertical="center", horizontal="left", wrap_text=False
            )
    _apply_readonly_fill(ws, headers, set())

    max_row = ws.max_row
    max_col = ws.max_column
    if max_row >= 2 and max_col >= 1:
        table = Table(
            displayName="SystemTable",
            ref=f"A1:{get_column_letter(max_col)}{max_row}",
        )
        style = TableStyleInfo(
            name="TableStyleLight9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        # Always refresh the table so its ref covers all rows/columns.
        if "SystemTable" in ws._tables:
            del ws._tables["SystemTable"]
        ws.add_table(table)


def _fallback_html_col_for_md(md_col: str, *, prefix: str) -> str:
    name = str(md_col or "")
    if prefix == "Text":
        if name.startswith("text_"):
            suffix = name[len("text_") :]
            return f"ishtml_{suffix}"
        if name.startswith("Text_") and name.endswith("_MD"):
            suffix = name[len("Text_") : -len("_MD")]
            return f"Text_{suffix}_IsHTML"
    if name.startswith(f"{prefix}_") and name.endswith("_MD"):
        suffix = name[len(prefix) + 1 : -len("_MD")]
        return f"{prefix}_{suffix}_IsHTML"
    return f"{prefix}_en_IsHTML"


def _find_base_text_col(
    headers: List[str],
    prefix: str,
    *,
    base_language: str | None = None,
) -> tuple[str, str]:
    """Find the first ``{prefix}_*_MD`` column and its ``_IsHTML`` companion.

    Returns ``(md_col, html_col)`` — e.g. ``("Text_cs_MD", "Text_cs_IsHTML")``.
    Falls back to ``{prefix}_en_MD`` when no match is found.
    """
    if prefix == "Text":
        lang_map = _question_text_lang_columns_from_headers(headers, include_legacy=True)
        if lang_map:
            preferred_lang = _normalize_language_code(base_language or "")
            if preferred_lang and preferred_lang in lang_map:
                md_col, html_col = lang_map[preferred_lang]
                return md_col, html_col or _fallback_html_col_for_md(
                    md_col, prefix=prefix
                )
            # No explicit base language match: pick the first discovered language
            # from workbook header order instead of assuming EN.
            first_lang = next(iter(lang_map))
            md_col, html_col = lang_map[first_lang]
            return md_col, html_col or _fallback_html_col_for_md(
                md_col, prefix=prefix
            )
        if base_language:
            return _question_text_md_column(base_language), _question_text_ishtml_column(
                base_language
            )
        raise ValueError(
            "Questions sheet is missing question text columns "
            "(expected at least one `text_<lang>` column)."
        )

    fallback_md = f"{prefix}_en_MD"
    fallback_html = f"{prefix}_en_IsHTML"
    for h in headers:
        if h.startswith(f"{prefix}_") and h.endswith("_MD"):
            suffix = h[len(prefix) + 1 : -len("_MD")]
            return h, f"{prefix}_{suffix}_IsHTML"
    return fallback_md, fallback_html


def _parse_question_config_json_strict(
    raw: object, *, qid: str, source: str
) -> dict[str, dict[str, object]]:
    text = str(raw or "").strip()
    if not text:
        return _question_config_dict(
            validation_settings={},
            randomization_settings={},
        )
    try:
        parsed = json.loads(text)
    except json.JSONDecodeError as exc:
        raise ValueError(
            f"Invalid {QUESTION_CONFIG_JSON_COLUMN} for {qid} ({source}): "
            f"{exc.msg} at line {exc.lineno}, column {exc.colno}."
        ) from exc
    if not isinstance(parsed, dict):
        raise ValueError(
            f"Invalid {QUESTION_CONFIG_JSON_COLUMN} for {qid} ({source}): expected a JSON object."
        )
    validation_raw = parsed.get("Validation")
    randomization_raw = parsed.get("Randomization")
    if validation_raw is None and randomization_raw is None:
        validation_raw = parsed
        randomization_raw = {}
    if validation_raw is None:
        validation_raw = {}
    if randomization_raw is None:
        randomization_raw = {}
    if not isinstance(validation_raw, dict):
        raise ValueError(
            f"Invalid {QUESTION_CONFIG_JSON_COLUMN} for {qid} ({source}): "
            "'Validation' must be a JSON object."
        )
    if not isinstance(randomization_raw, dict):
        raise ValueError(
            f"Invalid {QUESTION_CONFIG_JSON_COLUMN} for {qid} ({source}): "
            "'Randomization' must be a JSON object."
        )
    return _question_config_dict(
        validation_settings=validation_raw,
        randomization_settings=randomization_raw,
    )


def load_questions_from_workbook(
    xlsx_path: Path,
    *,
    base_language: str | None = None,
) -> Dict[str, QuestionRow]:
    """Read QuestionRow objects from an existing workbook.

    Parses the Questions sheet and returns a dictionary mapping each QID to its
    corresponding QuestionRow dataclass. Used by `preview_changes` and
    `apply_changes` to compare Excel wording against the cached survey JSON.

    Args:
        xlsx_path: Path to the workbook created by `qsync init`.
        base_language: Optional survey base language code (e.g., `EN`, `FR`).
            When provided, qsync resolves question text from that language column.

    Returns:
        Mapping of `QID -> QuestionRow`. Each QuestionRow contains:
        - `survey_id`: The Qualtrics survey ID.
        - `qid`: The question ID.
        - `block_name`: The Qualtrics block name.
        - `question_type`: Question type (MC, TE, Matrix, etc.).
        - `data_export_tag`: The DataExportTag / variable name.
        - `required_response`: Derived required-response marker.
        - `question_config_json`: Canonical question config JSON payload.
        - `force_response_mode`: Validation force mode (`OFF`, `ON`, `RequestResponse`).
        - `validation_type`: Validation type (`None`, `MinChoices`, etc.).
        - `validation_settings_json`: JSON payload for additional validation settings.
        - `randomization_type`: Randomization mode (`None`, `All`, `Subset`, etc.).
        - `randomization_settings_json`: JSON payload for additional randomization settings.
        - `text_en_md`: Base-language wording in Markdown or raw HTML.
        - `text_en_is_html`: True if `text_en_md` is raw HTML.

    Raises:
        FileNotFoundError: If the workbook does not exist.
        KeyError: If the Questions sheet is missing.

    Example:
        >>> from pathlib import Path
        >>> from qsync.excel_io import load_questions_from_workbook
        >>> questions = load_questions_from_workbook(Path("excel/SV_xxx.xlsx"))
        >>> for qid, row in questions.items():
        ...     print(f"{qid}: {row.data_export_tag}")
    """

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[QUESTION_SHEET]
    headers, data_rows = _iter_sheet_rows(ws)
    idx = {name: i for i, name in enumerate(headers)}
    has_legacy_config_columns = any(name in idx for name in LEGACY_QUESTION_CONFIG_COLUMNS)

    try:
        text_md_col, text_html_col = _find_base_text_col(
            headers,
            "Text",
            base_language=base_language,
        )
    except Exception as exc:
        raise ValueError(
            "Unable to resolve Questions text columns from workbook. "
            "Run `qsync items pull --survey-id ...` to refresh workbook columns "
            f"(workbook: {xlsx_path.name})."
        ) from exc
    if text_md_col not in idx:
        raise ValueError(
            "Questions sheet is missing the base text column "
            f"`{text_md_col}`. Run `qsync items pull --survey-id ...` to refresh "
            f"workbook columns (workbook: {xlsx_path.name})."
        )

    def _get(row, name, default=None):
        col = idx.get(name)
        if col is None or col >= len(row):
            return default
        cell = row[col]
        return cell.value if cell is not None else default

    result: Dict[str, QuestionRow] = {}
    for row in data_rows:
        qid_val = _get(row, "QID")
        if qid_val is None:
            continue
        qid = str(qid_val).strip()
        if not qid:
            continue
        config_raw = None
        if has_legacy_config_columns:
            config_raw = _build_question_config_json_from_legacy_values(
                force_response_mode=_get(row, "ForceResponseMode"),
                validation_type=_get(row, "ValidationType"),
                validation_settings_json=_get(row, "ValidationSettingsJSON"),
                randomization_type=_get(row, "RandomizationType"),
                randomization_settings_json=_get(row, "RandomizationSettingsJSON"),
            )
        else:
            config_raw = _get(row, QUESTION_CONFIG_JSON_COLUMN)
            if config_raw is None or str(config_raw).strip() == "":
                config_raw = _build_question_config_json_from_legacy_values(
                    force_response_mode="OFF",
                    validation_type="None",
                    validation_settings_json="",
                    randomization_type="None",
                    randomization_settings_json="",
                )
        config = _parse_question_config_json_strict(config_raw, qid=qid, source="workbook")
        validation = config.get("Validation", {})
        randomization = config.get("Randomization", {})
        force_mode = _normalize_force_response_mode(validation.get("ForceResponse"))
        validation_type = _normalize_validation_type(validation.get("Type"))
        validation_extras = _validation_settings_extra_dict(validation)
        if force_mode == "OFF":
            validation_extras.pop("ForceResponseType", None)
        randomization_type = _normalize_randomization_type(randomization.get("Type"))
        randomization_extras = _randomization_settings_extra_dict(randomization)
        config_text = json.dumps(
            config, ensure_ascii=True, sort_keys=True, separators=(",", ":")
        )
        qr = QuestionRow(
            survey_id=str(_get(row, "SurveyID") or "").strip(),
            qid=qid,
            block_name=str(_get(row, "BlockName") or "").strip(),
            question_type=str(_get(row, "QuestionType") or "").strip(),
            data_export_tag=str(_get(row, "DataExportTag") or "").strip(),
            required_response=_coerce_bool_cell(
                _get(row, "RequiredResponse"),
                default=_is_required_response(force_mode),
            ),
            question_config_json=config_text,
            force_response_mode=force_mode,
            validation_type=validation_type,
            validation_settings_json=_dump_validation_settings_json(validation_extras)
            or None,
            randomization_type=randomization_type,
            randomization_settings_json=(
                _dump_randomization_settings_json(randomization_extras) or None
            ),
            text_en_md=str(_get(row, text_md_col) or ""),
            text_en_is_html=_coerce_bool_cell(_get(row, text_html_col)),
            # Question wording is always editable via Excel; only options/subitems
            # are treated as externally managed via EXTERNALLY_MANAGED_TAGS.
            externally_managed_by=None,
        )
        result[qid] = qr
    return result


def load_options_from_workbook(xlsx_path: Path) -> Dict[Tuple[str, str], OptionRow]:
    """Read OptionRow objects from an existing workbook.

    Parses the Options sheet and returns a dictionary mapping each (QID, ChoiceId)
    tuple to its corresponding OptionRow dataclass. For Matrix questions, options
    represent the response scale (Answers); for MC/SC questions, options represent
    the choices.

    Args:
        xlsx_path: Path to the workbook created by `qsync init`.

    Returns:
        Mapping of `(QID, ChoiceId) -> OptionRow`. Each OptionRow contains:
        - `survey_id`: The Qualtrics survey ID.
        - `qid`: The question ID.
        - `choice_id`: The choice/answer ID within the question.
        - `question_type`: Question type (MC, Matrix, etc.).
        - `code`: The recode value (if set).
        - `label_en_md`: Base-language label in Markdown or raw HTML.
        - `label_en_is_html`: True if `label_en_md` is raw HTML.
        - `externally_managed_by`: Script path if managed externally (e.g., recognition).

    Raises:
        FileNotFoundError: If the workbook does not exist.

    Example:
        >>> from pathlib import Path
        >>> from qsync.excel_io import load_options_from_workbook
        >>> options = load_options_from_workbook(Path("excel/SV_xxx.xlsx"))
        >>> for (qid, choice_id), row in options.items():
        ...     print(f"{qid}/{choice_id}: {row.label_en_md[:30]}...")
    """

    wb = load_workbook(xlsx_path, data_only=True)
    if OPTIONS_SHEET not in wb.sheetnames:
        return {}
    ws = wb[OPTIONS_SHEET]
    headers, data_rows = _iter_sheet_rows(ws)
    idx = {name: i for i, name in enumerate(headers)}

    label_md_col, label_html_col = _find_base_text_col(headers, "Label")

    def _get(row, name, default=None):
        col = idx.get(name)
        if col is None or col >= len(row):
            return default
        cell = row[col]
        return cell.value if cell is not None else default

    result: Dict[Tuple[str, str], OptionRow] = {}
    for row in data_rows:
        qid_val = _get(row, "QID")
        choice_val = _get(row, "ChoiceId")
        if qid_val is None or choice_val is None:
            continue
        qid = str(qid_val).strip()
        choice_id = str(choice_val).strip()
        if not qid or not choice_id:
            continue
        export_tag = str(_get(row, "ExportTag") or "").strip()
        result[(qid, choice_id)] = OptionRow(
            survey_id=str(_get(row, "SurveyID") or "").strip(),
            qid=qid,
            choice_id=choice_id,
            question_type=str(_get(row, "QuestionType") or "").strip(),
            export_tag=export_tag,
            code=str(_get(row, "Code") or "").strip() or None,
            label_en_md=str(_get(row, label_md_col) or ""),
            label_en_is_html=bool(_get(row, label_html_col) or False),
            externally_managed_by=_is_externally_managed_question(export_tag),
        )
    return result


def load_subitems_from_workbook(xlsx_path: Path) -> Dict[Tuple[str, str], SubitemRow]:
    """Read SubitemRow objects from an existing workbook.

    Parses the Subitems sheet and returns a dictionary mapping each (QID, AnswerId)
    tuple to its corresponding SubitemRow dataclass. For Matrix questions, subitems
    represent the row statements (Choices); for other questions, they represent
    the Answers structure.

    Args:
        xlsx_path: Path to the workbook created by `qsync init`.

    Returns:
        Mapping of `(QID, AnswerId) -> SubitemRow`. Each SubitemRow contains:
        - `survey_id`: The Qualtrics survey ID.
        - `qid`: The question ID.
        - `answer_id`: The subitem/statement ID within the question.
        - `question_type`: Question type (Matrix, Slider, etc.).
        - `label_en_md`: Base-language label in Markdown or raw HTML.
        - `label_en_is_html`: True if `label_en_md` is raw HTML.
        - `field`: Field disambiguator (Answer | Label). Label rows are ignored here.

    Raises:
        FileNotFoundError: If the workbook does not exist.

    Example:
        >>> from pathlib import Path
        >>> from qsync.excel_io import load_subitems_from_workbook
        >>> subitems = load_subitems_from_workbook(Path("excel/SV_xxx.xlsx"))
        >>> for (qid, answer_id), row in subitems.items():
        ...     print(f"{qid}/{answer_id}: {row.label_en_md[:30]}...")
    """

    wb = load_workbook(xlsx_path, data_only=True)
    if SUBITEMS_SHEET not in wb.sheetnames:
        return {}
    ws = wb[SUBITEMS_SHEET]
    headers, data_rows = _iter_sheet_rows(ws)
    idx = {name: i for i, name in enumerate(headers)}

    label_md_col, label_html_col = _find_base_text_col(headers, "Label")

    def _get(row, name, default=None):
        col = idx.get(name)
        if col is None or col >= len(row):
            return default
        cell = row[col]
        return cell.value if cell is not None else default

    result: Dict[Tuple[str, str], SubitemRow] = {}
    for row in data_rows:
        qid_val = _get(row, "QID")
        answer_val = _get(row, "AnswerId")
        if qid_val is None or answer_val is None:
            continue
        field_val = _get(row, "Field")
        field = _normalize_subitem_field(field_val)
        if field == "Label":
            continue
        qid = str(qid_val).strip()
        answer_id = str(answer_val).strip()
        if not qid or not answer_id:
            continue
        result[(qid, answer_id)] = SubitemRow(
            survey_id=str(_get(row, "SurveyID") or "").strip(),
            qid=qid,
            answer_id=answer_id,
            field=field,
            question_type=str(_get(row, "QuestionType") or "").strip(),
            export_tag=str(_get(row, "ExportTag") or "").strip(),
            label_en_md=str(_get(row, label_md_col) or ""),
            label_en_is_html=bool(_get(row, label_html_col) or False),
        )
    return result


def load_sbs_columns_from_workbook(
    xlsx_path: Path,
) -> Dict[Tuple[str, str], SbsColumnRow]:
    """Read SbsColumnRow objects from an existing workbook."""

    wb = load_workbook(xlsx_path, data_only=True)
    if SBS_COLUMNS_SHEET not in wb.sheetnames:
        return {}
    ws = wb[SBS_COLUMNS_SHEET]
    headers, data_rows = _iter_sheet_rows(ws)
    idx = {name: i for i, name in enumerate(headers)}

    label_md_col, label_html_col = _find_base_text_col(headers, "Label")

    def _get(row, name, default=None):
        col = idx.get(name)
        if col is None or col >= len(row):
            return default
        cell = row[col]
        return cell.value if cell is not None else default

    result: Dict[Tuple[str, str], SbsColumnRow] = {}
    for row in data_rows:
        qid_val = _get(row, "QID")
        col_val = _get(row, "ColumnId")
        if qid_val is None or col_val is None:
            continue
        qid = str(qid_val).strip()
        column_id = str(col_val).strip()
        if not qid or not column_id:
            continue
        export_tag = str(_get(row, "ExportTag") or "").strip()
        result[(qid, column_id)] = SbsColumnRow(
            survey_id=str(_get(row, "SurveyID") or "").strip(),
            qid=qid,
            column_id=column_id,
            question_type=str(_get(row, "QuestionType") or "").strip(),
            export_tag=export_tag,
            label_en_md=str(_get(row, label_md_col) or ""),
            label_en_is_html=bool(_get(row, label_html_col) or False),
            externally_managed_by=_is_externally_managed_question(export_tag),
        )
    return result


def load_sbs_column_answers_from_workbook(
    xlsx_path: Path,
) -> Dict[Tuple[str, str, str], SbsColumnAnswerRow]:
    """Read SbsColumnAnswerRow objects from an existing workbook."""

    wb = load_workbook(xlsx_path, data_only=True)
    if SBS_COLUMN_ANSWERS_SHEET not in wb.sheetnames:
        return {}
    ws = wb[SBS_COLUMN_ANSWERS_SHEET]
    headers, data_rows = _iter_sheet_rows(ws)
    idx = {name: i for i, name in enumerate(headers)}

    label_md_col, label_html_col = _find_base_text_col(headers, "Label")

    def _get(row, name, default=None):
        col = idx.get(name)
        if col is None or col >= len(row):
            return default
        cell = row[col]
        return cell.value if cell is not None else default

    result: Dict[Tuple[str, str, str], SbsColumnAnswerRow] = {}
    for row in data_rows:
        qid_val = _get(row, "QID")
        col_val = _get(row, "ColumnId")
        ans_val = _get(row, "AnswerId")
        if qid_val is None or col_val is None or ans_val is None:
            continue
        qid = str(qid_val).strip()
        column_id = str(col_val).strip()
        answer_id = str(ans_val).strip()
        if not qid or not column_id or not answer_id:
            continue
        export_tag = str(_get(row, "ExportTag") or "").strip()
        result[(qid, column_id, answer_id)] = SbsColumnAnswerRow(
            survey_id=str(_get(row, "SurveyID") or "").strip(),
            qid=qid,
            column_id=column_id,
            answer_id=answer_id,
            question_type=str(_get(row, "QuestionType") or "").strip(),
            export_tag=export_tag,
            label_en_md=str(_get(row, label_md_col) or ""),
            label_en_is_html=bool(_get(row, label_html_col) or False),
            externally_managed_by=_is_externally_managed_question(export_tag),
        )
    return result


def load_embedded_data_from_workbook(xlsx_path: Path) -> List[EmbeddedDataRow]:
    """Read EmbeddedDataRow objects from an existing workbook."""

    wb = load_workbook(xlsx_path, data_only=True)
    if EMBEDDED_DATA_SHEET not in wb.sheetnames:
        return []
    ws = wb[EMBEDDED_DATA_SHEET]
    headers, data_rows = _iter_sheet_rows(ws)
    idx = {name: i for i, name in enumerate(headers)}

    def _get(row, name, default=None):
        col = idx.get(name)
        if col is None or col >= len(row):
            return default
        cell = row[col]
        return cell.value if cell is not None else default

    result: List[EmbeddedDataRow] = []
    for row in data_rows:
        field_val = _get(row, "Field")
        if field_val is None:
            continue
        field = str(field_val).strip()
        if not field:
            continue
        flow_val = _get(row, "FlowID")
        flow_id = str(flow_val).strip() if flow_val is not None else ""
        order_val = _get(row, "FlowOrder")
        try:
            flow_order = int(order_val)
        except (TypeError, ValueError):
            flow_order = 0
        value_val = _get(row, "Value")
        value = str(value_val) if value_val is not None else None
        ed_type = str(_get(row, "Type") or "").strip()
        written_by = str(_get(row, "WrittenByQIDs") or "").strip() or None
        result.append(
            EmbeddedDataRow(
                survey_id=str(_get(row, "SurveyID") or "").strip(),
                flow_id=flow_id or None,
                flow_order=flow_order,
                field=field,
                value=value,
                ed_type=ed_type,
                written_by_qids=written_by,
            )
        )
    return result


def question_row_to_html(q: QuestionRow) -> str:
    """Convert a QuestionRow's EN text to HTML according to the IsHTML flag."""

    text = q.text_en_md or ""
    if q.text_en_is_html:
        return normalize_text(text)
    return md_to_html(text)


def option_row_to_html(o: OptionRow) -> str:
    """Convert an OptionRow's EN label to HTML according to the IsHTML flag."""

    text = o.label_en_md or ""
    if o.label_en_is_html:
        return normalize_text(text)
    return md_to_html(text)


def subitem_row_to_html(s: SubitemRow) -> str:
    """Convert a SubitemRow's EN label to HTML according to the IsHTML flag."""

    text = s.label_en_md or ""
    if s.label_en_is_html:
        return normalize_text(text)
    return md_to_html(text)


def sbs_column_row_to_html(c: SbsColumnRow) -> str:
    """Convert an SbsColumnRow's EN label to HTML according to the IsHTML flag."""

    text = c.label_en_md or ""
    if c.label_en_is_html:
        return normalize_text(text)
    return md_to_html(text)


def sbs_column_answer_row_to_html(a: SbsColumnAnswerRow) -> str:
    """Convert an SbsColumnAnswerRow's EN label to HTML according to the IsHTML flag."""

    text = a.label_en_md or ""
    if a.label_en_is_html:
        return normalize_text(text)
    return md_to_html(text)
