"""Translation workbook helpers (Excel <-> translations JSON)."""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Dict, Iterable, List, Mapping, Sequence

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from . import excel_io
from .errors import QsyncValidationError
from .excel_io import (
    OPTIONS_SHEET,
    QUESTION_SHEET,
    SUBITEMS_SHEET,
    _iter_sheet_rows,
    _language_from_suffix,
    _language_suffix,
    _ordered_languages,
)
from .markdown_codec import (
    html_to_md,
    is_markdown_safe_html,
    md_to_html,
    normalize_text,
    should_treat_as_html,
)
from .translations import (
    TranslationDoctorReport,
    _check_large_deltas,
    _check_placeholders,
    _coverage_stats,
    _coverage_stats_with_allowed_empties,
    _normalize_language_list,
    load_local_map,
    normalize_translation_map,
    translation_map_path,
)

_TRANSLATION_KEY_RE = re.compile(
    r"^(?P<qid>QID[^_]+)_(?P<field>QuestionText|Choice|Answer)(?P<id>[0-9]+)?$"
)


def _format_cell_ref(sheet: str, row: int, col: int) -> str:
    return f"{sheet}!{get_column_letter(col)}{row}"


def _question_text_md_header(language: str) -> str:
    return excel_io._question_text_md_column(language)


def _question_text_html_header(language: str) -> str:
    return excel_io._question_text_ishtml_column(language)


def _resolve_question_headers_for_language(
    headers: Sequence[str],
    language: str,
) -> tuple[str | None, str | None]:
    suffix = _language_suffix(language)
    if not suffix:
        return None, None

    md_col = _question_text_md_header(language)
    html_col = _question_text_html_header(language)
    header_set = {str(h or "") for h in headers}
    if md_col in header_set:
        return md_col, html_col if html_col in header_set else None

    legacy_md = f"Text_{suffix}_MD"
    legacy_html = f"Text_{suffix}_IsHTML"
    if legacy_md in header_set:
        return legacy_md, legacy_html if legacy_html in header_set else None
    return None, None


def _build_header_index(ws) -> dict[str, int]:
    headers, _ = _iter_sheet_rows(ws)
    return {str(h or ""): idx + 1 for idx, h in enumerate(headers) if h}


def _build_row_index(ws, key_col: str, key2_col: str | None = None) -> dict:
    headers, data_rows = _iter_sheet_rows(ws)
    if not headers or key_col not in headers:
        return {}
    key_idx = headers.index(key_col)
    key2_idx = headers.index(key2_col) if key2_col and key2_col in headers else None
    out: dict = {}
    for row_idx, row in enumerate(data_rows, start=2):
        key = str(row[key_idx].value or "").strip()
        if not key:
            continue
        if key2_idx is None:
            out[key] = row_idx
            continue
        key2 = str(row[key2_idx].value or "").strip()
        if not key2:
            continue
        out[(key, key2)] = row_idx
    return out


def _locate_translation_key_cells(
    wb, *, translation_key: str, language: str
) -> list[str]:
    """Best-effort mapping from Qualtrics translation key -> workbook cell ref(s)."""

    match = _TRANSLATION_KEY_RE.match(str(translation_key or "").strip())
    if not match:
        return []
    qid = match.group("qid")
    field = match.group("field")
    item_id = match.group("id") or ""
    suffix = _language_suffix(language)
    if not suffix:
        return []

    refs: list[str] = []

    if field == "QuestionText":
        if QUESTION_SHEET in wb.sheetnames:
            ws = wb[QUESTION_SHEET]
            row_index = _build_row_index(ws, "QID")
            header_index = _build_header_index(ws)
            row = row_index.get(qid)
            col = header_index.get(_question_text_md_header(language))
            if not col:
                col = header_index.get(f"Text_{suffix}_MD")
            if row and col:
                refs.append(_format_cell_ref(QUESTION_SHEET, int(row), int(col)))
        return refs

    # Choice/Answer keys can refer to either Options or Subitems depending on question type.
    label_col = f"Label_{suffix}_MD"
    for sheet, key_col in ((OPTIONS_SHEET, "ChoiceId"), (SUBITEMS_SHEET, "AnswerId")):
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        header_index = _build_header_index(ws)
        if (
            "QID" not in header_index
            or key_col not in header_index
            or label_col not in header_index
        ):
            continue
        row_index = _build_row_index(ws, "QID", key_col)
        row = row_index.get((qid, item_id))
        col = header_index.get(label_col)
        if row and col:
            refs.append(_format_cell_ref(sheet, int(row), int(col)))

    return refs


def _summarize_empty_keys(
    wb,
    *,
    language: str,
    local_map: Mapping[str, str],
    allowed_empty_keys: set[str] | None = None,
    limit: int = 12,
) -> str:
    empties = [
        k
        for k, v in local_map.items()
        if not str(v or "").strip()
        and (not allowed_empty_keys or str(k) not in allowed_empty_keys)
    ]
    if not empties:
        return ""
    sample: list[str] = []
    for key in empties[:limit]:
        refs = _locate_translation_key_cells(
            wb, translation_key=str(key), language=language
        )
        if refs:
            sample.append(f"{key} ({refs[0]})")
        else:
            sample.append(str(key))
    suffix = ""
    if len(empties) > limit:
        suffix = f" … (+{len(empties) - limit} more)"
    return "Empty keys: " + ", ".join(sample) + suffix


def _languages_from_headers(headers: Iterable[str], prefix: str) -> List[str]:
    languages: List[str] = []
    seen: set[str] = set()
    for name in headers:
        header = str(name or "")
        if not header.startswith(f"{prefix}_") or not header.endswith("_MD"):
            continue
        suffix = header[len(prefix) + 1 : -len("_MD")]
        code = _language_from_suffix(suffix)
        if not code or code in seen:
            continue
        seen.add(code)
        languages.append(code)
    return languages


def _resolve_languages_from_workbook(wb) -> List[str]:
    languages: List[str] = []
    if QUESTION_SHEET in wb.sheetnames:
        headers, _ = _iter_sheet_rows(wb[QUESTION_SHEET])
        for lang in excel_io._question_text_lang_columns_from_headers(headers).keys():
            languages.append(lang)
    for sheet_name, prefix in ((OPTIONS_SHEET, "Label"), (SUBITEMS_SHEET, "Label")):
        if sheet_name not in wb.sheetnames:
            continue
        headers, _ = _iter_sheet_rows(wb[sheet_name])
        for lang in _languages_from_headers(headers, prefix):
            languages.append(lang)
    return _ordered_languages(languages)


def _translation_key_for_question(qid: str) -> str:
    return f"{qid}_QuestionText"


def _translation_key_for_option(qid: str, question_type: str, choice_id: str) -> str:
    if question_type == "Matrix":
        return f"{qid}_Answer{choice_id}"
    return f"{qid}_Choice{choice_id}"


def _translation_key_for_subitem(qid: str, question_type: str, answer_id: str) -> str:
    if question_type == "Matrix":
        return f"{qid}_Choice{answer_id}"
    return f"{qid}_Answer{answer_id}"


def _value_to_cell(value: str) -> tuple[str, bool]:
    if is_markdown_safe_html(value):
        return html_to_md(value), False
    if should_treat_as_html(value):
        return normalize_text(value), True
    return html_to_md(value), False


def _cell_to_value(text: str, is_html: bool) -> str:
    if is_html:
        return normalize_text(text)
    return md_to_html(text)


def _write_json(path: Path, payload: Mapping[str, str]) -> None:
    serialized = json.dumps(payload, indent=2, ensure_ascii=True, sort_keys=True) + "\n"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(serialized, encoding="utf-8")


def populate_workbook_from_translation_maps(
    survey_id: str,
    xlsx_path: Path,
    *,
    languages: Sequence[str] | None = None,
    overwrite: bool = False,
) -> List[str]:
    wb = load_workbook(xlsx_path)
    if languages is None:
        languages = _resolve_languages_from_workbook(wb)
    languages = _ordered_languages(languages)

    translation_maps: Dict[str, Dict[str, str]] = {}
    for lang in languages:
        local_map = load_local_map(survey_id, lang)
        if local_map is None:
            raise QsyncValidationError(
                error_id="QSYNC-TRANSLATIONS-WORKBOOK-001",
                problem=f"Missing translation map for {lang}.",
                why="Expected translations JSON on disk.",
                impact="Workbook could not be updated.",
                action="Translation maps are deprecated; use the workbook-based `qsync translations` workflow instead.",
                context={"survey_id": survey_id, "language": lang},
            )
        translation_maps[lang] = {str(k): str(v or "") for k, v in local_map.items()}

    if QUESTION_SHEET in wb.sheetnames:
        ws = wb[QUESTION_SHEET]
        headers, data_rows = _iter_sheet_rows(ws)
        idx = {name: i for i, name in enumerate(headers)}
        for row in data_rows:
            qid_val = row[idx.get("QID", -1)].value if "QID" in idx else None
            qtype_val = (
                row[idx.get("QuestionType", -1)].value if "QuestionType" in idx else ""
            )
            qid = str(qid_val or "").strip()
            if not qid:
                continue
            for lang in languages:
                md_col, html_col = _resolve_question_headers_for_language(headers, lang)
                if not md_col or md_col not in idx:
                    continue
                cell = row[idx[md_col]]
                if cell.value not in (None, "") and not overwrite:
                    continue
                key = _translation_key_for_question(qid)
                value = translation_maps.get(lang, {}).get(key)
                if value is None:
                    continue
                text, is_html = _value_to_cell(value)
                cell.value = text
                if html_col in idx:
                    row[idx[html_col]].value = bool(is_html)

    if OPTIONS_SHEET in wb.sheetnames:
        ws = wb[OPTIONS_SHEET]
        headers, data_rows = _iter_sheet_rows(ws)
        idx = {name: i for i, name in enumerate(headers)}
        for row in data_rows:
            qid_val = row[idx.get("QID", -1)].value if "QID" in idx else None
            choice_val = (
                row[idx.get("ChoiceId", -1)].value if "ChoiceId" in idx else None
            )
            qtype_val = (
                row[idx.get("QuestionType", -1)].value if "QuestionType" in idx else ""
            )
            qid = str(qid_val or "").strip()
            choice_id = str(choice_val or "").strip()
            if not qid or not choice_id:
                continue
            qtype = str(qtype_val or "").strip()
            for lang in languages:
                suffix = _language_suffix(lang)
                md_col = f"Label_{suffix}_MD"
                html_col = f"Label_{suffix}_IsHTML"
                if md_col not in idx:
                    continue
                cell = row[idx[md_col]]
                if cell.value not in (None, "") and not overwrite:
                    continue
                key = _translation_key_for_option(qid, qtype, choice_id)
                value = translation_maps.get(lang, {}).get(key)
                if value is None:
                    continue
                text, is_html = _value_to_cell(value)
                cell.value = text
                if html_col in idx:
                    row[idx[html_col]].value = bool(is_html)

    if SUBITEMS_SHEET in wb.sheetnames:
        ws = wb[SUBITEMS_SHEET]
        headers, data_rows = _iter_sheet_rows(ws)
        idx = {name: i for i, name in enumerate(headers)}
        for row in data_rows:
            qid_val = row[idx.get("QID", -1)].value if "QID" in idx else None
            answer_val = (
                row[idx.get("AnswerId", -1)].value if "AnswerId" in idx else None
            )
            qtype_val = (
                row[idx.get("QuestionType", -1)].value if "QuestionType" in idx else ""
            )
            qid = str(qid_val or "").strip()
            answer_id = str(answer_val or "").strip()
            if not qid or not answer_id:
                continue
            qtype = str(qtype_val or "").strip()
            for lang in languages:
                suffix = _language_suffix(lang)
                md_col = f"Label_{suffix}_MD"
                html_col = f"Label_{suffix}_IsHTML"
                if md_col not in idx:
                    continue
                cell = row[idx[md_col]]
                if cell.value not in (None, "") and not overwrite:
                    continue
                key = _translation_key_for_subitem(qid, qtype, answer_id)
                value = translation_maps.get(lang, {}).get(key)
                if value is None:
                    continue
                text, is_html = _value_to_cell(value)
                cell.value = text
                if html_col in idx:
                    row[idx[html_col]].value = bool(is_html)

    wb.save(xlsx_path)
    return list(languages)


def extract_translation_maps_from_workbook(
    survey_id: str,
    xlsx_path: Path,
    *,
    languages: Sequence[str] | None = None,
    allow_empty: bool = False,
) -> List[Path]:
    wb = load_workbook(xlsx_path, data_only=True)
    if languages is None:
        languages = _resolve_languages_from_workbook(wb)
    languages = _ordered_languages(languages)

    updates: Dict[str, Dict[str, str]] = {}
    for lang in languages:
        existing = load_local_map(survey_id, lang) or {}
        updates[lang] = {str(k): str(v or "") for k, v in existing.items()}

    if QUESTION_SHEET in wb.sheetnames:
        ws = wb[QUESTION_SHEET]
        headers, data_rows = _iter_sheet_rows(ws)
        idx = {name: i for i, name in enumerate(headers)}
        for row in data_rows:
            qid_val = row[idx.get("QID", -1)].value if "QID" in idx else None
            qid = str(qid_val or "").strip()
            if not qid:
                continue
            for lang in languages:
                md_col, html_col = _resolve_question_headers_for_language(headers, lang)
                if not md_col or md_col not in idx:
                    continue
                cell = row[idx[md_col]]
                raw = cell.value
                if raw is None or str(raw).strip() == "":
                    if not allow_empty:
                        continue
                    text = ""
                else:
                    text = str(raw)
                is_html = False
                if html_col in idx:
                    is_html = bool(row[idx[html_col]].value or False)
                value = _cell_to_value(text, is_html)
                key = _translation_key_for_question(qid)
                updates[lang][key] = value

    if OPTIONS_SHEET in wb.sheetnames:
        ws = wb[OPTIONS_SHEET]
        headers, data_rows = _iter_sheet_rows(ws)
        idx = {name: i for i, name in enumerate(headers)}
        for row in data_rows:
            qid_val = row[idx.get("QID", -1)].value if "QID" in idx else None
            choice_val = (
                row[idx.get("ChoiceId", -1)].value if "ChoiceId" in idx else None
            )
            qtype_val = (
                row[idx.get("QuestionType", -1)].value if "QuestionType" in idx else ""
            )
            qid = str(qid_val or "").strip()
            choice_id = str(choice_val or "").strip()
            if not qid or not choice_id:
                continue
            qtype = str(qtype_val or "").strip()
            for lang in languages:
                suffix = _language_suffix(lang)
                md_col = f"Label_{suffix}_MD"
                html_col = f"Label_{suffix}_IsHTML"
                if md_col not in idx:
                    continue
                cell = row[idx[md_col]]
                raw = cell.value
                if raw is None or str(raw).strip() == "":
                    if not allow_empty:
                        continue
                    text = ""
                else:
                    text = str(raw)
                is_html = False
                if html_col in idx:
                    is_html = bool(row[idx[html_col]].value or False)
                value = _cell_to_value(text, is_html)
                key = _translation_key_for_option(qid, qtype, choice_id)
                updates[lang][key] = value

    if SUBITEMS_SHEET in wb.sheetnames:
        ws = wb[SUBITEMS_SHEET]
        headers, data_rows = _iter_sheet_rows(ws)
        idx = {name: i for i, name in enumerate(headers)}
        for row in data_rows:
            qid_val = row[idx.get("QID", -1)].value if "QID" in idx else None
            answer_val = (
                row[idx.get("AnswerId", -1)].value if "AnswerId" in idx else None
            )
            qtype_val = (
                row[idx.get("QuestionType", -1)].value if "QuestionType" in idx else ""
            )
            qid = str(qid_val or "").strip()
            answer_id = str(answer_val or "").strip()
            if not qid or not answer_id:
                continue
            qtype = str(qtype_val or "").strip()
            for lang in languages:
                suffix = _language_suffix(lang)
                md_col = f"Label_{suffix}_MD"
                html_col = f"Label_{suffix}_IsHTML"
                if md_col not in idx:
                    continue
                cell = row[idx[md_col]]
                raw = cell.value
                if raw is None or str(raw).strip() == "":
                    if not allow_empty:
                        continue
                    text = ""
                else:
                    text = str(raw)
                is_html = False
                if html_col in idx:
                    is_html = bool(row[idx[html_col]].value or False)
                value = _cell_to_value(text, is_html)
                key = _translation_key_for_subitem(qid, qtype, answer_id)
                updates[lang][key] = value

    paths: List[Path] = []
    for lang, payload in updates.items():
        path = translation_map_path(survey_id, lang)
        _write_json(path, payload)
        paths.append(path)

    return paths


def read_translation_maps_from_workbook(
    survey_id: str,
    xlsx_path: Path,
    *,
    languages: Sequence[str] | None = None,
    allow_empty: bool = False,
) -> Dict[str, Dict[str, str]]:
    wb = load_workbook(xlsx_path, data_only=True)
    if languages is None:
        languages = _resolve_languages_from_workbook(wb)
    languages = _ordered_languages(languages)

    updates: Dict[str, Dict[str, str]] = {}
    for lang in languages:
        existing = load_local_map(survey_id, lang) or {}
        updates[lang] = {str(k): str(v or "") for k, v in existing.items()}

    if QUESTION_SHEET in wb.sheetnames:
        ws = wb[QUESTION_SHEET]
        headers, data_rows = _iter_sheet_rows(ws)
        idx = {name: i for i, name in enumerate(headers)}
        for row in data_rows:
            qid_val = row[idx.get("QID", -1)].value if "QID" in idx else None
            qtype_val = (
                row[idx.get("QuestionType", -1)].value if "QuestionType" in idx else ""
            )
            qid = str(qid_val or "").strip()
            if not qid:
                continue
            for lang in languages:
                md_col, html_col = _resolve_question_headers_for_language(headers, lang)
                if not md_col or md_col not in idx:
                    continue
                cell = row[idx[md_col]]
                raw = cell.value
                if raw is None or str(raw).strip() == "":
                    if not allow_empty:
                        continue
                    text = ""
                else:
                    text = str(raw)
                is_html = False
                if html_col in idx:
                    is_html = bool(row[idx[html_col]].value or False)
                key = _translation_key_for_question(qid)
                updates[lang][key] = _cell_to_value(text, is_html)

    if OPTIONS_SHEET in wb.sheetnames:
        ws = wb[OPTIONS_SHEET]
        headers, data_rows = _iter_sheet_rows(ws)
        idx = {name: i for i, name in enumerate(headers)}
        for row in data_rows:
            qid_val = row[idx.get("QID", -1)].value if "QID" in idx else None
            choice_val = (
                row[idx.get("ChoiceId", -1)].value if "ChoiceId" in idx else None
            )
            qtype_val = (
                row[idx.get("QuestionType", -1)].value if "QuestionType" in idx else ""
            )
            qid = str(qid_val or "").strip()
            choice_id = str(choice_val or "").strip()
            if not qid or not choice_id:
                continue
            qtype = str(qtype_val or "").strip()
            for lang in languages:
                suffix = _language_suffix(lang)
                md_col = f"Label_{suffix}_MD"
                html_col = f"Label_{suffix}_IsHTML"
                if md_col not in idx:
                    continue
                cell = row[idx[md_col]]
                raw = cell.value
                if raw is None or str(raw).strip() == "":
                    if not allow_empty:
                        continue
                    text = ""
                else:
                    text = str(raw)
                is_html = False
                if html_col in idx:
                    is_html = bool(row[idx[html_col]].value or False)
                key = _translation_key_for_option(qid, qtype, choice_id)
                updates[lang][key] = _cell_to_value(text, is_html)

    if SUBITEMS_SHEET in wb.sheetnames:
        ws = wb[SUBITEMS_SHEET]
        headers, data_rows = _iter_sheet_rows(ws)
        idx = {name: i for i, name in enumerate(headers)}
        for row in data_rows:
            qid_val = row[idx.get("QID", -1)].value if "QID" in idx else None
            answer_val = (
                row[idx.get("AnswerId", -1)].value if "AnswerId" in idx else None
            )
            qtype_val = (
                row[idx.get("QuestionType", -1)].value if "QuestionType" in idx else ""
            )
            qid = str(qid_val or "").strip()
            answer_id = str(answer_val or "").strip()
            if not qid or not answer_id:
                continue
            qtype = str(qtype_val or "").strip()
            for lang in languages:
                suffix = _language_suffix(lang)
                md_col = f"Label_{suffix}_MD"
                html_col = f"Label_{suffix}_IsHTML"
                if md_col not in idx:
                    continue
                cell = row[idx[md_col]]
                raw = cell.value
                if raw is None or str(raw).strip() == "":
                    if not allow_empty:
                        continue
                    text = ""
                else:
                    text = str(raw)
                is_html = False
                if html_col in idx:
                    is_html = bool(row[idx[html_col]].value or False)
                key = _translation_key_for_subitem(qid, qtype, answer_id)
                updates[lang][key] = _cell_to_value(text, is_html)

    return updates


def run_workbook_translation_doctor(
    survey_id: str,
    xlsx_path: Path,
    *,
    languages: Sequence[str] | None = None,
    base_language: str | None = None,
) -> TranslationDoctorReport:
    errors: list[str] = []
    warnings: list[str] = []
    coverage: dict[str, dict[str, int]] = {}

    wb_cells = load_workbook(xlsx_path, data_only=False)
    maps = read_translation_maps_from_workbook(
        survey_id, xlsx_path, languages=languages, allow_empty=True
    )

    base_lang = (
        _normalize_language_list([base_language or ""])[0] if base_language else ""
    )
    base_map = None
    allowed_empty_keys: set[str] = set()
    if base_lang:
        base_map = load_local_map(survey_id, base_lang)
        if base_map is None:
            warnings.append(
                f"[workbook:{base_lang}] Base language file missing; placeholder checks skipped."
            )
        else:
            allowed_empty_keys = {
                str(k)
                for k, v in base_map.items()
                if not isinstance(v, str) or not v.strip()
            }

    for lang, local_map in maps.items():
        try:
            normalize_translation_map(local_map, coerce_nulls=False)
        except QsyncValidationError as exc:
            errors.append(f"[workbook:{lang}] {exc.problem}")

        allowed = allowed_empty_keys if (base_map and lang != base_lang) else None
        coverage[lang] = (
            _coverage_stats(local_map)
            if not allowed
            else _coverage_stats_with_allowed_empties(
                local_map, allowed_empty_keys=allowed
            )
        )
        if base_lang and lang == base_lang:
            # Base language is not pushed via translations endpoint; keep doctor noise low.
            pass
        elif coverage[lang]["empty"] > 0:
            warnings.append(
                f"[workbook:{lang}] Coverage incomplete: {coverage[lang]['filled']}/{coverage[lang]['total']} filled."
            )
            summary = _summarize_empty_keys(
                wb_cells,
                language=lang,
                local_map=local_map,
                allowed_empty_keys=allowed,
            )
            if summary:
                warnings.append(f"[workbook:{lang}] {summary}")

        if base_map and lang != base_lang:
            ph_errors, ph_warnings = _check_placeholders(base_map, local_map, lang)
            errors.extend(f"[workbook]{msg}" for msg in ph_errors)
            warnings.extend(f"[workbook]{msg}" for msg in ph_warnings)

        if base_map and lang != base_lang:
            warnings.extend(
                f"[workbook]{msg}"
                for msg in _check_large_deltas(base_map, local_map, lang)
            )

    return TranslationDoctorReport(errors=errors, warnings=warnings, coverage=coverage)
